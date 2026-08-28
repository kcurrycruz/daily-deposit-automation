import ast
import subprocess
import sys
import unittest
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4


class MembershipPaymentTests(unittest.TestCase):
    def test_workflow_heading_html_escapes_visible_content(self):
        try:
            from app.ui_helpers import workflow_heading_html
        except ModuleNotFoundError:
            self.fail("workflow heading renderer is missing")

        rendered = workflow_heading_html(
            "Member & <Share>",
            'Use "QuickBooks" safely',
        )

        self.assertIn('class="hwfc-workflow-heading"', rendered)
        self.assertIn("Member &amp; &lt;Share&gt;", rendered)
        self.assertIn("Use &quot;QuickBooks&quot; safely", rendered)
        self.assertNotIn("Member & <Share>", rendered)

    def test_plan_guide_html_renders_each_plan_as_a_readable_card(self):
        try:
            from app.ui_helpers import plan_guide_html
        except ImportError:
            self.fail("plan guide card renderer is missing")
        from app.membership_payments import plan_reference_rows

        rendered = plan_guide_html(plan_reference_rows())

        self.assertEqual(rendered.count('class="hwfc-plan-card"'), 3)
        self.assertIn("1-Year Plan", rendered)
        self.assertIn("3-Year Plan", rendered)
        self.assertIn("5-Year Plan", rendered)
        for label, value in (
            ("Deposit", "$10.00"),
            ("Regular payment", "$8.45"),
            ("Principal", "$8.18"),
            ("Interest", "$0.27"),
            ("Number of payments", "11"),
            ("Total paid", "$102.95"),
        ):
            self.assertIn(label, rendered)
            self.assertIn(value, rendered)

    def test_deposit_download_details_requires_a_complete_iif_result(self):
        try:
            from app.ui_helpers import deposit_download_details
        except ImportError:
            self.fail("deposit download state helper is missing")

        self.assertIsNone(deposit_download_details(None))
        self.assertIsNone(deposit_download_details({"iif_path": Path("deposit.iif")}))
        self.assertEqual(
            deposit_download_details(
                {
                    "iif_path": Path("deposit_20260827.iif"),
                    "iif_bytes": b"IIF content",
                }
            ),
            {
                "file_name": "deposit_20260827.iif",
                "data": b"IIF content",
            },
        )

    def test_deposit_action_is_safe_before_files_are_uploaded(self):
        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))

        settlement_block = next(
            node
            for node in source_tree.body
            if isinstance(node, ast.If)
            and any(
                isinstance(child, ast.Name)
                and child.id == "run_clicked"
                and isinstance(child.ctx, ast.Store)
                for child in ast.walk(node)
            )
        )
        run_block = next(
            node
            for node in source_tree.body
            if isinstance(node, ast.If)
            and isinstance(node.test, ast.Name)
            and node.test.id == "run_clicked"
        )
        safe_defaults = [
            node
            for node in source_tree.body
            if isinstance(node, ast.Assign)
            and any(
                isinstance(target, ast.Name) and target.id == "run_clicked"
                for target in node.targets
            )
            and source_tree.body.index(node) < source_tree.body.index(settlement_block)
        ]

        namespace = {"settlement_file": None}
        exec(
            compile(
                ast.Module(
                    body=[*safe_defaults, settlement_block, run_block],
                    type_ignores=[],
                ),
                str(source_path),
                "exec",
            ),
            namespace,
        )

    def test_rejected_final_closeout_removes_generated_iif(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            normalize_closeout_payload,
            write_closeout_payload_file,
        )
        from app.membership_payments import write_membership_payments_file

        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        runtime_nodes = [
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef)
            and node.name in {"build_engine_command", "run_engine"}
        ]
        runtime_root = Path(__file__).parent / f"_rejected_closeout_runtime_{uuid4().hex}"
        input_dir = runtime_root / "input"
        iif_dir = runtime_root / "qb_imports"
        log_dir = runtime_root / "logs"
        temp_dir = runtime_root / "runtime"
        for folder in (input_dir, iif_dir, log_dir, temp_dir):
            folder.mkdir(parents=True, exist_ok=True)
        generated_iif = iif_dir / "deposit_20260828.iif"

        class UploadedWorkbook:
            name = "daily.xlsx"

            def getvalue(self):
                return b"workbook"

        class FakeSubprocess:
            def run(self, command, **_kwargs):
                generated_iif.write_bytes(b"rejected final iif")
                preview_path = Path(
                    command[command.index("--closeout-preview-output") + 1]
                )
                preview_path.write_text(
                    '{"remaining_after_approval": 1.0, "requires_approval": true}',
                    encoding="utf-8",
                )
                return SimpleNamespace(stdout="engine output", stderr="", returncode=0)

        namespace = {
            "Path": Path,
            "date": date,
            "json": __import__("json"),
            "sys": sys,
            "uuid4": uuid4,
            "subprocess": FakeSubprocess(),
            "ENGINE_PATH": source_path,
            "ROOT": runtime_root,
            "INPUT_DIR": input_dir,
            "QB_IMPORT_DIR": iif_dir,
            "LOG_DIR": log_dir,
            "RUNTIME_TEMP_DIR": temp_dir,
            "write_membership_payments_file": write_membership_payments_file,
            "normalize_closeout_payload": normalize_closeout_payload,
            "write_closeout_payload_file": write_closeout_payload_file,
            "parse_iif": lambda _path: (["parsed line"], "parsed dataframe"),
            "parse_validation": lambda _log_text, _lines: {"all_ok": True},
        }
        exec(
            compile(ast.Module(body=runtime_nodes, type_ignores=[]), str(source_path), "exec"),
            namespace,
        )
        closeout_payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {key: 0 for key in STANDARD_CLOSEOUT_ORDER},
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1,
            "approve_final_pos": False,
        }

        try:
            with self.assertRaisesRegex(ValueError, "approval is required"):
                namespace["run_engine"](
                    UploadedWorkbook(),
                    UploadedWorkbook(),
                    date(2026, 8, 28),
                    [],
                    "automatic",
                    "quickbooks",
                    None,
                    None,
                    None,
                    closeout_payload=closeout_payload,
                )
            self.assertFalse(generated_iif.exists())
        finally:
            for path in sorted(runtime_root.rglob("*"), reverse=True):
                if path.is_file():
                    path.unlink()
                else:
                    path.rmdir()
            runtime_root.rmdir()

    def test_failed_final_closeout_paths_remove_generated_iif(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            normalize_closeout_payload,
            write_closeout_payload_file,
        )
        from app.membership_payments import write_membership_payments_file

        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        runtime_nodes = [
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef)
            and node.name in {"build_engine_command", "run_engine"}
        ]
        closeout_payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {key: 0 for key in STANDARD_CLOSEOUT_ORDER},
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1,
            "approve_final_pos": False,
        }

        class UploadedWorkbook:
            name = "daily.xlsx"

            def getvalue(self):
                return b"workbook"

        cases = (
            ("subprocess_failure", 1, None, RuntimeError),
            ("missing_preview", 0, None, RuntimeError),
            ("malformed_preview", 0, "not JSON", RuntimeError),
        )
        for case_name, returncode, preview_text, error_type in cases:
            with self.subTest(case=case_name):
                runtime_root = (
                    Path(__file__).parent
                    / f"_failed_closeout_runtime_{uuid4().hex}"
                )
                input_dir = runtime_root / "input"
                iif_dir = runtime_root / "qb_imports"
                log_dir = runtime_root / "logs"
                temp_dir = runtime_root / "runtime"
                for folder in (input_dir, iif_dir, log_dir, temp_dir):
                    folder.mkdir(parents=True, exist_ok=True)
                generated_iif = iif_dir / "deposit_20260828.iif"

                class FakeSubprocess:
                    def run(self, command, **_kwargs):
                        generated_iif.write_bytes(b"failed final iif")
                        if preview_text is not None:
                            preview_path = Path(
                                command[command.index("--closeout-preview-output") + 1]
                            )
                            preview_path.write_text(preview_text, encoding="utf-8")
                        return SimpleNamespace(
                            stdout="engine output", stderr="", returncode=returncode
                        )

                namespace = {
                    "Path": Path,
                    "date": date,
                    "json": __import__("json"),
                    "sys": sys,
                    "uuid4": uuid4,
                    "subprocess": FakeSubprocess(),
                    "ENGINE_PATH": source_path,
                    "ROOT": runtime_root,
                    "INPUT_DIR": input_dir,
                    "QB_IMPORT_DIR": iif_dir,
                    "LOG_DIR": log_dir,
                    "RUNTIME_TEMP_DIR": temp_dir,
                    "write_membership_payments_file": write_membership_payments_file,
                    "normalize_closeout_payload": normalize_closeout_payload,
                    "write_closeout_payload_file": write_closeout_payload_file,
                    "parse_iif": lambda _path: (["parsed line"], "parsed dataframe"),
                    "parse_validation": lambda _log_text, _lines: {"all_ok": True},
                }
                exec(
                    compile(
                        ast.Module(body=runtime_nodes, type_ignores=[]),
                        str(source_path),
                        "exec",
                    ),
                    namespace,
                )

                try:
                    with self.assertRaises(error_type):
                        namespace["run_engine"](
                            UploadedWorkbook(),
                            UploadedWorkbook(),
                            date(2026, 8, 28),
                            [],
                            "automatic",
                            "quickbooks",
                            None,
                            None,
                            None,
                            closeout_payload=closeout_payload,
                        )
                    self.assertFalse(generated_iif.exists())
                finally:
                    for path in sorted(runtime_root.rglob("*"), reverse=True):
                        if path.is_file():
                            path.unlink()
                        else:
                            path.rmdir()
                    runtime_root.rmdir()

    def test_coupon_workflow_is_required_for_zero_bs_when_closeout_is_in_app(self):
        from app.closeout_reconciliation import coupon_workflow_is_required

        self.assertTrue(
            coupon_workflow_is_required(
                0,
                "Breakdown in app using Closeout Sheet",
            )
        )
        self.assertFalse(coupon_workflow_is_required(0, None))
        self.assertFalse(
            coupon_workflow_is_required(0, "Finish manually in QuickBooks")
        )
        self.assertTrue(coupon_workflow_is_required(10, None))

    def test_closeout_form_payload_preserves_order_defaults_and_confirmation(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            build_closeout_form_payload,
        )

        payload = build_closeout_form_payload(
            baselines={key: 10 for key in STANDARD_CLOSEOUT_ORDER},
            actuals={**{key: 10 for key in STANDARD_CLOSEOUT_ORDER}, "offline_zon": 0},
            reviewed=True,
            payroll=-4000,
            safe_type="shortage",
            safe_amount=25,
            plants_purchase=40,
            custom_tba=[{"memo": "Other", "amount": 3, "direction": "adds"}],
            final_total=1000,
            approve_final_pos=False,
        )

        self.assertEqual(list(payload["actuals"]), list(STANDARD_CLOSEOUT_ORDER))
        self.assertEqual(payload["payroll"], -4000.0)
        self.assertEqual(payload["safe"], {"type": "shortage", "amount": 25.0})
        self.assertEqual(payload["plants_purchase"], 40.0)
        self.assertEqual(
            payload["custom_tba"],
            [{"memo": "Other", "amount": 3.0, "direction": "adds"}],
        )
        self.assertTrue(payload["reviewed"])

    def test_closeout_form_requires_paper_review_confirmation(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            build_closeout_form_payload,
        )

        with self.assertRaisesRegex(ValueError, "paper Closeout Sheet"):
            build_closeout_form_payload(
                baselines={key: 10 for key in STANDARD_CLOSEOUT_ORDER},
                actuals={key: 10 for key in STANDARD_CLOSEOUT_ORDER},
                reviewed=False,
                payroll=0,
                safe_type="none",
                safe_amount=0,
                plants_purchase=0,
                custom_tba=[],
                final_total=1000,
                approve_final_pos=False,
            )

    def test_closeout_preview_fingerprint_invalidates_financial_changes_only(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            closeout_input_fingerprint,
        )

        payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {key: 10 for key in STANDARD_CLOSEOUT_ORDER},
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1000,
            "approve_final_pos": False,
        }
        original = closeout_input_fingerprint(payload)
        approved = {**payload, "approve_final_pos": True}
        changed = {**payload, "final_total": 1001}

        self.assertEqual(closeout_input_fingerprint(approved), original)
        self.assertNotEqual(closeout_input_fingerprint(changed), original)

    def test_closeout_preview_freshness_requires_matching_payload_fingerprint(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            closeout_input_fingerprint,
            closeout_preview_is_fresh,
        )

        payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {key: 10 for key in STANDARD_CLOSEOUT_ORDER},
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1000,
            "approve_final_pos": False,
        }
        saved = {
            "input_fingerprint": closeout_input_fingerprint(payload),
            "preview": {"remaining": 1.25},
        }

        self.assertTrue(closeout_preview_is_fresh(payload, saved))
        self.assertFalse(
            closeout_preview_is_fresh({**payload, "payroll": 4000}, saved)
        )
        self.assertFalse(closeout_preview_is_fresh(payload, None))

    def test_closeout_preview_fingerprint_includes_other_deposit_inputs(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            closeout_input_fingerprint,
        )

        payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {key: 10 for key in STANDARD_CLOSEOUT_ORDER},
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1000,
            "approve_final_pos": False,
        }
        original = closeout_input_fingerprint(
            payload,
            review_context={"membership_total": 8.45, "settlement": "first"},
        )

        self.assertNotEqual(
            closeout_input_fingerprint(
                payload,
                review_context={"membership_total": 16.90, "settlement": "first"},
            ),
            original,
        )
        self.assertNotEqual(
            closeout_input_fingerprint(
                payload,
                review_context={"membership_total": 8.45, "settlement": "second"},
            ),
            original,
        )

    def test_manual_closeout_payload_returns_final_result_without_preview(self):
        from app.closeout_reconciliation import (
            normalize_closeout_payload,
            write_closeout_payload_file,
        )
        from app.membership_payments import write_membership_payments_file

        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        runtime_nodes = [
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef)
            and node.name in {"build_engine_command", "run_engine"}
        ]
        runtime_root = Path(__file__).parent / f"_manual_closeout_runtime_{uuid4().hex}"
        input_dir = runtime_root / "input"
        iif_dir = runtime_root / "qb_imports"
        log_dir = runtime_root / "logs"
        temp_dir = runtime_root / "runtime"
        for folder in (input_dir, iif_dir, log_dir, temp_dir):
            folder.mkdir(parents=True, exist_ok=True)

        class UploadedWorkbook:
            name = "daily.xlsx"

            def getvalue(self):
                return b"workbook"

        class FakeSubprocess:
            def __init__(self):
                self.command = None

            def run(self, command, **_kwargs):
                self.command = command
                (iif_dir / "deposit_20260828.iif").write_bytes(b"legacy final iif")
                return SimpleNamespace(stdout="engine output", stderr="", returncode=0)

        fake_subprocess = FakeSubprocess()
        namespace = {
            "Path": Path,
            "date": date,
            "json": __import__("json"),
            "sys": sys,
            "uuid4": uuid4,
            "subprocess": fake_subprocess,
            "ENGINE_PATH": source_path,
            "ROOT": runtime_root,
            "INPUT_DIR": input_dir,
            "QB_IMPORT_DIR": iif_dir,
            "LOG_DIR": log_dir,
            "RUNTIME_TEMP_DIR": temp_dir,
            "write_membership_payments_file": write_membership_payments_file,
            "normalize_closeout_payload": normalize_closeout_payload,
            "write_closeout_payload_file": write_closeout_payload_file,
            "parse_iif": lambda _path: (["parsed line"], "parsed dataframe"),
            "parse_validation": lambda _log_text, _lines: {"all_ok": True},
        }
        exec(
            compile(ast.Module(body=runtime_nodes, type_ignores=[]), str(source_path), "exec"),
            namespace,
        )

        try:
            result = namespace["run_engine"](
                UploadedWorkbook(),
                UploadedWorkbook(),
                date(2026, 8, 28),
                [],
                "automatic",
                "quickbooks",
                None,
                None,
                None,
                closeout_payload={"mode": "manual"},
            )
        finally:
            for path in sorted(runtime_root.rglob("*"), reverse=True):
                if path.is_file():
                    path.unlink()
                else:
                    path.rmdir()
            runtime_root.rmdir()

        self.assertIn("--closeout-file", fake_subprocess.command)
        self.assertNotIn("--closeout-preview-output", fake_subprocess.command)
        self.assertFalse(result["preview_only"])
        self.assertIsNone(result["closeout_preview"])
        self.assertEqual(result["iif_bytes"], b"legacy final iif")

    def test_build_engine_command_passes_closeout_files(self):
        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        function_node = next(
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef) and node.name == "build_engine_command"
        )
        namespace = {"date": date, "Path": Path, "sys": sys}
        exec(
            compile(ast.Module(body=[function_node], type_ignores=[]), str(source_path), "exec"),
            namespace,
        )
        build_engine_command = namespace["build_engine_command"]

        command = build_engine_command(
            engine_path=Path("engine.py"),
            deposit_date=date(2026, 8, 28),
            membership_path=Path("members.json"),
            membership_mode="automatic",
            coupon_mode="closeout",
            coupon_closeout_total=188.25,
            coupon_ncg_total=152.25,
            coupon_mfg_total=36.00,
            closeout_path=Path("closeout.json"),
            closeout_preview_path=Path("preview.json"),
        )

        self.assertIn("--closeout-file", command)
        self.assertIn("closeout.json", command)
        self.assertIn("--closeout-preview-output", command)
        self.assertIn("preview.json", command)

    def test_engine_rejects_malformed_closeout_payload(self):
        engine_path = Path(__file__).parents[1] / "app" / "pos_to_quickbooks_v2.py"
        closeout_path = Path(__file__).parent / "_malformed_closeout_cli.json"
        try:
            closeout_path.write_text("{not valid JSON", encoding="utf-8")
            proc = subprocess.run(
                [
                    sys.executable,
                    str(engine_path),
                    "--date",
                    "08/28/26",
                    "--closeout-file",
                    str(closeout_path),
                ],
                cwd=engine_path.parents[1],
                capture_output=True,
                text=True,
                timeout=30,
            )
        finally:
            closeout_path.unlink(missing_ok=True)

        self.assertNotEqual(proc.returncode, 0)
        self.assertIn(
            "Closeout payload file contains malformed JSON",
            (proc.stdout or "") + (proc.stderr or ""),
        )

    def test_coupon_counter_reference_workbook_is_packaged(self):
        from pathlib import Path

        import openpyxl

        workbook_path = (
            Path(__file__).resolve().parents[1]
            / "assets"
            / "NCG-MFG Coupon Counter.xlsx"
        )
        self.assertTrue(workbook_path.is_file())
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
        try:
            self.assertIn("Template", workbook.sheetnames)
            sheet = workbook["Template"]
            self.assertEqual(
                [sheet.cell(1, column).value for column in range(2, 7)],
                ["NCG", "MFG", "VP", "MKTG", "SITKA"],
            )
        finally:
            workbook.close()

    def test_coupon_reconciliation_keeps_legacy_bs_process(self):
        try:
            from app.coupon_reconciliation import reconcile_coupon_receivable
        except ImportError as exc:
            self.fail(f"coupon reconciliation module is missing: {exc}")

        self.assertEqual(
            reconcile_coupon_receivable(181.50, mode="quickbooks"),
            {
                "bs_total": 181.50,
                "closeout_actual_total": None,
                "ncg_total": 181.50,
                "mfg_total": None,
                "difference": None,
            },
        )

    def test_coupon_reconciliation_builds_signed_closeout_differences(self):
        try:
            from app.coupon_reconciliation import reconcile_coupon_receivable
        except ImportError as exc:
            self.fail(f"coupon reconciliation module is missing: {exc}")

        cases = [
            (188.25, 152.25, 36.00, 6.75),
            (175.00, 150.00, 25.00, -6.50),
        ]
        for closeout, ncg, mfg, expected_difference in cases:
            with self.subTest(closeout=closeout):
                result = reconcile_coupon_receivable(
                    181.50,
                    mode="closeout",
                    closeout_actual_total=closeout,
                    ncg_total=ncg,
                    mfg_total=mfg,
                )
                self.assertEqual(result["difference"], expected_difference)
                self.assertEqual(result["ncg_total"], ncg)
                self.assertEqual(result["mfg_total"], mfg)

    def test_coupon_reconciliation_requires_counts_to_match_closeout_actual(self):
        try:
            from app.coupon_reconciliation import reconcile_coupon_receivable
        except ImportError as exc:
            self.fail(f"coupon reconciliation module is missing: {exc}")

        with self.assertRaisesRegex(
            ValueError,
            "NCG Coupons.*MFG Coupons.*Closeout Sheet Coupon Actual Total",
        ):
            reconcile_coupon_receivable(
                181.50,
                mode="closeout",
                closeout_actual_total=188.25,
                ncg_total=150.00,
                mfg_total=36.00,
            )

    def test_coupon_receivable_total_is_read_from_bs_code_908(self):
        from io import BytesIO

        import openpyxl

        try:
            from app.coupon_reconciliation import read_coupon_receivable_total
        except ImportError as exc:
            self.fail(f"coupon BS reader is missing: {exc}")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "082626 BS"
        sheet.append([908, "Dwr Vendor coupon", None, None, -181.50])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        self.assertEqual(
            read_coupon_receivable_total(output.getvalue(), "082626 BS"),
            181.50,
        )

    def test_result_actions_stay_outside_more_information_dropdown(self):
        import ast
        from pathlib import Path

        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        detail_dropdowns = []
        for node in ast.walk(source_tree):
            if not isinstance(node, ast.With) or not node.items:
                continue
            context = node.items[0].context_expr
            if not isinstance(context, ast.Call):
                continue
            function = context.func
            if not (
                isinstance(function, ast.Attribute)
                and function.attr == "expander"
                and context.args
                and isinstance(context.args[0], ast.Constant)
                and context.args[0].value == "More deposit information"
            ):
                continue
            detail_dropdowns.append(node)

        self.assertEqual(len(detail_dropdowns), 1)
        dropdown_source = ast.unparse(detail_dropdowns[0])
        self.assertIn("st.tabs", dropdown_source)
        self.assertNotIn("Download QuickBooks IIF", dropdown_source)
        self.assertNotIn("Run another deposit", dropdown_source)

    def test_coupon_closeout_ui_uses_direct_totals_without_stack_controls(self):
        import ast
        from pathlib import Path

        source = (
            Path(__file__).parents[1] / "streamlit_app.py"
        ).read_text(encoding="utf-8")
        required_labels = (
            "How should Coupons Receivable be handled?",
            "Finish manually in QuickBooks",
            "Breakdown in app using Closeout Sheet",
            "Closeout Sheet Coupon Actual Total",
            "NCG Coupons counted",
            "MFG Coupons counted",
            "Download Excel coupon counter",
        )
        for label in required_labels:
            with self.subTest(label=label):
                self.assertIn(label, source)

        for removed_control in (
            "How would you like to enter coupon counts?",
            "Count coupon stacks in app",
            "Add a stack",
            "Written stack total (optional)",
            "NCG quick amounts",
            "MFG + VP + MKTG + SITKA",
        ):
            with self.subTest(removed_control=removed_control):
                self.assertNotIn(removed_control, source)

        radio_options = {}
        for node in ast.walk(ast.parse(source)):
            if not isinstance(node, ast.Call):
                continue
            if not isinstance(node.func, ast.Attribute) or node.func.attr != "radio":
                continue
            if not node.args or not isinstance(node.args[0], ast.Constant):
                continue
            options_keyword = next(
                (keyword for keyword in node.keywords if keyword.arg == "options"),
                None,
            )
            if options_keyword and isinstance(options_keyword.value, ast.List):
                radio_options[node.args[0].value] = [
                    item.value for item in options_keyword.value.elts
                ]

        self.assertEqual(
            radio_options["How should these payments be handled?"],
            [
                "Breakdown in app using the Ownership Payments sheet",
                "Finish manually in QuickBooks",
            ],
        )
        self.assertEqual(
            radio_options["How should Coupons Receivable be handled?"],
            [
                "Breakdown in app using Closeout Sheet",
                "Finish manually in QuickBooks",
            ],
        )

    def test_app_passes_coupon_closeout_values_to_engine(self):
        import ast
        from pathlib import Path

        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        command_builder_node = next(
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef) and node.name == "build_engine_command"
        )
        command_builder_source = ast.unparse(command_builder_node)
        for flag in (
            "--coupon-mode",
            "--coupon-closeout-total",
            "--coupon-ncg-total",
            "--coupon-mfg-total",
        ):
            with self.subTest(flag=flag):
                self.assertIn(flag, command_builder_source)

    def test_workbook_validation_ignores_xxxxxx_discount_and_hash_tabs(self):
        import ast
        from io import BytesIO
        from pathlib import Path
        from typing import Optional

        import openpyxl

        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        function_node = next(
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef) and node.name == "detect_sheet_roles"
        )
        namespace = {"Optional": Optional}
        exec(compile(ast.Module(body=[function_node], type_ignores=[]), str(source_path), "exec"), namespace)

        workbook = openpyxl.Workbook()
        workbook.active.title = "XXXXXX Discounts"
        workbook.create_sheet("XXXXXX Hash")
        discounts = workbook.create_sheet("082626 Discounts")
        discounts.append(["Discounts by Shopper Level"])
        discounts.append([None, None, "Member Discounts"])
        hash_sheet = workbook.create_sheet("082626 Hash")
        hash_sheet.append([None, 23, "Refunded Discounts", None, None, None, 8, 6.96])
        hash_sheet.append([None, 32, "PASS THROUGH DONATIONS", None, None, None, 5, 5.00])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        roles = namespace["detect_sheet_roles"](output.getvalue())

        self.assertEqual(roles["discounts"], "082626 Discounts")
        self.assertEqual(roles["hash"], "082626 Hash")

    def test_discount_parser_prefers_the_dated_tab_over_xxxxxx_placeholder(self):
        from datetime import date
        from pathlib import Path

        import openpyxl

        from app import pos_to_quickbooks_v2 as engine

        workbook_path = Path(__file__).parent / "_dated_discounts_fixture.xlsx"
        workbook = openpyxl.Workbook()
        placeholder = workbook.active
        placeholder.title = "XXXXXX Discounts"
        dated = workbook.create_sheet("082626 Discounts")
        dated.append(["Discounts by Shopper Level"])
        dated.append(["Date: ", "8/26/2026", "to", "8/26/2026"])
        dated.append(["Target: ", "RAL", "Report all"])
        dated.append([None, None, "Description", None, None, None, None, None, "Qty", "Amount"])
        dated.append(["Shareholder", None, None, 2, None, None, None, 2365, 277.66])
        dated.append(["Senior NonMember", None, None, 3, None, None, None, 3294, 1646.30])
        dated.append(["Senior Share", None, None, 4, None, None, None, 3225, 1547.71])
        dated.append([None, None, "Member Discounts", None, None, None, None, 8814, 3471.67])
        workbook.save(workbook_path)
        workbook.close()
        try:
            discounts, grand_total = engine.parse_excel_discounts(
                workbook_path,
                date(2026, 8, 26),
            )
        finally:
            try:
                workbook_path.unlink()
            except PermissionError:
                pass

        self.assertEqual(grand_total, 3471.67)
        self.assertEqual(discounts["8512001 · Discount 2% - Owners"], 277.66)
        self.assertEqual(discounts["8511002 · Discount 8% - Senior Day"], 3194.01)

    def test_hash_parser_prefers_the_dated_tab_over_xxxxxx_placeholder(self):
        from datetime import date
        from pathlib import Path

        import openpyxl

        from app import pos_to_quickbooks_v2 as engine

        workbook_path = Path(__file__).parent / "_dated_hash_fixture.xlsx"
        workbook = openpyxl.Workbook()
        placeholder = workbook.active
        placeholder.title = "XXXXXX Hash"
        dated = workbook.create_sheet("082626 Hash")
        dated.append(["Sub-department Single Total"])
        dated.append(["Date: ", "8/26/2026", "to", "8/26/2026"])
        dated.append(["S-Dept.  ", 0, "to", 999999])
        dated.append(["Tlz.:", 6, "to", 6])
        dated.append(["Target: ", "RAL", "Report all"])
        dated.append([None, None, "Sub-Department", None, None, None, None, "Qty", "Amount"])
        dated.append([None, 23, "Refunded Discounts", None, None, None, 8, 6.96])
        dated.append([None, 32, "PASS THROUGH DONATIONS", None, None, None, 5, 5.00])
        workbook.save(workbook_path)
        workbook.close()
        try:
            parsed = engine.parse_hash_sheet(workbook_path, date(2026, 8, 26))
        finally:
            try:
                workbook_path.unlink()
            except PermissionError:
                pass

        self.assertEqual(parsed, (6.96, 5.00, 0.0))

    def test_hash_exact_amount_header_controls_paid_in_engine_and_iif(self):
        from datetime import date
        from pathlib import Path

        import openpyxl

        from app import pos_to_quickbooks_v2 as engine

        fixture_root = Path(__file__).parent / f"_hash_exact_amount_{uuid4().hex}"
        fixture_root.mkdir()
        workbook_path = fixture_root / "daily.xlsx"
        workbook = openpyxl.Workbook()
        hash_sheet = workbook.active
        hash_sheet.title = "082826 Hash"
        hash_sheet.append(["Code", "Description", "Net Amount", "Amount"])
        hash_sheet.append([34, "Paid-Ins", 999.99, 34.56])
        workbook.save(workbook_path)
        workbook.close()

        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = fixture_root
        engine.LOG_DIR = fixture_root
        engine.log.disabled = True
        try:
            refunded, pass_through, paid_in = engine.parse_hash_sheet(
                workbook_path, date(2026, 8, 28)
            )
            iif_path = engine.generate_iif(
                {},
                {},
                {},
                date(2026, 8, 28),
                paid_in_total=paid_in,
            )
            iif_text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in fixture_root.iterdir():
                generated_file.unlink()
            fixture_root.rmdir()

        self.assertEqual((refunded, pass_through, paid_in), (0.0, 0.0, 34.56))
        self.assertIn("4444 · TBA Purchases\t\t-34.56\tPAID IN:", iif_text)
        self.assertNotIn("999.99\tPAID IN:", iif_text)

    def test_discount_parser_falls_back_to_a_populated_custom_tab(self):
        from datetime import date
        from pathlib import Path

        import openpyxl

        from app import pos_to_quickbooks_v2 as engine

        workbook_path = Path(__file__).parent / "_custom_discounts_fixture.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "XXXXXX Discounts"
        custom = workbook.create_sheet("Daily Shopper Report")
        custom.append(["Discounts by Shopper Level"])
        custom.append(["Date: ", "8/26/2026", "to", "8/26/2026"])
        custom.append(["Target: ", "RAL", "Report all"])
        custom.append([None, None, "Description", None, None, None, None, None, "Qty", "Amount"])
        custom.append(["Shareholder", None, None, 2, None, None, None, 2365, 277.66])
        custom.append([None, None, "Member Discounts", None, None, None, None, 2365, 277.66])
        workbook.save(workbook_path)
        workbook.close()
        try:
            discounts, total = engine.parse_excel_discounts(
                workbook_path,
                date(2026, 8, 26),
            )
        finally:
            try:
                workbook_path.unlink()
            except PermissionError:
                pass

        self.assertEqual(total, 277.66)
        self.assertEqual(discounts["8512001 · Discount 2% - Owners"], 277.66)

    def test_hash_parser_falls_back_to_a_populated_custom_tab(self):
        from datetime import date
        from pathlib import Path

        import openpyxl

        from app import pos_to_quickbooks_v2 as engine

        workbook_path = Path(__file__).parent / "_custom_hash_fixture.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.title = "XXXXXX Hash"
        custom = workbook.create_sheet("Daily Special Items")
        custom.append(["Sub-department Single Total"])
        custom.append(["Date: ", "8/26/2026", "to", "8/26/2026"])
        custom.append(["S-Dept.  ", 0, "to", 999999])
        custom.append(["Tlz.:", 6, "to", 6])
        custom.append(["Target: ", "RAL", "Report all"])
        custom.append([None, None, "Sub-Department", None, None, None, None, "Qty", "Amount"])
        custom.append([None, 23, "Refunded Discounts", None, None, None, 8, 6.96])
        custom.append([None, 32, "PASS THROUGH DONATIONS", None, None, None, 5, 5.00])
        workbook.save(workbook_path)
        workbook.close()
        try:
            parsed = engine.parse_hash_sheet(workbook_path, date(2026, 8, 26))
        finally:
            try:
                workbook_path.unlink()
            except PermissionError:
                pass

        self.assertEqual(parsed, (6.96, 5.00, 0.0))

    def test_automatic_split_leaves_new_quickbooks_member_name_blank(self):
        from app.membership_payments import (
            build_membership_lines,
            membership_payment_from_entry,
        )

        try:
            payment = membership_payment_from_entry(
                member_name="This typed value must be ignored",
                member_number_status="No",
                member_number="",
                quickbooks_member_exists=False,
                payment_option="Paid in full — $100",
                amount=100.00,
            )
            lines = build_membership_lines([payment], handling_mode="automatic")
        except ValueError as exc:
            self.fail(f"new QuickBooks member was rejected: {exc}")

        self.assertEqual(payment["member_name"], "")
        self.assertEqual(lines[0]["name"], "")

    def test_member_payment_entry_requires_quickbooks_name_confirmation(self):
        from app.membership_payments import membership_payment_from_entry

        with self.assertRaisesRegex(
            ValueError,
            "Select Yes or No.*QuickBooks",
        ):
            membership_payment_from_entry(
                member_name="Existing Member",
                member_number_status="No",
                member_number="",
                payment_option="Paid in full — $100",
                amount=100.00,
            )

    def test_member_number_question_builds_assigned_and_pending_entries(self):
        try:
            from app.membership_payments import membership_payment_from_entry
        except ImportError as exc:
            self.fail(f"member payment entry builder is missing: {exc}")

        assigned = membership_payment_from_entry(
            member_name="Assigned Member",
            member_number_status="Yes",
            member_number="12345",
            quickbooks_member_exists=True,
            payment_option="Existing plan — 1 year",
            amount=8.45,
        )
        pending = membership_payment_from_entry(
            member_name="Pending Member",
            member_number_status="No",
            member_number="",
            quickbooks_member_exists=True,
            payment_option="New plan — 3 year",
            amount=15.00,
        )

        self.assertEqual(assigned["member_number"], "12345")
        self.assertFalse(assigned["member_number_pending"])
        self.assertEqual(assigned["payment_type"], "Existing plan")
        self.assertEqual(assigned["plan"], "1 year")
        self.assertEqual(pending["member_number"], "")
        self.assertTrue(pending["member_number_pending"])
        self.assertEqual(pending["payment_type"], "New plan")
        self.assertEqual(pending["plan"], "3 year")

    def test_saved_member_payment_can_be_removed_by_position(self):
        try:
            from app.membership_payments import remove_membership_payment
        except ImportError as exc:
            self.fail(f"saved member payment removal is missing: {exc}")

        payments = [
            {"member_name": "Keep First"},
            {"member_name": "Remove"},
            {"member_name": "Keep Last"},
        ]

        self.assertEqual(
            remove_membership_payment(payments, 1),
            [
                {"member_name": "Keep First"},
                {"member_name": "Keep Last"},
            ],
        )

    def test_pending_member_number_rejects_a_typed_number(self):
        from app.membership_payments import build_membership_lines

        with self.assertRaisesRegex(
            ValueError,
            "Enter a member number or select Member # pending, not both",
        ):
            build_membership_lines([{
                "member_name": "New Member",
                "member_number": "12345",
                "member_number_pending": True,
                "payment_type": "Existing plan",
                "plan": "1 year",
                "amount": 8.45,
            }])

    def test_pending_member_number_uses_pending_memo(self):
        from app.membership_payments import build_membership_lines

        try:
            lines = build_membership_lines([{
                "member_name": "New Member",
                "member_number": "",
                "member_number_pending": True,
                "payment_type": "Existing plan",
                "plan": "1 year",
                "amount": 8.45,
            }])
        except ValueError as exc:
            self.fail(f"pending member number was rejected: {exc}")

        self.assertEqual(lines[0]["memo"], "Share Installments - Paid #Pending")
        self.assertEqual(lines[1]["memo"], "Share Installments - Paid #Pending")

    def test_paid_in_full_does_not_require_member_number_or_plan(self):
        from app.membership_payments import build_membership_lines

        try:
            lines = build_membership_lines([{
                "member_name": "Fully Paid Member",
                "member_number": "",
                "payment_type": "Paid in full",
                "plan": "5 year",
                "amount": 100.00,
            }])
        except ValueError as exc:
            self.fail(f"paid-in-full member number should be optional: {exc}")

        self.assertEqual(lines, [{
            "account": "6100000 · Member Shares (Paid-In Equity)",
            "name": "Fully Paid Member",
            "memo": "Member Shares - Paid",
            "class_name": "",
            "amount": 100.00,
        }])

    def test_combined_payment_option_prevents_a_plan_for_paid_in_full(self):
        try:
            from app.membership_payments import payment_fields_from_option
        except ImportError as exc:
            self.fail(f"combined payment option mapping is missing: {exc}")

        self.assertEqual(
            payment_fields_from_option("Paid in full — $100"),
            {"payment_type": "Paid in full", "plan": ""},
        )
        self.assertEqual(
            payment_fields_from_option("New plan — 3 year"),
            {"payment_type": "New plan", "plan": "3 year"},
        )
        self.assertEqual(
            payment_fields_from_option("Existing plan — 5 year"),
            {"payment_type": "Existing plan", "plan": "5 year"},
        )

    def test_entering_paid_in_full_resets_stale_quickbooks_name_state(self):
        from app.membership_payments import quickbooks_name_state_for_payment_option

        self.assertEqual(
            quickbooks_name_state_for_payment_option(
                "Paid in full — $100", "New plan — 1 year", "Yes", "Stale Name"
            ),
            ("No", ""),
        )

    def test_rerunning_paid_in_full_preserves_deliberate_quickbooks_name(self):
        from app.membership_payments import quickbooks_name_state_for_payment_option

        self.assertEqual(
            quickbooks_name_state_for_payment_option(
                "Paid in full — $100",
                "Paid in full — $100",
                "Yes",
                "  Karl Chester Cruz  ",
            ),
            ("Yes", "Karl Chester Cruz"),
        )

    def test_non_paid_in_full_rerun_preserves_current_quickbooks_name_state(self):
        from app.membership_payments import quickbooks_name_state_for_payment_option

        self.assertEqual(
            quickbooks_name_state_for_payment_option(
                "Existing plan — 1 year", "New plan — 3 year", None, ""
            ),
            (None, ""),
        )

    def test_leaving_paid_in_full_clears_quickbooks_name_default(self):
        from app.membership_payments import quickbooks_name_state_for_payment_option

        self.assertEqual(
            quickbooks_name_state_for_payment_option(
                "New plan — 5 year", "Paid in full — $100", "No", "Karl Chester Cruz"
            ),
            (None, ""),
        )

    def test_quickbooks_name_state_adapter_enters_paid_in_full_with_defaults(self):
        from app.membership_payments import apply_quickbooks_name_option_state

        state = {
            "entry_previous_payment_option": "New plan — 1 year",
            "entry_quickbooks_name_status": "Yes",
            "entry_member_name": "Stale Name",
        }

        result = apply_quickbooks_name_option_state(
            state,
            "entry",
            "Paid in full — $100",
        )

        self.assertEqual(result, ("No", ""))
        self.assertEqual(
            state,
            {
                "entry_previous_payment_option": "Paid in full — $100",
                "entry_quickbooks_name_status": "No",
                "entry_member_name": "",
            },
        )

    def test_quickbooks_name_state_adapter_preserves_deliberate_paid_in_full_name(self):
        from app.membership_payments import apply_quickbooks_name_option_state

        state = {
            "entry_previous_payment_option": "Paid in full — $100",
            "entry_quickbooks_name_status": "Yes",
            "entry_member_name": "  Karl Chester Cruz  ",
        }

        result = apply_quickbooks_name_option_state(
            state,
            "entry",
            "Paid in full — $100",
        )

        self.assertEqual(result, ("Yes", "Karl Chester Cruz"))
        self.assertEqual(state["entry_previous_payment_option"], "Paid in full — $100")
        self.assertEqual(state["entry_quickbooks_name_status"], "Yes")
        self.assertEqual(state["entry_member_name"], "Karl Chester Cruz")

    def test_quickbooks_name_state_adapter_leaves_paid_in_full_with_blank_state(self):
        from app.membership_payments import apply_quickbooks_name_option_state

        state = {
            "entry_previous_payment_option": "Paid in full — $100",
            "entry_quickbooks_name_status": "No",
            "entry_member_name": "Stale Name",
        }

        result = apply_quickbooks_name_option_state(
            state,
            "entry",
            "New plan — 5 year",
        )

        self.assertEqual(result, (None, ""))
        self.assertEqual(
            state,
            {
                "entry_previous_payment_option": "New plan — 5 year",
                "entry_quickbooks_name_status": None,
                "entry_member_name": "",
            },
        )

    def test_blank_dynamic_editor_row_does_not_require_a_payment_option(self):
        from app.membership_payments import prepare_membership_editor_rows

        try:
            prepared = prepare_membership_editor_rows(
                [{"payment_option": float("nan"), "interest_periods": None}],
                allow_interest_override=False,
            )
        except ValueError as exc:
            self.fail(f"blank dynamic editor row was treated as a payment: {exc}")

        self.assertNotIn("payment_type", prepared[0])
        self.assertNotIn("plan", prepared[0])

    def test_editor_rows_refresh_only_to_autofill_paid_in_full(self):
        try:
            from app.membership_payments import normalize_membership_editor_rows
        except ImportError as exc:
            self.fail(f"membership editor normalization is missing: {exc}")

        ordinary_rows = [{
            "member_name": "Still Typing",
            "payment_option": "Existing plan — 1 year",
            "amount": 8.45,
        }]
        paid_in_full_rows = [{
            "member_name": "Fully Paid Member",
            "payment_option": "Paid in full — $100",
            "amount": None,
        }]

        normalized, refresh_required = normalize_membership_editor_rows(ordinary_rows)
        self.assertEqual(normalized, ordinary_rows)
        self.assertFalse(refresh_required)

        normalized, refresh_required = normalize_membership_editor_rows(paid_in_full_rows)
        self.assertEqual(normalized[0]["amount"], 100.00)
        self.assertTrue(refresh_required)

    def test_plan_reference_rows_match_the_staff_payment_guide(self):
        try:
            from app.membership_payments import plan_reference_rows
        except ImportError as exc:
            self.fail(f"staff plan reference is missing: {exc}")

        self.assertEqual(plan_reference_rows(), [
            {
                "Plan": "1 year",
                "Deposit": 10.00,
                "Total Paid": 102.95,
                "Payments": 11,
                "Installment": 8.45,
                "Principal": 8.18,
                "Interest": 0.27,
            },
            {
                "Plan": "3 year",
                "Deposit": 15.00,
                "Total Paid": 109.14,
                "Payments": 6,
                "Installment": 15.69,
                "Principal": 14.17,
                "Interest": 1.52,
            },
            {
                "Plan": "5 year",
                "Deposit": 10.00,
                "Total Paid": 115.50,
                "Payments": 10,
                "Installment": 10.55,
                "Principal": 9.00,
                "Interest": 1.55,
            },
        ])

    def test_membership_choice_requires_a_selection_and_maps_to_engine_mode(self):
        try:
            from app.membership_payments import membership_mode_from_choice
        except ImportError as exc:
            self.fail(f"membership workflow choice mapping is missing: {exc}")

        self.assertIsNone(membership_mode_from_choice(None))
        self.assertEqual(
            membership_mode_from_choice(
                "Breakdown in app using the Ownership Payments sheet"
            ),
            "automatic",
        )
        self.assertEqual(
            membership_mode_from_choice("Finish manually in QuickBooks"),
            "manual",
        )

    def test_hidden_payoff_override_is_cleared_before_automatic_split(self):
        try:
            from app.membership_payments import prepare_membership_editor_rows
        except ImportError as exc:
            self.fail(f"membership editor row preparation is missing: {exc}")

        rows = [{
            "member_name": "A Member",
            "member_number": "12345",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 20.00,
            "interest_periods": 1,
        }]

        self.assertEqual(
            prepare_membership_editor_rows(rows, allow_interest_override=False)[0]["interest_periods"],
            None,
        )
        self.assertEqual(
            prepare_membership_editor_rows(rows, allow_interest_override=True)[0]["interest_periods"],
            1,
        )

    def test_manual_quickbooks_mode_posts_one_balancing_member_share_line(self):
        from app.membership_payments import build_membership_lines

        try:
            lines = build_membership_lines(
                [],
                expected_subscription_total=8.45,
                handling_mode="manual",
            )
        except TypeError as exc:
            self.fail(f"manual QuickBooks handling mode is missing: {exc}")

        self.assertEqual(lines, [{
            "account": "6100000 · Member Shares (Paid-In Equity)",
            "name": "",
            "memo": "Member Shares - Paid",
            "class_name": "",
            "amount": 8.45,
        }])

    def test_automatic_mode_still_requires_member_details(self):
        from app.membership_payments import build_membership_lines

        try:
            with self.assertRaisesRegex(ValueError, "no membership payments were supplied"):
                build_membership_lines(
                    [],
                    expected_subscription_total=8.45,
                    handling_mode="automatic",
                )
        except TypeError as exc:
            self.fail(f"automatic membership handling mode is missing: {exc}")

    def test_subscription_action_status_distinguishes_clear_and_action_required(self):
        try:
            from app.membership_payments import subscription_action_status
        except ImportError as exc:
            self.fail(f"subscription action status helper is missing: {exc}")

        self.assertEqual(
            subscription_action_status(0),
            {
                "needs_action": False,
                "title": "No Subscription Revenue",
                "message": "No member-share action is needed for this deposit.",
            },
        )
        self.assertEqual(
            subscription_action_status(8.45),
            {
                "needs_action": True,
                "title": "Subscription Revenue found: $8.45",
                "message": (
                    "Choose automatic splitting or finish manually in QuickBooks "
                    "before building the deposit."
                ),
            },
        )

    def test_iif_delimiters_are_rejected_in_member_identity(self):
        from app.membership_payments import build_membership_lines

        base = {
            "member_name": "A Member",
            "member_number": "12345",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 8.45,
        }
        invalid_values = (
            ("member_name", "A\tMember", "tabs or line breaks"),
            ("member_name", "A\nMember", "tabs or line breaks"),
            ("member_number", "12\r345", "tabs or line breaks"),
            ("member_number", "12-345", "digits only"),
            ("member_number", "１２３４５", "digits only"),
        )

        for field, value, message in invalid_values:
            with self.subTest(field=field, value=value):
                payment = dict(base)
                payment[field] = value
                with self.assertRaisesRegex(ValueError, message):
                    build_membership_lines([payment])

    def test_exclusive_run_lock_blocks_overlap_and_releases(self):
        from pathlib import Path

        from app.membership_payments import exclusive_run_lock

        lock_path = Path(__file__).parent / "_deposit_run.lock"
        lock_path.unlink(missing_ok=True)
        try:
            with exclusive_run_lock(lock_path):
                self.assertTrue(lock_path.exists())
                with self.assertRaisesRegex(RuntimeError, "Another deposit is currently running"):
                    with exclusive_run_lock(lock_path):
                        pass
            self.assertTrue(lock_path.exists())
            with exclusive_run_lock(lock_path):
                self.assertTrue(lock_path.exists())
        finally:
            lock_path.unlink(missing_ok=True)

    def test_abandoned_run_lock_is_reusable_and_stale_membership_files_are_cleaned_up(self):
        import os
        from pathlib import Path
        import time

        from app.membership_payments import exclusive_run_lock, write_membership_payments_file

        folder = Path(__file__).parent / "_stale_membership_output"
        folder.mkdir(exist_ok=True)
        lock_path = folder / "deposit.lock"
        stale_json = folder / "membership_payments_stale.json"
        lock_path.write_text("abandoned", encoding="utf-8")
        stale_json.write_text("[]", encoding="utf-8")
        old_time = time.time() - 7200
        os.utime(lock_path, (old_time, old_time))
        os.utime(stale_json, (old_time, old_time))
        created = None
        try:
            with exclusive_run_lock(lock_path, stale_seconds=60):
                self.assertTrue(lock_path.exists())
            created = write_membership_payments_file(folder, [], stale_seconds=60)
            self.assertFalse(stale_json.exists())
            self.assertTrue(created.exists())
        finally:
            lock_path.unlink(missing_ok=True)
            stale_json.unlink(missing_ok=True)
            if created is not None:
                created.unlink(missing_ok=True)
            folder.rmdir()

    def test_two_processes_cannot_both_take_over_an_abandoned_lock_file(self):
        import os
        from pathlib import Path
        import subprocess
        import sys
        import time

        test_folder = Path(__file__).parent
        lock_path = test_folder / "_cross_process_deposit.lock"
        start_path = test_folder / "_cross_process_start"
        release_path = test_folder / "_cross_process_release"
        result_paths = [test_folder / f"_cross_process_result_{index}" for index in range(2)]
        for path in [lock_path, start_path, release_path, *result_paths]:
            path.unlink(missing_ok=True)
        lock_path.write_text("abandoned", encoding="utf-8")
        old_time = time.time() - 7200
        os.utime(lock_path, (old_time, old_time))

        child_code = """
import sys
import time
from pathlib import Path
from app.membership_payments import exclusive_run_lock
lock_path, start_path, release_path, result_path = map(Path, sys.argv[1:])
deadline = time.time() + 10
while not start_path.exists() and time.time() < deadline:
    time.sleep(0.01)
try:
    with exclusive_run_lock(lock_path):
        result_path.write_text('acquired', encoding='utf-8')
        while not release_path.exists() and time.time() < deadline:
            time.sleep(0.01)
except RuntimeError:
    result_path.write_text('blocked', encoding='utf-8')
"""
        processes = [
            subprocess.Popen(
                [sys.executable, "-c", child_code, lock_path, start_path, release_path, result_path],
                cwd=Path(__file__).parents[1],
            )
            for result_path in result_paths
        ]
        try:
            start_path.write_text("start", encoding="utf-8")
            deadline = time.time() + 10
            while not all(path.exists() for path in result_paths) and time.time() < deadline:
                time.sleep(0.02)
            outcomes = [path.read_text(encoding="utf-8") for path in result_paths]
            self.assertCountEqual(outcomes, ["acquired", "blocked"])
        finally:
            release_path.write_text("release", encoding="utf-8")
            for process in processes:
                try:
                    process.wait(timeout=10)
                except subprocess.TimeoutExpired:
                    process.terminate()
                    process.wait(timeout=5)
            for path in [lock_path, start_path, release_path, *result_paths]:
                path.unlink(missing_ok=True)

    def test_nonzero_subscription_total_explains_required_membership_input(self):
        from app.membership_payments import build_membership_lines

        with self.assertRaisesRegex(ValueError, r"--membership-payments-file"):
            build_membership_lines([], expected_subscription_total=10.00)

    def test_one_year_installment_builds_principal_and_interest_lines(self):
        try:
            from app.membership_payments import build_membership_lines
        except ModuleNotFoundError as exc:
            self.fail(f"membership payment feature is missing: {exc}")

        lines = build_membership_lines(
            [
                {
                    "member_name": "Tara Caruso",
                    "member_number": "22206",
                    "payment_type": "Existing plan",
                    "plan": "1 year",
                    "amount": 8.45,
                }
            ]
        )

        self.assertEqual(
            lines,
            [
                {
                    "account": "1260000 · Member Shares Receivable",
                    "name": "Tara Caruso",
                    "memo": "Share Installments - Paid #22206",
                    "class_name": "",
                    "amount": 8.18,
                },
                {
                    "account": "9104000 · Interest Income",
                    "name": "Tara Caruso",
                    "memo": "Share Installments - Paid #22206",
                    "class_name": "Admin",
                    "amount": 0.27,
                },
            ],
        )

    def test_irregular_one_year_payment_charges_only_complete_periods(self):
        from app.membership_payments import build_membership_lines

        lines = build_membership_lines(
            [{
                "member_name": "Tara Caruso",
                "member_number": "#22206",
                "payment_type": "Existing plan",
                "plan": "1 year",
                "amount": 20.00,
            }]
        )

        self.assertEqual(lines[0]["amount"], 19.46)
        self.assertEqual(lines[1]["amount"], 0.54)
        self.assertEqual(lines[0]["memo"], "Share Installments - Paid #22206")


    def test_five_year_installment_uses_five_year_interest(self):
        from app.membership_payments import build_membership_lines

        lines = build_membership_lines([{
            "member_name": "Will Travers",
            "member_number": "21916",
            "payment_type": "Existing plan",
            "plan": "5 year",
            "amount": 10.55,
        }])

        self.assertEqual(lines[0]["amount"], 9.00)
        self.assertEqual(lines[1]["amount"], 1.55)
        self.assertEqual(lines[1]["class_name"], "Admin")


    def test_three_year_installment_uses_three_year_interest(self):
        from app.membership_payments import build_membership_lines

        try:
            lines = build_membership_lines([{
                "member_name": "A Member",
                "member_number": "30001",
                "payment_type": "Existing plan",
                "plan": "3 year",
                "amount": 15.69,
            }])
        except KeyError as exc:
            self.fail(f"3-year plan is missing: {exc}")

        self.assertEqual(lines[0]["amount"], 14.17)
        self.assertEqual(lines[1]["amount"], 1.52)


    def test_interest_period_override_handles_payoff_adjustment(self):
        from app.membership_payments import build_membership_lines

        lines = build_membership_lines([{
            "member_name": "Tara Caruso",
            "member_number": "22206",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 20.00,
            "interest_periods": 1,
        }])

        self.assertEqual(lines[0]["amount"], 19.73)
        self.assertEqual(lines[1]["amount"], 0.27)


    def test_new_three_year_plan_creates_receivable_and_interest_free_deposit(self):
        from app.membership_payments import build_membership_lines

        lines = build_membership_lines([{
            "member_name": "A New Member",
            "member_number": "30002",
            "payment_type": "New plan",
            "plan": "3 year",
            "amount": 15.00,
        }])

        self.assertEqual(
            lines,
            [
                {
                    "account": "6100000 · Member Shares (Paid-In Equity)",
                    "name": "A New Member",
                    "memo": "Member Shares - Receivable",
                    "class_name": "",
                    "amount": 100.00,
                },
                {
                    "account": "1260000 · Member Shares Receivable",
                    "name": "A New Member",
                    "memo": "Member Shares - Receivable",
                    "class_name": "",
                    "amount": -100.00,
                },
                {
                    "account": "1260000 · Member Shares Receivable",
                    "name": "A New Member",
                    "memo": "Share Installments - Paid #30002",
                    "class_name": "",
                    "amount": 15.00,
                },
            ],
        )


    def test_new_one_and_five_year_plans_create_receivable_with_ten_dollar_deposit(self):
        from app.membership_payments import build_membership_lines

        for plan in ("1 year", "5 year"):
            with self.subTest(plan=plan):
                try:
                    lines = build_membership_lines([{
                        "member_name": "A New Member",
                        "member_number": "40001",
                        "payment_type": "New plan",
                        "plan": plan,
                        "amount": 10.00,
                    }])
                except KeyError as exc:
                    self.fail(f"new {plan} plan is missing its deposit rule: {exc}")

                self.assertEqual([line["amount"] for line in lines], [100.00, -100.00, 10.00])
                self.assertEqual(lines[2]["memo"], "Share Installments - Paid #40001")


    def test_new_plan_can_include_deposit_and_first_installment(self):
        from app.membership_payments import build_membership_lines

        lines = build_membership_lines([{
            "member_name": "A New Member",
            "member_number": "40002",
            "payment_type": "New plan",
            "plan": "1 year",
            "amount": 18.45,
        }])

        self.assertEqual([line["amount"] for line in lines], [100.00, -100.00, 18.18, 0.27])


    def test_paid_in_full_posts_one_hundred_to_member_shares_equity(self):
        from app.membership_payments import build_membership_lines

        try:
            lines = build_membership_lines([{
                "member_name": "Paid Member",
                "member_number": "50001",
                "payment_type": "Paid in full",
                "plan": "",
                "amount": 100.00,
            }])
        except KeyError as exc:
            self.fail(f"paid-in-full path is missing: {exc}")

        self.assertEqual(lines, [{
            "account": "6100000 · Member Shares (Paid-In Equity)",
            "name": "Paid Member",
            "memo": "Member Shares - Paid",
            "class_name": "",
            "amount": 100.00,
        }])


    def test_payment_total_must_match_subscription_revenue(self):
        from app.membership_payments import build_membership_lines

        payment = {
            "member_name": "Tara Caruso",
            "member_number": "22206",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 8.45,
        }

        try:
            with self.assertRaisesRegex(
                ValueError,
                r"Entered membership payments \(\$8\.45\) must equal Subscription Revenue \(\$10\.00\)",
            ):
                build_membership_lines([payment], expected_subscription_total=10.00)
        except TypeError as exc:
            self.fail(f"subscription reconciliation is missing: {exc}")


    def test_blank_interest_override_uses_automatic_periods(self):
        from app.membership_payments import build_membership_lines

        for blank_value in (None, float("nan")):
            with self.subTest(blank_value=blank_value):
                try:
                    lines = build_membership_lines([{
                        "member_name": "Tara Caruso",
                        "member_number": "22206",
                        "payment_type": "Existing plan",
                        "plan": "1 year",
                        "amount": 16.90,
                        "interest_periods": blank_value,
                    }])
                except (TypeError, ValueError) as exc:
                    self.fail(f"blank interest override should use automatic periods: {exc}")

                self.assertEqual(lines[0]["amount"], 16.36)
                self.assertEqual(lines[1]["amount"], 0.54)


    def test_automatic_interest_periods_are_capped_at_full_plan_schedule(self):
        from app.membership_payments import build_membership_lines

        lines = build_membership_lines([{
            "member_name": "Payoff Member",
            "member_number": "60001",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 200.00,
        }])

        self.assertEqual(lines[0]["amount"], 197.03)
        self.assertEqual(lines[1]["amount"], 2.97)


    def test_invalid_membership_rows_are_rejected(self):
        from app.membership_payments import build_membership_lines

        base = {
            "member_name": "A Member",
            "member_number": "70001",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 8.45,
        }
        cases = [
            ({**base, "member_name": ""}, r"Member name is required"),
            ({**base, "member_number": ""}, r"Member number is required"),
            ({**base, "payment_type": "Mystery"}, r"Payment type must be"),
            ({**base, "plan": "2 year"}, r"Plan must be"),
            ({**base, "amount": 0}, r"Amount must be greater than zero"),
            (
                {**base, "payment_type": "Paid in full", "plan": "", "amount": 99},
                r"Paid in full must be exactly \$100\.00",
            ),
            (
                {**base, "payment_type": "New plan", "plan": "3 year", "amount": 10},
                r"New 3 year plan payment must include the \$15\.00 deposit",
            ),
            ({**base, "interest_periods": 12}, r"Interest periods must be between 0 and 11"),
        ]

        for payment, message in cases:
            with self.subTest(payment=payment):
                try:
                    with self.assertRaisesRegex(ValueError, message):
                        build_membership_lines([payment])
                except (KeyError, TypeError) as exc:
                    self.fail(f"invalid input was not validated: {exc}")


    def test_generate_iif_writes_member_principal_and_admin_interest_lines(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        payment = {
            "member_name": "Tara Caruso",
            "member_number": "22206",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 8.45,
        }

        temp_dir = Path(__file__).parent / "_membership_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            try:
                iif_path = engine.generate_iif(
                    {}, {}, {}, date(2026, 8, 24),
                    bs_data={"subscription": 8.45},
                    membership_payments=[payment],
                )
            except TypeError as exc:
                self.fail(f"IIF membership integration is missing: {exc}")
            iif_text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        self.assertIn(
            "SPL\tDEPOSIT\t08/24/2026\t1260000 · Member Shares Receivable\t"
            "Tara Caruso\t-8.18\tShare Installments - Paid #22206\t",
            iif_text,
        )
        self.assertIn(
            "SPL\tDEPOSIT\t08/24/2026\t9104000 · Interest Income\t"
            "Tara Caruso\t-0.27\tShare Installments - Paid #22206\tAdmin",
            iif_text,
        )

    def test_generate_iif_keeps_legacy_coupon_receivable_process(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        temp_dir = Path(__file__).parent / "_legacy_coupon_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            iif_path = engine.generate_iif(
                {}, {}, {}, date(2026, 8, 26),
                bs_data={"vendor_coupon": 181.50},
            )
            iif_text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        self.assertIn(
            "1250000 · Coupons Receivable\t\t181.50\tNCG Coupons",
            iif_text,
        )
        self.assertIn(
            "1250000 · Coupons Receivable\t\t\tMFG Coupons",
            iif_text,
        )
        self.assertNotIn("Over/Short per Closeout Sheet - Coupon", iif_text)

    def test_iif_omits_empty_calculated_lines_but_keeps_manual_placeholders(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        temp_dir = Path(__file__).parent / "_clean_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            empty_path = engine.generate_iif(
                {}, {}, {}, date(2026, 8, 24),
                bs_data={},
            )
            empty_text = empty_path.read_text(encoding="utf-8")
            populated_path = engine.generate_iif(
                {},
                {
                    "8511002 · Discount 8% - Senior Day": 8.00,
                    "8512005 · Discount 8% - College Day": 4.00,
                },
                {},
                date(2026, 8, 26),
                bs_data={"donation": 5.00, "paid_out": 3.00},
                paid_in_total=12.00,
            )
            populated_text = populated_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        for conditional_text in (
            "Sales - Frozen Foods",
            "Discount 8% - Senior Day",
            "Discount 8% - College Day",
            "Outreach - Donations",
            "PAID IN:",
            "PAID OUT:",
        ):
            with self.subTest(conditional_text=conditional_text):
                self.assertNotIn(conditional_text, empty_text)

        for placeholder_text in (
            "MFG Coupons",
            "InHouse:",
            "Over/Short per Closeout Sheet",
            "Over/Short per POS (to = POS total)",
        ):
            with self.subTest(placeholder_text=placeholder_text):
                self.assertIn(placeholder_text, empty_text)

        for populated_line in (
            "Discount 8% - Senior Day\t\t8.00\tPdOut -",
            "Discount 8% - College Day\t\t4.00\tPdOut -",
            "Outreach - Donations\t\t5.00\t",
            "TBA Purchases\t\t-12.00\tPAID IN:",
            "TBA Purchases\t\t3.00\tPAID OUT:",
        ):
            with self.subTest(populated_line=populated_line):
                self.assertIn(populated_line, populated_text)

    def _generate_closeout_fixture(self, closeout_payload):
        import json
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        temp_dir = Path(__file__).parent / "_closeout_iif_output"
        temp_dir.mkdir(exist_ok=True)
        preview_path = temp_dir / "closeout_preview.json"
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            iif_path = engine.generate_iif(
                {},
                {},
                {},
                date(2026, 8, 28),
                bs_data={
                    "cash": 100.00,
                    "check": 50.00,
                    "donation": 20.00,
                    "charge": 10.00,
                    "offline_credit_card": -12.00,
                    "vendor_coupon": 181.50,
                    "paid_out": 40.00,
                },
                pass_through_total=7.00,
                paid_in_total=30.00,
                misc_tba_lines=[("Existing misc item", 3.00)],
                coupon_mode="closeout",
                coupon_closeout_total=188.25,
                coupon_ncg_total=152.25,
                coupon_mfg_total=36.00,
                closeout_payload=closeout_payload,
                closeout_preview_path=preview_path,
            )
            text = iif_path.read_text(encoding="utf-8")
            preview = (
                json.loads(preview_path.read_text(encoding="utf-8"))
                if preview_path.exists()
                else None
            )
            return text, preview
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

    def test_generate_iif_applies_closeout_actuals_memos_signs_and_order(self):
        payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {
                "cash": 110,
                "checks": 45,
                "donation": 25,
                "charge_house": 8,
                "offline_zon": 0,
                "vendor_coupons": 188.25,
                "paid_out": 50,
                "paid_in": 35,
            },
            "payroll": -4000,
            "safe": {"type": "shortage", "amount": 10},
            "plants_purchase": 20,
            "custom_tba": [
                {"memo": "Other paper item", "amount": 5, "direction": "adds"}
            ],
            "final_total": 5000,
            "approve_final_pos": True,
        }

        text, preview = self._generate_closeout_fixture(payload)

        expected_adjustments = [
            ("-10.00", "Over/Short per Closeout Sheet - Cash"),
            ("5.00", "Over/Short per Closeout Sheet - Check"),
            ("-5.00", "Over/Short per Closeout Sheet - Donation"),
            ("2.00", "Over/Short per Closeout Sheet - Charge (House)"),
            ("12.00", "Over/Short per Closeout Sheet - Offline Zon"),
            ("-6.75", "Over/Short per Closeout Sheet - Coupon"),
            ("-10.00", "Over/Short per Closeout Sheet - Paid Out"),
            ("5.00", "Over/Short per Closeout Sheet - Paid In"),
        ]
        adjustment_positions = []
        for amount, memo in expected_adjustments:
            with self.subTest(memo=memo):
                line = f"8314000 · FE - Cash Over/Shorts\t\t{amount}\t{memo}\tAdmin"
                self.assertIn(line, text)
                adjustment_positions.append(text.index(line))
        self.assertEqual(adjustment_positions, sorted(adjustment_positions))
        self.assertEqual(text.count("Over/Short per Closeout Sheet - Coupon"), 1)

        for detail_line in (
            "4160000 · Charitable Donations Payable\t\t-7.00\tCharity/Pass through Donations (Round up)",
            "8506000 · Outreach - Donations\t\t25.00\t",
            "4444 · TBA Purchases\t\t8.00\tInHouse:",
            "4444 · TBA Purchases\t\t-35.00\tPAID IN:",
            "4444 · TBA Purchases\t\t50.00\tPAID OUT:",
        ):
            with self.subTest(detail_line=detail_line):
                self.assertIn(detail_line, text)
        self.assertNotIn("Offline Credit Card:", text)

        iif_lines = text.splitlines()
        inhouse_position = next(
            index
            for index, line in enumerate(iif_lines)
            if "\t4444 · TBA Purchases\t\t8.00\tInHouse:\t" in line
        )
        inhouse_breakdown_rows = [
            line.split("\t")[3:8]
            for line in iif_lines[inhouse_position + 1:inhouse_position + 6]
        ]
        self.assertEqual(
            inhouse_breakdown_rows,
            [["4444 · TBA Purchases", "", "", "", ""]] * 5,
        )

        ordered_memos = [
            "Over/Short per Closeout Sheet - Paid In",
            "Payroll - Check Cashing",
            "Safe Shortage Cash Taken from Deposit",
            "Plants Dept - Market Purchases",
            "Over/Short per POS (to = POS total)",
            "Other paper item",
            "Existing misc item",
        ]
        memo_positions = [text.index(memo) for memo in ordered_memos]
        self.assertEqual(memo_positions, sorted(memo_positions))
        self.assertIn("1140000 · Cash Drawers/Safe\t\t4000.00", text)
        self.assertIn("1130000 · Petty Cash\t\t20.00", text)
        self.assertIn("4444 · TBA Purchases\t\t-5.00\tOther paper item", text)
        self.assertIn(
            "8314000 · FE - Cash Over/Shorts\t\t-9243.50\t"
            "Over/Short per POS (to = POS total)",
            text,
        )

        self.assertEqual(preview["provisional_total"], -4243.5)
        self.assertEqual(preview["final_total"], 5000.0)
        self.assertEqual(preview["remaining"], 9243.5)
        self.assertEqual(preview["remaining_after_approval"], 0.0)
        self.assertFalse(preview["requires_approval"])
        self.assertEqual(
            preview["final_pos_line"],
            {
                "kind": "final_pos",
                "account": "8314000 · FE - Cash Over/Shorts",
                "memo": "Over/Short per POS (to = POS total)",
                "qb_effect": 9243.5,
                "iif_amount": -9243.5,
            },
        )
        self.assertEqual([row["key"] for row in preview["standard_rows"]], [
            "cash",
            "checks",
            "donation",
            "charge_house",
            "offline_zon",
            "vendor_coupons",
            "paid_out",
            "paid_in",
        ])
        self.assertEqual(
            [row["kind"] for row in preview["misc_rows"]],
            ["payroll", "safe_shortage", "plants_purchase", "custom_tba"],
        )

    def test_generate_iif_manual_closeout_mode_is_byte_equivalent_to_legacy(self):
        legacy = self._generate_closeout_fixture(None)[0]
        manual = self._generate_closeout_fixture({"mode": "manual"})[0]

        self.assertEqual(manual, legacy)

    def test_generate_iif_writes_coupon_closeout_breakdown_and_signed_difference(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        temp_dir = Path(__file__).parent / "_coupon_closeout_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            try:
                positive_path = engine.generate_iif(
                    {}, {}, {}, date(2026, 8, 26),
                    bs_data={
                        "vendor_coupon": 181.50,
                        "visa_mc": 100.00,
                        "offline_credit_card": -44.07,
                    },
                    misc_tba_lines=[("Unique unmapped account", 12.34)],
                    settlement_data={"visa_mc": 101.00},
                    coupon_mode="closeout",
                    coupon_closeout_total=188.25,
                    coupon_ncg_total=152.25,
                    coupon_mfg_total=36.00,
                )
                positive_text = positive_path.read_text(encoding="utf-8")
                negative_path = engine.generate_iif(
                    {}, {}, {}, date(2026, 8, 27),
                    bs_data={"vendor_coupon": 181.50},
                    coupon_mode="closeout",
                    coupon_closeout_total=175.00,
                    coupon_ncg_total=150.00,
                    coupon_mfg_total=25.00,
                )
                negative_text = negative_path.read_text(encoding="utf-8")
            except TypeError as exc:
                self.fail(f"coupon closeout IIF integration is missing: {exc}")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        self.assertIn(
            "1250000 · Coupons Receivable\t\t152.25\tNCG Coupons",
            positive_text,
        )
        self.assertIn(
            "1250000 · Coupons Receivable\t\t36.00\tMFG Coupons",
            positive_text,
        )
        self.assertIn(
            "8314000 · FE - Cash Over/Shorts\t\t-6.75\t"
            "Over/Short per Closeout Sheet - Coupon\tAdmin",
            positive_text,
        )
        self.assertIn(
            "8314000 · FE - Cash Over/Shorts\t\t6.50\t"
            "Over/Short per Closeout Sheet - Coupon\tAdmin",
            negative_text,
        )
        card_adjustment_position = positive_text.index(
            "VISA/MC - Difference between First Data vs BS"
        )
        coupon_adjustment_position = positive_text.index(
            "Over/Short per Closeout Sheet - Coupon"
        )
        unique_tba_position = positive_text.index("Unique unmapped account")
        offline_tba_position = positive_text.index("Offline Credit Card:")
        self.assertLess(card_adjustment_position, coupon_adjustment_position)
        self.assertLess(coupon_adjustment_position, unique_tba_position)
        self.assertLess(unique_tba_position, offline_tba_position)

    def test_bs_penny_sign_is_preserved_and_offline_credit_is_bottom_tba(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        temp_dir = Path(__file__).parent / "_bs_sign_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            negative_path = engine.generate_iif(
                {}, {}, {}, date(2026, 8, 24),
                bs_data={"penny_round": -0.03, "offline_credit_card": -44.07},
            )
            negative_text = negative_path.read_text(encoding="utf-8")
            positive_path = engine.generate_iif(
                {}, {}, {}, date(2026, 8, 25),
                bs_data={"penny_round": 0.03},
            )
            positive_text = positive_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        self.assertIn(
            "9107000 · Miscellaneous Income\t\t0.03\t"
            "Penny Round Up for Cash Transactions",
            negative_text,
        )
        self.assertIn(
            "9107000 · Miscellaneous Income\t\t-0.03\t"
            "Penny Round Up for Cash Transactions",
            positive_text,
        )
        offline_line = (
            "4444 · TBA Purchases\t\t44.07\tOffline Credit Card:"
        )
        self.assertIn(offline_line, negative_text)
        generated_splits = [
            line for line in negative_text.splitlines()
            if line.startswith("SPL\t")
        ]
        self.assertIn(
            offline_line,
            generated_splits[-1],
        )

    def test_parse_bs_maps_offline_credit_card_as_a_negative_unique_item(self):
        from datetime import date
        from pathlib import Path

        import openpyxl

        from app import pos_to_quickbooks_v2 as engine

        workbook_path = Path(__file__).parent / "_offline_credit_bs.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "082426 BS"
        sheet.append([1334, "Dwr Offline Credit card", None, None, 44.07, None, "D"])
        workbook.save(workbook_path)
        workbook.close()
        try:
            parsed = engine.parse_bs_sheet(workbook_path, date(2026, 8, 24))
        finally:
            workbook_path.unlink(missing_ok=True)

        self.assertEqual(parsed["offline_credit_card"], -44.07)

    def test_generate_iif_keeps_multiple_members_and_new_plan_offsets_separate(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        payments = [
            {
                "member_name": "Paid Member",
                "member_number": "11111",
                "payment_type": "Paid in full",
                "plan": "",
                "amount": 100.00,
            },
            {
                "member_name": "New Member",
                "member_number": "22222",
                "payment_type": "New plan",
                "plan": "5 year",
                "amount": 10.00,
            },
        ]

        temp_dir = Path(__file__).parent / "_multiple_members_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            iif_path = engine.generate_iif(
                {}, {}, {}, date(2026, 8, 24),
                bs_data={"subscription": 110.00},
                membership_payments=payments,
            )
            iif_text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        self.assertIn(
            "6100000 · Member Shares (Paid-In Equity)\tPaid Member\t-100.00\t"
            "Member Shares - Paid",
            iif_text,
        )
        self.assertIn(
            "6100000 · Member Shares (Paid-In Equity)\tNew Member\t-100.00\t"
            "Member Shares - Receivable",
            iif_text,
        )
        self.assertIn(
            "1260000 · Member Shares Receivable\tNew Member\t100.00\t"
            "Member Shares - Receivable",
            iif_text,
        )
        self.assertIn(
            "1260000 · Member Shares Receivable\tNew Member\t-10.00\t"
            "Share Installments - Paid #22222",
            iif_text,
        )

    def test_generate_iif_manual_mode_uses_one_unnamed_equity_line(self):
        from datetime import date
        from pathlib import Path

        from app import pos_to_quickbooks_v2 as engine

        temp_dir = Path(__file__).parent / "_manual_membership_iif_output"
        temp_dir.mkdir(exist_ok=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = temp_dir
        engine.LOG_DIR = temp_dir
        engine.log.disabled = True
        try:
            try:
                iif_path = engine.generate_iif(
                    {}, {}, {}, date(2026, 8, 24),
                    bs_data={"subscription": 8.45},
                    membership_payments=[],
                    membership_mode="manual",
                )
            except TypeError as exc:
                self.fail(f"manual membership IIF mode is missing: {exc}")
            iif_text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            for generated_file in temp_dir.iterdir():
                generated_file.unlink()
            temp_dir.rmdir()

        self.assertIn(
            "6100000 · Member Shares (Paid-In Equity)\t\t-8.45\tMember Shares - Paid",
            iif_text,
        )
        self.assertNotIn("1260000 · Member Shares Receivable", iif_text)
        self.assertNotIn("9104000 · Interest Income", iif_text)


    def test_membership_payment_file_loads_manual_app_rows(self):
        import json
        from pathlib import Path

        try:
            from app.membership_payments import load_membership_payments_file
        except ImportError as exc:
            self.fail(f"membership payment file loader is missing: {exc}")

        expected = [{
            "member_name": "Tara Caruso",
            "member_number": "22206",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 8.45,
            "interest_periods": None,
        }]
        path = Path(__file__).parent / "_membership_rows.json"
        path.write_text(json.dumps(expected), encoding="utf-8")
        try:
            self.assertEqual(load_membership_payments_file(path), expected)
        finally:
            path.unlink(missing_ok=True)


    def test_membership_payment_file_must_contain_a_list(self):
        from pathlib import Path

        from app.membership_payments import load_membership_payments_file

        path = Path(__file__).parent / "_invalid_membership_rows.json"
        path.write_text('{"member_name": "not a list"}', encoding="utf-8")
        try:
            with self.assertRaisesRegex(ValueError, "must contain a list"):
                load_membership_payments_file(path)
        finally:
            path.unlink(missing_ok=True)


    def test_subscription_total_is_read_from_balance_sheet_code_3420(self):
        from io import BytesIO

        import openpyxl

        try:
            from app.membership_payments import read_subscription_total
        except ImportError as exc:
            self.fail(f"subscription total reader is missing: {exc}")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "082426 BS"
        sheet.append([3420, "Subscription Revenue", None, None, -28.45])
        content = BytesIO()
        workbook.save(content)

        self.assertEqual(read_subscription_total(content.getvalue(), "082426 BS"), 28.45)

    def test_subscription_total_rejects_missing_or_malformed_balance_sheet_data(self):
        from io import BytesIO

        import openpyxl

        from app.membership_payments import read_subscription_total

        missing_bs_workbook = openpyxl.Workbook()
        missing_bs_workbook.active.title = "Sales"
        missing_bs_content = BytesIO()
        missing_bs_workbook.save(missing_bs_content)

        with self.assertRaisesRegex(ValueError, "Balance Sheet"):
            read_subscription_total(missing_bs_content.getvalue())

        for malformed_amount in ("not a dollar amount", "NaN", "Infinity", "-Infinity"):
            with self.subTest(malformed_amount=malformed_amount):
                malformed_workbook = openpyxl.Workbook()
                malformed_sheet = malformed_workbook.active
                malformed_sheet.title = "082426 BS"
                malformed_sheet.append(
                    [3420, "Subscription Revenue", None, None, malformed_amount]
                )
                malformed_content = BytesIO()
                malformed_workbook.save(malformed_content)

                with self.assertRaisesRegex(ValueError, "3420"):
                    read_subscription_total(malformed_content.getvalue(), "082426 BS")

    def test_valid_balance_sheet_without_3420_means_no_subscription_activity(self):
        from io import BytesIO

        import openpyxl

        from app.membership_payments import read_subscription_total

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "082426 BS"
        sheet.append([901, "Cash", None, None, 250.00])
        content = BytesIO()
        workbook.save(content)

        self.assertEqual(read_subscription_total(content.getvalue(), "082426 BS"), 0.0)


    def test_interest_override_cannot_exceed_automatic_period_count(self):
        from app.membership_payments import build_membership_lines

        payment = {
            "member_name": "A Member",
            "member_number": "80001",
            "payment_type": "Existing plan",
            "plan": "1 year",
            "amount": 8.45,
            "interest_periods": 2,
        }
        with self.assertRaisesRegex(ValueError, "automatic count of 1"):
            build_membership_lines([payment])


    def test_membership_payment_files_are_unique_per_app_run(self):
        from pathlib import Path

        try:
            from app.membership_payments import (
                load_membership_payments_file,
                write_membership_payments_file,
            )
        except ImportError as exc:
            self.fail(f"unique membership payment writer is missing: {exc}")

        folder = Path(__file__).parent / "_membership_json_output"
        folder.mkdir(exist_ok=True)
        payments = [{"member_name": "A Member", "amount": 8.45}]
        paths = []
        try:
            paths = [
                write_membership_payments_file(folder, payments),
                write_membership_payments_file(folder, payments),
            ]
            self.assertNotEqual(paths[0], paths[1])
            self.assertEqual(load_membership_payments_file(paths[0]), payments)
            self.assertEqual(load_membership_payments_file(paths[1]), payments)
        finally:
            for path in paths:
                path.unlink(missing_ok=True)
            folder.rmdir()


    def test_membership_editor_key_changes_for_a_different_workbook(self):
        try:
            from app.membership_payments import membership_editor_key
        except ImportError as exc:
            self.fail(f"membership editor key helper is missing: {exc}")

        first = membership_editor_key(b"workbook one", 0)
        same = membership_editor_key(b"workbook one", 0)
        different = membership_editor_key(b"workbook two", 0)
        reset = membership_editor_key(b"workbook one", 1)

        self.assertEqual(first, same)
        self.assertNotEqual(first, different)
        self.assertNotEqual(first, reset)


if __name__ == "__main__":
    unittest.main()
