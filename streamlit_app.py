"""
HWFC Daily Deposit - Streamlit UI v3 · SIDEBAR HISTORY BUILD 2026-08-25
Honest Weight Food Co-op

Drop-in replacement for streamlit_app.py.

Goals:
- Warm Honest Weight / grocery co-op visual identity
- Keep the existing deposit calculation engine as the source of truth
- Show a full post-run reconciliation:
  * Sales
  * Discounts
  * Balance Sheet / tender lines
  * HASH checks
  * IIF debit / credit balance
  * Full QuickBooks IIF preview
"""

from __future__ import annotations

import html
import json
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Optional
from uuid import uuid4

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from app.activity_breakdowns import (
    activity_actuals,
    activity_closeout_ready,
    activity_workflow_keys,
    append_activity_entry,
    normalize_activity_payload,
    normalize_activity_section,
    read_activity_source_totals,
    write_activity_payload_file,
)
from app.coupon_reconciliation import (
    read_coupon_receivable_total,
    reconcile_coupon_receivable,
)
from app.deposit_workflow import (
    STEP_CLOSEOUT,
    STEP_COUPONS,
    STEP_MEMBER_SHARES,
    active_deposit_step,
    complete_deposit_step,
    deposit_workflow_complete,
    deposit_step_rows,
    edit_deposit_step,
    normalize_step_completions,
    required_deposit_steps,
)
from app.closeout_reconciliation import (
    STANDARD_CLOSEOUT_ORDER,
    STANDARD_METADATA,
    build_closeout_form_payload,
    closeout_input_fingerprint,
    closeout_preview_is_fresh,
    default_closeout_actuals,
    normalize_closeout_payload,
    read_closeout_baselines,
    write_closeout_payload_file,
)
from app.membership_payments import (
    PAYMENT_OPTIONS,
    apply_membership_amount_option_state,
    apply_quickbooks_name_option_state,
    build_membership_lines,
    exclusive_run_lock,
    membership_editor_key,
    membership_mode_from_choice,
    membership_payment_from_entry,
    plan_reference_rows,
    read_subscription_total,
    remove_membership_payment,
    subscription_action_status,
    write_membership_payments_file,
)
from app.guided_deposit_state import (
    hydrate_reopened_closeout_state,
    reopen_step_for_edit,
    recover_completed_membership_state,
    resolve_activity_detection_workflow,
    save_activity_transition,
    save_manual_member_share_transition,
    save_member_share_transition,
)
from app.guided_step_ui import (
    closeout_review_blockers,
    queue_breakdown_scroll,
    queue_continue_scroll,
    render_card_settlement_verification,
    render_breakdown_scroll_target,
    render_deposit_step_panels,
    render_prepare_iif_action,
)
from app.ui_helpers import (
    deposit_download_details,
    plan_guide_html,
)

# ---------------------------------------------------------------------
# Self-contained UI helpers and SOP content
# ---------------------------------------------------------------------

EASTERN_TZ = ZoneInfo("America/New_York")

def format_history_run_time(value: str, include_date: bool = False) -> str:
    """Display stored Run History timestamps in HWFC local Eastern Time.

    Existing history timestamps are stored without a timezone by the Streamlit
    server. Treat those naive values as UTC, then convert only the display to
    America/New_York. Timezone-aware timestamps are converted directly.
    """
    try:
        dt = datetime.fromisoformat(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        local_dt = dt.astimezone(EASTERN_TZ)
        return local_dt.strftime("%m/%d/%Y %I:%M %p" if include_date else "%I:%M %p")
    except Exception:
        return "—"

def build_history_option_label(record: dict) -> str:
    """Return a compact label for the Run History select box."""
    try:
        report_label = datetime.fromisoformat(record.get("report_date", "")).strftime("%m/%d/%Y")
    except Exception:
        report_label = record.get("report_date", "—") or "—"

    run_label = format_history_run_time(record.get("run_at", ""))

    status = record.get("status", "—") or "—"
    return f"{report_label} · {status} · {run_label}"


SOP_STEPS = [
    {
        "title": "Step 1 · Open the Daily Deposit template",
        "body": r"""
Open the approved Excel template from:

`S:\Finance & Payroll Forms\Finance Work Files\Daily Deposit`

File name: **TEMPLATE - SubDept Single Total Report**

Use this workbook as the master file for the deposit you are preparing. Do not build the deposit from a blank workbook.
""".strip(),
    },
    {
        "title": "Step 2 · Import daily sales from SMS",
        "body": """
In **SMS (POS System)** go to:

**Reports → Sub-department Single Total**

1. Change the report date to the date of the deposit you are preparing.
2. Leave the **Sub-department Range** as-is.
3. Set the **Totalizer Range** to:
   - Start: **3 (Net Sales)**
   - End: **4 (Interstore Sales)**
4. Select **Launch**.
5. Export the report to Excel.
6. In the exported Excel report, copy the sales data beginning with the first sub-department code and continuing through the final department shown for the day. In the example below, this is the highlighted report area from **Columns B–H**.
7. Open the Daily Deposit template and go to the **SubDept Single** tab.
8. Paste the copied sales data starting in **cell A1**.
9. Confirm that the pasted table begins with the same first row as the exported report and that all rows through the final department were included.

Do not paste over template formulas outside of the intended data-entry area.
""".strip(),
    },
    {
        "title": "Step 2a · Import store coupons from SMS",
        "body": """
In **SMS (POS System)** go to:

**Reports → Sub-department Single Total**

1. Change the report date to the date of the deposit.
2. Leave the **Sub-department Range** as-is.
3. Set the **Totalizer Range** to:
   - Start: **3542 (Elect. Store Coupon Distribution)**
   - End: **3542 (Elect. Store Coupon Distribution)**
4. Select **Launch**.
5. Export the report to Excel.
6. Copy **Columns B–H**, starting with the sub-department codes and ending with Quantity.
7. Paste those columns into **SubDept Coupon (Local Discount)** in the Daily Deposit template.

### Sales check
After Steps 2 and 2a, the **SubDept Sales Report** should be complete. Confirm that the green-filled **Sales Total** in the template matches the total at the bottom of the SMS **Sub-department Single Total** report **exactly**. If the totals do not match, stop and correct the workbook before continuing.
""".strip(),
    },
    {
        "title": "Step 3 · Enter Milk Bottle Returns",
        "body": """
In **SMS (POS System)** go to:

**Reports → SubDepartments → Items → Item Multi Totals by Sub-department**

1. Change the report date to the corresponding deposit date.
2. Set the **Sub Department Range** to:
   - Start: **27 (Store Coupons)**
   - End: **27 (Store Coupons)**
3. Select **Launch**.
4. Scroll through the report and find all Milk Bottle Return items, such as:
   - **MILK BOTTLE 1 - RETURNED**
   - **MILK BOTTLE 2 - RETURNED**
   - any additional **MILK BOTTLE ... - RETURNED** lines for that day
5. Add together all Milk Bottle Returns for the day.
6. Enter the total into cell **M1** on **SubDept Sales Report**.
""".strip(),
    },
    {
        "title": "Step 4 · Import the Discounts worksheet",
        "body": """
In **SMS (POS System)** go to:

**Reports → Customers → Totals → Sub-department Discounts by Shopper**

1. Change the report date to the deposit date.
2. Select **Launch**.
3. Export the report to Excel.
4. In Excel, right-click the exported worksheet tab and choose **Move or Copy**.
5. Move or copy the entire worksheet into the Daily Deposit template workbook.
6. Place it in the **`XXXXXX Discounts`** position/tab.
7. After the worksheet has been copied into the template, rename the tab so `XXXXXX` matches the Daily Deposit date you are working on.

Example: for an August 21, 2026 deposit, rename the tab **`082126 Discounts`**.
""".strip(),
    },
    {
        "title": "Step 5 · Import the HASH worksheet",
        "body": """
The HASH report supplies **Refunded Discounts, Pass Through/Charity Donations, and Paid-Ins**.

In **SMS (POS System)** go to:

**Reports → Sub-department Single Total**

1. Change the report date to the deposit date.
2. Set the **Totalizer Range** to:
   - Start: **6 (HASH Sales)**
   - End: **6 (HASH Sales)**
3. Select **Launch**.
4. Export the report to Excel.
5. In Excel, right-click the exported worksheet tab and choose **Move or Copy**.
6. Move or copy the entire worksheet into the Daily Deposit template workbook.
7. Place it in the **`XXXXXX HASH`** position/tab.
8. After the worksheet has been copied into the template, rename the tab so `XXXXXX` matches the Daily Deposit date you are working on.

Example: for an August 21, 2026 deposit, rename the tab **`082126 HASH`**.
""".strip(),
    },
    {
        "title": "Step 6 · Import the Balance Sheet (BS)",
        "body": """
In **SMS (POS System)** go to:

**Reports → Balance Sheet**

1. Change the report date to the deposit date.
2. Select **Launch**.
3. Export the report to Excel.
4. In Excel, right-click the exported worksheet tab and choose **Move or Copy**.
5. Move or copy the entire worksheet into the Daily Deposit template workbook.
6. Place it in the **`XXXXXX BS`** position/tab.
7. After the worksheet has been copied into the template, rename the tab so `XXXXXX` matches the Daily Deposit date you are working on.

Example: for an August 21, 2026 deposit, rename the tab **`082126 BS`**.
""".strip(),
    },
    {
        "title": "Step 7 · Final Workbook Check",
        "body": """
Before uploading the workbook to the Daily Deposit app, confirm all of the following:

- **SubDept Sales Report** is complete.
- **SubDept Coupon (Local Discount)** is complete.
- The green **Sales Total** matches the SMS Sub-department Single Total report exactly.
- Milk Bottle Returns are entered in **M1**.
- The dated **Discounts** tab is present.
- The dated **HASH** tab is present.
- The dated **BS** tab is present.
- All tabs are for the same deposit date.

Then compare your workbook against the approved example shown below. The workbook should follow the same tab structure before you run the deposit automation.
""".strip(),
    },
]

# ---------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent

def resolve_engine_path() -> Path:
    """Find the deposit engine in the repo root, including known Codespaces filenames."""
    candidates = [
        ROOT / "app" / "pos_to_quickbooks_v2.py",
        ROOT / "pos_to_quickbooks_v2.py",
        ROOT / "pos_to_quickbooks_v2_CODESPACES_CONTENT_DETECTION.py",
        ROOT / "pos_to_quickbooks_v2_CODESPACES_HASH_FINAL_FIX.py",
        ROOT / "pos_to_quickbooks_v2_CODESPACES_ROLLBACK_STABLE.py",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]

ENGINE_PATH = resolve_engine_path()
INPUT_DIR = ROOT / "input" / "daily_reports"
QB_IMPORT_DIR = ROOT / "output" / "qb_imports"
LOG_DIR = ROOT / "logs"
HISTORY_DIR = ROOT / "output" / "history"
HISTORY_FILE = HISTORY_DIR / "run_history.json"
HISTORY_UPLOAD_DIR = HISTORY_DIR / "uploads"
HISTORY_IIF_DIR = HISTORY_DIR / "iif"
RUNTIME_TEMP_DIR = Path(tempfile.gettempdir()) / "hwfc_daily_deposit"
RUN_LOCK_PATH = RUNTIME_TEMP_DIR / "deposit_run.lock"

for folder in (
    INPUT_DIR,
    QB_IMPORT_DIR,
    LOG_DIR,
    HISTORY_DIR,
    HISTORY_UPLOAD_DIR,
    HISTORY_IIF_DIR,
    RUNTIME_TEMP_DIR,
):
    folder.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------
# Page / brand
# ---------------------------------------------------------------------

st.set_page_config(
    page_title="HWFC Daily Deposit",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    :root {
        --hwfc-cream: #0E1117;
        --hwfc-paper: #161B22;
        --hwfc-forest: #315F3A;
        --hwfc-leaf: #78A85B;
        --hwfc-sage: #243126;
        --hwfc-gold: #C7952B;
        --hwfc-clay: #A95F3B;
        --hwfc-brown: #D6C5A8;
        --hwfc-ink: #F4F1E8;
        --hwfc-muted: #A7B0A3;
        --hwfc-green-soft: #16251A;
        --hwfc-red-soft: #2A1717;
        --hwfc-yellow-soft: #2B2515;
        --hwfc-border: #30363D;
    }

    .stApp {
        background:
            radial-gradient(circle at 8% 4%, rgba(120,168,91,.10), transparent 22rem),
            linear-gradient(180deg, #0E1117 0%, #0B0F14 100%);
        color: var(--hwfc-ink);
    }

    .block-container {
        max-width: 1180px;
        padding-top: 2rem;
        padding-bottom: 4rem;
    }

    h1, h2, h3 {
        color: #F4F1E8;
        letter-spacing: -0.02em;
    }

    .hwfc-hero {
        background:
            linear-gradient(120deg, rgba(47,82,51,.97), rgba(68,101,57,.96)),
            #2F5233;
        border: 1px solid rgba(255,255,255,.10);
        border-radius: 22px;
        padding: 26px 30px;
        box-shadow: 0 14px 34px rgba(54, 65, 43, .14);
        margin-bottom: 18px;
        position: relative;
        overflow: hidden;
    }

    .hwfc-hero:after {
        content: "🌿";
        position: absolute;
        right: 28px;
        top: 8px;
        font-size: 82px;
        opacity: .11;
        transform: rotate(-12deg);
    }

    .hwfc-kicker {
        color: #DDE7D5;
        font-size: .78rem;
        font-weight: 800;
        letter-spacing: .16em;
        text-transform: uppercase;
        margin-bottom: 6px;
    }

    .hwfc-title {
        color: #FFFDF8;
        font-family: Georgia, 'Times New Roman', serif;
        font-size: 2.15rem;
        font-weight: 700;
        line-height: 1.08;
        margin: 0;
    }

    .hwfc-subtitle {
        color: #E9EEE3;
        margin-top: 8px;
        font-size: .98rem;
    }

    .hwfc-stepbar {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 8px;
        margin: 16px 0 22px;
    }

    .hwfc-step {
        border: 1px solid var(--hwfc-border);
        border-radius: 999px;
        background: #161B22;
        padding: 9px 12px;
        text-align: center;
        color: #D6C5A8;
        font-size: .82rem;
        font-weight: 700;
    }

    .hwfc-step.active {
        color: white;
        background: var(--hwfc-forest);
        border-color: var(--hwfc-forest);
    }

    .hwfc-section-label {
        font-size: .75rem;
        color: var(--hwfc-muted);
        font-weight: 800;
        letter-spacing: .12em;
        text-transform: uppercase;
        margin: 8px 0 6px;
    }

    .hwfc-workflow-heading {
        background: linear-gradient(90deg, rgba(47,82,51,.48), rgba(22,27,34,.94));
        border: 1px solid #315F3A;
        border-left: 5px solid #78A85B;
        border-radius: 14px;
        padding: 15px 18px;
        margin: 28px 0 18px;
    }

    .hwfc-workflow-heading-title {
        color: #FFFDF8;
        font-size: 1.45rem;
        font-weight: 850;
        line-height: 1.2;
    }

    .hwfc-workflow-heading-sub {
        color: #C7D2C2;
        font-size: .92rem;
        margin-top: 4px;
    }

    .hwfc-deposit-stepper {
        display: flex;
        align-items: flex-start;
        overflow-x: auto;
        padding: 8px 2px 14px;
        margin: 2px 0 0;
        scrollbar-width: thin;
    }

    .hwfc-stepper-item {
        flex: 1 0 130px;
        min-width: 130px;
        position: relative;
        text-align: center;
    }

    .hwfc-stepper-track {
        align-items: center;
        display: flex;
        height: 38px;
        justify-content: center;
        position: relative;
    }

    .hwfc-stepper-item:not(:last-child) .hwfc-stepper-track::after {
        background: #38414C;
        content: "";
        height: 3px;
        left: calc(50% + 17px);
        position: absolute;
        right: calc(-50% + 17px);
        top: calc(50% - 1px);
    }

    .hwfc-stepper-item.is-complete .hwfc-stepper-track::after {
        background: #78A85B;
    }

    .hwfc-stepper-bubble {
        align-items: center;
        background: #161B22;
        border: 2px solid #56606C;
        border-radius: 50%;
        color: #C7D2C2;
        display: inline-flex;
        font-size: .88rem;
        font-weight: 850;
        height: 34px;
        justify-content: center;
        position: relative;
        width: 34px;
        z-index: 1;
    }

    .hwfc-stepper-item.is-complete .hwfc-stepper-bubble {
        background: #315F3A;
        border-color: #78A85B;
        color: #FFFDF8;
    }

    .hwfc-stepper-item.is-current .hwfc-stepper-bubble {
        background: #78A85B;
        border-color: #A9D18E;
        box-shadow: 0 0 0 5px rgba(120,168,91,.18);
        color: #0F1410;
    }

    .hwfc-stepper-label {
        display: flex;
        flex-direction: column;
        gap: 2px;
        margin: 7px auto 0;
        max-width: 150px;
    }

    .hwfc-stepper-label span {
        color: #D6C5A8;
        font-size: .68rem;
        font-weight: 850;
        letter-spacing: .08em;
        text-transform: uppercase;
    }

    .hwfc-stepper-label strong {
        color: #FFFDF8;
        font-size: .86rem;
        line-height: 1.2;
    }

    .hwfc-stepper-item.is-pending .hwfc-stepper-label strong {
        color: #87919D;
    }

    @media (max-width: 700px) {
        .hwfc-stepper-item {
            flex-basis: 112px;
            min-width: 112px;
        }

        .hwfc-stepper-label strong {
            font-size: .78rem;
        }
    }

    .hwfc-plan-guide-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
        gap: 12px;
        margin: 10px 0 18px;
    }

    .hwfc-plan-card {
        background: #161B22;
        border: 1px solid var(--hwfc-border);
        border-top: 4px solid #78A85B;
        border-radius: 14px;
        padding: 14px 16px 10px;
        box-shadow: 0 6px 18px rgba(0, 0, 0, .16);
    }

    .hwfc-plan-card-title {
        color: #FFFDF8;
        font-size: 1.08rem;
        font-weight: 850;
        margin-bottom: 8px;
    }

    .hwfc-plan-card-row {
        display: flex;
        justify-content: space-between;
        gap: 14px;
        padding: 6px 0;
        border-bottom: 1px solid rgba(255,255,255,.07);
        color: #B8C2B5;
        font-size: .88rem;
    }

    .hwfc-plan-card-row:last-child {
        border-bottom: 0;
    }

    .hwfc-plan-card-row strong {
        color: #F4F1E8;
        white-space: nowrap;
    }

    .hwfc-result {
        border-radius: 18px;
        padding: 18px 20px;
        margin: 8px 0 18px;
        border: 1px solid;
    }

    .hwfc-result.good {
        background: var(--hwfc-green-soft);
        border-color: #315F3A;
    }

    .hwfc-result.warn {
        background: var(--hwfc-yellow-soft);
        border-color: #8D7427;
    }

    .hwfc-result.bad {
        background: var(--hwfc-red-soft);
        border-color: #8D3A33;
    }

    .hwfc-result-title {
        font-size: 1.18rem;
        font-weight: 850;
        margin-bottom: 2px;
        color: #EAF4E7;
    }

    .hwfc-result-sub {
        color: var(--hwfc-muted);
        font-size: .92rem;
    }

    .hwfc-card {
        background: #161B22;
        border: 1px solid var(--hwfc-border);
        border-radius: 16px;
        padding: 16px 18px;
        min-height: 118px;
        box-shadow: 0 8px 22px rgba(0, 0, 0, .18);
    }

    .hwfc-card-label {
        font-size: .72rem;
        font-weight: 800;
        letter-spacing: .11em;
        text-transform: uppercase;
        color: var(--hwfc-muted);
        margin-bottom: 8px;
    }

    .hwfc-card-value {
        font-family: Georgia, 'Times New Roman', serif;
        font-size: 1.58rem;
        font-weight: 700;
        color: #EAF4E7;
        line-height: 1.05;
    }

    .hwfc-card-foot {
        color: var(--hwfc-muted);
        font-size: .80rem;
        margin-top: 7px;
    }

    .hwfc-match {
        font-weight: 800;
        color: #3E713D;
    }

    .hwfc-mismatch {
        font-weight: 800;
        color: #A34A32;
    }

    .hwfc-equation {
        padding: 16px 18px;
        border-radius: 14px;
        background: #161B22;
        border: 1px solid #30363D;
        font-family: Georgia, 'Times New Roman', serif;
        color: #D6C5A8;
        margin: 10px 0 14px;
    }

    .hwfc-footer {
        margin-top: 32px;
        padding-top: 14px;
        border-top: 1px solid var(--hwfc-border);
        color: var(--hwfc-muted);
        font-size: .8rem;
    }

    div[data-testid="stFileUploader"] {
        background: #161B22;
        border: 1px dashed #49604B;
        border-radius: 14px;
        padding: 4px 12px;
    }

    div[data-testid="stDateInput"] > div,
    div[data-testid="stFileUploader"] section {
        border-radius: 12px;
    }

    div[data-baseweb="input"],
    div[data-baseweb="select"],
    div[data-testid="stFileUploader"] section,
    div[data-testid="stExpander"],
    div[data-testid="stExpander"] details {
        background: #161B22 !important;
        color: #F4F1E8 !important;
        border-color: #30363D !important;
    }

    label, p, span, .stMarkdown, .stCaption {
        color: #E8ECE7;
    }

    div[data-testid="stAppViewContainer"],
    div[data-testid="stHeader"],
    section[data-testid="stSidebar"] {
        background-color: #0E1117 !important;
    }

    div[data-testid="stFileUploaderDropzone"],
    div[data-testid="stFileUploaderFile"],
    div[data-testid="stMetric"],
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #161B22 !important;
        color: #F4F1E8 !important;
        border-color: #30363D !important;
    }

    div[data-testid="stAlert"] {
        color: #F4F1E8;
    }

    .stButton > button,
    .stDownloadButton > button {
        border-radius: 12px !important;
        font-weight: 800 !important;
        min-height: 44px;
    }

    .stButton > button[kind="primary"] {
        background: var(--hwfc-forest) !important;
        border-color: var(--hwfc-forest) !important;
        color: white !important;
    }

    .stDownloadButton > button {
        background: var(--hwfc-forest) !important;
        border-color: var(--hwfc-forest) !important;
        color: white !important;
    }

    div[data-testid="stDataFrame"] {
        border: 1px solid var(--hwfc-border);
        border-radius: 12px;
        overflow: hidden;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
    }

    .stTabs [data-baseweb="tab"] {
        border-radius: 10px 10px 0 0;
        padding-left: 14px;
        padding-right: 14px;
        font-weight: 750;
    }

    .hwfc-mini-card {
        background: #161B22;
        border: 1px solid var(--hwfc-border);
        border-radius: 12px;
        padding: 12px 14px;
        min-height: 74px;
    }
    .hwfc-mini-label {
        color: var(--hwfc-muted);
        font-size: .68rem;
        font-weight: 800;
        letter-spacing: .09em;
        text-transform: uppercase;
        margin-bottom: 5px;
    }
    .hwfc-mini-value {
        color: #F4F1E8;
        font-size: .95rem;
        font-weight: 750;
        line-height: 1.25;
        word-break: break-word;
    }
    .hwfc-check-card {
        background: #14251B;
        border: 1px solid #315F3A;
        border-radius: 11px;
        padding: 10px 11px;
        min-height: 74px;
    }
    .hwfc-check-title {
        color: #EAF4E7;
        font-size: .84rem;
        font-weight: 800;
        margin-bottom: 5px;
    }
    .hwfc-check-sheet {
        color: #C4CFC2;
        font-size: .75rem;
        line-height: 1.25;
        word-break: break-word;
    }
    .hwfc-source-strip {
        background: #151F18;
        border: 1px solid #315F3A;
        border-radius: 10px;
        padding: 9px 12px;
        color: #DDE7D5;
        font-size: .82rem;
        font-weight: 650;
    }

    .hwfc-tip-card {
        background: #161B22;
        border: 1px solid #30363D;
        border-left: 3px solid #49604B;
        border-radius: 10px;
        padding: 13px 15px;
        margin: 9px 0;
    }
    .hwfc-tip-card.attention { border-left-color: #C7952B; }
    .hwfc-tip-card.info { border-left-color: #5E7F93; }
    .hwfc-tip-card.success { border-left-color: #78A85B; }
    .hwfc-tip-title {
        color: #F4F1E8;
        font-weight: 800;
        font-size: .92rem;
        margin-bottom: 5px;
    }
    .hwfc-tip-body {
        color: #D9DFD7;
        font-size: .86rem;
        line-height: 1.5;
    }
    .hwfc-tip-body strong { color: #F4F1E8; }
    .hwfc-tip-example {
        margin-top: 7px;
        color: #DDE7D5;
        font-weight: 750;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

MONEY_RE = r"-?\$?\s*[\(\-]?\s*[\d,]+(?:\.\d{1,2})?\s*\)?"

DISCOUNT_PREFIXES = (
    "8511",
    "8512",
    "8140",
    "8423",
)

TENDER_KEYWORDS = (
    "cash",
    "check",
    "visa",
    "master",
    "amex",
    "american express",
    "discover",
    "debit",
    "ebt",
    "food stamp",
    "snap",
    "gift card",
    "credit card",
)

BALANCE_SHEET_KEYWORDS = (
    "sales tax",
    "bottle",
    "charity",
    "donation",
    "prepaid",
    "coupon",
    "member share",
    "receivable",
    "paid in",
    "paid-in",
    "cash over",
    "cash short",
    "penny",
    "nickel",
    "dufb",
    "double up",
    "outreach",
    "paid out",
    "paid-out",
)

def money(value: Optional[float]) -> str:
    if value is None:
        return "—"
    value = float(value)
    if value < 0:
        return f"(${abs(value):,.2f})"
    return f"${value:,.2f}"

def abs_money(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"${abs(float(value)):,.2f}"

def parse_money(text: str) -> Optional[float]:
    if not text:
        return None
    s = str(text).strip().replace("$", "").replace(",", "").replace(" ", "")
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    return -abs(v) if neg else v

def last_amount_after_label(log_text: str, labels: list[str]) -> Optional[float]:
    for label in labels:
        matches = re.findall(
            rf"{re.escape(label)}\s*:?\s*({MONEY_RE})",
            log_text,
            flags=re.IGNORECASE,
        )
        if matches:
            return parse_money(matches[-1])
    return None

def section_status(log_text: str, section_name: str) -> Optional[bool]:
    upper = log_text.upper()
    start = upper.rfind(section_name.upper())
    if start < 0:
        return None
    chunk = upper[start : start + 1600]
    if "MISMATCH" in chunk or "FAILED" in chunk or "ERROR" in chunk:
        return False
    if "MATCH" in chunk or "OK TO IMPORT" in chunk:
        return True
    return None

def detect_sheet_roles(
    upload_bytes: bytes,
    preferred_date=None,
) -> dict[str, Optional[str]]:
    try:
        from io import BytesIO
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(upload_bytes), read_only=True, data_only=True)
        detected = {"sales": None, "coupons": None, "discounts": None, "bs": None, "hash": None}
        preferred_markers = set()
        if preferred_date is not None:
            preferred_markers = {
                preferred_date.strftime(pattern).casefold()
                for pattern in (
                    "%m%d%y",
                    "%m%d%Y",
                    "%m-%d-%y",
                    "%m-%d-%Y",
                    "%m_%d_%y",
                    "%m_%d_%Y",
                )
            }
        ordered_sheet_names = sorted(
            wb.sheetnames,
            key=lambda name: (
                0
                if preferred_markers
                and any(marker in name.casefold() for marker in preferred_markers)
                else 1,
                wb.sheetnames.index(name),
            ),
        )

        for name in ordered_sheet_names:
            low = name.strip().lower()
            if "xxxxxx" in low:
                continue
            if detected["hash"] is None and "hash" in low:
                detected["hash"] = name
                continue
            if detected["coupons"] is None and "coupon" in low:
                detected["coupons"] = name
                continue
            if detected["bs"] is None and (low == "bs" or low.endswith(" bs") or "balance sheet" in low):
                detected["bs"] = name
                continue
            if detected["discounts"] is None and "discount" in low and "coupon" not in low:
                detected["discounts"] = name
                continue
            if detected["sales"] is None and ("subdept sales" in low or "sales report" in low):
                detected["sales"] = name

        previews = {}
        for name in ordered_sheet_names:
            if "xxxxxx" in name.casefold():
                continue
            ws = wb[name]
            parts = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True):
                parts.extend(str(v).strip() for v in row if v is not None)
            previews[name] = " ".join(parts).lower()

        used = {v for v in detected.values() if v}

        if detected["sales"] is None:
            for name, preview in previews.items():
                if name in used:
                    continue
                markers = ("subdept sales report", "sub-department", "sub department", "subdept")
                if sum(m in preview for m in markers) >= 2:
                    detected["sales"] = name
                    used.add(name)
                    break

        if detected["coupons"] is None:
            for name, preview in previews.items():
                if name in used:
                    continue
                if "store coupon" in preview or "local discount" in preview:
                    detected["coupons"] = name
                    used.add(name)
                    break

        if detected["discounts"] is None:
            for name, preview in previews.items():
                if name in used:
                    continue
                markers = ("member discounts", "shopper level", "discounts by shopper level", "senior", "owner")
                if ("member discounts" in preview or "discounts by shopper level" in preview or sum(m in preview for m in markers) >= 3):
                    detected["discounts"] = name
                    used.add(name)
                    break

        if detected["hash"] is None:
            for name, preview in previews.items():
                if name in used:
                    continue
                markers = ("refunded discounts", "pass through donations", "paid-ins", "paid ins")
                if sum(m in preview for m in markers) >= 2:
                    detected["hash"] = name
                    used.add(name)
                    break

        if detected["bs"] is None:
            for name, preview in previews.items():
                if name in used:
                    continue
                markers = (
                    "taxes", "sales tax", "charity", "visa", "mastercard",
                    "amex", "discover", "debit", "cash", "bottle", "nickel round", "prepaid",
                )
                if sum(m in preview for m in markers) >= 4:
                    detected["bs"] = name
                    used.add(name)
                    break

        wb.close()
        return detected
    except Exception:
        return {"sales": None, "coupons": None, "discounts": None, "bs": None, "hash": None}

def validate_settlement_processed_net_header(upload_bytes: bytes) -> tuple[bool, Optional[str]]:
    try:
        from io import BytesIO
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(upload_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
                labels = [str(v or "").strip().lower() for v in row]
                if "network" in labels and "processed net amount" in labels:
                    return True, sheet_name
        return False, None
    except Exception:
        return False, None

def detect_workbook_dates(upload_bytes: bytes) -> dict:
    try:
        from collections import Counter
        from io import BytesIO
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(upload_bytes), read_only=True, data_only=True)
        dates_by_sheet = {}

        def parse_date_value(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            if value is None:
                return None

            s = str(value).strip()
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y", "%m.%d.%Y", "%m.%d.%y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        def date_from_sheet_name(name):
            compact = re.search(r"(?<!\d)(\d{5,6})(?!\d)", name)
            if compact:
                digits = compact.group(1)
                candidates = [digits]
                if len(digits) == 5:
                    candidates.append("0" + digits)
                for candidate in candidates:
                    try:
                        return datetime.strptime(candidate, "%m%d%y").date()
                    except ValueError:
                        pass

            separated = re.search(r"(?<!\d)(\d{1,2})[-_/](\d{1,2})[-_/](\d{2,4})(?!\d)", name)
            if separated:
                m, d, y = separated.groups()
                y = ("20" + y) if len(y) == 2 else y
                try:
                    return date(int(y), int(m), int(d))
                except ValueError:
                    pass
            return None

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found = None
            rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True))
            for row in rows:
                for idx, value in enumerate(row):
                    label = str(value).strip().lower() if value is not None else ""
                    if label in {"date", "date:"} or label.startswith("date:"):
                        if ":" in label and label.split(":", 1)[1].strip():
                            found = parse_date_value(label.split(":", 1)[1].strip())
                        if found is None:
                            for offset in (1, 2, 3):
                                if idx + offset < len(row):
                                    found = parse_date_value(row[idx + offset])
                                    if found is not None:
                                        break
                        if found is not None:
                            break
                if found is not None:
                    break

            if found is None:
                found = date_from_sheet_name(sheet_name)

            if found is not None:
                dates_by_sheet[sheet_name] = found

        if not dates_by_sheet:
            return {"detected_date": None, "dates_by_sheet": {}, "has_mismatch": False, "unique_dates": [], "source_sheet": None}

        sales_sheet = next((name for name in wb.sheetnames if "subdept sales" in name.lower() or "sales report" in name.lower()), None)

        source_sheet = sales_sheet if sales_sheet in dates_by_sheet else None
        if source_sheet:
            detected_date = dates_by_sheet[source_sheet]
        else:
            counts = Counter(dates_by_sheet.values())
            detected_date = counts.most_common(1)[0][0]
            source_sheet = next(name for name, dt in dates_by_sheet.items() if dt == detected_date)

        unique_dates = sorted(set(dates_by_sheet.values()))
        return {
            "detected_date": detected_date,
            "dates_by_sheet": dates_by_sheet,
            "has_mismatch": len(unique_dates) > 1,
            "unique_dates": unique_dates,
            "source_sheet": source_sheet,
        }
    except Exception:
        return {"detected_date": None, "dates_by_sheet": {}, "has_mismatch": False, "unique_dates": [], "source_sheet": None}

# ---------------------------------------------------------------------
# Reset / run history helpers
# ---------------------------------------------------------------------

def load_run_history() -> list[dict]:
    if not HISTORY_FILE.exists():
        return []
    try:
        data = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            return []
        return sorted(data, key=lambda item: item.get("run_at", ""), reverse=True)
    except Exception:
        return []

def save_run_history(records: list[dict]) -> None:
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_FILE.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

def _safe_history_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(name).name).strip("._")
    return cleaned or "uploaded_workbook.xlsx"

def archive_run(uploaded_file, settlement_file, result: dict, report_date: date, roles: dict, date_info: dict) -> dict:
    run_at = datetime.now()
    stamp = run_at.strftime("%Y%m%d_%H%M%S_%f")
    upload_name = _safe_history_name(uploaded_file.name)
    upload_path = HISTORY_UPLOAD_DIR / f"{stamp}_{upload_name}"
    settlement_name = _safe_history_name(settlement_file.name) if settlement_file else None
    settlement_path = HISTORY_UPLOAD_DIR / f"{stamp}_settlement_{settlement_name}" if settlement_name else None
    iif_name = Path(result["iif_path"]).name
    iif_path = HISTORY_IIF_DIR / f"{stamp}_{iif_name}"

    upload_path.write_bytes(uploaded_file.getvalue())
    if settlement_path is not None:
        settlement_path.write_bytes(settlement_file.getvalue())
    iif_path.write_bytes(result["iif_bytes"])

    v = result.get("validation", {})
    record = {
        "id": stamp,
        "run_at": run_at.isoformat(timespec="seconds"),
        "report_date": report_date.isoformat(),
        "uploaded_filename": uploaded_file.name,
        "settlement_filename": settlement_file.name if settlement_file else None,
        "archived_upload": str(upload_path),
        "archived_settlement": str(settlement_path) if settlement_path else None,
        "iif_filename": iif_name,
        "archived_iif": str(iif_path),
        "status": "Passed" if v.get("all_ok") else "Review",
        "sales_status": "MATCH" if v.get("sales_ok") is True else ("REVIEW" if v.get("sales_ok") is False else "N/A"),
        "discount_status": "MATCH" if v.get("discounts_ok") is True else ("REVIEW" if v.get("discounts_ok") is False else "N/A"),
        "hash_status": "MATCH" if v.get("hash_ok") is True else ("REVIEW" if v.get("hash_ok") is False else "N/A"),
        "iif_status": "MATCH" if v.get("iif_ok") is True else "REVIEW",
        "card_settlement_status": "MATCH" if v.get("card_settlement_ok") is True else "REVIEW",
        "date_mismatch": bool(date_info.get("has_mismatch", False)),
        "sheet_roles": {k: v for k, v in roles.items() if v},
    }

    records = load_run_history()
    records.insert(0, record)
    save_run_history(records[:250])
    return record

def reset_current_work() -> None:
    for key in ("run_result", "run_date", "run_filename", "run_date_mismatch", "run_settlement_filename", "last_history_id"):
        st.session_state.pop(key, None)
    st.session_state["file_uploader_key"] = st.session_state.get("file_uploader_key", 0) + 1

@dataclass
class IIFLine:
    line_type: str
    trns_type: str
    date: str
    account: str
    name: str
    amount: Optional[float]
    memo: str
    qb_class: str

def parse_iif(path: Path) -> tuple[list[IIFLine], pd.DataFrame]:
    lines: list[IIFLine] = []
    headers: dict[str, list[str]] = {}

    with path.open("r", encoding="utf-8-sig", errors="replace") as f:
        for raw in f:
            row = raw.rstrip("\n\r").split("\t")
            if not row:
                continue

            first = row[0].strip()
            if first.startswith("!"):
                headers[first[1:].upper()] = [x.strip().upper() for x in row]
                continue

            line_type = first.upper()
            if line_type not in {"TRNS", "SPL"}:
                continue

            header = headers.get(line_type)
            if header and len(header) == len(row):
                data = {key.lstrip("!"): val for key, val in zip(header, row)}
                trns_type = data.get("TRNSTYPE", "")
                dt = data.get("DATE", "")
                account = data.get("ACCNT", "")
                name = data.get("NAME", "")
                amount = parse_money(data.get("AMOUNT", ""))
                memo = data.get("MEMO", "")
                qb_class = data.get("CLASS", "")
            else:
                trns_type = row[1].strip() if len(row) > 1 else ""
                dt = row[2].strip() if len(row) > 2 else ""
                account = row[3].strip() if len(row) > 3 else ""
                name = row[4].strip() if len(row) > 4 else ""
                amount = parse_money(row[5]) if len(row) > 5 else None
                memo = row[6].strip() if len(row) > 6 else ""
                qb_class = row[7].strip() if len(row) > 7 else ""

            lines.append(IIFLine(line_type, trns_type, dt, account, name, amount, memo, qb_class))

    df = pd.DataFrame(
        [{"Type": x.line_type, "Account": x.account, "Memo": x.memo, "Amount": x.amount, "Name": x.name, "Class": x.qb_class} for x in lines]
    )
    return lines, df

def account_code(account: str) -> str:
    m = re.match(r"\s*(\d+)", account or "")
    return m.group(1) if m else ""

def classify_line(line: IIFLine) -> str:
    acct = (line.account or "").lower()
    memo = (line.memo or "").lower()
    code = account_code(line.account)

    if line.line_type == "TRNS":
        return "Deposit"
    if code.startswith("711"):
        return "Sales"
    if code.startswith(DISCOUNT_PREFIXES):
        return "Discounts"
    if code == "8515000" or "store coupon" in memo:
        return "Store Coupons"

    combined = f"{acct} {memo}"
    if any(k in combined for k in TENDER_KEYWORDS):
        return "Tenders"
    if any(k in combined for k in BALANCE_SHEET_KEYWORDS):
        return "Balance Sheet"
    return "Other"

def build_detail_df(lines: list[IIFLine], categories: set[str]) -> pd.DataFrame:
    rows = []
    for x in lines:
        cat = classify_line(x)
        if cat not in categories:
            continue
        if x.amount is None and not x.memo and not x.account:
            continue
        rows.append({"Category": cat, "Account": x.account, "Memo": x.memo or "—", "Amount": x.amount})
    return pd.DataFrame(rows)

def parse_card_settlement_rows(log_text: str) -> list[dict]:
    pattern = re.compile(
        r"CARD SETTLEMENT \| (?P<tender>[^|]+?) \| "
        r"Settlement=(?P<settlement>-?[0-9.]+) \| "
        r"BS=(?P<bs>-?[0-9.]+) \| Difference=(?P<diff>-?[0-9.]+) \| "
        r"Adjustment=(?P<adjustment>-?[0-9.]+) \| "
        r"(?P<status>MATCH|MISMATCH)",
        flags=re.IGNORECASE,
    )
    rows = []
    for match in pattern.finditer(log_text or ""):
        rows.append({
            "Tender": match.group("tender").strip(),
            "Daily Card Settlement": float(match.group("settlement")),
            "BS": float(match.group("bs")),
            "Difference": float(match.group("diff")),
            "Adjustment": float(match.group("adjustment")),
            "Status": match.group("status").upper(),
            "Resolution": "Matched" if match.group("status").upper() == "MATCH" else "Adjusted to BS via 8314000",
        })
    return rows

def parse_validation(log_text: str, lines: list[IIFLine]) -> dict:
    sales_ok = section_status(log_text, "SALES CHECK")
    discounts_ok = section_status(log_text, "DISCOUNTS CHECK")
    hash_ok = section_status(log_text, "HASH SALES")

    gross_sales = last_amount_after_label(log_text, ["Gross Sales"])
    store_coupons = last_amount_after_label(log_text, ["Store Coupons"])
    owner_apprec = last_amount_after_label(log_text, ["Owner Apprec", "Owner Appreciation"])
    milk_bottle = last_amount_after_label(log_text, ["Milk Btl", "Milk Bottle Returns", "Milk Bottle Return"])
    script_net = last_amount_after_label(log_text, ["Script Net", "Script Net Sales"])
    excel_sales = last_amount_after_label(log_text, ["Excel Sales", "Excel Sales Total"])
    script_discounts = last_amount_after_label(log_text, ["Script Discounts"])
    excel_discounts = last_amount_after_label(log_text, ["Excel Disc Total", "Excel Discount Total"])
    refunded = last_amount_after_label(log_text, ["Refunded Discounts"])
    pass_through = last_amount_after_label(log_text, ["Pass Thru Donations", "Pass Through Donations"])
    hash_script = last_amount_after_label(log_text, ["Script Total"])
    hash_excel = last_amount_after_label(log_text, ["Hash Sales 6 Total", "HASH Sales 6 Total"])

    trns_amounts = [x.amount for x in lines if x.line_type == "TRNS" and x.amount is not None]
    deposit_total = sum(trns_amounts) if trns_amounts else None

    positive = round(sum(x.amount for x in lines if x.amount is not None and x.amount > 0), 2)
    negative = round(abs(sum(x.amount for x in lines if x.amount is not None and x.amount < 0)), 2)
    iif_difference = round(abs(positive - negative), 2)
    iif_ok = iif_difference < 0.02

    warning_count = sum(
        1 for ln in log_text.splitlines()
        if "WARNING" in ln.upper() or "MISMATCH" in ln.upper() or "FAILED" in ln.upper()
    )

    card_settlement_rows = parse_card_settlement_rows(log_text)
    card_settlement_ok = bool(card_settlement_rows) and all(r["Status"] == "MATCH" for r in card_settlement_rows)

    checks = [x for x in (sales_ok, discounts_ok, hash_ok, iif_ok, card_settlement_ok) if x is not None]
    all_ok = bool(checks) and all(checks)

    return {
        "sales_ok": sales_ok,
        "discounts_ok": discounts_ok,
        "hash_ok": hash_ok,
        "iif_ok": iif_ok,
        "all_ok": all_ok,
        "warning_count": warning_count,
        "gross_sales": gross_sales,
        "store_coupons": store_coupons,
        "owner_apprec": owner_apprec,
        "milk_bottle": milk_bottle,
        "script_net": script_net,
        "excel_sales": excel_sales,
        "script_discounts": script_discounts,
        "excel_discounts": excel_discounts,
        "refunded": refunded,
        "pass_through": pass_through,
        "hash_script": hash_script,
        "hash_excel": hash_excel,
        "deposit_total": deposit_total,
        "positive_total": positive,
        "negative_total": negative,
        "iif_difference": iif_difference,
        "card_settlement_rows": card_settlement_rows,
        "card_settlement_ok": card_settlement_ok,
    }

def build_deposit_summary(lines: list[IIFLine], validation: dict) -> dict[str, Optional[float]]:
    def total_for(*, account_starts: str | None = None, memo_contains: str | None = None) -> Optional[float]:
        matches = []
        for line in lines:
            if line.line_type != "SPL" or line.amount is None:
                continue
            if account_starts and not (line.account or "").startswith(account_starts):
                continue
            if memo_contains and memo_contains.lower() not in (line.memo or "").lower():
                continue
            matches.append(float(line.amount))
        return round(sum(matches), 2) if matches else None

    store_coupons = total_for(account_starts="8515000", memo_contains="Store Coupons")
    owner_apprec = total_for(account_starts="8512006")
    refunded = total_for(memo_contains="Refunded Discounts")
    milk_bottle = total_for(account_starts="1311100", memo_contains="Milk Bottle Return")

    card_adjustment_iif = 0.0
    card_adjustment_found = False
    for line in lines:
        if line.line_type != "SPL" or line.amount is None:
            continue
        if not (line.account or "").startswith("8314000"):
            continue
        if "difference between first data vs bs" not in (line.memo or "").lower():
            continue
        card_adjustment_iif += float(line.amount)
        card_adjustment_found = True

    card_adjustment = round(-card_adjustment_iif, 2) if card_adjustment_found else 0.0

    return {
        "Store Coupons": abs(store_coupons) if store_coupons is not None else None,
        "Owner Appreciation": abs(owner_apprec) if owner_apprec is not None else None,
        "Milk Bottle Returns": abs(milk_bottle) if milk_bottle is not None else None,
        "Refunded Discounts": abs(refunded) if refunded is not None else None,
        "Pass Through Donations": abs(float(validation.get("pass_through"))) if validation.get("pass_through") is not None else None,
        "Card Settlement Adjustment": card_adjustment,
        "IIF Difference": float(validation.get("iif_difference", 0.0)),
    }

def build_engine_command(
    *,
    engine_path: Path,
    deposit_date: date,
    membership_path: Path,
    membership_mode: str,
    coupon_mode: str,
    coupon_closeout_total: float | None,
    coupon_ncg_total: float | None,
    coupon_mfg_total: float | None,
    activity_path: Path | None = None,
    closeout_path: Path | None = None,
    closeout_preview_path: Path | None = None,
) -> list[str]:
    """Build the engine command without performing subprocess or file I/O."""
    command = [
        sys.executable,
        str(engine_path),
        "--date",
        deposit_date.strftime("%m/%d/%y"),
        "--membership-payments-file",
        str(membership_path),
        "--membership-mode",
        membership_mode,
        "--coupon-mode",
        coupon_mode,
    ]
    if coupon_mode == "closeout":
        command.extend([
            "--coupon-closeout-total",
            f"{coupon_closeout_total:.2f}",
            "--coupon-ncg-total",
            f"{coupon_ncg_total:.2f}",
            "--coupon-mfg-total",
            f"{coupon_mfg_total:.2f}",
        ])
    if activity_path is not None:
        command.extend(["--activity-breakdowns-file", str(activity_path)])
    if closeout_path is not None:
        command.extend(["--closeout-file", str(closeout_path)])
        if closeout_preview_path is not None:
            command.extend([
                "--closeout-preview-output",
                str(closeout_preview_path),
            ])
    return command


def run_engine(
    uploaded_file,
    settlement_file,
    deposit_date: date,
    membership_payments: list[dict],
    membership_mode: str,
    coupon_mode: str,
    coupon_closeout_total: float | None,
    coupon_ncg_total: float | None,
    coupon_mfg_total: float | None,
    activity_payload: dict | None = None,
    closeout_payload: dict | None = None,
    preview_only: bool = False,
) -> dict:
    if not ENGINE_PATH.exists():
        raise FileNotFoundError(
            "Deposit engine is missing from this Streamlit repository. "
            "Upload pos_to_quickbooks_v2.py into the SAME GitHub folder as streamlit_app.py, "
            "commit it, and let Streamlit redeploy."
        )

    ext = Path(uploaded_file.name).suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise ValueError("Please upload an .xlsx or .xlsm workbook.")

    safe_name = f"SubDept Single Total Report {deposit_date.strftime('%m-%d-%y')}{ext}"
    input_path = INPUT_DIR / safe_name
    input_path.write_bytes(uploaded_file.getvalue())

    settlement_ext = Path(settlement_file.name).suffix.lower()
    if settlement_ext not in {".xlsx", ".xlsm"}:
        raise ValueError("Please upload the Daily Card Settlement Report as .xlsx or .xlsm.")
    settlement_name = f"{deposit_date.strftime('%m.%d.%Y')}_Daily Card Settlement Report{settlement_ext}"
    settlement_path = INPUT_DIR / settlement_name
    settlement_path.write_bytes(settlement_file.getvalue())

    owned_iif_path = QB_IMPORT_DIR / f"deposit_{deposit_date.strftime('%Y%m%d')}.iif"
    if owned_iif_path.exists():
        owned_iif_path.unlink()

    membership_path = None
    activity_path = None
    closeout_path = None
    closeout_preview_path = None
    closeout_final_run = False
    engine_invocation_attempted = False
    final_result_validated = False
    try:
        normalized_closeout_payload = (
            normalize_closeout_payload(closeout_payload)
            if closeout_payload is not None else None
        )
        if (
            preview_only
            and (
                normalized_closeout_payload is None
                or normalized_closeout_payload["mode"] != "closeout"
            )
        ):
            raise ValueError("Closeout preview requires a Closeout payload.")
        closeout_final_run = (
            normalized_closeout_payload is not None
            and normalized_closeout_payload["mode"] == "closeout"
            and not preview_only
        )
        membership_path = write_membership_payments_file(
            RUNTIME_TEMP_DIR, membership_payments
        )
        if activity_payload is not None:
            normalized_activity_payload = normalize_activity_payload(activity_payload)
            activity_path = write_activity_payload_file(
                RUNTIME_TEMP_DIR, normalized_activity_payload
            )
        if normalized_closeout_payload is not None:
            closeout_path = write_closeout_payload_file(
                RUNTIME_TEMP_DIR, normalized_closeout_payload
            )
            if normalized_closeout_payload["mode"] == "closeout":
                closeout_preview_path = RUNTIME_TEMP_DIR / (
                    f"closeout_preview_{uuid4().hex}.json"
                )
        cmd = build_engine_command(
            engine_path=ENGINE_PATH,
            deposit_date=deposit_date,
            membership_path=membership_path,
            membership_mode=membership_mode,
            coupon_mode=coupon_mode,
            coupon_closeout_total=coupon_closeout_total,
            coupon_ncg_total=coupon_ncg_total,
            coupon_mfg_total=coupon_mfg_total,
            activity_path=activity_path,
            closeout_path=closeout_path,
            closeout_preview_path=closeout_preview_path,
        )
        engine_invocation_attempted = True
        proc = subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True, timeout=180)
        log_text = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")

        for status_path in [LOG_DIR / "last_run_status.txt", QB_IMPORT_DIR / "last_run_status.txt"]:
            if status_path.exists():
                try:
                    status_text = status_path.read_text(encoding="utf-8", errors="replace")
                    if status_text.strip():
                        log_text += "\n\n--- STATUS SUMMARY ---\n" + status_text
                except Exception:
                    pass
                break

        if proc.returncode != 0:
            raise RuntimeError(log_text.strip() or f"Deposit engine exited with code {proc.returncode}.")

        closeout_preview = None
        if closeout_preview_path is not None:
            if not closeout_preview_path.exists():
                raise RuntimeError("The automation finished without a Closeout preview.")
            try:
                closeout_preview = json.loads(
                    closeout_preview_path.read_text(encoding="utf-8")
                )
            except json.JSONDecodeError:
                raise RuntimeError("The automation wrote a malformed Closeout preview.") from None
            if not isinstance(closeout_preview, dict):
                raise RuntimeError("The automation wrote an invalid Closeout preview.")

        if preview_only:
            return {
                "input_path": input_path,
                "settlement_path": settlement_path,
                "closeout_preview": closeout_preview,
                "preview_only": True,
                "log_text": log_text,
            }

        if closeout_preview is not None:
            try:
                remaining_after_approval = float(
                    closeout_preview["remaining_after_approval"]
                )
            except (KeyError, TypeError, ValueError):
                raise ValueError("Closeout preview is missing a valid approval balance.") from None
            if (
                closeout_preview.get("requires_approval") is not False
                or remaining_after_approval != 0.0
            ):
                raise ValueError(
                    "Closeout Sheet approval is required before generating a downloadable IIF."
                )

        resolved_iif_path = owned_iif_path
        if not resolved_iif_path.exists():
            candidates = sorted(QB_IMPORT_DIR.glob("deposit_*.iif"), key=lambda p: p.stat().st_mtime, reverse=True)
            if candidates:
                resolved_iif_path = candidates[0]

        if not resolved_iif_path.exists():
            raise RuntimeError(
                "The automation finished, but no IIF file was found in output/qb_imports.\n\n"
                + log_text[-4000:]
            )

        lines, iif_df = parse_iif(resolved_iif_path)
        validation = parse_validation(log_text, lines)

        result = {
            "input_path": input_path,
            "settlement_path": settlement_path,
            "iif_path": resolved_iif_path,
            "iif_bytes": resolved_iif_path.read_bytes(),
            "lines": lines,
            "iif_df": iif_df,
            "validation": validation,
            "closeout_preview": closeout_preview,
            "preview_only": False,
            "log_text": log_text,
        }
        final_result_validated = True
        return result
    finally:
        if membership_path is not None:
            membership_path.unlink(missing_ok=True)
        if activity_path is not None:
            activity_path.unlink(missing_ok=True)
        if closeout_path is not None:
            closeout_path.unlink(missing_ok=True)
        if closeout_preview_path is not None:
            closeout_preview_path.unlink(missing_ok=True)
        if preview_only or (
            closeout_final_run
            and engine_invocation_attempted
            and not final_result_validated
        ):
            owned_iif_path.unlink(missing_ok=True)

def card(label: str, value: str, foot: str = ""):
    st.markdown(
        f"""
        <div class="hwfc-card">
          <div class="hwfc-card-label">{html.escape(label)}</div>
          <div class="hwfc-card-value">{html.escape(value)}</div>
          <div class="hwfc-card-foot">{foot}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def status_word(value: Optional[bool]) -> str:
    if value is True:
        return '<span class="hwfc-match">✓</span>'
    if value is False:
        return '<span class="hwfc-mismatch">✕</span>'
    return '<span style="color:#777">—</span>'

# ---------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------

st.markdown(
    """
    <div class="hwfc-hero">
      <div class="hwfc-kicker">Honest Weight Food Co-op · Finance</div>
      <div class="hwfc-title">Daily Deposit Reconciliation</div>
      <div class="hwfc-subtitle">
        Upload the completed daily workbook, validate the full deposit, then review the QuickBooks entry before import.
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

if "file_uploader_key" not in st.session_state:
    st.session_state["file_uploader_key"] = 0

action_left, action_right = st.columns([0.80, 0.20])
with action_right:
    if st.button("↻ Start Over", use_container_width=True, help="Clear the current upload and results. Run History is preserved."):
        reset_current_work()
        st.rerun()

has_results = "run_result" in st.session_state

st.markdown(
    f"""
    <div class="hwfc-stepbar">
      <div class="hwfc-step {'active' if not has_results else ''}">1 · Upload</div>
      <div class="hwfc-step {'active' if not has_results else ''}">2 · Validate</div>
      <div class="hwfc-step {'active' if has_results else ''}">3 · Review</div>
      <div class="hwfc-step {'active' if has_results else ''}">4 · Download</div>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.expander("📘 Daily Workbook SOP", expanded=False):
    st.markdown(
        """
        ### How to Build the Daily Workbook
        Follow these steps in order for each deposit date. The Daily Deposit template is the master workbook; SMS reports supply the source data that is copied or moved into it.

        **Important:** keep every report on the same deposit date, and do not continue past the Sales check if the totals do not match exactly.
        """
    )

    for index, step in enumerate(SOP_STEPS):
        expanded = index == 0
        with st.expander(step["title"], expanded=expanded):
            st.markdown(step["body"])

            if step["title"].startswith("Step 1 ·"):
                step1_folder_image = ROOT / "assets" / "step1_daily_deposit_folder.png"
                if step1_folder_image.exists():
                    st.image(
                        str(step1_folder_image),
                        caption="Step 1 · Daily Deposit folder — open TEMPLATE - SubDept Single Total Report.",
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Step 1 Daily Deposit folder image is not installed. "
                        "Add assets/step1_daily_deposit_folder.png to show it here."
                    )

            if step["title"].startswith("Step 2 ·"):
                step2a_export_image = ROOT / "assets" / "step2a_sms_sales_export.png"
                step2a_paste_image = ROOT / "assets" / "step2a_subdept_single_paste.png"

                if step2a_export_image.exists():
                    st.image(
                        str(step2a_export_image),
                        caption=(
                            "Step 2 · SMS export: after launching the Sub-department Single Total report, "
                            "export it to Excel and copy the report data from the first sub-department through the final department."
                        ),
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Step 2 SMS export example is not installed. "
                        "Add assets/step2a_sms_sales_export.png to show it here."
                    )

                if step2a_paste_image.exists():
                    st.image(
                        str(step2a_paste_image),
                        caption=(
                            "Step 2 · Daily Deposit template: paste the exported sales data starting in cell A1 "
                            "of the SubDept Single tab."
                        ),
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Step 2 paste example is not installed. "
                        "Add assets/step2a_subdept_single_paste.png to show it here."
                    )

                st.warning(
                    "**⚠️ Unexpected / Unique Item**\n\n"
                    "The example report contains **23 · Refunded Discounts**. This is not a normal Sales item and "
                    "would ordinarily be expected in the **HASH** process. If an unexpected item appears in the "
                    "Sub-department Single Total report, do not assume it should simply be kept or deleted. "
                    "Drill into the activity in SMS to determine why it appeared, confirm whether it is also represented "
                    "in the HASH report, and ask the Finance team for help if the source is unclear before completing the deposit.",
                    icon="⚠️",
                )

            if step["title"].startswith("Step 2a"):
                sales_check_images = [
                    ROOT / "assets" / "sub_department_sales_report.png",
                    ROOT / "assets" / "department_sales_summary_report.png",
                ]

                available_sales_check_images = [
                    image_path for image_path in sales_check_images if image_path.exists()
                ]

                if available_sales_check_images:
                    st.caption(
                        "Sales-check example from SMS. "
                        "The source report is shown in multiple images because the report is longer than one screen. "
                        "The highlighted total at the bottom must match the green Sales Total in the Daily Deposit workbook exactly."
                    )

                    image_captions = {
                        "sub_department_sales_report.png":
                            "SMS Sub-department Single Total Report · Part 1",
                        "department_sales_summary_report.png":
                            "SMS Sub-department Single Total Report · Part 2 · Verify the highlighted Total",
                    }

                    for image_path in available_sales_check_images:
                        st.image(
                            str(image_path),
                            caption=image_captions.get(image_path.name, "SMS Sales-check example"),
                            use_container_width=True,
                        )

                    missing_images = [
                        image_path.name for image_path in sales_check_images if not image_path.exists()
                    ]

                    if missing_images:
                        st.warning(
                            "One Sales Check example image is missing from the assets folder:\n\n"
                            + "\n".join(f"• {name}" for name in missing_images),
                            icon="⚠️",
                        )
                else:
                    st.error(
                        "Sales Check example images could not be found.\n\n"
                        "The app expects these files inside the assets folder:\n\n"
                        "• sub_department_sales_report.png\n\n"
                        "• department_sales_summary_report.png",
                        icon="🚫",
                    )

            if step["title"].startswith("Step 3"):
                milk_bottle_returns_example = ROOT / "assets" / "milk_bottle_returns_example.png"
                if milk_bottle_returns_example.exists():
                    st.caption(
                        "Milk Bottle Returns example from SMS. Add together the highlighted return amounts for all "
                        "Milk Bottle Return items, then enter the combined total into cell M1 on SubDept Sales Report."
                    )
                    st.image(
                        str(milk_bottle_returns_example),
                        caption=(
                            "Step 3 · Add all highlighted Milk Bottle Return amounts before entering the total in M1."
                        ),
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Milk Bottle Returns example image is not installed. "
                        "Add assets/milk_bottle_returns_example.png to show it here."
                    )

            if step["title"].startswith("Step 4"):
                step4_discounts_image = ROOT / "assets" / "step4_discounts_move_copy.png"
                if step4_discounts_image.exists():
                    st.image(
                        str(step4_discounts_image),
                        caption=(
                            "Step 4 · Move or copy the exported Discounts worksheet into the XXXXXX Discounts "
                            "location, then rename it to the current deposit date."
                        ),
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Step 4 Discounts Move or Copy example is not installed. "
                        "Add assets/step4_discounts_move_copy.png to show it here."
                    )

            if step["title"].startswith("Step 5"):
                step5_hash_image = ROOT / "assets" / "step5_hash_move_copy.png"
                if step5_hash_image.exists():
                    st.image(
                        str(step5_hash_image),
                        caption=(
                            "Step 5 · Move or copy the exported HASH worksheet into the XXXXXX HASH "
                            "location, then rename it to the current deposit date."
                        ),
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Step 5 HASH Move or Copy example is not installed. "
                        "Add assets/step5_hash_move_copy.png to show it here."
                    )

            if step["title"].startswith("Step 6"):
                step6_bs_image = ROOT / "assets" / "step6_bs_move_copy.png"
                if step6_bs_image.exists():
                    st.image(
                        str(step6_bs_image),
                        caption=(
                            "Step 6 · Move or copy the exported Balance Sheet worksheet into the XXXXXX BS "
                            "location, then rename it to the current deposit date."
                        ),
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "Step 6 Balance Sheet Move or Copy example is not installed. "
                        "Add assets/step6_bs_move_copy.png to show it here."
                    )

            if step["title"].startswith("Step 7"):
                daily_workbook_example = ROOT / "assets" / "daily_workbook_example.png"
                daily_workbook_example_pdf = ROOT / "assets" / "daily_workbook_example.pdf"
                if daily_workbook_example.exists():
                    st.image(
                        str(daily_workbook_example),
                        caption="Approved Daily Workbook example",
                        use_container_width=True,
                    )
                elif daily_workbook_example_pdf.exists():
                    st.caption("Use the approved PDF in the assets folder for the final visual comparison.")
                    st.download_button(
                        "Open approved workbook example PDF",
                        data=daily_workbook_example_pdf.read_bytes(),
                        file_name=daily_workbook_example_pdf.name,
                        mime="application/pdf",
                        key="daily_workbook_example_pdf",
                        use_container_width=True,
                    )
                else:
                    st.info(
                        "The Daily Workbook example is not installed in the repo assets folder. "
                        "Add assets/daily_workbook_example.png or assets/daily_workbook_example.pdf to show it here."
                    )

    st.markdown("### Daily Card Settlement Report")
    st.markdown(
        """
        After the Daily Workbook is complete, prepare the separate **Daily Card Settlement Report** for the same date. The automation uses only the **Processed Net Amount** column for **VISA/MC, Discover, AMEX, Debit Card, and EBT Cash/Food Stamp**.
        """
    )

    with st.expander("View Daily Card Settlement Example", expanded=False):
        st.caption(
            "Your Daily Card Settlement Report should look like this. "
            "The automation uses the Processed Net Amount column."
        )
        daily_card_settlement_example = ROOT / "assets" / "daily_card_settlement_example.png"
        if daily_card_settlement_example.exists():
            st.image(str(daily_card_settlement_example), use_container_width=True)
        else:
            st.info("Daily Card Settlement example image is not installed in the repo assets folder.")

    st.markdown(
        """
        ### Before You Run
        Confirm that the Daily Workbook and Daily Card Settlement Report are for the same date, all five required workbook roles are detected, and the settlement report shows as verified. Then select **Validate & Build Deposit**.

        The app compares each card settlement amount with the matching BS tender total. Review any red **✕** before importing the IIF into QuickBooks.
        """
    )


# ---------------------------------------------------------------------
# Main-page tips and known exceptions
# ---------------------------------------------------------------------

with st.expander("💡 Tips & Known Exceptions · WIP", expanded=False):
    st.caption(
        "Working guidance for unusual situations. This section will continue to grow as Finance documents more exceptions."
    )

    st.markdown(
        """
        <div class="hwfc-tip-card attention">
          <div class="hwfc-tip-title">⚠️ Unique / unrecognized items → TBA</div>
          <div class="hwfc-tip-body">
            Any unique item the automation does not recognize is coded as <strong>TBA</strong> at the bottom of the generated IIF.
            Review those lines and change them to the correct QuickBooks account before final posting. If the correct coding is unclear,
            research the SMS/source reports and ask Finance before posting.
          </div>
        </div>

        <div class="hwfc-tip-card info">
          <div class="hwfc-tip-title">🔎 Unexpected SMS items</div>
          <div class="hwfc-tip-body">
            If an item appears in a report where it normally does not belong, investigate the source activity in SMS before manually
            changing the workbook. For example, if Refunded Discounts appears in the Step 2 Sales report, confirm whether it is also
            represented in HASH and involve Finance if the reason is unclear.
          </div>
        </div>

        <div class="hwfc-tip-card info">
          <div class="hwfc-tip-title">💵 Paid Out</div>
          <div class="hwfc-tip-body">
            Paid Out does not appear every day. When it is detected on the <strong>Balance Sheet</strong>, the automation carries that
            amount into the generated IIF, similar to <strong>Paid-Ins</strong> and <strong>Pass Through Donations</strong>.
            Paid Out is treated as a <strong>negative amount</strong>, so it reduces the QuickBooks deposit total.
            <div class="hwfc-tip-example">Example: $47.06 Paid Out → -$47.06 deposit effect</div>
          </div>
        </div>

        <div class="hwfc-tip-card attention">
          <div class="hwfc-tip-title">⚠️ Important Disclosures &amp; Automation Limitations</div>
          <div class="hwfc-tip-body">
            <strong>1. SMS-Based Automation</strong><br>
            The automated IIF is built from <strong>SMS data</strong>. It does not automatically include every actual daily adjustment
            documented on the <strong>Store Closeout</strong> sheet provided by the Front End Manager.<br><br>
            <strong>2. Store Closeout Review Is Required</strong><br>
            Import the generated IIF into <strong>QuickBooks first</strong>, then compare the deposit to the Store Closeout and manually
            adjust the deposit in QuickBooks for Cash Over / Short, additional cash differences, Plants / Dept. Market Purchases,
            Payroll cash activity, Comments / Notes / Issues from Front End, safe overage or shortage, and any other documented amount
            that changes the actual daily deposit.<br><br>
            <strong>3. Automation Does Not Replace Final Review</strong><br>
            A successfully generated and balanced IIF does not necessarily mean the final QuickBooks deposit matches the actual Store
            Closeout. The deposit is not complete until required Store Closeout adjustments are made in QuickBooks and the final deposit
            has been reviewed for accuracy.
          </div>
        </div>

        <div class="hwfc-tip-card success">
          <div class="hwfc-tip-title">✅ QuickBooks final review</div>
          <div class="hwfc-tip-body">
            Review all TBA lines, confirm Sales / Discounts / HASH checks, review card settlement differences, import the IIF, apply the
            Store Closeout adjustments directly in QuickBooks, and confirm the final deposit is correct before posting.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        **Future tips to document**
        - Date mismatch handling
        - When to stop and ask Finance
        - Common TBA mappings once approved
        - QuickBooks pre-post review reminders
        """
    )


# ---------------------------------------------------------------------
# Sidebar: optional guidance and run history
# ---------------------------------------------------------------------

with st.sidebar:
    st.markdown("## 🌿 HWFC Daily Deposit")
    st.caption("Prior deposit records")

    with st.expander("📜 Run History", expanded=False):
        st.caption("Open only when you need to review a prior deposit.")

        history_records = load_run_history()
        if not history_records:
            st.caption("No completed runs yet.")
        else:
            selectable_history = history_records[:25]
            history_ids = [record.get("id", str(idx)) for idx, record in enumerate(selectable_history)]
            history_by_id = dict(zip(history_ids, selectable_history))

            selected_history_id = st.selectbox(
                "Prior deposit",
                options=history_ids,
                format_func=lambda record_id: build_history_option_label(history_by_id[record_id]),
                key="run_history_selection",
            )
            record = history_by_id[selected_history_id]

            try:
                report_label = datetime.fromisoformat(record.get("report_date", "")).strftime("%m/%d/%Y")
            except Exception:
                report_label = record.get("report_date", "—")
            run_label = format_history_run_time(record.get("run_at", ""), include_date=True)

            status_icon = "✓" if record.get("status") == "Passed" else "⚠"
            st.markdown(
                f"**{status_icon} {html.escape(str(report_label))} · "
                f"{html.escape(str(record.get('status', '—')))}**"
            )
            st.caption(f"Run {run_label}")

            history_checks = [
                ("Sales", record.get("sales_status", "N/A")),
                ("Discounts", record.get("discount_status", "N/A")),
                ("HASH", record.get("hash_status", "N/A")),
                ("IIF", record.get("iif_status", "N/A")),
                ("Card Settlement", record.get("card_settlement_status", "N/A")),
            ]

            status_text = []
            for label, status in history_checks:
                icon = "✓" if status == "MATCH" else ("⚠" if status == "REVIEW" else "—")
                status_text.append(f"{icon} {label}")
            st.caption("  ·  ".join(status_text))

            with st.expander("Run details", expanded=False):
                st.caption(f"Workbook: {record.get('uploaded_filename', '—')}")
                st.caption(f"Card Settlement: {record.get('settlement_filename', '—')}")
                if record.get("date_mismatch"):
                    st.warning("This run had a workbook date mismatch warning.", icon="⚠️")

            with st.expander("Files from this run", expanded=False):
                archived_upload = Path(record.get("archived_upload", ""))
                archived_settlement = Path(record.get("archived_settlement", ""))
                archived_iif = Path(record.get("archived_iif", ""))

                if archived_upload.is_file():
                    st.download_button(
                        "Download Daily Workbook",
                        data=archived_upload.read_bytes(),
                        file_name=record.get("uploaded_filename", archived_upload.name),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"history_upload_{selected_history_id}",
                        use_container_width=True,
                    )

                if archived_settlement.is_file():
                    st.download_button(
                        "Download Card Settlement",
                        data=archived_settlement.read_bytes(),
                        file_name=record.get("settlement_filename", archived_settlement.name),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"history_settlement_{selected_history_id}",
                        use_container_width=True,
                    )

                if archived_iif.is_file():
                    st.download_button(
                        "Download IIF",
                        data=archived_iif.read_bytes(),
                        file_name=record.get("iif_filename", archived_iif.name),
                        mime="text/plain",
                        key=f"history_iif_{selected_history_id}",
                        use_container_width=True,
                    )

# ---------------------------------------------------------------------
# Input area
# ---------------------------------------------------------------------

setup_col, workbook_col, settlement_col = st.columns([0.22, 0.39, 0.39], gap="medium")

roles = {}
date_info = {"detected_date": None, "dates_by_sheet": {}, "has_mismatch": False, "unique_dates": [], "source_sheet": None}
deposit_date = None

with setup_col:
    st.markdown('<div class="hwfc-section-label">Report date</div>', unsafe_allow_html=True)
    date_placeholder = st.empty()

with workbook_col:
    st.markdown('<div class="hwfc-section-label">Daily workbook</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Upload completed SubDept workbook",
        type=["xlsx", "xlsm"],
        label_visibility="collapsed",
        help="Workbook should contain Sales, Coupons, Discounts, BS, and HASH data.",
        key=f"daily_workbook_{st.session_state['file_uploader_key']}",
    )

with settlement_col:
    st.markdown('<div class="hwfc-section-label">Card settlement</div>', unsafe_allow_html=True)
    settlement_file = st.file_uploader(
        "Upload Daily Card Settlement Report",
        type=["xlsx", "xlsm"],
        label_visibility="collapsed",
        help="Uses ONLY Processed Net Amount for VISA/MC, Discover, AMEX, Debit Card, and EBT.",
        key=f"card_settlement_{st.session_state['file_uploader_key']}",
    )

settlement_date_info = None
settlement_date_mismatch = False
settlement_source_ok = False
settlement_source_sheet = None

if uploaded:
    upload_bytes = uploaded.getvalue()
    date_info = detect_workbook_dates(upload_bytes)
    deposit_date = date_info["detected_date"]
    roles = detect_sheet_roles(upload_bytes, preferred_date=deposit_date)

    with setup_col:
        if deposit_date is not None:
            date_placeholder.markdown(
                f'<div class="hwfc-mini-card"><div class="hwfc-mini-label">Detected</div><div class="hwfc-mini-value">📅 {deposit_date.strftime("%m/%d/%Y")}</div></div>',
                unsafe_allow_html=True,
            )
        else:
            date_placeholder.error("Report date not detected", icon="⚠️")

    if date_info["has_mismatch"]:
        detail_lines = [f"**{sheet}:** {dt.strftime('%m/%d/%Y')}" for sheet, dt in date_info["dates_by_sheet"].items()]
        source = date_info.get("source_sheet") or "workbook"
        st.warning(
            "**DATE MISMATCH WARNING**\n\n"
            + "The workbook contains more than one report date. "
            + f"The deposit will use **{deposit_date.strftime('%m/%d/%Y')}** from **{source}**. "
            + "You can still run the deposit, but review the dates first.\n\n"
            + "  \n".join(detail_lines),
            icon="⚠️",
        )

    with st.expander("Workbook validation", expanded=True):
        cols = st.columns(5)
        labels = [
            ("Sales", roles.get("sales")),
            ("Coupons", roles.get("coupons")),
            ("Discounts", roles.get("discounts")),
            ("Balance Sheet", roles.get("bs")),
            ("HASH", roles.get("hash")),
        ]
        for col, (label, sheet_name) in zip(cols, labels):
            with col:
                if sheet_name:
                    st.markdown(
                        f'<div class="hwfc-check-card"><div class="hwfc-check-title">✓ {html.escape(label)}</div><div class="hwfc-check-sheet">{html.escape(sheet_name)}</div></div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.warning(f"{label}\n\nNot detected", icon="⚠️")

    if settlement_file is not None:
        settlement_source_ok, settlement_source_sheet = validate_settlement_processed_net_header(
            settlement_file.getvalue()
        )
        try:
            from io import BytesIO
            import openpyxl

            swb = openpyxl.load_workbook(
                BytesIO(settlement_file.getvalue()),
                read_only=True,
                data_only=True,
            )
            sws = swb[swb.sheetnames[0]]
            raw_settlement_date = None
            for row in sws.iter_rows(
                min_row=1,
                max_row=min(sws.max_row, 12),
                values_only=True,
            ):
                for idx, value in enumerate(row):
                    if (
                        str(value or "").strip().lower() == "date"
                        and idx + 1 < len(row)
                    ):
                        raw_settlement_date = row[idx + 1]
                        break
                if raw_settlement_date is not None:
                    break
            if isinstance(raw_settlement_date, datetime):
                settlement_date_info = raw_settlement_date.date()
            elif isinstance(raw_settlement_date, date):
                settlement_date_info = raw_settlement_date
        except Exception as exc:
            st.warning(
                f"Could not read the Daily Card Settlement Report date: {exc}",
                icon="⚠️",
            )

        settlement_date_mismatch = render_card_settlement_verification(
            st,
            source_ok=settlement_source_ok,
            settlement_date=settlement_date_info,
            deposit_date=deposit_date,
        )

    missing_roles = [k for k, v in roles.items() if not v]
else:
    with setup_col:
        date_placeholder.markdown(
            '<div class="hwfc-mini-card"><div class="hwfc-mini-label">Detected</div><div class="hwfc-mini-value">Upload workbook</div></div>',
            unsafe_allow_html=True,
        )
    missing_roles = []

subscription_total = 0.0
membership_payments: list[dict] = []
membership_valid = True
membership_mode = "automatic"
coupon_bs_total = 0.0
coupon_valid = True
coupon_mode = "quickbooks"
coupon_closeout_total = 0.0
coupon_ncg_total = 0.0
coupon_mfg_total = 0.0
activity_source_totals = {key: 0.0 for key in ("donation", "paid_out", "paid_in")}
activity_payload = {
    key: {"mode": "quickbooks", "rows": []}
    for key in ("donation", "paid_out", "paid_in")
}
activity_valid = True
activity_detection_valid = True
closeout_payload = None
closeout_valid = False

if uploaded:
    try:
        subscription_total = read_subscription_total(upload_bytes, roles.get("bs"))
    except Exception as exc:
        membership_valid = False
        st.error(f"Could not read Subscription Revenue from the Balance Sheet: {exc}", icon="🚫")
    try:
        coupon_bs_total = read_coupon_receivable_total(upload_bytes, roles.get("bs"))
    except Exception as exc:
        coupon_valid = False
        st.error(f"Could not read Coupons Receivable from the Balance Sheet: {exc}", icon="🚫")

if uploaded and membership_valid:
    subscription_status = subscription_action_status(subscription_total)
    if subscription_status["show_banner"]:
        status_text = (
            f"**{subscription_status['title']}** — "
            f"{subscription_status['message']}"
        )
        if subscription_status["needs_action"]:
            st.warning(status_text, icon="⚠️")
        else:
            st.success(status_text, icon="✅")

closeout_workbook_key = (
    membership_editor_key(upload_bytes, st.session_state["file_uploader_key"])
    if uploaded
    else None
)
closeout_payload_key = f"closeout_payload_{closeout_workbook_key}"
closeout_preview_key = f"closeout_preview_{closeout_workbook_key}"
closeout_hydration_key = f"closeout_form_needs_hydration_{closeout_workbook_key}"
closeout_choice_state = (
    st.session_state.get(f"closeout_handling_{closeout_workbook_key}")
    if closeout_workbook_key is not None
    else None
)

if uploaded:
    try:
        activity_source_totals = read_activity_source_totals(
            upload_bytes,
            roles.get("bs"),
            roles.get("hash"),
        )
    except Exception as exc:
        activity_detection_valid = False
        st.warning(
            "Optional Donations, Paid Out, and Paid In totals could not be read. "
            f"The existing QuickBooks process remains available. Details: {exc}",
            icon="⚠️",
        )

required_steps = ()
step_completions = {}
active_step = None
workflow_completion_key = None
workflow_requirements_key = None
workflow_blocked = False
guided_workflow_ready = False
active_step_content = None
if uploaded:
    detected_required_steps = None
    if activity_detection_valid:
        detected_required_steps = required_deposit_steps(
            subscription_total,
            activity_source_totals,
            coupon_bs_total,
        )
    workflow_completion_key = f"deposit_step_completions_{closeout_workbook_key}"
    workflow_requirements_key = (
        f"deposit_required_steps_{closeout_workbook_key}"
    )
    workflow_state = resolve_activity_detection_workflow(
        detection_valid=activity_detection_valid,
        detected_required_steps=detected_required_steps,
        saved_required_steps=st.session_state.get(workflow_requirements_key),
        saved_completions=st.session_state.get(workflow_completion_key),
    )
    workflow_blocked = workflow_state["blocked"]
    required_steps = workflow_state["required_steps"]
    step_completions = workflow_state["completions"]
    if activity_detection_valid:
        st.session_state[workflow_requirements_key] = required_steps
        st.session_state[workflow_completion_key] = step_completions
    elif not workflow_blocked:
        st.session_state[workflow_completion_key] = step_completions

    if workflow_blocked:
        st.markdown("## Today’s Deposit Steps")
        st.error(
            "Today’s Deposit Steps are blocked until Donations, Paid Out, and Paid In "
            "totals can be read successfully. Reload the workbook and try again.",
            icon="🚫",
        )
    else:
        active_step = active_deposit_step(required_steps, step_completions)

        st.markdown("## Today’s Deposit Steps")
        active_step_content, edited_step = render_deposit_step_panels(
            st,
            deposit_step_rows(required_steps, step_completions),
            edit_key_prefix=f"edit_deposit_step_{closeout_workbook_key}",
        )
        if edited_step is not None:
            st.session_state[workflow_completion_key] = reopen_step_for_edit(
                st.session_state,
                required_steps=required_steps,
                completions=step_completions,
                step=edited_step,
                workbook_key=closeout_workbook_key,
            )
            if edited_step != STEP_CLOSEOUT:
                st.session_state.pop(closeout_preview_key, None)
                st.session_state[closeout_hydration_key] = True
            st.rerun()

        step_completions = normalize_step_completions(
            required_steps,
            st.session_state.get(workflow_completion_key),
        )
        guided_workflow_ready = deposit_workflow_complete(
            required_steps,
            step_completions,
        ) and activity_detection_valid
        if guided_workflow_ready:
            st.success(
                "All deposit steps are complete. Validate and prepare the QuickBooks IIF below.",
                icon="✅",
            )

membership_choice_key = f"membership_handling_{closeout_workbook_key}"
membership_scroll_request_key = f"membership_scroll_{closeout_workbook_key}"
membership_continue_scroll_key = f"membership_continue_scroll_{closeout_workbook_key}"
membership_saved_choice_key = f"membership_saved_choice_{closeout_workbook_key}"
saved_payments_key = f"membership_saved_payments_{closeout_workbook_key}"
if active_step == STEP_MEMBER_SHARES and membership_choice_key not in st.session_state:
    canonical_membership_choice = st.session_state.get(
        membership_saved_choice_key
    )
    if membership_mode_from_choice(canonical_membership_choice) is not None:
        st.session_state[membership_choice_key] = canonical_membership_choice

if subscription_total > 0 and active_step == STEP_MEMBER_SHARES:
    active_step_panel = active_step_content.container()
    active_step_panel.__enter__()
    handling_choice = st.radio(
        "How should these payments be handled?",
        options=[
            "Breakdown in app using the Ownership Payments sheet",
            "Finish manually in QuickBooks",
        ],
        horizontal=True,
        index=None,
        key=membership_choice_key,
        on_change=queue_breakdown_scroll,
        args=(
            st.session_state,
            membership_choice_key,
            membership_scroll_request_key,
        ),
    )
    membership_mode = membership_mode_from_choice(handling_choice)

    if membership_mode is None:
        membership_valid = False
        st.caption("Select how you want to handle member shares before building the deposit.")
    elif membership_mode == "manual":
        st.info(
            f"The app will post ${subscription_total:,.2f} as one unnamed Member Shares line. "
            "In QuickBooks, create or select the member name and complete the principal/interest split.",
            icon="ℹ️",
        )
    else:
        st.caption(
            "Enter each payment below. Existing members must use the exact name shown in QuickBooks. "
            "For a new member, leave the QuickBooks name blank and assign it after import."
        )
        st.info(
            "**Use the Ownership Payments sheet**\n\n"
            "Before splitting automatically, use the Ownership Payments sheet to find "
            "the member, member number (or select Member # pending when the sheet "
            "says \"New\"), plan type (1, 3, or 5 year), and amount paid. For an "
            "existing member, use the name exactly as it appears in QuickBooks.",
            icon="📄",
        )
        st.caption("Deposits are interest-free; installment payments include interest.")
        st.markdown("#### Plan Payment Guide")
        st.markdown(
            plan_guide_html(plan_reference_rows()),
            unsafe_allow_html=True,
        )
        with st.expander("View Ownership Payments Sheet Example"):
            st.image(
                str(Path(__file__).parent / "assets" / "ownership_payments_example.png"),
                caption="Use the actual paper sheet for the current deposit.",
            )
        show_payoff_adjustment = st.checkbox(
            "Advanced: adjust interest periods for a payoff",
            value=False,
            help="Most deposits do not need this. Leave it off to calculate interest automatically.",
            key=f"membership_payoff_{membership_editor_key(upload_bytes, st.session_state['file_uploader_key'])}",
        )
        entry_base_key = membership_editor_key(
            upload_bytes,
            st.session_state["file_uploader_key"],
        )
        render_breakdown_scroll_target(
            st,
            components.html,
            st.session_state,
            target_id="member-share-breakdown",
            request_key=membership_scroll_request_key,
        )
        saved_payments_key = f"membership_saved_payments_{entry_base_key}"
        entry_version_key = f"membership_entry_version_{entry_base_key}"
        if saved_payments_key not in st.session_state:
            st.session_state[saved_payments_key] = []
        if entry_version_key not in st.session_state:
            st.session_state[entry_version_key] = 0

        entry_version = st.session_state[entry_version_key]
        entry_key = f"membership_entry_{entry_base_key}_{entry_version}"
        st.markdown("**Add a member payment**")
        entry_columns = st.columns(
            [2.2, 2.0, 1.4, 1.1, 1.1],
            vertical_alignment="bottom",
        )
        with entry_columns[0]:
            payment_option = st.selectbox(
                "Payment Option",
                options=list(PAYMENT_OPTIONS),
                key=f"{entry_key}_payment_option",
                help="Paid in full has no plan selection because the share is fully paid.",
            )
            quickbooks_name_status, saved_quickbooks_name = (
                apply_quickbooks_name_option_state(
                    st.session_state,
                    entry_key,
                    payment_option,
                )
            )
            apply_membership_amount_option_state(
                st.session_state,
                entry_key,
                payment_option,
            )
        with entry_columns[1]:
            quickbooks_status_key = f"{entry_key}_quickbooks_name_status"
            quickbooks_name_key = f"{entry_key}_member_name"
            if quickbooks_name_status == "No":
                member_name_label = "Member Name: New"
            elif quickbooks_name_status == "Yes" and saved_quickbooks_name:
                member_name_label = "Member Name: Set"
            else:
                member_name_label = "Member Name"
            with st.popover(member_name_label, use_container_width=True):
                quickbooks_name_status = st.radio(
                    "Does this member already exist in QuickBooks?",
                    options=["Yes", "No"],
                    index=None,
                    horizontal=True,
                    key=quickbooks_status_key,
                )
                if quickbooks_name_status == "Yes":
                    member_name = st.text_input(
                        "Enter the exact QuickBooks member name",
                        key=quickbooks_name_key,
                    )
                    st.caption(
                        "Use the name exactly as it appears in QuickBooks. For example, "
                        "the sheet may say Karl Cruz while QuickBooks says Karl Chester Cruz."
                    )
                elif quickbooks_name_status == "No":
                    member_name = ""
                    st.caption(
                        "Name entry is disabled. The QuickBooks NAME field will stay blank "
                        "so the new member can be assigned after import."
                    )
                else:
                    member_name = ""
        quickbooks_member_exists = (
            True if quickbooks_name_status == "Yes"
            else False if quickbooks_name_status == "No"
            else None
        )

        member_number_status = None
        member_number = ""
        with entry_columns[2]:
            if payment_option == "Paid in full — $100":
                st.text_input(
                    "Member #",
                    value="Not required",
                    disabled=True,
                    key=f"{entry_key}_paid_in_full_number",
                )
                member_number_status = "No"
            else:
                number_status_key = f"{entry_key}_member_number_status"
                number_value_key = f"{entry_key}_member_number"
                selected_status = st.session_state.get(number_status_key)
                selected_number = str(st.session_state.get(number_value_key) or "").strip()
                if selected_status == "No":
                    member_number_label = "Member #: Pending"
                elif selected_status == "Yes" and selected_number:
                    member_number_label = f"Member #: {selected_number}"
                else:
                    member_number_label = "Member #"
                with st.popover(member_number_label, use_container_width=True):
                    member_number_status = st.radio(
                        "Does this member have a member number?",
                        options=["Yes", "No"],
                        index=None,
                        horizontal=True,
                        key=number_status_key,
                    )
                    if member_number_status == "Yes":
                        member_number = st.text_input(
                            "Enter member number",
                            key=number_value_key,
                            help="Digits only; do not include the # symbol.",
                        )
                    elif member_number_status == "No":
                        st.caption("The QuickBooks memo will use #Pending.")

        with entry_columns[3]:
            if payment_option == "Paid in full — $100":
                amount = st.number_input(
                    "Amount",
                    value=100.00,
                    format="%.2f",
                    disabled=True,
                    key=f"{entry_key}_paid_in_full_amount",
                )
            else:
                amount = st.number_input(
                    "Amount",
                    min_value=0.00,
                    step=0.01,
                    format="%.2f",
                    key=f"{entry_key}_amount",
                )
        with entry_columns[4]:
            add_payment_clicked = st.button(
                "+ Add payment",
                type="secondary",
                use_container_width=True,
                key=f"{entry_key}_add",
            )

        interest_periods = None
        if show_payoff_adjustment:
            interest_periods = st.number_input(
                "Interest Periods (optional payoff override)",
                min_value=0,
                step=1,
                value=None,
                key=f"{entry_key}_interest_periods",
            )

        if add_payment_clicked:
            try:
                new_payment = membership_payment_from_entry(
                    member_name=member_name,
                    member_number_status=member_number_status,
                    member_number=member_number,
                    quickbooks_member_exists=quickbooks_member_exists,
                    payment_option=payment_option,
                    amount=amount,
                    interest_periods=interest_periods,
                )
                build_membership_lines([new_payment], handling_mode="automatic")
            except ValueError as exc:
                st.error(str(exc), icon="🚫")
            else:
                st.session_state[saved_payments_key] = [
                    *st.session_state[saved_payments_key],
                    new_payment,
                ]
                st.session_state[entry_version_key] += 1
                queue_continue_scroll(
                    st.session_state,
                    membership_continue_scroll_key,
                )
                st.rerun()

        saved_payments = st.session_state[saved_payments_key]
        if saved_payments:
            st.markdown("**Added member payments**")
            payment_headers = st.columns([2.2, 1.2, 2.0, 1.0, 0.8])
            payment_headers[0].caption("Member Name")
            payment_headers[1].caption("Member #")
            payment_headers[2].caption("Payment Option")
            payment_headers[3].caption("Amount")
        for payment_index, payment in enumerate(saved_payments):
            if payment["payment_type"] == "Paid in full":
                saved_option = "Paid in full — $100"
                saved_number = "Not required"
            else:
                saved_option = f"{payment['payment_type']} — {payment['plan']}"
                saved_number = (
                    "Pending" if payment.get("member_number_pending")
                    else payment.get("member_number", "")
                )
            payment_columns = st.columns([2.2, 1.2, 2.0, 1.0, 0.8])
            payment_columns[0].write(
                payment["member_name"] or "New member — assign in QuickBooks"
            )
            payment_columns[1].write(saved_number)
            payment_columns[2].write(saved_option)
            payment_columns[3].write(f"${float(payment['amount']):,.2f}")
            if payment_columns[4].button(
                "Remove",
                key=f"remove_membership_{entry_base_key}_{payment_index}",
            ):
                st.session_state[saved_payments_key] = remove_membership_payment(
                    saved_payments,
                    payment_index,
                )
                st.rerun()

        membership_payments.extend(
            dict(payment) for payment in st.session_state[saved_payments_key]
        )

        render_breakdown_scroll_target(
            st,
            components.html,
            st.session_state,
            target_id="member-share-save-and-continue",
            request_key=membership_continue_scroll_key,
        )
        try:
            membership_preview = build_membership_lines(
                membership_payments,
                expected_subscription_total=subscription_total,
                handling_mode=membership_mode,
            )
            st.success(f"Ready — payments total ${subscription_total:,.2f}.", icon="✅")
            preview_frame = pd.DataFrame(membership_preview).rename(columns={
                "name": "Member Name",
                "account": "QuickBooks Account",
                "memo": "Memo",
                "class_name": "Class",
                "amount": "QuickBooks Amount",
            })
            with st.expander("Review QuickBooks breakdown"):
                st.dataframe(
                    preview_frame,
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        "QuickBooks Amount": st.column_config.NumberColumn(format="$%.2f"),
                    },
                )
        except ValueError as exc:
            membership_valid = False
            st.warning(str(exc), icon="⚠️")
    active_step_panel.__exit__(None, None, None)

if STEP_MEMBER_SHARES in required_steps:
    saved_choice = st.session_state.get(membership_saved_choice_key)
    if active_step == STEP_MEMBER_SHARES:
        saved_choice = st.session_state.get(membership_choice_key)
    saved_payments = st.session_state.get(saved_payments_key, [])
    if STEP_MEMBER_SHARES in step_completions:
        recovered_membership = recover_completed_membership_state(
            required_steps,
            step_completions,
            saved_choice,
            saved_payments,
            subscription_total=subscription_total,
        )
        if recovered_membership["needs_review"]:
            st.session_state[workflow_completion_key] = recovered_membership[
                "completions"
            ]
            st.error(recovered_membership["error"], icon="🚫")
            st.rerun()
        membership_mode = recovered_membership["membership_mode"]
        membership_payments = recovered_membership["membership_payments"]
    else:
        membership_mode = membership_mode_from_choice(saved_choice)
        membership_payments = (
            [dict(payment) for payment in saved_payments]
            if isinstance(saved_payments, list)
            else []
        )

if active_step == STEP_MEMBER_SHARES and membership_mode == "manual":
    manual_transition = save_manual_member_share_transition(
        required_steps,
        step_completions,
        membership_saved_choice_key,
    )
    st.session_state.update(manual_transition["saved_payload"])
    st.session_state[workflow_completion_key] = manual_transition[
        "completions"
    ]
    st.rerun()

if active_step == STEP_MEMBER_SHARES and membership_mode == "automatic" and membership_valid:
    if st.button("Save Member Share Payments & Continue", type="primary"):
        try:
            save_transition = save_member_share_transition(
                required_steps,
                step_completions,
                saved_payments_key,
                membership_payments,
                subscription_total=subscription_total,
            )
        except ValueError as exc:
            membership_valid = False
            st.error(str(exc), icon="🚫")
        else:
            st.session_state.update(save_transition["saved_payload"])
            st.session_state[workflow_completion_key] = save_transition[
                "completions"
            ]
            st.rerun()

activity_labels = {
    "donation": ("Donations", "Balance Sheet Donation (code 1122)"),
    "paid_out": ("Paid Out", "Balance Sheet Paid Out (code 1114)"),
    "paid_in": ("Paid In", "HASH Paid-Ins (code 34)"),
}
activity_save_labels = {
    "donation": "Save Donations & Continue",
    "paid_in": "Save Paid In & Continue",
    "paid_out": "Save Paid Out & Continue",
}
activity_valid = activity_detection_valid
for activity_key in activity_workflow_keys(activity_source_totals):
    if activity_key not in step_completions:
        continue
    saved_section_key = (
        f"activity_saved_section_{activity_key}_{closeout_workbook_key}"
    )
    try:
        activity_payload[activity_key] = normalize_activity_section(
            activity_key,
            st.session_state[saved_section_key],
        )
    except (KeyError, TypeError, ValueError) as exc:
        st.session_state[workflow_completion_key] = edit_deposit_step(
            required_steps, step_completions, activity_key
        )
        activity_title = activity_labels[activity_key][0]
        st.error(
            f"Saved {activity_title} details need to be reviewed again: {exc}",
            icon="🚫",
        )
        st.rerun()

for activity_key in activity_workflow_keys(activity_source_totals):
    if active_step != activity_key:
        continue
    active_step_panel = active_step_content.container()
    active_step_panel.__enter__()
    activity_title, activity_source_label = activity_labels[activity_key]
    source_total = float(activity_source_totals[activity_key])
    st.caption(f"{activity_source_label}: ${source_total:,.2f}")
    activity_choice_key = (
        f"activity_handling_{activity_key}_{closeout_workbook_key}"
    )
    activity_scroll_request_key = (
        f"activity_scroll_{activity_key}_{closeout_workbook_key}"
    )
    activity_continue_scroll_key = (
        f"activity_continue_scroll_{activity_key}_{closeout_workbook_key}"
    )
    handling_choice = st.radio(
        f"How should {activity_title} be handled?",
        options=["Breakdown in app", "Finish manually in QuickBooks"],
        horizontal=True,
        index=None,
        key=activity_choice_key,
        on_change=queue_breakdown_scroll,
        args=(
            st.session_state,
            activity_choice_key,
            activity_scroll_request_key,
        ),
    )
    if handling_choice is None:
        activity_valid = False
        st.caption(f"Select how you want to handle {activity_title} before building the deposit.")
        continue
    if handling_choice == "Finish manually in QuickBooks":
        activity_payload[activity_key] = {"mode": "quickbooks", "rows": []}
        saved_section_key = (
            f"activity_saved_section_{activity_key}_{closeout_workbook_key}"
        )
        st.session_state[saved_section_key] = activity_payload[activity_key]
        st.info(
            f"The current {activity_title} process stays unchanged. Complete its detail and the Closeout Sheet in QuickBooks.",
            icon="ℹ️",
        )
        st.session_state[workflow_completion_key] = complete_deposit_step(
            required_steps, step_completions, activity_key, "quickbooks"
        )
        st.rerun()

    render_breakdown_scroll_target(
        st,
        components.html,
        st.session_state,
        target_id=f"{activity_key}-breakdown",
        request_key=activity_scroll_request_key,
    )

    if activity_key == "paid_in":
        saved_rows_key = f"activity_saved_rows_{activity_key}_{closeout_workbook_key}"
        entry_version_key = f"activity_entry_version_{activity_key}_{closeout_workbook_key}"
        st.session_state.setdefault(saved_rows_key, [])
        st.session_state.setdefault(entry_version_key, 0)
        entry_key = (
            f"activity_entry_{activity_key}_{closeout_workbook_key}_"
            f"{st.session_state[entry_version_key]}"
        )

        st.markdown("**Add a Paid In item**")
        entry_columns = st.columns(
            [1.2, 2.0, 1.2, 1.0, 1.0],
            vertical_alignment="bottom",
        )
        item_type = entry_columns[0].selectbox(
            "Type",
            options=["ESP Deposit", "Outreach", "Other"],
            key=f"{entry_key}_type",
        )
        if item_type == "ESP Deposit":
            original_date = entry_columns[1].date_input(
                "Original ESP Deposit Date",
                value=deposit_date or date.today(),
                key=f"{entry_key}_date",
            )
            initials = entry_columns[2].text_input(
                "Initials",
                key=f"{entry_key}_initials",
            )
            new_entry = {
                "type": "esp",
                "original_date": original_date,
                "initials": initials,
            }
        else:
            memo = entry_columns[1].text_input(
                "Description / Memo",
                key=f"{entry_key}_memo",
            )
            if item_type == "Outreach":
                entry_columns[2].caption("Posts to 8505000 · Outreach")
                new_entry = {"type": "outreach", "memo": memo}
            else:
                entry_columns[2].caption("Posts to TBA Purchases")
                new_entry = {"type": "other", "memo": memo}
        amount = entry_columns[3].number_input(
            "Amount",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            key=f"{entry_key}_amount",
        )
        add_entry_clicked = entry_columns[4].button(
            "+ Add Paid In item",
            type="secondary",
            use_container_width=True,
            key=f"{entry_key}_add",
        )
        if add_entry_clicked:
            try:
                updated_rows = append_activity_entry(
                    "paid_in",
                    st.session_state[saved_rows_key],
                    {**new_entry, "amount": float(amount)},
                )
            except ValueError as exc:
                st.error(str(exc), icon="🚫")
            else:
                st.session_state[saved_rows_key] = updated_rows
                st.session_state[entry_version_key] += 1
                queue_continue_scroll(
                    st.session_state,
                    activity_continue_scroll_key,
                )
                st.rerun()

        raw_rows = list(st.session_state[saved_rows_key])
        if raw_rows:
            st.markdown("**Added Paid In items**")
            saved_headers = st.columns([1.2, 2.0, 1.2, 1.0, 1.0])
            saved_headers[0].caption("Type")
            saved_headers[1].caption("Description / Memo")
            saved_headers[2].caption("Account")
            saved_headers[3].caption("Amount")
        for row_index, saved_row in enumerate(raw_rows):
            saved_columns = st.columns([1.2, 2.0, 1.2, 1.0, 1.0])
            if saved_row["type"] == "esp":
                saved_columns[0].write("ESP Deposit")
                saved_columns[1].write(
                    f"{saved_row['original_date']} · {saved_row['initials']}"
                )
                saved_columns[2].write("Misc. Receivable")
            elif saved_row["type"] == "outreach":
                saved_columns[0].write("Outreach")
                saved_columns[1].write(saved_row["memo"])
                saved_columns[2].write("8505000 · Outreach")
            else:
                saved_columns[0].write("Other")
                saved_columns[1].write(saved_row["memo"])
                saved_columns[2].write("TBA Purchases")
            saved_columns[3].write(f"${float(saved_row['amount']):,.2f}")
            if saved_columns[4].button(
                "Remove",
                key=(
                    f"remove_activity_{activity_key}_{closeout_workbook_key}_"
                    f"{row_index}"
                ),
            ):
                st.session_state[saved_rows_key] = [
                    row
                    for index, row in enumerate(raw_rows)
                    if index != row_index
                ]
                st.rerun()
    else:
        row_ids_key = f"activity_row_ids_{activity_key}_{closeout_workbook_key}"
        if row_ids_key not in st.session_state:
            st.session_state[row_ids_key] = [uuid4().hex]
        raw_rows = []
        for row_number, row_id in enumerate(st.session_state[row_ids_key], start=1):
            st.markdown(f"**{activity_title} item {row_number}**")
            if activity_key == "donation":
                row_columns = st.columns([1.5, 1.8, 1.2, 0.9, 0.35])
                given_key = f"activity_{activity_key}_{row_id}_given_to"
                purpose_key = f"activity_{activity_key}_{row_id}_purpose"
                manager_key = f"activity_{activity_key}_{row_id}_manager"
                amount_key = f"activity_{activity_key}_{row_id}_amount"
                given_to = row_columns[0].text_input("Given To", key=given_key)
                purpose = row_columns[1].text_input("For", key=purpose_key)
                manager = row_columns[2].text_input("Manager Approval", key=manager_key)
                amount = row_columns[3].number_input(
                    "Amount",
                    min_value=0.0,
                    step=0.01,
                    format="%.2f",
                    key=amount_key,
                )
                widget_keys = (given_key, purpose_key, manager_key, amount_key)
                raw_rows.append(
                    {
                        "given_to": given_to,
                        "purpose": purpose,
                        "manager": manager,
                        "amount": float(amount),
                    }
                )
            else:
                row_columns = st.columns([1.15, 1.8, 1.0, 0.9, 0.35])
                type_key = f"activity_{activity_key}_{row_id}_type"
                amount_key = f"activity_{activity_key}_{row_id}_amount"
                item_type = row_columns[0].selectbox(
                    "Type",
                    options=["ESP Deposit", "Outreach", "Other"],
                    key=type_key,
                )
                if item_type == "ESP Deposit":
                    date_key = f"activity_{activity_key}_{row_id}_date"
                    initials_key = f"activity_{activity_key}_{row_id}_initials"
                    original_date = row_columns[1].date_input(
                        "Original ESP Deposit Date",
                        value=deposit_date or date.today(),
                        key=date_key,
                    )
                    initials = row_columns[2].text_input("Initials", key=initials_key)
                    widget_keys = (type_key, date_key, initials_key, amount_key)
                    raw_row = {
                        "type": "esp",
                        "original_date": original_date,
                        "initials": initials,
                    }
                else:
                    memo_key = f"activity_{activity_key}_{row_id}_memo"
                    memo = row_columns[1].text_input("Description / Memo", key=memo_key)
                    if item_type == "Outreach":
                        row_columns[2].caption("Posts to 8505000 · Outreach")
                        raw_type = "outreach"
                    else:
                        row_columns[2].caption("Posts to TBA Purchases")
                        raw_type = "other"
                    widget_keys = (type_key, memo_key, amount_key)
                    raw_row = {"type": raw_type, "memo": memo}
                amount = row_columns[3].number_input(
                    "Amount",
                    min_value=0.0,
                    step=0.01,
                    format="%.2f",
                    key=amount_key,
                )
                raw_rows.append({**raw_row, "amount": float(amount)})

            delete_key = f"delete_activity_{activity_key}_{row_id}"
            if row_columns[4].button(
                "×",
                key=delete_key,
                help=f"Delete {activity_title} item {row_number}",
                disabled=len(st.session_state[row_ids_key]) == 1,
            ):
                st.session_state[row_ids_key] = [
                    existing_id
                    for existing_id in st.session_state[row_ids_key]
                    if existing_id != row_id
                ]
                for widget_key in (*widget_keys, delete_key):
                    st.session_state.pop(widget_key, None)
                st.rerun()

        if st.button(
            f"+ Add {activity_title} item",
            key=f"add_activity_{activity_key}_{closeout_workbook_key}",
            type="secondary",
        ):
            st.session_state[row_ids_key].append(uuid4().hex)
            queue_continue_scroll(
                st.session_state,
                activity_continue_scroll_key,
            )
            st.rerun()

    render_breakdown_scroll_target(
        st,
        components.html,
        st.session_state,
        target_id=f"{activity_key}-save-and-continue",
        request_key=activity_continue_scroll_key,
    )
    try:
        activity_payload[activity_key] = normalize_activity_section(
            activity_key,
            {"mode": "app", "rows": raw_rows},
        )
        actual_total = round(
            sum(row["amount"] for row in activity_payload[activity_key]["rows"]),
            2,
        )
        discrepancy = round(float(actual_total) - source_total, 2)
        st.success(
            f"{activity_title} actual is ${actual_total:,.2f}. "
            f"Closeout actual minus system total: {discrepancy:+,.2f}.",
            icon="✅",
        )
        if st.button(
            activity_save_labels[activity_key],
            type="primary",
            key=f"save_activity_{activity_key}_{closeout_workbook_key}",
        ):
            saved_section_key = (
                f"activity_saved_section_{activity_key}_{closeout_workbook_key}"
            )
            save_transition = save_activity_transition(
                required_steps,
                step_completions,
                activity_key,
                saved_section_key,
                activity_payload[activity_key],
            )
            activity_payload[activity_key] = save_transition["saved_payload"][
                saved_section_key
            ]
            st.session_state.update(save_transition["saved_payload"])
            st.session_state[workflow_completion_key] = save_transition[
                "completions"
            ]
            st.rerun()
    except ValueError as exc:
        activity_valid = False
        activity_payload[activity_key] = {"mode": "app", "rows": raw_rows}
        st.caption(str(exc))
    active_step_panel.__exit__(None, None, None)

coupon_saved_key = f"coupon_saved_payload_{closeout_workbook_key}"
if STEP_COUPONS in required_steps and STEP_COUPONS in step_completions:
    try:
        saved_coupon_payload = st.session_state[coupon_saved_key]
        coupon_mode = saved_coupon_payload["mode"]
        coupon_closeout_total = float(saved_coupon_payload["closeout_total"])
        coupon_ncg_total = float(saved_coupon_payload["ncg_total"])
        coupon_mfg_total = float(saved_coupon_payload["mfg_total"])
        reconcile_coupon_receivable(
            coupon_bs_total,
            mode=coupon_mode,
            closeout_actual_total=coupon_closeout_total,
            ncg_total=coupon_ncg_total,
            mfg_total=coupon_mfg_total,
        )
    except (KeyError, TypeError, ValueError) as exc:
        st.session_state[workflow_completion_key] = edit_deposit_step(
            required_steps, step_completions, STEP_COUPONS
        )
        st.error(f"Saved Coupons details need to be reviewed again: {exc}", icon="🚫")
        st.rerun()

if uploaded and STEP_COUPONS in required_steps and active_step == STEP_COUPONS:
    active_step_panel = active_step_content.container()
    active_step_panel.__enter__()
    st.caption(f"Balance Sheet Coupons Receivable (code 908): ${coupon_bs_total:,.2f}")
    coupon_choice_key = f"coupon_handling_{closeout_workbook_key}"
    coupon_scroll_request_key = f"coupon_scroll_{closeout_workbook_key}"
    coupon_handling_choice = st.radio(
        "How should Coupons Receivable be handled?",
        options=[
            "Breakdown in app using Closeout Sheet",
            "Finish manually in QuickBooks",
        ],
        horizontal=True,
        index=None,
        key=coupon_choice_key,
        on_change=queue_breakdown_scroll,
        args=(
            st.session_state,
            coupon_choice_key,
            coupon_scroll_request_key,
        ),
    )

    if coupon_handling_choice is None:
        coupon_valid = False
        st.caption("Select how you want to handle Coupons Receivable before building the deposit.")
    elif coupon_handling_choice == "Finish manually in QuickBooks":
        coupon_mode = "quickbooks"
        coupon_payload = {
            "mode": coupon_mode,
            "closeout_total": 0.0,
            "ncg_total": 0.0,
            "mfg_total": 0.0,
        }
        st.session_state[coupon_saved_key] = coupon_payload
        st.session_state[workflow_completion_key] = complete_deposit_step(
            required_steps, step_completions, STEP_COUPONS, "quickbooks"
        )
        st.rerun()
    else:
        coupon_mode = "closeout"
        st.caption(
            "Use the Excel coupon counter if needed, then enter the final NCG and MFG totals below."
        )
        coupon_reference_path = Path(__file__).parent / "assets" / "NCG-MFG Coupon Counter.xlsx"
        if coupon_reference_path.exists():
            with st.expander("Coupon counter reference"):
                st.download_button(
                    "Download Excel coupon counter",
                    data=coupon_reference_path.read_bytes(),
                    file_name="NCG-MFG Coupon Counter.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_coupon_counter_reference",
                )

        render_breakdown_scroll_target(
            st,
            components.html,
            st.session_state,
            target_id="coupon-breakdown",
            request_key=coupon_scroll_request_key,
        )
        coupon_closeout_total = st.number_input(
            "Closeout Sheet Coupon Actual Total",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            key=f"coupon_closeout_{closeout_workbook_key}",
        )

        direct_columns = st.columns(2)
        coupon_ncg_total = direct_columns[0].number_input(
            "NCG Coupons counted",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            key=f"coupon_ncg_{closeout_workbook_key}",
        )
        coupon_mfg_total = direct_columns[1].number_input(
            "MFG Coupons counted",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            key=f"coupon_mfg_{closeout_workbook_key}",
        )

        try:
            coupon_preview = reconcile_coupon_receivable(
                coupon_bs_total,
                mode=coupon_mode,
                closeout_actual_total=coupon_closeout_total,
                ncg_total=coupon_ncg_total,
                mfg_total=coupon_mfg_total,
            )
            discrepancy = coupon_preview["difference"]
            st.success(
                f"Reconciled — counted coupons total ${coupon_closeout_total:,.2f}. "
                f"Closeout Sheet minus Balance Sheet: {discrepancy:+,.2f}.",
                icon="✅",
            )
            if st.button("Save Coupons & Continue", type="primary"):
                coupon_payload = {
                    "mode": coupon_mode,
                    "closeout_total": float(coupon_closeout_total),
                    "ncg_total": float(coupon_ncg_total),
                    "mfg_total": float(coupon_mfg_total),
                }
                st.session_state[coupon_saved_key] = coupon_payload
                st.session_state[workflow_completion_key] = complete_deposit_step(
                    required_steps, step_completions, STEP_COUPONS, "app"
                )
                st.rerun()
        except ValueError as exc:
            coupon_valid = False
            st.warning(str(exc), icon="⚠️")
    active_step_panel.__exit__(None, None, None)

if uploaded and STEP_CLOSEOUT in step_completions:
    try:
        closeout_payload = normalize_closeout_payload(
            st.session_state[closeout_payload_key]
        )
    except (KeyError, TypeError, ValueError) as exc:
        st.session_state[workflow_completion_key] = edit_deposit_step(
            required_steps,
            step_completions,
            STEP_CLOSEOUT,
        )
        st.error(f"Saved Closeout Sheet details need to be reviewed again: {exc}", icon="🚫")
        st.rerun()
    else:
        closeout_valid = True

if uploaded and active_step == STEP_CLOSEOUT:
    active_step_panel = active_step_content.container()
    active_step_panel.__enter__()
    if st.session_state.pop(closeout_hydration_key, False):
        hydrate_reopened_closeout_state(
            st.session_state,
            payload_key=closeout_payload_key,
            preview_key=closeout_preview_key,
            workbook_key=closeout_workbook_key,
        )
    closeout_choice_key = f"closeout_handling_{closeout_workbook_key}"
    closeout_continue_scroll_key = (
        f"closeout_continue_scroll_{closeout_workbook_key}"
    )
    closeout_scroll_request_key = f"closeout_scroll_{closeout_workbook_key}"
    closeout_choice = st.radio(
        "How should the Closeout Sheet be handled?",
        options=[
            "Breakdown in app using Closeout Sheet",
            "Finish manually in QuickBooks",
        ],
        horizontal=True,
        index=None,
        key=closeout_choice_key,
        on_change=queue_breakdown_scroll,
        args=(
            st.session_state,
            closeout_choice_key,
            closeout_scroll_request_key,
        ),
    )

    if closeout_choice is None:
        st.caption("Select how you want to handle the Closeout Sheet before building the deposit.")
    elif closeout_choice == "Finish manually in QuickBooks":
        closeout_payload = {"mode": "manual"}
        closeout_valid = activity_valid
        st.session_state[closeout_payload_key] = closeout_payload
        st.session_state[workflow_completion_key] = complete_deposit_step(
            required_steps,
            step_completions,
            STEP_CLOSEOUT,
            "quickbooks",
        )
        st.rerun()
    else:
        render_breakdown_scroll_target(
            st,
            components.html,
            st.session_state,
            target_id="closeout-breakdown",
            request_key=closeout_scroll_request_key,
        )
        st.markdown("### Closeout Sheet reconciliation")
        st.caption(
            "Enter positive amounts exactly as printed on the paper Closeout Sheet. "
            "The app handles which amounts add to or remove from the deposit."
        )
        coupon_link_ready = (
            STEP_COUPONS not in required_steps
            or (coupon_mode == "closeout" and coupon_valid)
        )
        if not coupon_link_ready:
            st.warning(
                "Choose Breakdown in app using Closeout Sheet for Coupons Receivable first. "
                "Vendor Coupons must use the reconciled NCG + MFG total.",
                icon="⚠️",
            )
        activity_link_ready = activity_valid and activity_closeout_ready(
            activity_payload,
            activity_source_totals,
        )
        if not activity_link_ready:
            st.warning(
                "Every detected Donations, Paid Out, and Paid In section must use Breakdown in app before the Closeout Sheet can be completed here.",
                icon="⚠️",
            )

        closeout_baselines = None
        try:
            closeout_baselines = read_closeout_baselines(
                upload_bytes,
                roles.get("bs"),
                roles.get("hash"),
            )
        except Exception as exc:
            st.error(f"Could not read Closeout baselines: {exc}", icon="🚫")

        current_closeout_payload = None
        closeout_form_error = None
        if closeout_baselines is not None:
            counted_coupon_total = (
                float(coupon_ncg_total or 0) + float(coupon_mfg_total or 0)
                if coupon_mode == "closeout"
                else 0.0
            )
            closeout_defaults = default_closeout_actuals(
                closeout_baselines,
                counted_coupon_total,
            )
            locked_activity_actuals = (
                activity_actuals(activity_payload)
                if activity_link_ready
                else {}
            )
            closeout_actuals = {}
            header_columns = st.columns([1.6, 1.1, 1.3, 1.1, 0.9])
            for column, heading in zip(
                header_columns,
                ["Category", "System / BS", "Actual", "Difference", "Status"],
            ):
                column.caption(heading)
            for field in STANDARD_CLOSEOUT_ORDER:
                row_columns = st.columns([1.6, 1.1, 1.3, 1.1, 0.9])
                label = STANDARD_METADATA[field]["label"]
                baseline = float(closeout_baselines[field])
                row_columns[0].write(label)
                row_columns[1].write(f"${baseline:,.2f}")
                if field == "vendor_coupons":
                    actual = counted_coupon_total
                    row_columns[2].write(f"${actual:,.2f} (NCG + MFG)")
                elif locked_activity_actuals.get(field) is not None:
                    actual = float(locked_activity_actuals[field])
                    row_columns[2].write(f"${actual:,.2f} (breakdown)")
                else:
                    actual = row_columns[2].number_input(
                        f"{label} actual",
                        min_value=0.0,
                        value=float(closeout_defaults[field]),
                        step=0.01,
                        format="%.2f",
                        label_visibility="collapsed",
                        key=f"closeout_actual_{field}_{closeout_workbook_key}",
                    )
                closeout_actuals[field] = float(actual)
                difference = round(float(actual) - baseline, 2)
                row_columns[3].write(f"{difference:+,.2f}")
                row_columns[4].write("Match" if difference == 0 else "Review")

            reviewed_closeout = True

            st.markdown("#### Other Closeout Sheet activity")
            payroll_choice = st.selectbox(
                "Payroll - Check Cashing",
                options=["None", "Adds $4,000", "Removes $4,000"],
                key=f"closeout_payroll_{closeout_workbook_key}",
                help="Use the exact $4,000 choice shown on the paper Closeout Sheet.",
            )
            payroll_value = {
                "None": 0.0,
                "Adds $4,000": 4000.0,
                "Removes $4,000": -4000.0,
            }[payroll_choice]

            plants_purchase = st.number_input(
                "Plants Dept - Market Purchases",
                min_value=0.0,
                value=0.0,
                step=0.01,
                format="%.2f",
                help="Enter a positive amount. This always removes money from the deposit.",
                key=f"closeout_plants_{closeout_workbook_key}",
            )

            safe_columns = st.columns(2)
            safe_choice = safe_columns[0].selectbox(
                "Safe cash",
                options=["None", "Overage", "Shortage"],
                key=f"closeout_safe_type_{closeout_workbook_key}",
                help=(
                    "Overage means cash was added to the deposit. "
                    "Shortage means cash was taken from the deposit."
                ),
            )
            safe_type = safe_choice.casefold()
            safe_amount_key = f"closeout_safe_amount_{closeout_workbook_key}"
            if safe_type == "none":
                st.session_state[safe_amount_key] = 0.0
            safe_amount_entered = safe_columns[1].number_input(
                "Safe amount",
                min_value=0.0,
                value=0.0,
                step=0.01,
                format="%.2f",
                disabled=safe_type == "none",
                key=safe_amount_key,
            )
            safe_amount = 0.0 if safe_type == "none" else float(safe_amount_entered)

            custom_ids_key = f"closeout_custom_ids_{closeout_workbook_key}"
            custom_next_key = f"closeout_custom_next_{closeout_workbook_key}"
            st.session_state.setdefault(custom_ids_key, [])
            st.session_state.setdefault(custom_next_key, 0)
            custom_tba = []
            if st.button(
                "+ Add other item",
                type="secondary",
                key=f"closeout_add_custom_{closeout_workbook_key}",
            ):
                next_id = st.session_state[custom_next_key]
                st.session_state[custom_ids_key] = [
                    *st.session_state[custom_ids_key],
                    next_id,
                ]
                st.session_state[custom_next_key] = next_id + 1
                queue_continue_scroll(
                    st.session_state,
                    closeout_continue_scroll_key,
                )
                st.rerun()
            for custom_id in list(st.session_state[custom_ids_key]):
                custom_columns = st.columns([2.2, 1.1, 1.4, 0.8])
                memo_key = f"closeout_custom_memo_{closeout_workbook_key}_{custom_id}"
                amount_key = f"closeout_custom_amount_{closeout_workbook_key}_{custom_id}"
                direction_key = f"closeout_custom_direction_{closeout_workbook_key}_{custom_id}"
                memo = custom_columns[0].text_input(
                    "Memo",
                    key=memo_key,
                    label_visibility="collapsed",
                    placeholder="What is this item?",
                )
                amount = custom_columns[1].number_input(
                    "Amount",
                    min_value=0.0,
                    value=0.0,
                    step=0.01,
                    format="%.2f",
                    key=amount_key,
                    label_visibility="collapsed",
                )
                direction_label = custom_columns[2].selectbox(
                    "Direction",
                    options=["Adds to deposit", "Removes from deposit"],
                    key=direction_key,
                    label_visibility="collapsed",
                )
                if custom_columns[3].button(
                    "Remove",
                    key=f"closeout_remove_custom_{closeout_workbook_key}_{custom_id}",
                ):
                    st.session_state[custom_ids_key] = [
                        row_id
                        for row_id in st.session_state[custom_ids_key]
                        if row_id != custom_id
                    ]
                    for widget_key in (memo_key, amount_key, direction_key):
                        st.session_state.pop(widget_key, None)
                    st.rerun()
                custom_tba.append(
                    {
                        "memo": memo,
                        "amount": float(amount),
                        "direction": (
                            "adds" if direction_label == "Adds to deposit" else "removes"
                        ),
                    }
                )

            st.markdown("**Final Closeout Sheet Deposit Total — Required**")
            final_closeout_total = st.number_input(
                "Final Closeout Sheet Deposit Total",
                min_value=0.0,
                value=0.0,
                step=0.01,
                format="%.2f",
                key=f"closeout_final_total_{closeout_workbook_key}",
                label_visibility="collapsed",
            )

            try:
                current_closeout_payload = build_closeout_form_payload(
                    baselines=closeout_baselines,
                    actuals=closeout_actuals,
                    reviewed=reviewed_closeout,
                    payroll=payroll_value,
                    safe_type=safe_type,
                    safe_amount=safe_amount,
                    plants_purchase=plants_purchase,
                    custom_tba=custom_tba,
                    final_total=final_closeout_total,
                    approve_final_pos=False,
                )
            except ValueError as exc:
                closeout_form_error = str(exc)

            saved_closeout_preview = st.session_state.get(closeout_preview_key)
            final_approval_key = f"closeout_approve_final_{closeout_workbook_key}"
            closeout_review_context = {
                "deposit_date": deposit_date.isoformat() if deposit_date else None,
                "membership_mode": membership_mode,
                "membership_payments": membership_payments,
                "coupon_mode": coupon_mode,
                "coupon_closeout_total": coupon_closeout_total,
                "coupon_ncg_total": coupon_ncg_total,
                "coupon_mfg_total": coupon_mfg_total,
                "activity_payload": activity_payload,
                "settlement_key": (
                    membership_editor_key(
                        settlement_file.getvalue(),
                        st.session_state["file_uploader_key"],
                    )
                    if settlement_file is not None
                    else None
                ),
            }
            preview_is_fresh = bool(
                current_closeout_payload is not None
                and activity_valid
                and activity_link_ready
                and closeout_preview_is_fresh(
                    current_closeout_payload,
                    saved_closeout_preview,
                    review_context=closeout_review_context,
                )
            )
            review_settlement_ok = bool(
                settlement_file is not None
                and validate_settlement_processed_net_header(
                    settlement_file.getvalue()
                )[0]
            )
            review_blockers = closeout_review_blockers(
                form_error=closeout_form_error,
                coupon_ready=coupon_link_ready,
                activity_ready=activity_link_ready,
                activity_valid=activity_valid,
                membership_valid=membership_valid,
                deposit_date_ready=deposit_date is not None,
                settlement_ready=review_settlement_ok,
            )
            review_help = (
                "Review Closeout is unavailable until:\n\n- "
                + "\n- ".join(review_blockers)
                if review_blockers
                else "Review the full deposit against the Closeout Sheet."
            )
            review_clicked = st.button(
                "Review Closeout",
                type="secondary",
                disabled=bool(review_blockers),
                help=review_help,
                key=f"review_closeout_{closeout_workbook_key}",
            )
            if review_clicked:
                try:
                    with st.spinner("Reviewing the full deposit against the Closeout Sheet..."):
                        with exclusive_run_lock(RUN_LOCK_PATH):
                            review_result = run_engine(
                                uploaded,
                                settlement_file,
                                deposit_date,
                                membership_payments,
                                membership_mode,
                                coupon_mode,
                                coupon_closeout_total,
                                coupon_ncg_total,
                                coupon_mfg_total,
                                activity_payload=activity_payload,
                                closeout_payload=current_closeout_payload,
                                preview_only=True,
                            )
                    st.session_state[closeout_preview_key] = {
                        "input_fingerprint": closeout_input_fingerprint(
                            current_closeout_payload,
                            review_context=closeout_review_context,
                        ),
                        "preview": review_result["closeout_preview"],
                    }
                    st.session_state[final_approval_key] = False
                    st.session_state[closeout_payload_key] = current_closeout_payload
                    st.rerun()
                except Exception as exc:
                    st.error("The Closeout review could not be completed.")
                    st.code(str(exc), language="text")

            approve_final_pos = False
            if preview_is_fresh:
                closeout_preview = saved_closeout_preview["preview"]
                st.markdown("#### Closeout review")
                standard_display = pd.DataFrame(closeout_preview.get("standard_rows", []))
                if not standard_display.empty:
                    standard_display = standard_display.rename(
                        columns={
                            "label": "Category",
                            "baseline": "System / BS",
                            "actual": "Actual",
                            "difference": "Difference",
                        }
                    )
                    st.dataframe(
                        standard_display[["Category", "System / BS", "Actual", "Difference"]],
                        hide_index=True,
                        use_container_width=True,
                    )
                generated_rows = []
                for row in closeout_preview.get("standard_rows", []):
                    if float(row.get("adjustment_qb_effect", 0) or 0) != 0:
                        generated_rows.append(
                            {
                                "Account": row.get("adjustment_account"),
                                "Memo": row.get("adjustment_memo"),
                                "Effect": row.get("adjustment_qb_effect"),
                            }
                        )
                for row in closeout_preview.get("misc_rows", []):
                    generated_rows.append(
                        {
                            "Account": row.get("account"),
                            "Memo": row.get("memo"),
                            "Effect": row.get("qb_effect"),
                        }
                    )
                if generated_rows:
                    st.dataframe(
                        pd.DataFrame(generated_rows),
                        hide_index=True,
                        use_container_width=True,
                    )
                total_columns = st.columns(3)
                total_columns[0].metric(
                    "Generated deposit",
                    f"${float(closeout_preview['provisional_total']):,.2f}",
                )
                total_columns[1].metric(
                    "Paper Closeout total",
                    f"${float(closeout_preview['final_total']):,.2f}",
                )
                remaining = float(closeout_preview["remaining"])
                total_columns[2].metric("Remaining", f"{remaining:+,.2f}")
                if remaining != 0:
                    st.warning(
                        f"The reviewed deposit still differs by {remaining:+,.2f}. "
                        "Approve the exact final POS adjustment only after checking the paper sheet.",
                        icon="⚠️",
                    )
                    approve_final_pos = st.checkbox(
                        "Add final POS adjustment",
                        value=False,
                        key=final_approval_key,
                    )
                current_closeout_payload = {
                    **current_closeout_payload,
                    "approve_final_pos": approve_final_pos,
                }
                closeout_payload = normalize_closeout_payload(current_closeout_payload)
                closeout_valid = bool(
                    (remaining == 0 or approve_final_pos)
                    and coupon_link_ready
                    and activity_link_ready
                    and activity_valid
                )
                st.session_state[closeout_payload_key] = closeout_payload
                render_breakdown_scroll_target(
                    st,
                    components.html,
                    st.session_state,
                    target_id="closeout-save-and-continue",
                    request_key=closeout_continue_scroll_key,
                )
                if closeout_valid and st.button(
                    "Save Closeout Sheet & Continue",
                    type="primary",
                    use_container_width=True,
                ):
                    st.session_state[workflow_completion_key] = complete_deposit_step(
                        required_steps,
                        step_completions,
                        STEP_CLOSEOUT,
                        "app",
                    )
                    st.rerun()
            elif closeout_form_error:
                st.caption(closeout_form_error)
    active_step_panel.__exit__(None, None, None)

run_clicked = False

step_completions = normalize_step_completions(
    required_steps,
    st.session_state.get(workflow_completion_key),
)
guided_workflow_ready = deposit_workflow_complete(
    required_steps,
    step_completions,
) and activity_detection_valid

download_details = deposit_download_details(
    st.session_state.get("run_result")
)
run_clicked = render_prepare_iif_action(
    st,
    visible=bool(uploaded is not None and guided_workflow_ready),
    download_details=download_details,
    disabled=(
        settlement_file is None
        or deposit_date is None
        or not settlement_source_ok
        or not membership_valid
        or not coupon_valid
        or not activity_valid
        or not closeout_valid
    ),
)

if run_clicked:
    try:
        with st.spinner("Reading workbook, running deposit automation, and reconciling QuickBooks lines..."):
            with exclusive_run_lock(RUN_LOCK_PATH):
                result = run_engine(
                    uploaded,
                    settlement_file,
                    deposit_date,
                    membership_payments,
                    membership_mode,
                    coupon_mode,
                    coupon_closeout_total,
                    coupon_ncg_total,
                    coupon_mfg_total,
                    activity_payload=activity_payload,
                    closeout_payload=closeout_payload,
                )
                st.session_state["run_result"] = result
                st.session_state["run_date"] = deposit_date
                st.session_state["run_filename"] = uploaded.name
                st.session_state["run_settlement_filename"] = settlement_file.name
                st.session_state["run_date_mismatch"] = date_info.get("has_mismatch", False)
                history_record = archive_run(
                    uploaded, settlement_file, result, deposit_date, roles, date_info
                )
                st.session_state["last_history_id"] = history_record["id"]
        st.rerun()
    except Exception as exc:
        st.error("The deposit could not be completed.")
        st.code(str(exc), language="text")

# ---------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------

if "run_result" in st.session_state:
    result = st.session_state["run_result"]
    v = result["validation"]
    lines = result["lines"]
    iif_df = result["iif_df"]
    run_date = st.session_state.get("run_date") or deposit_date or date.today()

    st.markdown("---")

    if v["all_ok"]:
        result_class = "good"
        title = "🟢 Deposit balanced"
        sub = "Core reconciliation checks passed. The IIF is ready to download above; review the detail below as needed."
    elif v["warning_count"] > 0:
        result_class = "warn"
        title = "🟡 Deposit needs review"
        sub = "The file was generated, but one or more checks or warnings need attention before QuickBooks import."
    else:
        result_class = "bad"
        title = "🔴 Deposit does not fully reconcile"
        sub = "Review the validation differences below before importing the IIF."

    st.markdown(
        f"""
        <div class="hwfc-result {result_class}">
          <div class="hwfc-result-title">{title}</div>
          <div class="hwfc-result-sub">
            {run_date.strftime('%A, %B %d, %Y')} · {html.escape(st.session_state.get('run_filename', 'Uploaded workbook'))}
          </div>
          <div class="hwfc-result-sub" style="margin-top:5px">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        card("Gross sales", abs_money(v["gross_sales"]), status_word(v["sales_ok"]))
    with c2:
        card("Discounts", abs_money(v["script_discounts"]), status_word(v["discounts_ok"]))
    with c3:
        card("Net sales", abs_money(v["script_net"]), status_word(v["sales_ok"]))
    with c4:
        card("Deposit total", abs_money(v["deposit_total"]), "QuickBooks TRNS amount")

    c5, c6, c7, c8 = st.columns(4)
    with c5:
        card("Excel sales total", abs_money(v["excel_sales"]), "Workbook J3 comparison")
    with c6:
        card("Excel discount total", abs_money(v["excel_discounts"]), "Workbook discount report")
    with c7:
        card("HASH Sales 6", abs_money(v["hash_excel"]), status_word(v["hash_ok"]))
    with c8:
        card("IIF difference", money(v["iif_difference"]), status_word(v["iif_ok"]))

    if v["script_net"] is not None and v["excel_sales"] is not None:
        sales_diff = round(abs(abs(v["script_net"]) - abs(v["excel_sales"])), 2)
        st.markdown(
            f"""
            <div class="hwfc-equation">
              <strong>Sales reconciliation:</strong>
              Script Net Sales {abs_money(v['script_net'])}
              &nbsp; vs. &nbsp;
              Excel Sales {abs_money(v['excel_sales'])}
              &nbsp; → &nbsp;
              <strong>Difference {money(sales_diff)}</strong>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.subheader("Validation checks")
    validation_rows = [
        {"Check": "Sales total", "Source / comparison": "Script Net Sales vs Excel Sales Total", "Status": "✓" if v["sales_ok"] is True else ("✕" if v["sales_ok"] is False else "—")},
        {"Check": "Discount total", "Source / comparison": "Script Discounts vs Excel Discount Total", "Status": "✓" if v["discounts_ok"] is True else ("✕" if v["discounts_ok"] is False else "—")},
        {"Check": "HASH sales", "Source / comparison": "Script HASH total vs HASH Sales 6", "Status": "✓" if v["hash_ok"] is True else ("✕" if v["hash_ok"] is False else "—")},
        {"Check": "IIF balance", "Source / comparison": "Positive vs negative IIF amounts", "Status": "✓" if v["iif_ok"] else "✕"},
        {"Check": "Card settlement", "Source / comparison": "Daily Card Settlement Report vs BS tender totals", "Status": "✓" if v.get("card_settlement_ok") else "✕"},
    ]
    validation_df = pd.DataFrame(validation_rows)
    validation_styled = validation_df.style.map(
        lambda value: (
            "color: #78A85B; font-weight: 900; font-size: 1.05rem;" if value == "✓"
            else "color: #E05A4F; font-weight: 900; font-size: 1.05rem;" if value == "✕"
            else ""
        ),
        subset=["Status"],
    )
    st.dataframe(validation_styled, use_container_width=True, hide_index=True)

    st.subheader("Card Settlement Reconciliation")
    settlement_rows = v.get("card_settlement_rows", [])
    if settlement_rows:
        settlement_df = pd.DataFrame(settlement_rows)
        if "Status" in settlement_df.columns:
            settlement_df["Status"] = settlement_df["Status"].map(lambda value: "✓" if str(value).upper() == "MATCH" else "✕")
        for col in ["Daily Card Settlement", "BS", "Difference", "Adjustment"]:
            settlement_df[col] = settlement_df[col].map(lambda x: f"(${abs(x):,.2f})" if x < 0 else f"${x:,.2f}")
        settlement_styled = settlement_df.style.map(
            lambda value: (
                "color: #78A85B; font-weight: 900; font-size: 1.05rem;" if value == "✓"
                else "color: #E05A4F; font-weight: 900; font-size: 1.05rem;" if value == "✕"
                else ""
            ),
            subset=["Status"] if "Status" in settlement_df.columns else None,
        )
        st.dataframe(settlement_styled, use_container_width=True, hide_index=True)
        if v.get("card_settlement_ok"):
            st.success("All five card settlement amounts match the BS control totals.", icon="✅")
        else:
            st.warning(
                "One or more card settlement amounts differ from the BS tab. "
                "The IIF uses the Daily Card Settlement Report amounts, then posts the signed difference "
                "to 8314000 · FE - Cash Over/Shorts so each tender ties back to the BS control total.",
                icon="⚠️",
            )
    else:
        st.warning("No card settlement reconciliation was found in the engine output.", icon="⚠️")

    with st.expander("More deposit information", expanded=False):
        overview_tab, sales_tab, bs_tab, qb_tab, log_tab = st.tabs(
            ["🌿 Overview", "🛒 Sales & Discounts", "💰 Balance Sheet & Tenders", "📘 QuickBooks Preview", "🧾 Run Log"]
        )

    with overview_tab:
        st.subheader("Deposit Summary")
        summary = build_deposit_summary(lines, v)
        primary_items = [
            ("Store Coupons", summary["Store Coupons"]),
            ("Owner Appreciation", summary["Owner Appreciation"]),
            ("Refunded Discounts", summary["Refunded Discounts"]),
            ("Pass Through Donations", summary["Pass Through Donations"]),
            ("Card Settlement Adjustment", summary["Card Settlement Adjustment"]),
            ("IIF Difference", summary["IIF Difference"]),
        ]

        visible_items = [(label, value) for label, value in primary_items if value is not None]
        for row_start in range(0, len(visible_items), 3):
            row_items = visible_items[row_start:row_start + 3]
            metric_cols = st.columns(len(row_items))
            for metric_col, (label, value) in zip(metric_cols, row_items):
                with metric_col:
                    if label == "Card Settlement Adjustment":
                        st.metric(label, money(value))
                    else:
                        st.metric(label, abs_money(value))

        st.markdown(
            f"""
            <div class="hwfc-equation">
              <strong>IIF balance:</strong>
              Positive amounts {money(v['positive_total'])}
              &nbsp; | &nbsp;
              Negative amounts {money(v['negative_total'])}
              &nbsp; | &nbsp;
              Difference <strong>{money(v['iif_difference'])}</strong>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with sales_tab:
        sales_df = build_detail_df(lines, {"Sales"})
        disc_df = build_detail_df(lines, {"Discounts", "Store Coupons"})

        st.subheader("Sales")
        if sales_df.empty:
            st.info("No sales lines were detected in the generated IIF.")
        else:
            display = sales_df.copy()
            display["Amount"] = display["Amount"].map(lambda x: f"${abs(x):,.2f}" if pd.notna(x) else "")
            st.dataframe(display, use_container_width=True, hide_index=True)
            sales_total = sales_df["Amount"].dropna().abs().sum()
            st.caption(f"Sales lines shown: {len(sales_df)} · Total: ${sales_total:,.2f}")

        st.subheader("Discounts & Store Coupons")
        if disc_df.empty:
            st.info("No discount lines were detected in the generated IIF.")
        else:
            display = disc_df.copy()
            display["Amount"] = display["Amount"].map(lambda x: money(x) if pd.notna(x) else "")
            st.dataframe(display, use_container_width=True, hide_index=True)

    with bs_tab:
        tender_df = build_detail_df(lines, {"Tenders"})
        bs_df = build_detail_df(lines, {"Balance Sheet", "Other"})

        st.subheader("Tender lines")
        if tender_df.empty:
            st.info("No tender lines were classified from the IIF.")
        else:
            display = tender_df.copy()
            display["Amount"] = display["Amount"].map(lambda x: money(x) if pd.notna(x) else "")
            st.dataframe(display, use_container_width=True, hide_index=True)

        st.subheader("Balance Sheet / supporting lines")
        if bs_df.empty:
            st.info("No additional Balance Sheet lines were detected.")
        else:
            display = bs_df.copy()
            display["Amount"] = display["Amount"].map(lambda x: money(x) if pd.notna(x) else "")
            st.dataframe(display, use_container_width=True, hide_index=True)

    with qb_tab:
        st.subheader("QuickBooks IIF preview")
        st.caption("This is the actual generated transaction detail that will be imported into QuickBooks.")

        if iif_df.empty:
            st.info("No TRNS/SPL lines were found in the IIF.")
        else:
            preview = iif_df.copy()
            preview["Amount"] = preview["Amount"].map(lambda x: money(x) if pd.notna(x) else "")
            st.dataframe(preview, use_container_width=True, hide_index=True, height=520)

    with log_tab:
        st.caption("Full engine output for troubleshooting and audit review.")
        st.code(result["log_text"], language="text")

    _, another_col = st.columns([3, 1])
    with another_col:
        if st.button(
            "Run another deposit",
            type="secondary",
            use_container_width=True,
        ):
            reset_current_work()
            st.rerun()

st.markdown(
    """
    <div class="hwfc-footer">
      Honest Weight Food Co-op · Daily Deposit Automation ·
      Always review the generated deposit in QuickBooks before final posting.
    </div>
    """,
    unsafe_allow_html=True,
)
