"""
HWFC  —  Daily Deposit Automation  v5.0
========================================
Generates a QuickBooks IIF that imports directly into
Banking -> Make Deposits -> 1120200 · NBT Bank - Operating Account

RUN:
  py C:\\POS_Automation\\pos_to_quickbooks_v2.py
"""

import sys, re, logging, csv, argparse, json
from datetime import date, datetime, timedelta
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

try:
    from .membership_payments import build_membership_lines, load_membership_payments_file
    from .coupon_reconciliation import reconcile_coupon_receivable
    from .closeout_reconciliation import (
        build_misc_adjustments,
        build_standard_reconciliation,
        calculate_final_pos_adjustment,
        load_closeout_payload_file,
        normalize_closeout_payload,
    )
except ImportError:
    from membership_payments import build_membership_lines, load_membership_payments_file
    from coupon_reconciliation import reconcile_coupon_receivable
    from closeout_reconciliation import (
        build_misc_adjustments,
        build_standard_reconciliation,
        calculate_final_pos_adjustment,
        load_closeout_payload_file,
        normalize_closeout_payload,
    )


def get_project_root() -> Path:
    """Return repository root for Codespaces/source runs."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent


PROJECT_ROOT = get_project_root()
INPUT_DIR = PROJECT_ROOT / "input" / "daily_reports"
QB_IMPORT_DIR = PROJECT_ROOT / "output" / "qb_imports"
SUMMARY_DIR = PROJECT_ROOT / "output" / "summaries"
LOG_DIR = PROJECT_ROOT / "logs"

for _folder in (INPUT_DIR, QB_IMPORT_DIR, SUMMARY_DIR, LOG_DIR):
    _folder.mkdir(parents=True, exist_ok=True)

CONFIG = {
    "pos_export_folder": str(INPUT_DIR),
    "output_folder": str(QB_IMPORT_DIR),
    "company_name":      "HWFC",
    "base_excel_path": str(INPUT_DIR),
    "deposit_account":   "1120200 · NBT Bank - Operating Account",
    "cc_pattern": ["Settlement", "Batch", "Commerce", "CCC", "BusinessTrack"],
}

SUBDEPT_TO_QB = {
    110:  "7110110 · Sales - Packaged Grocery - NT",
    120:  "7110120 · Sales - Dairy/Refrigerated",
    130:  "7110130 · Sales - Frozen Foods",
    140:  "7110141 · Sales - Bulk",
    142:  "7110142 · Sales - Bulk Herbs",
    150:  "7110150 · Sales - Bread",
    160:  "7110160 · Sales - Beer",
    180:  "7110180 · Sales - Pets",
    190:  "7110190 · Sales - General Merchandise",
    210:  "7110210 · Sales - In-House Deli",
    220:  "7110220 · Sales - In-House Bakery",
    230:  "7110210 · Sales - In-House Deli|Catering",
    235:  "7110235 · Sales - Sushi",
    240:  "7110240 · Sales - Java & Juice",
    250:  "7110250 · Sales - Co-op Cafe (ESP)",
    310:  "7110310 · Sales - Cheese",
    320:  "7110320 · Sales - Meat",
    330:  "7110330 · Sales - Specialty Foods",
    340:  "7110340 · Sales - Fish & Seafood",
    350:  "7110350 · Sales - Specialty Mercantile",
    410:  "7110410 · Sales - Produce",
    420:  "7110420 · Sales - Gardening/Plants",
    510:  "7110510 · Sales - Personal Care Taxable",
    520:  "7110520 · Sales - Vitamins & Supplements",
    550:  "7110550 · Sales - Magazines",
    1300: "7111300 · Promotional Sales",
    27:   "8515000 · Marketing - Coupons, Store",
    28:   "8320000 · Store Supplies",
    38:   "4150300 · NYS Paper Bag Fees Payable",
    170:  None,
}

INCOME_ACCOUNTS = {
    "7110110 · Sales - Packaged Grocery - NT",
    "7110120 · Sales - Dairy/Refrigerated",
    "7110130 · Sales - Frozen Foods",
    "7110141 · Sales - Bulk",
    "7110142 · Sales - Bulk Herbs",
    "7110150 · Sales - Bread",
    "7110160 · Sales - Beer",
    "7110180 · Sales - Pets",
    "7110190 · Sales - General Merchandise",
    "7110210 · Sales - In-House Deli",
    "7110220 · Sales - In-House Bakery",
    "7110235 · Sales - Sushi",
    "7110240 · Sales - Java & Juice",
    "7110250 · Sales - Co-op Cafe (ESP)",
    "7110310 · Sales - Cheese",
    "7110320 · Sales - Meat",
    "7110330 · Sales - Specialty Foods",
    "7110340 · Sales - Fish & Seafood",
    "7110350 · Sales - Specialty Mercantile",
    "7110410 · Sales - Produce",
    "7110420 · Sales - Gardening/Plants",
    "7110510 · Sales - Personal Care Taxable",
    "7110520 · Sales - Vitamins & Supplements",
    "7110550 · Sales - Magazines",
    "7111300 · Promotional Sales",
}

# Level 15 is Student Discount Sun and posts to College Day.
SHOPPER_LEVEL_TO_QB = {
    2:  "8512001 · Discount 2% - Owners",
    3:  "8511001 · Discount 2% - Senior Non Owner",
    4:  "8511003 · Discount 5% - Senior Owners",
    5:  "8140010 · Monthly Time Discount (8%)",
    6:  "8140026 · Weekly Time Discount (24%)",
    7:  "8423100 · Discount - Staff  (24%)",
    8:  "8512002 · Discount 2% - Visiting Coop",
    9:  "8512003 · Discount 8% - Vendors",
    10: "8511001 · Discount 2% - Senior Non Owner",
    15: "8512005 · Discount 8% - College Day",
    19: "8512007 · Discount 15% - Non-profit",
}

CC_ACCOUNT = "1240001 · Credit Card Payments Receivable"
CC_TYPE_MAP = {
    "Visa/MC":             ["visa", "mastercard", "mc", "visa/mc"],
    "Discover":            ["discover"],
    "AMEX":                ["amex", "american express"],
    "Debit Card":          ["debit", "pin debit"],
    "EBT Cash/Food Stamp": ["ebt", "food stamp", "snap"],
}

output_dir = QB_IMPORT_DIR
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


def clean_amount(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip().replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_rows(filepath: Path) -> list:
    rows = []
    with open(filepath, encoding="utf-8-sig", errors="replace", newline="") as f:
        for row in csv.reader(f):
            rows.append([cell.strip() for cell in row])
    return rows


def parse_sms_file(filepath: Path) -> tuple:
    rows = read_rows(filepath)
    log.info(f"  Parsing {filepath.name} ({len(rows)} rows)")
    sales_raw, discount_raw = {}, {}
    mode = None
    in_data = False

    for row in rows:
        if not any(row):
            continue
        joined = ",".join(row)
        if "Sub-department Single Total" in joined or "Sub-Department Single Total" in joined:
            mode, in_data = "subdept", False
            continue
        if "Discounts by Shopper Level" in joined:
            mode, in_data = "discount", False
            continue
        if mode == "subdept" and "Sub-Department" in joined:
            in_data = True
            continue
        if mode == "discount" and "Description" in joined:
            in_data = True
            continue
        if not in_data:
            continue
        low = joined.lower()
        if any(x in low for x in ["total", "printed", "grand total", "member discount", "honest weight", "date:", "target:", "tlz.", "s-dept"]):
            continue

        if mode == "subdept":
            dept_num = None
            for cell in row:
                if re.match(r"^\d+$", cell):
                    try:
                        dept_num = int(cell)
                        break
                    except ValueError:
                        pass
            if dept_num is None:
                continue
            amt = 0.0
            for col_idx in [7, 8, 6, 9]:
                if col_idx < len(row) and row[col_idx]:
                    candidate = clean_amount(row[col_idx])
                    if candidate != 0.0:
                        amt = candidate
                        break
            if amt == 0.0:
                if dept_num == 170:
                    log.info("    Dept 170 Adjustments — skipped")
                continue
            sales_raw[dept_num] = round(sales_raw.get(dept_num, 0.0) + amt, 2)

        elif mode == "discount":
            code = None
            for cell in row:
                if re.match(r"^\d{1,2}$", cell):
                    try:
                        code = int(cell)
                        break
                    except ValueError:
                        pass
            if code is None:
                continue
            amt = 0.0
            for col_idx in [8, 7, 9]:
                if col_idx < len(row) and row[col_idx]:
                    candidate = clean_amount(row[col_idx])
                    if candidate != 0.0:
                        amt = candidate
                        break
            if amt == 0.0:
                continue
            discount_raw[code] = round(discount_raw.get(code, 0.0) + amt, 2)

    sales_totals = {}
    log.info("\n  ── Sales ──")
    for num, amt in sorted(sales_raw.items()):
        qb = SUBDEPT_TO_QB.get(num)
        if qb is None:
            if num in SUBDEPT_TO_QB:
                log.info(f"    Dept {num:>5}  SKIPPED")
            else:
                log.warning(f"    Dept {num:>5}  ${amt:,.2f}  NOT IN MAP")
            continue
        log.info(f"    Dept {num:>5}  ${amt:>10,.2f}  →  {qb}")
        sales_totals[qb] = round(sales_totals.get(qb, 0.0) + amt, 2)

    discount_totals = {}
    log.info("\n  ── Discounts ──")
    for code, amt in sorted(discount_raw.items()):
        qb = SHOPPER_LEVEL_TO_QB.get(code)
        if qb:
            log.info(f"    Level {code:>2}  ${amt:>9,.2f}  →  {qb}")
            discount_totals[qb] = round(discount_totals.get(qb, 0.0) + amt, 2)
        else:
            log.warning(f"    Level {code}  ${amt:,.2f}  NOT MAPPED")

    return sales_totals, discount_totals


def parse_cc_file(filepath: Path) -> dict:
    rows = read_rows(filepath)
    log.info(f"  Parsing CC: {filepath.name}")
    header_idx = None
    for i, row in enumerate(rows):
        joined = ",".join(row).lower()
        if any(x in joined for x in ["card type", "cardtype", "tender", "payment type"]):
            header_idx = i
            break
    if header_idx is None:
        log.warning(f"  Could not find header in {filepath.name}")
        return {}
    headers = [h.lower() for h in rows[header_idx]]
    type_idx = next((i for i, h in enumerate(headers) if any(x in h for x in ["card type", "cardtype", "tender", "type"])), None)
    amt_idx = next((i for i, h in enumerate(headers) if any(x in h for x in ["net", "amount", "total", "sales"])), None)
    if type_idx is None or amt_idx is None:
        log.warning(f"  CC columns not found. Headers: {headers}")
        return {}
    raw_totals = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= max(type_idx, amt_idx):
            continue
        card = row[type_idx].lower().strip()
        amt = clean_amount(row[amt_idx])
        if not card or amt == 0:
            continue
        raw_totals[card] = round(raw_totals.get(card, 0.0) + amt, 2)
    totals = {}
    log.info("\n  ── Credit Cards ──")
    for card_raw, amt in raw_totals.items():
        label = next((lbl for lbl, kws in CC_TYPE_MAP.items() if any(kw in card_raw for kw in kws)), card_raw.title())
        key = f"{CC_ACCOUNT}|{label}"
        log.info(f"    {label:30s}  ${amt:>10,.2f}")
        totals[key] = round(totals.get(key, 0.0) + amt, 2)
    return totals


def parse_coupon_file(filepath: Path) -> tuple:
    rows = read_rows(filepath)
    log.info(f"  Parsing coupon file: {filepath.name}")
    per_dept, grand_total, in_data = {}, 0.0, False
    for row in rows:
        if not any(row):
            continue
        joined = ",".join(row)
        if "Sub-Department" in joined:
            in_data = True
            continue
        if not in_data:
            continue
        low = joined.lower()
        if "total" in low and "printed" not in low:
            amt = clean_amount(row[7]) if len(row) > 7 and row[7] else 0.0
            if abs(amt) > 0:
                grand_total = abs(amt)
                log.info(f"    Coupon grand total: ${grand_total:.2f}")
            continue
        if "printed" in low:
            continue
        if len(row) > 7:
            try:
                dept_num = int(str(row[1]).strip())
            except (ValueError, TypeError):
                continue
            amt = clean_amount(row[7])
            if abs(amt) > 0:
                per_dept[dept_num] = round(per_dept.get(dept_num, 0.0) + abs(amt), 2)
                log.info(f"    Coupon dept {dept_num:>5}: ${abs(amt):.2f}")
    return per_dept, grand_total


MISC_SUBDEPTS = {
    170: "Adjustments", 22: "Coop Scoop Ad payment", 23: "Refunded Discounts",
    24: "Bottle Deposits", 25: "Gift Certificates", 26: "Share Payment",
    29: "Building Blocks", 30: "Groupon", 31: "Crowd Savings",
    32: "Pass Through Donations", 33: "Bag Credits", 34: "Paid-Ins",
    35: "Owner Appreciation 5%", 37: "Staff Appreciation", 50: "Envirotokens",
    260: "Maria College Cafe", 530: "Herbs", 540: "Books",
    560: "Candles/Incense/Baskets", 999: "UnAssigned",
}

def parse_excel_report(filepath: Path) -> tuple:
    import openpyxl
    log.info(f"  Reading Excel report: {filepath.name}")
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        if "WinError 2" in str(e) or "cannot find the file" in str(e).lower() or "openpyx" in str(e).lower():
            raise RuntimeError(
                f"Cannot read Excel file — please CLOSE IT IN EXCEL first, then run again.\n"
                f"File: {filepath.name}"
            ) from e
        raise

    if 'SubDept Sales Report' not in wb.sheetnames:
        log.error(f"  Sheet 'SubDept Sales Report' not found in {filepath.name}")
        log.error(f"  Available sheets: {wb.sheetnames}")
        return {}, [], 0.0, 0.0, 0.0

    ws = wb['SubDept Sales Report']
    sales = {}
    misc_lines = []
    milk_bottle_return = 0.0
    store_coupons_amt = 0.0
    owner_apprec_amt = 0.0
    sales_total_xl = 0.0
    pass_through_total = 0.0
    dust_bunnies_total = 0.0
    milk_bottles_returns = 0.0
    refunded_discounts = 0.0
    hash_sales_total = 0.0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            val = row[12] if len(row) > 12 else None
            if val is not None:
                try:
                    milk_bottle_return = -abs(float(val))
                    log.info(f"    Milk Bottle Return (M1): ${milk_bottle_return:.2f}")
                except (ValueError, TypeError):
                    pass
            val_m1_mbr = row[12] if len(row) > 12 else None
            if val_m1_mbr is not None:
                try:
                    milk_bottles_returns = float(val_m1_mbr)
                    log.info(f"    Milk Bottles Returns (M1): ${milk_bottles_returns:.2f}")
                except (ValueError, TypeError):
                    pass
            val_o1 = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
            val_p1 = row[15] if len(row) > 15 else None
            if val_p1 is not None:
                try:
                    p1_val = float(val_p1)
                    if "refunded" in val_o1:
                        log.info(f"    Refunded Discounts found in O1/P1 (${p1_val:.2f}) — HASH tab is authoritative")
                    elif "hash" in val_o1:
                        hash_sales_total = p1_val
                        log.info(f"    Hash Sales (O1/P1): ${p1_val:.2f}")
                    elif "pass through" in val_o1 or "donation" in val_o1:
                        log.info(f"    Pass Through found in O1/P1 (${p1_val:.2f}) — HASH tab is authoritative")
                    elif "dust" in val_o1:
                        dust_bunnies_total = p1_val
                        log.info(f"    Dust Bunnies (O1/P1): ${p1_val:.2f}")
                except (ValueError, TypeError):
                    pass
            continue

        if i == 2:
            val_j = row[9] if len(row) > 9 else None
            val_m = row[12] if len(row) > 12 else None
            if val_j is not None:
                try:
                    owner_apprec_amt = float(val_j)
                    log.info(f"    Owner Apprec 5% (J2): ${owner_apprec_amt:.2f}")
                except (ValueError, TypeError):
                    pass
            if val_m is not None:
                try:
                    store_coupons_amt = float(val_m)
                    log.info(f"    Store Coupons (M2): ${store_coupons_amt:.2f}")
                except (ValueError, TypeError):
                    pass
            val_o2 = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
            val_p2 = row[15] if len(row) > 15 else None
            if val_p2 is not None:
                try:
                    p2_val = float(val_p2)
                    if "refunded" in val_o2:
                        log.info(f"    Refunded Discounts found in O2/P2 (${p2_val:.2f}) — HASH tab is authoritative")
                    elif "hash" in val_o2:
                        hash_sales_total = p2_val
                        log.info(f"    Hash Sales (O2/P2): ${p2_val:.2f}")
                    elif "pass through" in val_o2 or "donation" in val_o2:
                        log.info(f"    Pass Through found in O2/P2 (${p2_val:.2f}) — HASH tab is authoritative")
                    elif "dust" in val_o2:
                        dust_bunnies_total = p2_val
                        log.info(f"    Dust Bunnies (O2/P2): ${p2_val:.2f}")
                except (ValueError, TypeError):
                    pass
            continue

        if i == 3:
            val_j3 = row[9] if len(row) > 9 else None
            if val_j3 is not None:
                try:
                    sales_total_xl = round(float(val_j3), 2)
                    log.info(f"    Excel Sales Total (J3): ${sales_total_xl:,.2f}")
                except (ValueError, TypeError):
                    pass
            val_o3 = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
            val_p3 = row[15] if len(row) > 15 else None
            if val_p3 is not None:
                try:
                    p3_val = float(val_p3)
                    if "dust" in val_o3:
                        dust_bunnies_total = p3_val
                        log.info(f"    Dust Bunnies (O3/P3): ${p3_val:.2f}")
                    elif "refunded" in val_o3:
                        log.info(f"    Refunded Discounts found in O3/P3 (${p3_val:.2f}) — HASH tab is authoritative")
                    elif "hash" in val_o3:
                        hash_sales_total = p3_val
                        log.info(f"    Hash Sales (O3/P3): ${p3_val:.2f}")
                    elif "pass through" in val_o3 or "donation" in val_o3:
                        log.info(f"    Pass Through found in O3/P3 (${p3_val:.2f}) — HASH tab is authoritative")
                except (ValueError, TypeError):
                    pass
            continue

        dept_num = row[0]
        amount = row[6] if len(row) > 6 else None
        if dept_num is None:
            continue
        try:
            dept_num = int(dept_num)
        except (ValueError, TypeError):
            continue
        if amount is None:
            amount = 0.0
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.0

        qb = SUBDEPT_TO_QB.get(dept_num)
        if qb is None:
            if dept_num in MISC_SUBDEPTS:
                if amount != 0.0:
                    memo = MISC_SUBDEPTS[dept_num]
                    misc_lines.append((memo, round(amount, 2)))
                    log.info(f"    Dept {dept_num:>5}  ${amount:>10,.2f}  → TBA Purchases ({memo})")
            elif dept_num in SUBDEPT_TO_QB:
                log.info(f"    Dept {dept_num:>5}  SKIPPED")
            else:
                if amount != 0.0:
                    misc_lines.append((f"Dept {dept_num}", round(amount, 2)))
                    log.warning(f"    Dept {dept_num:>5}  ${amount:,.2f}  → TBA Purchases (unmapped dept — add to MISC_SUBDEPTS)")
                else:
                    log.info(f"    Dept {dept_num:>5}  zero — skipped")
            continue

        log.info(f"    Dept {dept_num:>5}  ${amount:>10,.2f}  →  {qb}")
        sales[qb] = round(sales.get(qb, 0.0) + amount, 2)

    log.info(f"  Excel: {len(sales)} accounts mapped, {len(misc_lines)} misc TBA lines")
    return sales, misc_lines, milk_bottle_return, store_coupons_amt, owner_apprec_amt, sales_total_xl, pass_through_total, dust_bunnies_total, milk_bottles_returns, refunded_discounts, hash_sales_total


def _dated_report_sheet(wb, report_date, report_label: str):
    date_names = {
        f"{report_date.strftime('%m%d%y')} {report_label}".casefold(),
        f"{report_date.month}{report_date.strftime('%d%y')} {report_label}".casefold(),
    }
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().casefold() in date_names:
            return wb[sheet_name], sheet_name
    return None, None


def _content_report_sheet(wb, required_phrases: tuple[str, ...]):
    for sheet_name in wb.sheetnames:
        if "xxxxxx" in sheet_name.casefold():
            continue
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row, 25),
            values_only=True,
        ):
            joined = " ".join(str(value or "") for value in row).casefold()
            if any(phrase in joined for phrase in required_phrases):
                return sheet, sheet_name
    return None, None


def parse_hash_sheet(filepath: Path, report_date) -> tuple:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws, found_tab = _dated_report_sheet(wb, report_date, "Hash")
    if ws is None:
        ws, found_tab = _content_report_sheet(
            wb,
            ("refunded discounts", "pass through donations", "paid-ins", "paid ins"),
        )

    if ws is None:
        log.warning(f"  No HASH sheet found in {filepath.name}")
        wb.close()
        return 0.0, 0.0, 0.0

    log.info(f"  Reading HASH sheet: '{found_tab}'")
    amount_col = None
    amount_header_row = None
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        for idx, value in enumerate(row):
            if value is None:
                continue
            label = str(value).strip().casefold()
            if label == "amount":
                amount_col = idx
                amount_header_row = row_num
                break
        if amount_col is not None:
            break

    if amount_col is None:
        log.warning("  HASH Amount column not found — HASH values cannot be imported safely.")
        wb.close()
        return 0.0, 0.0, 0.0

    log.info(f"  HASH Amount header found at row {amount_header_row}, column {amount_col + 1}")
    refunded_discounts = 0.0
    pass_through_total = 0.0
    paid_in_total = 0.0
    target_codes = {23: "refunded", 32: "pass_through", 34: "paid_in"}

    for excel_row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        code = None
        for idx in range(min(4, len(row))):
            value = row[idx]
            if value is None:
                continue
            try:
                candidate_code = int(float(value))
            except (TypeError, ValueError):
                continue
            if candidate_code in target_codes:
                code = candidate_code
                break
        if code is None:
            continue
        if len(row) <= amount_col:
            log.warning(f"    HASH row {excel_row_num} code {code}: Amount column missing")
            continue
        raw_amount = row[amount_col]
        try:
            amount = round(float(raw_amount), 2)
        except (TypeError, ValueError):
            amount = None
        if amount is None:
            fallback_values = []
            for idx, value in enumerate(row):
                if idx < 3 or value is None:
                    continue
                try:
                    num = float(value)
                except (TypeError, ValueError):
                    continue
                if abs(num - code) < 0.000001:
                    continue
                fallback_values.append((idx, num))
            if fallback_values:
                amount = round(fallback_values[1][1] if len(fallback_values) >= 2 else fallback_values[0][1], 2)
        if amount is None:
            log.warning(f"    HASH row {excel_row_num} code {code}: could not read Amount")
            continue
        row_type = target_codes[code]
        if row_type == "refunded":
            refunded_discounts = amount
            log.info(f"    HASH code 23 Refunded Discounts: ${amount:,.2f}")
        elif row_type == "pass_through":
            pass_through_total = amount
            log.info(f"    HASH code 32 Pass Through Donations: ${amount:,.2f}")
        elif row_type == "paid_in":
            paid_in_total = amount
            log.info(f"    HASH code 34 Paid-Ins: ${amount:,.2f}")

    log.info(f"  HASH values used: Refunded=${refunded_discounts:,.2f}, PassThrough=${pass_through_total:,.2f}, PaidIn=${paid_in_total:,.2f}")
    wb.close()
    return refunded_discounts, pass_through_total, paid_in_total


def parse_excel_discounts(filepath: Path, report_date) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws, found_tab = _dated_report_sheet(wb, report_date, "Discounts")
    if ws is None:
        ws, found_tab = _content_report_sheet(wb, ("discounts by shopper level",))
    if ws is None:
        log.warning(f"  No discounts sheet found in {filepath.name}")
        log.warning(f"  Available: {wb.sheetnames}")
        wb.close()
        return {}, 0.0

    log.info(f"  Reading discounts from sheet: '{found_tab}'")
    WEDNESDAY_SENIOR_LEVELS = {3, 4, 10}
    WEDNESDAY_SENIOR_ACCT = "8511002 · Discount 8% - Senior Day"
    is_wednesday = report_date.weekday() == 2
    if is_wednesday:
        log.info(f"  Wednesday — combining Senior levels 3, 4, 10 → {WEDNESDAY_SENIOR_ACCT}")

    discounts = {}
    grand_total = 0.0
    for row in ws.iter_rows(values_only=True):
        if row[2] is not None and "member discount" in str(row[2]).lower():
            try:
                grand_total = abs(float(str(row[8])))
                log.info(f"    Discounts grand total: ${grand_total:,.2f}")
            except (ValueError, TypeError):
                pass
            continue

        code = row[3] if len(row) > 3 else None
        amt = row[8] if len(row) > 8 else None
        desc_candidates = []
        for idx in (2, 4, 1, 5):
            if len(row) > idx and row[idx] is not None:
                val = str(row[idx]).strip()
                if val and not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", val):
                    desc_candidates.append(val)
        shopper_desc = " | ".join(dict.fromkeys(desc_candidates)) if desc_candidates else "(no description found)"

        if code is None or amt is None:
            continue
        try:
            code = int(float(str(code)))
            amt = float(str(amt))
        except (ValueError, TypeError):
            continue
        if amt == 0:
            continue

        if is_wednesday and code in WEDNESDAY_SENIOR_LEVELS:
            discounts[WEDNESDAY_SENIOR_ACCT] = round(discounts.get(WEDNESDAY_SENIOR_ACCT, 0.0) + abs(amt), 2)
            log.info(f"    Level {code:>2}  ${amt:>9,.2f}  →  {WEDNESDAY_SENIOR_ACCT} (Senior Day)")
            continue

        qb = SHOPPER_LEVEL_TO_QB.get(code)
        if qb:
            discounts[qb] = round(discounts.get(qb, 0.0) + abs(amt), 2)
            log.info(f"    Level {code:>2}  ${amt:>9,.2f}  →  {qb}")
        else:
            if code == 1:
                log.warning(f"    Level 1  ${amt:,.2f}  NOT MAPPED — POS description: {shopper_desc}")
                log.warning("    Level 1 mapping candidate: 8512006 · Discount 5% - Owner buy Local (UNCONFIRMED — do not post automatically yet)")
            else:
                log.warning(f"    Level {code}  ${amt:,.2f}  NOT MAPPED — POS description: {shopper_desc}")

    log.info(f"  Discounts: {len(discounts)} accounts, grand total=${grand_total:,.2f}")
    wb.close()
    return discounts, grand_total



def parse_card_settlement_report(filepath: Path) -> dict:
    """Read the Daily Card Settlement Report.

    The Processed Net Amount is the bank-settlement source of truth for the
    five tender lines imported to QuickBooks. VISA and MASTERCARD are combined
    into VISA/MC. Labels are matched case-insensitively so minor formatting
    changes such as trailing asterisks do not break the parser.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    report_date = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        for idx, value in enumerate(row):
            label = str(value or "").strip().lower()
            if label == "date" and idx + 1 < len(row):
                raw = row[idx + 1]
                if isinstance(raw, datetime):
                    report_date = raw.date()
                elif isinstance(raw, date):
                    report_date = raw
                elif raw:
                    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
                        try:
                            report_date = datetime.strptime(str(raw).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                if report_date:
                    break
        if report_date:
            break

    header_row = None
    network_col = None
    net_col = None
    for rnum, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True),
        start=1,
    ):
        normalized = [str(v or "").strip().lower() for v in row]
        for idx, label in enumerate(normalized):
            if label == "network":
                network_col = idx
            if label == "processed net amount":
                net_col = idx
        if network_col is not None and net_col is not None:
            header_row = rnum
            break

    if header_row is None:
        raise ValueError(
            f"Daily Card Settlement Report columns not found in {filepath.name}. "
            "Expected exact headers: Network and Processed Net Amount. "
            "No other amount column will be used."
        )

    log.info("  Card Settlement source column: Processed Net Amount")

    totals = {
        "visa_mc": 0.0,
        "discover": 0.0,
        "amex": 0.0,
        "debit": 0.0,
        "ebt": 0.0,
    }

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= max(network_col, net_col):
            continue
        raw_network = str(row[network_col] or "").strip()
        network = re.sub(r"[^a-z0-9]+", "", raw_network.lower())
        try:
            amount = round(float(row[net_col]), 2)
        except (TypeError, ValueError):
            continue

        if network.startswith("visa") or network.startswith("mastercard") or network == "mc":
            totals["visa_mc"] = round(totals["visa_mc"] + amount, 2)
        elif network.startswith("discover"):
            totals["discover"] = round(totals["discover"] + amount, 2)
        elif network.startswith("amex") or network.startswith("americanexpress"):
            totals["amex"] = round(totals["amex"] + amount, 2)
        elif network.startswith("debit"):
            totals["debit"] = round(totals["debit"] + amount, 2)
        elif network.startswith("ebt"):
            totals["ebt"] = round(totals["ebt"] + amount, 2)

    totals["report_date"] = report_date
    totals["source_file"] = filepath.name

    log.info(
        "  Card Settlement: "
        f"VISA/MC=${totals['visa_mc']:,.2f} "
        f"Discover=${totals['discover']:,.2f} "
        f"AMEX=${totals['amex']:,.2f} "
        f"Debit=${totals['debit']:,.2f} "
        f"EBT=${totals['ebt']:,.2f}"
    )
    return totals

def parse_bs_sheet(filepath: Path, report_date) -> dict:
    import openpyxl
    mmddyy = f"{report_date.month}{report_date.strftime('%d%y')}"
    tab_candidates = [f"{mmddyy} BS", f"0{mmddyy} BS", report_date.strftime("%m%d%y BS")]
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    for candidate in tab_candidates:
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            log.info(f"  Reading BS sheet: '{candidate}'")
            break
    if ws is None:
        for sheet in wb.sheetnames:
            if sheet.upper().endswith(" BS"):
                ws = wb[sheet]
                log.info(f"  Reading BS sheet: '{sheet}'")
                break
    if ws is None:
        log.warning(f"  No BS sheet found in {filepath.name}")
        wb.close()
        return {}

    bs = {
        "sales_tax": 0.0, "bottle_sales": 0.0, "milk_bottle_fee": 0.0,
        "milk_bottle_return": 0.0, "bottle_return": 0.0, "charity": 0.0,
        "prepaid_increase": 0.0, "penny_round": 0.0, "visa_mc": 0.0,
        "amex": 0.0, "discover": 0.0, "debit": 0.0, "ebt_cash": 0.0,
        "ebt_food": 0.0, "dufb": 0.0, "cash": 0.0, "check": 0.0,
        "vendor_coupon": 0.0, "charge": 0.0, "prepaid_card": 0.0,
        "donation": 0.0, "subscription": 0.0, "paid_out": 0.0,
        "offline_credit_card": 0.0,
    }

    for row in ws.iter_rows(values_only=True):
        code = row[0]
        amt = row[4] if len(row) > 4 else None
        bal = row[7] if len(row) > 7 else None
        def to_float(v):
            try:
                return float(v)
            except Exception:
                return 0.0

        if row[1] == "Total" and row[2] == "Taxes":
            bs["sales_tax"] = to_float(bal)
            log.info(f"    Sales Tax: ${bs['sales_tax']:,.2f}")
        if code == 39: bs["bottle_sales"] = to_float(amt)
        if code == 40: bs["milk_bottle_fee"] = to_float(amt)
        if code == 205: bs["charity"] = to_float(amt)
        if code == 207: bs["penny_round"] = to_float(amt)
        if code == 208: bs["prepaid_increase"] = to_float(amt)
        if code == 910: bs["milk_bottle_return"] = to_float(amt)
        if code == 911: bs["bottle_return"] = to_float(amt)
        if code == 901: bs["cash"] = to_float(amt)
        if code == 902: bs["check"] = to_float(amt)
        if code == 903: bs["debit"] = to_float(amt)
        if code == 920: bs["ebt_cash"] = to_float(amt)
        if code == 921: bs["ebt_food"] = to_float(amt)
        if code == 928: bs["dufb"] = to_float(amt)
        if code == 930: bs["visa_mc"] = round(bs["visa_mc"] + to_float(amt), 2)
        if code == 931: bs["visa_mc"] = round(bs["visa_mc"] + to_float(amt), 2)
        if code == 932: bs["amex"] = to_float(amt)
        if code == 933: bs["discover"] = to_float(amt)
        if code == 934 and not bs["offline_credit_card"]:
            bs["offline_credit_card"] = -abs(to_float(amt))
        if code == 1334:
            bs["offline_credit_card"] = -abs(to_float(amt))
        if code == 980: bs["prepaid_card"] = to_float(amt)
        if code == 1117: bs["prepaid_card"] = round(bs["prepaid_card"] + to_float(amt), 2)
        if code == 906: bs["charge"] = to_float(amt)
        if code == 908: bs["vendor_coupon"] = to_float(amt)
        if code == 1114:
            bs["paid_out"] = to_float(amt)
            log.info(f"    Paid Out (1114 PkUp Paid out): ${bs['paid_out']:,.2f} — reduces the QuickBooks deposit")
        if code == 1122: bs["donation"] = to_float(amt)
        if code == 3420: bs["subscription"] = to_float(amt)

    log.info(f"  BS: Tax=${bs['sales_tax']:,.2f} BottleSales=${bs['bottle_sales']:,.2f} Fee=${bs['milk_bottle_fee']:,.2f} Charity=${bs['charity']:,.2f} Visa/MC=${bs['visa_mc']:,.2f} AMEX=${bs['amex']:,.2f} Discover=${bs['discover']:,.2f} Debit=${bs['debit']:,.2f}")
    wb.close()
    return bs


def find_todays_files(deposit_date=None):
    folder = Path(CONFIG["pos_export_folder"])
    selected_date = deposit_date or date.today()
    date_patterns = [
        selected_date.strftime("%Y-%m-%d"), selected_date.strftime("%m-%d-%Y"),
        selected_date.strftime("%m%d%Y"), selected_date.strftime("%Y%m%d"),
        selected_date.strftime("%m%d%y"), selected_date.strftime("%m-%d-%y"),
        selected_date.strftime("%-m.%-d.%Y") if sys.platform != "win32" else selected_date.strftime("%#m.%#d.%Y"),
    ]
    all_files = [f for f in folder.iterdir() if f.is_file() and not f.name.startswith("~$")]
    dated_files = [f for f in all_files if any(p in f.name for p in date_patterns)]
    candidates = dated_files if dated_files else all_files
    if not candidates:
        raise FileNotFoundError(f"No input files found in {folder}. Upload the day's files into input/daily_reports and run again.")

    sms_files, cc_file, coupon_files, excel_files, settlement_files = [], None, [], [], []
    for f in candidates:
        suffix = f.suffix.lower()
        name_up = f.name.upper()
        if suffix in {".xlsx", ".xlsm"}:
            is_settlement = "DAILY CARD SETTLEMENT" in name_up or "CARD SETTLEMENT" in name_up
            if not is_settlement:
                try:
                    import openpyxl
                    wb_probe = openpyxl.load_workbook(f, read_only=True, data_only=True)
                    ws_probe = wb_probe[wb_probe.sheetnames[0]]
                    probe = " ".join(
                        str(ws_probe.cell(r, c).value or "")
                        for r in range(1, min(ws_probe.max_row, 8) + 1)
                        for c in range(1, min(ws_probe.max_column, 6) + 1)
                    ).upper()
                    is_settlement = "DAILY CARD SETTLEMENT REPORT" in probe and "PROCESSED NET AMOUNT" in probe
                except Exception:
                    is_settlement = False
            if is_settlement:
                settlement_files.append(f)
                log.info(f"  Card settlement : {f.name}")
            else:
                excel_files.append(f)
                log.info(f"  Excel report    : {f.name}")
            continue
        if suffix != ".csv":
            log.info(f"  Skipping unsupported file: {f.name}")
            continue
        try:
            peek = f.read_text(encoding="utf-8-sig", errors="replace")[:2000].upper()
        except Exception:
            peek = ""
        if any(p.upper() in name_up or p.upper() in peek for p in CONFIG["cc_pattern"]):
            cc_file = f
            log.info(f"  CC settlement  : {f.name}")
        elif "COUPON" in name_up:
            coupon_files.append(f)
            log.info(f"  Coupon file    : {f.name}")
        elif "DISCOUNTS BY SHOPPER LEVEL" in peek:
            sms_files.append(f)
            log.info(f"  SMS discounts  : {f.name}")
        elif "SUB-DEPARTMENT SINGLE TOTAL" in peek or "SUB-DEPARTMENT" in peek:
            sms_files.append(f)
            log.info(f"  SMS sales      : {f.name}")
        else:
            log.info(f"  CSV not recognized: {f.name}")

    if len(excel_files) > 1:
        dated_excel = [f for f in excel_files if any(p in f.name for p in date_patterns)]
        excel_files = dated_excel or [max(excel_files, key=lambda p: p.stat().st_mtime)]
        if not dated_excel:
            log.info(f"  Multiple Excel files found; using newest: {excel_files[0].name}")
    if len(settlement_files) > 1:
        matching = []
        for f in settlement_files:
            try:
                parsed = parse_card_settlement_report(f)
                if parsed.get("report_date") == selected_date:
                    matching.append(f)
            except Exception:
                pass
        settlement_files = matching or [max(settlement_files, key=lambda p: p.stat().st_mtime)]
    if not sms_files:
        log.info("  No SMS CSV needed if Excel report contains sales/discount tabs.")
    return sms_files, cc_file, coupon_files, excel_files, settlement_files

def spl(date_str, acct, name, amount, memo, class_name=""):
    amt_str = f"{amount:.2f}" if amount is not None else ""
    return f"SPL\tDEPOSIT\t{date_str}\t{acct}\t{name}\t{amt_str}\t{memo}\t{class_name}"


def build_card_settlement_adjustments(settlement_data: dict, bs_data: dict) -> list[dict]:
    """Return signed adjustments needed to bring settlement amounts back to BS totals.

    QuickBooks display adjustment = Processed Net Amount - BS.
    These tender lines display as credits/negative charges in QuickBooks, so:
      - when BS is greater than settlement, the QB adjustment is negative;
      - when settlement is greater than BS, the QB adjustment is positive.
    Differences under two cents are treated as matched and do not create a line.
    """
    if not settlement_data:
        return []

    bs_ebt = round(abs(bs_data.get("ebt_cash", 0.0)) + abs(bs_data.get("ebt_food", 0.0)), 2)
    checks = [
        ("VISA/MC", "visa_mc", round(abs(bs_data.get("visa_mc", 0.0)), 2)),
        ("Discover", "discover", round(abs(bs_data.get("discover", 0.0)), 2)),
        ("AMEX", "amex", round(abs(bs_data.get("amex", 0.0)), 2)),
        ("Debit Card", "debit", round(abs(bs_data.get("debit", 0.0)), 2)),
        ("EBT Cash/Food Stamp", "ebt", bs_ebt),
    ]

    adjustments = []
    for label, key, bs_value in checks:
        if key not in settlement_data:
            continue
        settlement_value = round(abs(settlement_data.get(key, 0.0)), 2)
        adjustment = round(settlement_value - bs_value, 2)
        if abs(adjustment) < 0.02:
            continue
        adjustments.append({
            "label": label,
            "key": key,
            "settlement": settlement_value,
            "bs": bs_value,
            "adjustment": adjustment,
            "memo": f"{label} - Difference between First Data vs BS",
        })
    return adjustments


def generate_iif(sales: dict, discounts: dict, cc: dict, report_date: date, owner_local_amt: float = 0.0, per_dept_coupons: dict = None, milk_bottle_return: float = 0.0, store_coupons_xl: float = 0.0, owner_apprec_xl: float = 0.0, misc_tba_lines: list = None, excel_sales_total: float = 0.0, excel_discount_total: float = 0.0, bs_data: dict = None, pass_through_total: float = 0.0, dust_bunnies_total: float = 0.0, milk_bottles_returns: float = 0.0, refunded_discounts: float = 0.0, hash_sales_total: float = 0.0, paid_in_total: float = 0.0, settlement_data: dict = None, membership_payments: list = None, membership_mode: str = "automatic", coupon_mode: str = "quickbooks", coupon_closeout_total: float | None = None, coupon_ncg_total: float | None = None, coupon_mfg_total: float | None = None, closeout_payload: dict | None = None, closeout_preview_path: Path | None = None) -> Path:
    date_str = report_date.strftime("%m/%d/%Y")
    deposit_acct = CONFIG["deposit_account"]
    iif_path = output_dir / f"deposit_{report_date.strftime('%Y%m%d')}.iif"
    if misc_tba_lines is None:
        misc_tba_lines = []
    if bs_data is None:
        bs_data = {}
    if settlement_data is None:
        settlement_data = {}
    if membership_payments is None:
        membership_payments = []

    normalized_closeout = None
    if closeout_payload is not None:
        candidate_closeout = normalize_closeout_payload(closeout_payload)
        if candidate_closeout["mode"] == "closeout":
            normalized_closeout = candidate_closeout

    closeout_rows = []
    closeout_rows_by_key = {}
    if normalized_closeout is not None:
        closeout_rows = build_standard_reconciliation(
            {
                "cash": abs(bs_data.get("cash", 0.0)),
                "checks": abs(bs_data.get("check", 0.0)),
                "donation": abs(bs_data.get("donation", 0.0)),
                "charge_house": abs(bs_data.get("charge", 0.0)),
                "offline_zon": abs(bs_data.get("offline_credit_card", 0.0)),
                "vendor_coupons": abs(bs_data.get("vendor_coupon", 0.0)),
                "paid_out": abs(bs_data.get("paid_out", 0.0)),
                "paid_in": abs(paid_in_total),
            },
            normalized_closeout["actuals"],
        )
        closeout_rows_by_key = {row["key"]: row for row in closeout_rows}

    def closeout_actual(key, legacy_value):
        row = closeout_rows_by_key.get(key)
        return row["actual"] if row else legacy_value

    def bs(key, default=None):
        v = bs_data.get(key, 0.0)
        if v != 0.0:
            return abs(v)
        return default

    def bs_signed(key, default=None):
        v = bs_data.get(key, 0.0)
        if v != 0.0:
            return v
        return default

    def s(key):
        if key not in sales:
            return None
        return -abs(sales[key])

    def d(key):
        return abs(discounts[key]) if key in discounts else None

    def has_amount(value):
        return value is not None and value != 0

    lines = [
        "!TRNS\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO\tCLASS",
        "!SPL\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO\tCLASS",
        "!ENDTRNS",
        "",
    ]
    spls = []

    SALES_ORDER = [
        ("7110110 · Sales - Packaged Grocery - NT", ""),
        ("7110120 · Sales - Dairy/Refrigerated", ""),
        ("7110130 · Sales - Frozen Foods", ""),
        ("7110141 · Sales - Bulk", ""),
        ("7110142 · Sales - Bulk Herbs", ""),
        ("7110150 · Sales - Bread", ""),
        ("7110160 · Sales - Beer", ""),
        ("7110180 · Sales - Pets", ""),
        ("7110190 · Sales - General Merchandise", ""),
        ("7110210 · Sales - In-House Deli", ""),
        ("7110220 · Sales - In-House Bakery", ""),
        ("7110210 · Sales - In-House Deli", "Catering"),
        ("7110235 · Sales - Sushi", ""),
        ("7110240 · Sales - Java & Juice", ""),
        ("7110250 · Sales - Co-op Cafe (ESP)", "ESP Co-op Cafe"),
        ("7110310 · Sales - Cheese", ""),
        ("7110320 · Sales - Meat", ""),
        ("7110330 · Sales - Specialty Foods", ""),
        ("7110340 · Sales - Fish & Seafood", ""),
        ("7110350 · Sales - Specialty Mercantile", ""),
        ("7110410 · Sales - Produce", ""),
        ("7110420 · Sales - Gardening/Plants", ""),
        ("7110510 · Sales - Personal Care Taxable", ""),
        ("7110520 · Sales - Vitamins & Supplements", ""),
        ("7110550 · Sales - Magazines", ""),
        ("7111300 · Promotional Sales", ""),
        ("8320000 · Store Supplies", "HWFC Grocery Paper Bags"),
    ]

    spl_total = 0.0
    seen_deli = False
    for acct, memo in SALES_ORDER:
        if acct == "7110210 · Sales - In-House Deli":
            if not seen_deli:
                amt = s(acct)
                seen_deli = True
            else:
                catering_key = "7110210 · Sales - In-House Deli|Catering"
                amt = -abs(sales[catering_key]) if catering_key in sales else None
        elif acct in INCOME_ACCOUNTS:
            amt = s(acct)
        else:
            amt = -(abs(sales[acct])) if acct in sales else None
        if not has_amount(amt):
            continue
        spl_total += amt
        spls.append(spl(date_str, acct, "", amt, memo))

    bag_amt = -(abs(sales.get("4150300 · NYS Paper Bag Fees Payable", 0))) or None
    if has_amount(bag_amt):
        spl_total += bag_amt
        spls.append(spl(date_str, "4150300 · NYS Paper Bag Fees Payable", "", bag_amt, "NYS-Albany County Paper Bag Fees"))

    if store_coupons_xl != 0.0:
        coupon_amt = abs(store_coupons_xl)
        log.info(f"    Coupons from Excel M2: ${store_coupons_xl:.2f}")
    else:
        raw_coupon = sales.get("8515000 · Marketing - Coupons, Store", 0)
        coupon_amt = abs(raw_coupon) if raw_coupon != 0 else None
    if has_amount(coupon_amt):
        spl_total += coupon_amt
        spls.append(spl(date_str, "8515000 · Marketing - Coupons, Store", "", coupon_amt, "Store Coupons"))

    DISCOUNT_ORDER = [
        ("8512006 · Discount 5% - Owner buy Local", "PdOut -"),
        ("8512001 · Discount 2% - Owners", "PdOut -"),
        ("8511001 · Discount 2% - Senior Non Owner", "PdOut -"),
        ("8511002 · Discount 8% - Senior Day", "PdOut -"),
        ("8511003 · Discount 5% - Senior Owners", "PdOut -"),
        ("8140010 · Monthly Time Discount (8%)", "PdOut -"),
        ("8140026 · Weekly Time Discount (24%)", "PdOut -"),
        ("8423100 · Discount - Staff  (24%)", "PdOut -"),
        ("8512002 · Discount 2% - Visiting Coop", "PdOut -"),
        ("8512003 · Discount 8% - Vendors", "PdOut -"),
        ("8512005 · Discount 8% - College Day", "PdOut -"),
        ("8512007 · Discount 15% - Non-profit", "PdOut-"),
        ("8140026 · Weekly Time Discount (24%)", "Refunded Discounts", refunded_discounts if refunded_discounts else None),
    ]
    for entry in DISCOUNT_ORDER:
        acct = entry[0]
        memo = entry[1]
        override_amt = entry[2] if len(entry) > 2 else None
        if override_amt is not None:
            amt = -override_amt
        elif acct == "8512006 · Discount 5% - Owner buy Local":
            source = owner_apprec_xl if owner_apprec_xl != 0 else owner_local_amt
            amt = abs(source) if source != 0 else None
        elif memo == "Refunded Discounts":
            amt = None
        else:
            amt = d(acct)
        if not has_amount(amt):
            continue
        spl_total += amt
        spls.append(spl(date_str, acct, "", amt, memo))

    ebt_cash = bs_data.get("ebt_cash", 0.0)
    ebt_food = bs_data.get("ebt_food", 0.0)
    bs_ebt = round(abs(ebt_cash) + abs(ebt_food), 2) or None
    charity_bs_amt = round(abs(bs_data.get("charity", 0.0)), 2)
    pass_through_amt = round(abs(pass_through_total), 2)
    charity_combined = round(charity_bs_amt + pass_through_amt, 2)
    log.info(f"    Charity mapping: BS Charity=${charity_bs_amt:,.2f} + HASH Pass Through=${pass_through_amt:,.2f} = ${charity_combined:,.2f}")
    log.info(f"    Paid-In mapping: HASH Paid-Ins=${abs(paid_in_total):,.2f} → 4444 · TBA Purchases / PAID IN:")

    coupon_reconciliation = reconcile_coupon_receivable(
        bs("vendor_coupon", 0.0),
        mode=coupon_mode,
        closeout_actual_total=coupon_closeout_total,
        ncg_total=coupon_ncg_total,
        mfg_total=coupon_mfg_total,
    )
    coupon_ncg_source = (
        -coupon_reconciliation["ncg_total"]
        if coupon_reconciliation["ncg_total"]
        else None
    )
    coupon_mfg_source = (
        -coupon_reconciliation["mfg_total"]
        if coupon_reconciliation["mfg_total"]
        else None
    )
    coupon_difference_source = coupon_reconciliation["difference"]
    if coupon_difference_source == 0:
        coupon_difference_source = None

    if normalized_closeout is not None:
        counted_coupon_actual = (
            Decimal(str(coupon_reconciliation["ncg_total"] or 0))
            + Decimal(str(coupon_reconciliation["mfg_total"] or 0))
        ).quantize(Decimal("0.01"))
        closeout_coupon_actual = Decimal(
            str(closeout_actual("vendor_coupons", 0))
        ).quantize(Decimal("0.01"))
        if counted_coupon_actual != closeout_coupon_actual:
            raise ValueError(
                "Closeout Vendor Coupons actual must equal the counted NCG and MFG coupons"
            )

    bs_ebt_compare = round(abs(bs_data.get("ebt_cash", 0.0)) + abs(bs_data.get("ebt_food", 0.0)), 2)
    tender_checks = [
        ("VISA/MC", "visa_mc", round(abs(bs_data.get("visa_mc", 0.0)), 2)),
        ("Discover", "discover", round(abs(bs_data.get("discover", 0.0)), 2)),
        ("AMEX", "amex", round(abs(bs_data.get("amex", 0.0)), 2)),
        ("Debit Card", "debit", round(abs(bs_data.get("debit", 0.0)), 2)),
        ("EBT Cash/Food Stamp", "ebt", bs_ebt_compare),
    ]
    card_adjustments = build_card_settlement_adjustments(settlement_data, bs_data)
    card_adjustments_by_key = {row["key"]: row for row in card_adjustments}

    if settlement_data:
        log.info("")
        log.info("  ─────────────────────────────────────────────────────────")
        log.info("  CARD SETTLEMENT RECONCILIATION")
        log.info("  ─────────────────────────────────────────────────────────")
        for label, key, bs_value in tender_checks:
            settlement_value = round(abs(settlement_data.get(key, 0.0)), 2)
            signed_adjustment = round(settlement_value - bs_value, 2)
            difference = round(abs(signed_adjustment), 2)
            status = "MATCH" if difference < 0.02 else "MISMATCH"
            log.info(
                f"  CARD SETTLEMENT | {label} | Settlement={settlement_value:.2f} | "
                f"BS={bs_value:.2f} | Difference={difference:.2f} | "
                f"Adjustment={signed_adjustment:.2f} | {status}"
            )
        log.info("  ─────────────────────────────────────────────────────────")
        log.info("")

    def tender_source(key, fallback):
        if settlement_data and key in settlement_data:
            return abs(settlement_data.get(key, 0.0)) or None
        return fallback

    membership_lines = build_membership_lines(
        membership_payments,
        expected_subscription_total=abs(bs_data.get("subscription", 0.0)),
        handling_mode=membership_mode,
    )
    for membership_line in membership_lines:
        iif_amt = -membership_line["amount"]
        spl_total += iif_amt
        spls.append(
            spl(
                date_str,
                membership_line["account"],
                membership_line["name"],
                iif_amt,
                membership_line["memo"],
                membership_line["class_name"],
            )
        )
        log.info(
            f"    Member share: {membership_line['name']} | "
            f"{membership_line['account']} | {membership_line['memo']} | "
            f"${membership_line['amount']:,.2f}"
        )

    donation_source = closeout_actual("donation", bs("donation"))
    charge_house_source = closeout_actual("charge_house", bs("charge"))
    paid_in_source = closeout_actual(
        "paid_in", abs(paid_in_total) if paid_in_total else None
    )
    paid_out_source = closeout_actual("paid_out", bs("paid_out"))

    MANUAL_LINES = [
        ("4150100 · Sales Tax Payable", "New York State Sales Tax", "", bs("sales_tax")),
        ("1311100 · Inventory - Bottles Deposit", "", "Bottle Sales", bs("bottle_sales")),
        ("1311100 · Inventory - Bottles Deposit", "", "Milk Bottle Fee", bs("milk_bottle_fee")),
        ("1311100 · Inventory - Bottles Deposit", "", "Milk Bottle Return", -(abs(milk_bottle_return) + bs("milk_bottle_return", 0.0)) if (milk_bottle_return or bs("milk_bottle_return")) else None),
        ("1311100 · Inventory - Bottles Deposit", "", "Bottle Return", -bs("bottle_return") if bs("bottle_return") else None),
        ("4160000 · Charitable Donations Payable", "", "Charity/Pass through Donations (Round up)", charity_combined or None),
        ("9107000 · Miscellaneous Income", "", "Penny Round Up for Cash Transactions", bs_signed("penny_round")),
        ("4160500 · Gift Cards - Sold - Old/Vantiv", "", "Gift cards sold", bs("prepaid_increase") if bs("prepaid_increase") else None),
        ("1230400 · Due From Double Up Food Bucks", "", "Double Up Food Bucks Customer Spending", -bs("dufb") if bs("dufb") else None),
        ("4160510 · Gift Cards- Redeemed-Old/Vantiv", "", "Gift cards redeemed", -bs("prepaid_card") if bs("prepaid_card") else None),
        ("1250000 · Coupons Receivable", "", "NCG Coupons", coupon_ncg_source),
        ("1250000 · Coupons Receivable", "", "MFG Coupons", coupon_mfg_source),
        ("4444 · TBA Purchases", "", "InHouse:", -charge_house_source if charge_house_source else None),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("8506000 · Outreach - Donations", "", "", -donation_source if donation_source else None),
        ("4444 · TBA Purchases", "", "PAID IN:", paid_in_source if paid_in_source else None),
        # Paid Out is stored as a positive BS pickup amount, but it reduces the QuickBooks deposit.
        # MANUAL_LINES inverts the source amount for IIF sign convention, so pass a negative source
        # value here to produce a positive IIF SPL that displays as a negative deposit line in QB.
        ("4444 · TBA Purchases", "", "PAID OUT:", -paid_out_source if paid_out_source else None),
        ("1240001 · Credit Card Payments Receivable", "", "Visa/MC", -tender_source("visa_mc", bs("visa_mc")) if tender_source("visa_mc", bs("visa_mc")) else None),
        ("1240001 · Credit Card Payments Receivable", "", "Discover", -tender_source("discover", bs("discover")) if tender_source("discover", bs("discover")) else None),
        ("1240001 · Credit Card Payments Receivable", "", "AMEX", -tender_source("amex", bs("amex")) if tender_source("amex", bs("amex")) else None),
        ("1240001 · Credit Card Payments Receivable", "", "Debit Card", -tender_source("debit", bs("debit")) if tender_source("debit", bs("debit")) else None),
        ("1240001 · Credit Card Payments Receivable", "", "EBT Cash/Food Stamp", -tender_source("ebt", bs_ebt) if tender_source("ebt", bs_ebt) else None),
        ("8314000 · FE - Cash Over/Shorts", "", "Over/Short per Closeout Sheet"),
        ("8314000 · FE - Cash Over/Shorts", "", "Over/Short per POS (to = POS total)"),
    ]

    for entry in MANUAL_LINES:
        acct, name, memo = entry[0], entry[1], entry[2]
        amt = entry[3] if len(entry) > 3 else None
        class_name = entry[4] if len(entry) > 4 else ""
        if normalized_closeout is not None and (
            acct == "8314000 · FE - Cash Over/Shorts"
            and memo in {
                "Over/Short per Closeout Sheet",
                "Over/Short per POS (to = POS total)",
            }
        ):
            continue
        keep_blank_placeholder = (
            memo == "MFG Coupons"
            or (acct == "4444 · TBA Purchases" and memo in {"InHouse:", ""})
            or (
                acct == "8314000 · FE - Cash Over/Shorts"
                and memo in {
                    "Over/Short per Closeout Sheet",
                    "Over/Short per POS (to = POS total)",
                }
            )
        )
        if has_amount(amt):
            iif_amt = -amt
            spl_total += iif_amt
        elif keep_blank_placeholder:
            iif_amt = None
        else:
            continue
        spls.append(spl(date_str, acct, name, iif_amt, memo, class_name))

    # Card settlement differences are posted at the bottom of the deposit.
    # The desired QuickBooks display adjustment is Processed Net Amount - BS.
    # The IIF amount is inverted because QuickBooks flips the sign on deposit SPL lines.
    for adjustment_row in card_adjustments:
        qb_adjustment = adjustment_row["adjustment"]
        iif_amt = -qb_adjustment
        spl_total += iif_amt
        spls.append(
            spl(
                date_str,
                "8314000 · FE - Cash Over/Shorts",
                "",
                iif_amt,
                adjustment_row["memo"],
            )
        )
        log.info(
            f"    Card settlement adjustment: {adjustment_row['label']} "
            f"Settlement=${adjustment_row['settlement']:,.2f} "
            f"BS=${adjustment_row['bs']:,.2f} "
            f"Adjustment=${qb_adjustment:,.2f} → 8314000"
        )

    # Keep legacy coupon-only reconciliation byte-compatible when the reviewed
    # Closeout workflow is not active.
    if normalized_closeout is None and coupon_mode == "closeout":
        iif_amt = (
            -coupon_difference_source
            if coupon_difference_source is not None
            else None
        )
        if iif_amt is not None:
            spl_total += iif_amt
        spls.append(
            spl(
                date_str,
                "8314000 · FE - Cash Over/Shorts",
                "",
                iif_amt,
                "Over/Short per Closeout Sheet - Coupon",
                "Admin",
            )
        )

    if normalized_closeout is None and misc_tba_lines:
        for memo, amount in misc_tba_lines:
            iif_amt = -amount
            spl_total += iif_amt
            spls.append(spl(date_str, "4444 · TBA Purchases", "", iif_amt, memo))
            log.info(f"    TBA Purchases: {memo} = ${amount:.2f}")

    # Offline Credit Card is a unique BS item. Keep it separate from gift cards
    # and place it at the bottom as a negative QuickBooks TBA line.
    offline_credit_card = bs_data.get("offline_credit_card", 0.0)
    if normalized_closeout is not None:
        offline_credit_card = -closeout_actual(
            "offline_zon", abs(offline_credit_card)
        )
    if normalized_closeout is None and offline_credit_card:
        iif_amt = -offline_credit_card
        spl_total += iif_amt
        spls.append(
            spl(
                date_str,
                "4444 · TBA Purchases",
                "",
                iif_amt,
                "Offline Credit Card:",
            )
        )
        log.info(
            f"    Offline Credit Card: ${offline_credit_card:,.2f} → TBA Purchases"
        )

    if normalized_closeout is not None:
        for row in closeout_rows:
            qb_effect = row["adjustment_qb_effect"]
            if row["managed_externally"]:
                qb_effect = coupon_difference_source
            if not has_amount(qb_effect):
                continue
            iif_amount = -qb_effect
            spl_total += iif_amount
            spls.append(
                spl(
                    date_str,
                    row["adjustment_account"],
                    "",
                    iif_amount,
                    row["adjustment_memo"],
                    "Admin",
                )
            )

        misc_rows = build_misc_adjustments(normalized_closeout)
        custom_tba_rows = []
        for row in misc_rows:
            if row["kind"] == "custom_tba":
                custom_tba_rows.append(
                    {
                        "account": row["account"],
                        "name": "",
                        "memo": row["memo"],
                        "iif_amount": row["iif_amount"],
                        "class_name": "",
                    }
                )
                continue
            spl_total += row["iif_amount"]
            spls.append(
                spl(
                    date_str,
                    row["account"],
                    "",
                    row["iif_amount"],
                    row["memo"],
                )
            )

        existing_misc_tba_rows = [
            {
                "account": "4444 · TBA Purchases",
                "name": "",
                "memo": memo,
                "iif_amount": -amount,
                "class_name": "",
            }
            for memo, amount in misc_tba_lines
        ]
        offline_tba_rows = []
        if offline_credit_card:
            offline_tba_rows.append(
                {
                    "account": "4444 · TBA Purchases",
                    "name": "",
                    "memo": "Offline Credit Card:",
                    "iif_amount": -offline_credit_card,
                    "class_name": "",
                }
            )
        pending_tba_rows = (
            custom_tba_rows
            + existing_misc_tba_rows
            + offline_tba_rows
        )
        pending_tba_iif_total = sum(
            (
                Decimal(str(row["iif_amount"]))
                for row in pending_tba_rows
                if row["iif_amount"] is not None
            ),
            Decimal("0"),
        )
        provisional_total = -(
            Decimal(str(spl_total)) + pending_tba_iif_total
        ).quantize(Decimal("0.01"))
        final_balance = calculate_final_pos_adjustment(
            provisional_total,
            normalized_closeout["final_total"],
            normalized_closeout["approve_final_pos"],
        )
        final_pos_line = final_balance["line"]
        if final_pos_line is not None:
            spl_total += final_pos_line["iif_amount"]
            spls.append(
                spl(
                    date_str,
                    final_pos_line["account"],
                    "",
                    final_pos_line["iif_amount"],
                    final_pos_line["memo"],
                )
            )

        for row in pending_tba_rows:
            if row["iif_amount"] is not None:
                spl_total += row["iif_amount"]
            spls.append(
                spl(
                    date_str,
                    row["account"],
                    row["name"],
                    row["iif_amount"],
                    row["memo"],
                    row["class_name"],
                )
            )

        if closeout_preview_path is not None:
            preview = {
                "standard_rows": closeout_rows,
                "misc_rows": misc_rows,
                "provisional_total": final_balance["provisional_total"],
                "final_total": final_balance["final_total"],
                "remaining": final_balance["remaining"],
                "remaining_after_approval": (
                    0.0 if final_pos_line is not None else final_balance["remaining"]
                ),
                "requires_approval": final_balance["requires_approval"],
                "final_pos_line": final_pos_line,
            }
            preview_path = Path(closeout_preview_path)
            preview_path.parent.mkdir(parents=True, exist_ok=True)
            temporary_preview_path = preview_path.with_name(
                f".{preview_path.name}.tmp"
            )
            temporary_preview_path.write_text(
                json.dumps(preview, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            temporary_preview_path.replace(preview_path)

    spl_total = round(spl_total, 2)
    trns_amt = round(-spl_total, 2)

    COUPON_ACCT = "8515000 · Marketing - Coupons, Store"
    sales_acct_set = set(sales.keys()) - {COUPON_ACCT}
    gross_sales_spl = 0.0
    for line in spls:
        parts = line.split("\t")
        if len(parts) < 6 or not parts[5].strip():
            continue
        acct = parts[3].strip()
        if acct in sales_acct_set:
            try:
                gross_sales_spl += abs(float(parts[5]))
            except ValueError:
                pass
    if misc_tba_lines:
        for memo, amount in misc_tba_lines:
            gross_sales_spl += amount
    gross_sales_spl = round(gross_sales_spl, 2)

    coupon_deduct = abs(store_coupons_xl) if store_coupons_xl else abs(sales.get("8515000 · Marketing - Coupons, Store", 0))
    owner_apprec_deduct = abs(owner_apprec_xl) if owner_apprec_xl else 0.0
    milk_returns_deduct = abs(milk_bottle_return) if milk_bottle_return else abs(milk_bottles_returns) if milk_bottles_returns else 0.0
    net_sales_check = round(gross_sales_spl - coupon_deduct - owner_apprec_deduct - milk_returns_deduct, 2)

    if excel_sales_total != 0.0:
        diff = round(abs(excel_sales_total - net_sales_check), 2)
        log.info("")
        log.info("  ─────────────────────────────────────────")
        log.info("  SALES CHECK")
        log.info("  ─────────────────────────────────────────")
        log.info(f"  Gross Sales:        ${gross_sales_spl:>12,.2f}")
        log.info(f"  Store Coupons:     -${coupon_deduct:>12,.2f}")
        log.info(f"  Owner Apprec:      -${owner_apprec_deduct:>12,.2f}")
        if milk_returns_deduct:
            log.info(f"  Milk Btl Returns:  -${milk_returns_deduct:>12,.2f}")
        log.info("  ─────────────────────────────────────────")
        log.info(f"  Script Net Sales:   ${net_sales_check:>12,.2f}")
        log.info(f"  Excel Sales Total:  ${excel_sales_total:>12,.2f}")
        if diff < 0.02:
            log.info("  RESULT: ✓ MATCH — OK to import!")
        else:
            log.info(f"  Difference:         ${diff:>12,.2f}")
            log.warning("  RESULT: ⚠ MISMATCH — Check your Excel file before importing!")
        log.info("  ─────────────────────────────────────────")
        log.info("")

    script_hash = round(abs(refunded_discounts) + abs(pass_through_total), 2)
    excel_hash = round(abs(hash_sales_total), 2)
    hash_diff = round(abs(excel_hash - script_hash), 2)
    log.info("  ─────────────────────────────────────────")
    log.info("  HASH SALES 6 CHECK")
    log.info("  ─────────────────────────────────────────")
    log.info(f"  Refunded Discounts:  ${abs(refunded_discounts):>10,.2f}")
    log.info(f"  Pass Thru Donations: ${abs(pass_through_total):>10,.2f}")
    log.info("  ─────────────────────────────────────────")
    log.info(f"  Script Total:        ${script_hash:>10,.2f}")
    log.info(f"  Hash Sales 6 Total:  ${excel_hash:>10,.2f}")
    if hash_sales_total == 0.0:
        log.warning("  RESULT: ⚠ NO HASH SALES 6 TOTAL FOUND IN EXCEL — verify manually")
    elif hash_diff < 0.02:
        log.info("  RESULT: ✓ MATCH — OK to import!")
    else:
        log.info(f"  Difference:          ${hash_diff:>10,.2f}")
        log.warning("  RESULT: ⚠ MISMATCH — Check before importing!")
    log.info("  ─────────────────────────────────────────")
    log.info("")

    disc_acct_names = set(discounts.keys())
    disc_spl_total = 0.0
    for line in spls:
        parts = line.split("\t")
        if len(parts) < 6 or not parts[5].strip():
            continue
        acct = parts[3].strip()
        memo = parts[6].strip() if len(parts) > 6 else ""
        if acct in disc_acct_names and "refunded" not in memo.lower():
            try:
                disc_spl_total += abs(float(parts[5]))
            except ValueError:
                pass
    disc_spl_total = round(disc_spl_total, 2)

    if excel_discount_total != 0.0:
        diff2 = round(abs(excel_discount_total - disc_spl_total), 2)
        log.info("  ─────────────────────────────────────────────")
        log.info("  DISCOUNTS CHECK")
        log.info("  ─────────────────────────────────────────────")
        log.info(f"  Script Discounts:   ${disc_spl_total:>12,.2f}")
        log.info(f"  Excel Disc Total:   ${excel_discount_total:>12,.2f}")
        if diff2 < 0.02:
            log.info("  RESULT: ✓ MATCH — OK to import!")
        else:
            log.info(f"  Difference:         ${diff2:>12,.2f}")
            log.warning("  RESULT: ⚠ MISMATCH — Check your discounts sheet before importing!")
        log.info("  ─────────────────────────────────────────────")
        log.info("")

    status_lines = [
        f"HWFC Daily Deposit — {report_date.strftime('%A, %B %d, %Y')}",
        f"Run time: {datetime.now().strftime('%I:%M %p')}",
        "",
    ]
    if excel_sales_total != 0.0:
        sales_diff = round(abs(excel_sales_total - net_sales_check), 2)
        if sales_diff < 0.02:
            status_lines.append(f"  SALES:     ✓ MATCH   ${net_sales_check:,.2f}")
        else:
            status_lines.append(f"  SALES:     ⚠ MISMATCH  Script=${net_sales_check:,.2f}  Excel=${excel_sales_total:,.2f}  (off by ${sales_diff:,.2f})")
    if excel_discount_total != 0.0:
        disc_spl = round(sum(abs(v) for v in discounts.values()), 2)
        disc_diff = round(abs(excel_discount_total - disc_spl), 2)
        if disc_diff < 0.02:
            status_lines.append(f"  DISCOUNTS: ✓ MATCH   ${disc_spl:,.2f}")
        else:
            status_lines.append(f"  DISCOUNTS: ⚠ MISMATCH  Script=${disc_spl:,.2f}  Excel=${excel_discount_total:,.2f}  (off by ${disc_diff:,.2f})")
    if hash_sales_total != 0.0:
        if hash_diff < 0.02:
            status_lines.append(f"  HASH SALES: ✓ MATCH   ${script_hash:,.2f}")
        else:
            status_lines.append(f"  HASH SALES: ⚠ MISMATCH  Script=${script_hash:,.2f}  Excel=${excel_hash:,.2f}  (off by ${hash_diff:,.2f})")
    else:
        status_lines.append("  HASH SALES: ⚠ NO EXCEL TOTAL FOUND")
    status_lines.extend(["", f"  IIF file: {iif_path}", "  QB import: File → Utilities → Import → IIF Files", ""])

    overall_ok = all([
        abs(excel_sales_total - net_sales_check) < 0.02 if excel_sales_total else True,
        abs(excel_discount_total - round(sum(abs(v) for v in discounts.values()), 2)) < 0.02 if excel_discount_total else True,
        hash_diff < 0.02 if hash_sales_total else False,
    ])
    if overall_ok:
        status_lines.append("  ✓ ALL CHECKS PASSED — Safe to import into QuickBooks!")
    else:
        status_lines.append("  ⚠ ISSUES FOUND — Review before importing into QuickBooks!")

    status_path = LOG_DIR / "last_run_status.txt"
    try:
        status_path.write_text("\n".join(status_lines), encoding="utf-8")
        log.info(f"  Status → {status_path}")
    except Exception as e:
        log.warning(f"  Could not write status file: {e}")

    lines.append(f"TRNS\tDEPOSIT\t{date_str}\t{deposit_acct}\t\t{trns_amt:.2f}\tDeposit\t")
    lines.extend(spls)
    lines.append("ENDTRNS")
    try:
        with open(iif_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        log.info(f"\n  IIF written → {iif_path}")
        log.info(f"  Auto-filled total: ${trns_amt:,.2f}")
    except Exception as e:
        log.error(f"  FAILED to write file: {e}")
        log.error(f"  Attempted path: {iif_path}")
        raise
    return iif_path


def write_excel_summary(sales, discounts, cc, report_date: date) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    xlsx_path = SUMMARY_DIR / f"deposit_summary_{report_date.strftime('%Y%m%d')}.xlsx"
    wb = openpyxl.Workbook()

    def fill_sheet(ws, data: dict, title: str):
        ws.title = title
        ws.append(["QB Account", "Memo", "Amount"])
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        total = 0.0
        for key, amt in sorted(data.items()):
            acct = key.split("|")[0]
            memo = key.split("|")[1] if "|" in key else ""
            ws.append([acct, memo, amt])
            total += amt

        ws.append(["TOTAL", "", round(total, 2)])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(width, 55)

    fill_sheet(wb.active,         sales,     "Sales")
    fill_sheet(wb.create_sheet(), discounts, "Discounts")
    fill_sheet(wb.create_sheet(), cc,        "Credit Cards")

    try:
        wb.save(xlsx_path)
        log.info(f"  Excel  → {xlsx_path}")
    except Exception as e:
        if "WinError 2" in str(e) or "openpyx" in str(e).lower():
            log.warning(f"  Could not write Excel summary (file may be open) — skipping")
        else:
            log.warning(f"  Could not write Excel summary: {e}")
    return xlsx_path


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    log.info("=" * 65)
    log.info(f"  HWFC Daily Deposit Automation   {datetime.now():%Y-%m-%d %H:%M:%S}")
    log.info("=" * 65)

    try:
        today = date.today()

        parser = argparse.ArgumentParser(description="HWFC Daily Deposit Automation")
        parser.add_argument("--date", dest="deposit_date",
                            help="Deposit date in MM/DD/YY or MM/DD/YYYY format")
        parser.add_argument("--auto", action="store_true",
                            help="Use yesterday without prompting")
        parser.add_argument("--membership-payments-file",
                            help="JSON file containing manually entered membership payments")
        parser.add_argument("--membership-mode", choices=("automatic", "manual"),
                            default="automatic",
                            help="Split member payments automatically or finish them manually in QuickBooks")
        parser.add_argument("--coupon-mode", choices=("quickbooks", "closeout"),
                            default="quickbooks",
                            help="Keep the current coupon process or use the Closeout Sheet breakdown")
        parser.add_argument("--coupon-closeout-total", type=float,
                            help="Closeout Sheet Coupon Actual Total")
        parser.add_argument("--coupon-ncg-total", type=float,
                            help="Counted NCG coupon total")
        parser.add_argument("--coupon-mfg-total", type=float,
                            help="Counted MFG coupon total")
        parser.add_argument("--closeout-file",
                            help="Validated Closeout Sheet JSON payload")
        parser.add_argument("--closeout-preview-output",
                            help="Path for Closeout preview JSON")
        args, _unknown = parser.parse_known_args()

        membership_payments = (
            load_membership_payments_file(args.membership_payments_file)
            if args.membership_payments_file else []
        )
        closeout_payload = (
            load_closeout_payload_file(args.closeout_file)
            if args.closeout_file else None
        )

        if args.deposit_date:
            parsed = None
            for fmt in ("%m/%d/%y", "%m/%d/%Y", "%m-%d-%y", "%m-%d-%Y", "%m%d%y", "%m%d%Y"):
                try:
                    parsed = datetime.strptime(args.deposit_date, fmt).date()
                    break
                except ValueError:
                    continue
            if parsed is None:
                raise ValueError(
                    f"Could not parse date '{args.deposit_date}'. Use MM/DD/YY, for example 08/20/26."
                )
            yesterday = parsed
            log.info(f"  Using requested date: {yesterday.strftime('%B %d, %Y')}")
        else:
            yesterday = today - timedelta(days=1)
            log.info(f"  Using yesterday: {yesterday.strftime('%B %d, %Y')}")

        log.info(f"Scanning {CONFIG['pos_export_folder']} ...")
        sms_files, cc_file, coupon_files, excel_files, settlement_files = find_todays_files(yesterday)

        sales     = {}
        discounts = {}
        cc        = {}
        owner_local_amt = 0.0

        # Prefer Excel report for sales if available (more accurate net amounts)
        milk_bottle_return = 0.0
        store_coupons_xl   = 0.0
        owner_apprec_xl    = 0.0
        misc_tba_lines     = []
        excel_sales_total  = 0.0
        pass_through_total   = 0.0
        dust_bunnies_total   = 0.0
        milk_bottles_returns = 0.0
        refunded_discounts   = 0.0
        hash_sales_total     = 0.0
        refunded_discounts   = 0.0
        paid_in_total        = 0.0
        bs_data            = {}
        settlement_data    = {}

        if excel_files:
            for f in excel_files:
                xl_sales, xl_misc, mb_ret, sc_amt, oa_amt, xl_total, pt_total, db_total, mbr, rd, hs = parse_excel_report(f)
                for k, v in xl_sales.items():
                    sales[k] = round(sales.get(k, 0.0) + v, 2)
                misc_tba_lines.extend(xl_misc)
                milk_bottle_return   += mb_ret
                store_coupons_xl     += sc_amt
                owner_apprec_xl      += oa_amt
                excel_sales_total    += xl_total
                pass_through_total   += pt_total
                dust_bunnies_total   += db_total
                milk_bottles_returns += mbr
                refunded_discounts   += rd
                hash_sales_total     += hs
            log.info(f"  Using Excel report for sales ({len(excel_files)} file(s))")

            # Also read BS sheet
            for f in excel_files:
                bs_data = parse_bs_sheet(f, yesterday)

            # HASH tab is authoritative for Refunded Discounts,
            # Pass Through Donations, and Paid-Ins.
            refunded_discounts = 0.0
            pass_through_total = 0.0
            paid_in_total = 0.0
            for f in excel_files:
                hash_refunded, hash_pass_through, hash_paid_in = parse_hash_sheet(f, yesterday)
                refunded_discounts += hash_refunded
                pass_through_total += hash_pass_through
                paid_in_total += hash_paid_in
        else:
            log.warning("  No Excel report found — falling back to SMS CSV files")

        # Read discounts from Excel file if available, else fall back to SMS CSV
        excel_discount_total = 0.0
        if excel_files:
            for f in excel_files:
                d_map, d_total = parse_excel_discounts(f, yesterday)
                for k, v in d_map.items():
                    discounts[k] = round(discounts.get(k, 0.0) + v, 2)
                excel_discount_total += d_total
            if not discounts:
                log.warning("  No discounts found in Excel — falling back to SMS CSV")
                for f in sms_files:
                    _, d_map = parse_sms_file(f)
                    for k, v in d_map.items():
                        discounts[k] = round(discounts.get(k, 0.0) + v, 2)
        else:
            for f in sms_files:
                _, d_map = parse_sms_file(f)
                for k, v in d_map.items():
                    discounts[k] = round(discounts.get(k, 0.0) + v, 2)

        if settlement_files:
            # The Daily Card Settlement Report is the bank-received source of truth
            # for VISA/MC, Discover, AMEX, Debit, and EBT.
            settlement_data = parse_card_settlement_report(settlement_files[0])
            settlement_report_date = settlement_data.get("report_date")
            if settlement_report_date and settlement_report_date != yesterday:
                log.warning(
                    f"  CARD SETTLEMENT DATE MISMATCH: settlement={settlement_report_date:%m/%d/%Y} "
                    f"deposit={yesterday:%m/%d/%Y}"
                )
        else:
            log.warning("  No Daily Card Settlement Report found — card tender lines will fall back to BS values.")

        per_dept_coupons = {}
        for f in coupon_files:
            dept_map, grand_total = parse_coupon_file(f)
            owner_local_amt += grand_total
            for k, v in dept_map.items():
                per_dept_coupons[k] = round(per_dept_coupons.get(k, 0.0) + v, 2)

        if cc_file:
            cc = parse_cc_file(cc_file)
        else:
            log.warning("\n  No CC file found — add Commerce Control Center CSV to folder.")

        if not any([sales, discounts, cc]):
            log.error("Nothing mapped. Check files and try again.")
            sys.exit(1)

        iif_path = generate_iif(
            sales, discounts, cc, yesterday, owner_local_amt, per_dept_coupons,
            milk_bottle_return, store_coupons_xl, owner_apprec_xl, misc_tba_lines,
            excel_sales_total, excel_discount_total, bs_data, pass_through_total,
            dust_bunnies_total, milk_bottles_returns, refunded_discounts,
            hash_sales_total, paid_in_total, settlement_data, membership_payments,
            args.membership_mode,
            coupon_mode=args.coupon_mode,
            coupon_closeout_total=args.coupon_closeout_total,
            coupon_ncg_total=args.coupon_ncg_total,
            coupon_mfg_total=args.coupon_mfg_total,
            closeout_payload=closeout_payload,
            closeout_preview_path=(
                Path(args.closeout_preview_output)
                if args.closeout_preview_output and closeout_payload is not None else None
            ),
        )
        try:
            xlsx_path = write_excel_summary(sales, discounts, cc, yesterday)
        except Exception as e:
            log.warning(f"  Excel summary skipped: {e}")
            xlsx_path = None

        log.info("")
        log.info("✓ COMPLETE")
        log.info(f"  Sales accounts   : {len(sales)}")
        log.info(f"  Discount accounts: {len(discounts)}")
        log.info(f"  CC card types    : {len(cc)}")
        log.info(f"  IIF  → {iif_path}")
        if xlsx_path:
            log.info(f"  XLSX → {xlsx_path}")
        log.info("")
        log.info("  QB import: File → Utilities → Import → IIF Files")
        log.info("  Then check: Banking → Make Deposits")
        log.info("")
        log.info("  Still manual: Sales Tax, Gift Cards,")
        log.info("  Bottle Deposits, NCG Coupons, Petty Cash, Donations,")
        log.info("  In-House Purchases, Cash Over/Short")

    except FileNotFoundError as e:
        log.error(str(e))
        sys.exit(1)
    except Exception as e:
        log.exception(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
