import ast
from io import BytesIO
import shutil
import sys
import unittest
from datetime import date
from pathlib import Path
from uuid import uuid4


class ActivityBreakdownTests(unittest.TestCase):
    def activity_api(self):
        try:
            from app.activity_breakdowns import (
                activity_actuals,
                append_activity_entry,
                activity_closeout_ready,
                activity_workflow_keys,
                build_activity_lines,
                load_activity_payload_file,
                normalize_activity_payload,
                normalize_activity_section,
                read_activity_source_totals,
                write_activity_payload_file,
            )
        except ImportError as exc:
            self.fail(f"activity breakdown feature is missing: {exc}")
        return {
            "actuals": activity_actuals,
            "append_entry": append_activity_entry,
            "closeout_ready": activity_closeout_ready,
            "workflow_keys": activity_workflow_keys,
            "build_lines": build_activity_lines,
            "load": load_activity_payload_file,
            "normalize": normalize_activity_payload,
            "normalize_section": normalize_activity_section,
            "read_sources": read_activity_source_totals,
            "write": write_activity_payload_file,
        }

    def complete_payload(self):
        return {
            "donation": {
                "mode": "app",
                "rows": [
                    {
                        "given_to": "AME Zion Church",
                        "purpose": "Food Pantry",
                        "manager": "Steoria",
                        "amount": 100,
                    },
                    {
                        "given_to": "Disability Empower Network",
                        "purpose": "Self Defense Event",
                        "manager": "Anestia",
                        "amount": 100,
                    },
                ],
            },
            "paid_out": {
                "mode": "app",
                "rows": [
                    {
                        "type": "esp",
                        "original_date": "2026-08-18",
                        "initials": "MR",
                        "amount": 184.35,
                    },
                    {
                        "type": "other",
                        "memo": "Register correction",
                        "amount": 5.25,
                    },
                ],
            },
            "paid_in": {
                "mode": "app",
                "rows": [
                    {
                        "type": "esp",
                        "original_date": "2026-08-19",
                        "initials": "MR",
                        "amount": 100.15,
                    },
                    {
                        "type": "other",
                        "memo": "Returned drawer cash",
                        "amount": 10,
                    },
                ],
            },
        }

    def test_donation_rows_build_marketing_lines_with_negative_effects(self):
        api = self.activity_api()

        lines = api["build_lines"](self.complete_payload())

        self.assertEqual(
            lines["donation"],
            [
                {
                    "account": "8506000 · Outreach - Donations",
                    "memo": "Given to AME Zion Church for Food Pantry - Steoria",
                    "class_name": "Marketing",
                    "qb_effect": -100.0,
                },
                {
                    "account": "8506000 · Outreach - Donations",
                    "memo": (
                        "Given to Disability Empower Network for "
                        "Self Defense Event - Anestia"
                    ),
                    "class_name": "Marketing",
                    "qb_effect": -100.0,
                },
            ],
        )

    def test_paid_in_and_out_apply_esp_and_other_mappings_with_fixed_signs(self):
        api = self.activity_api()

        lines = api["build_lines"](self.complete_payload())

        self.assertEqual(
            lines["paid_in"],
            [
                {
                    "account": "1230000 · Miscellaneous Receivable",
                    "memo": "PAID IN: 8/19's ESP Deposit - MR",
                    "class_name": "",
                    "qb_effect": 100.15,
                },
                {
                    "account": "4444 · TBA Purchases",
                    "memo": "PAID IN: Returned drawer cash",
                    "class_name": "",
                    "qb_effect": 10.0,
                },
            ],
        )
        self.assertEqual(
            lines["paid_out"],
            [
                {
                    "account": "1230000 · Miscellaneous Receivable",
                    "memo": "PAID OUT: 8/18's ESP Deposit - MR",
                    "class_name": "",
                    "qb_effect": -184.35,
                },
                {
                    "account": "4444 · TBA Purchases",
                    "memo": "PAID OUT: Register correction",
                    "class_name": "",
                    "qb_effect": -5.25,
                },
            ],
        )

    def test_paid_in_outreach_requires_memo_and_posts_positive_to_outreach(self):
        api = self.activity_api()
        payload = self.complete_payload()
        payload["paid_in"] = {
            "mode": "app",
            "rows": [
                {
                    "type": "outreach",
                    "memo": "Community event repayment",
                    "amount": 42.50,
                }
            ],
        }

        lines = api["build_lines"](payload)

        self.assertEqual(
            lines["paid_in"],
            [
                {
                    "account": "8505000 · Outreach",
                    "memo": "PAID IN: Community event repayment",
                    "class_name": "",
                    "qb_effect": 42.50,
                }
            ],
        )

    def test_paid_out_outreach_requires_memo_and_posts_negative_to_outreach(self):
        api = self.activity_api()
        payload = self.complete_payload()
        payload["paid_out"] = {
            "mode": "app",
            "rows": [
                {
                    "type": "outreach",
                    "memo": "Community event advance",
                    "amount": 42.50,
                }
            ],
        }

        lines = api["build_lines"](payload)

        self.assertEqual(
            lines["paid_out"],
            [
                {
                    "account": "8505000 · Outreach",
                    "memo": "PAID OUT: Community event advance",
                    "class_name": "",
                    "qb_effect": -42.50,
                }
            ],
        )

    def test_actuals_are_locked_to_the_sum_of_each_app_breakdown(self):
        api = self.activity_api()

        self.assertEqual(
            api["actuals"](self.complete_payload()),
            {"donation": 200.0, "paid_out": 189.6, "paid_in": 110.15},
        )

    def test_manual_modes_discard_rows_and_produce_no_detail_lines(self):
        api = self.activity_api()
        payload = {
            key: {"mode": "quickbooks", "rows": [{"amount": 99}]}
            for key in ("donation", "paid_out", "paid_in")
        }

        normalized = api["normalize"](payload)

        self.assertEqual(
            normalized,
            {
                key: {"mode": "quickbooks", "rows": []}
                for key in ("donation", "paid_out", "paid_in")
            },
        )
        self.assertEqual(
            api["build_lines"](normalized),
            {"donation": [], "paid_out": [], "paid_in": []},
        )

    def test_invalid_breakdown_rows_are_rejected(self):
        api = self.activity_api()
        invalid_rows = (
            (
                "donation",
                {"given_to": "", "purpose": "Food", "manager": "AB", "amount": 5},
                "Given To is required",
            ),
            (
                "donation",
                {"given_to": "Church", "purpose": "Food\nPantry", "manager": "AB", "amount": 5},
                "cannot contain tabs or line breaks",
            ),
            (
                "paid_in",
                {"type": "esp", "original_date": "", "initials": "MR", "amount": 5},
                "Original ESP deposit date is required",
            ),
            (
                "paid_out",
                {"type": "other", "memo": "", "amount": 5},
                "memo is required",
            ),
            (
                "paid_out",
                {"type": "other", "memo": "Correction", "amount": 0},
                "amount must be greater than zero",
            ),
        )
        for category, row, message in invalid_rows:
            with self.subTest(category=category, row=row):
                payload = {
                    key: {"mode": "quickbooks", "rows": []}
                    for key in ("donation", "paid_out", "paid_in")
                }
                payload[category] = {"mode": "app", "rows": [row]}
                with self.assertRaisesRegex(ValueError, message):
                    api["normalize"](payload)

    def test_payload_files_are_unique_and_round_trip_canonical_data(self):
        api = self.activity_api()
        folder = Path(__file__).parents[1] / "runtime" / f"activity_test_{uuid4().hex}"
        folder.mkdir(parents=True)
        try:
            first = api["write"](folder, self.complete_payload())
            second = api["write"](folder, self.complete_payload())
            self.assertNotEqual(first, second)
            self.assertEqual(api["load"](first), api["normalize"](self.complete_payload()))
            self.assertEqual(api["load"](second), api["normalize"](self.complete_payload()))
        finally:
            shutil.rmtree(folder)

    def test_only_nonzero_activity_workflows_appear_in_approved_order(self):
        api = self.activity_api()

        self.assertEqual(
            api["workflow_keys"](
                {"paid_in": 10, "donation": -5, "paid_out": 0}
            ),
            ("donation", "paid_in"),
        )
        self.assertEqual(
            api["workflow_keys"](
                {"paid_in": 10, "donation": 5, "paid_out": 2}
            ),
            ("donation", "paid_in", "paid_out"),
        )
        self.assertEqual(
            api["workflow_keys"](
                {"paid_in": 0, "donation": 0, "paid_out": 0}
            ),
            (),
        )

    def test_paid_in_entry_is_validated_then_appended_without_changing_saved_rows(self):
        api = self.activity_api()
        saved_rows = [
            {
                "type": "other",
                "memo": "First receipt",
                "amount": 25.0,
            }
        ]

        updated_rows = api["append_entry"](
            "paid_in",
            saved_rows,
            {
                "type": "esp",
                "original_date": date(2026, 8, 28),
                "initials": "MR",
                "amount": 75,
            },
        )

        self.assertEqual(
            updated_rows,
            [
                {
                    "type": "other",
                    "memo": "First receipt",
                    "amount": 25.0,
                },
                {
                    "type": "esp",
                    "original_date": "2026-08-28",
                    "initials": "MR",
                    "amount": 75.0,
                },
            ],
        )
        self.assertEqual(len(saved_rows), 1)

    def test_manual_detected_activity_blocks_in_app_closeout(self):
        api = self.activity_api()
        payload = self.complete_payload()
        sources = {"donation": 200, "paid_out": 0, "paid_in": 110.15}

        self.assertTrue(api["closeout_ready"](payload, sources))
        payload["paid_in"] = {"mode": "quickbooks", "rows": []}
        self.assertFalse(api["closeout_ready"](payload, sources))
        sources["paid_in"] = 0
        self.assertTrue(api["closeout_ready"](payload, sources))

    def test_activity_source_reader_does_not_require_hash_amount_header(self):
        import openpyxl

        api = self.activity_api()

        workbook = openpyxl.Workbook()
        balance_sheet = workbook.active
        balance_sheet.title = "BS"
        balance_sheet.append(["Code", None, None, None, "Amount"])
        balance_sheet.append([1122, None, None, None, -200])
        balance_sheet.append([1114, None, None, None, -50])
        hash_sheet = workbook.create_sheet("HASH")
        hash_sheet.append(["Code", "Description", "Value"])
        hash_sheet.append([34, "Paid In", 125])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        self.assertEqual(
            api["read_sources"](output.getvalue(), "BS", "HASH"),
            {"donation": 200.0, "paid_out": 50.0, "paid_in": 0.0},
        )

    def test_activity_source_reader_extracts_only_required_totals(self):
        import openpyxl

        api = self.activity_api()

        workbook = openpyxl.Workbook()
        balance_sheet = workbook.active
        balance_sheet.title = "BS"
        balance_sheet.append(["Code", None, None, None, "Amount"])
        balance_sheet.append([1122, None, None, None, -200])
        balance_sheet.append([1114, None, None, None, -50])
        hash_sheet = workbook.create_sheet("HASH")
        hash_sheet.append(["Code", "Description", "Amount"])
        hash_sheet.append([34, "Paid In", 125])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        self.assertEqual(
            api["read_sources"](output.getvalue(), "BS", "HASH"),
            {"donation": 200.0, "paid_out": 50.0, "paid_in": 125.0},
        )

    def test_activity_source_reader_uses_hash_paid_in_fallback_amount(self):
        import openpyxl

        api = self.activity_api()
        workbook = openpyxl.Workbook()
        balance_sheet = workbook.active
        balance_sheet.title = "BS"
        hash_sheet = workbook.create_sheet("HASH")
        hash_sheet.append(
            ["Code", "Description", "Notes", "Qty", "Value", None, None, None, "Amount"]
        )
        hash_sheet.append([34, "Paid-Ins", None, 1, 184.35, None, None, None, None])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        self.assertEqual(
            api["read_sources"](output.getvalue(), "BS", "HASH")["paid_in"],
            184.35,
        )

    def test_each_activity_section_validates_without_other_section_state(self):
        api = self.activity_api()

        self.assertEqual(
            api["normalize_section"](
                "paid_in",
                {
                    "mode": "app",
                    "rows": [
                        {
                            "type": "other",
                            "memo": "Returned cash",
                            "amount": 25,
                        }
                    ],
                },
            ),
            {
                "mode": "app",
                "rows": [
                    {
                        "type": "other",
                        "memo": "Returned cash",
                        "amount": 25.0,
                    }
                ],
            },
        )

    def test_engine_writes_detail_lines_and_closeout_differences(self):
        from app import pos_to_quickbooks_v2 as engine

        payload = self.complete_payload()
        actuals = self.activity_api()["actuals"](payload)
        closeout_actuals = {
            "cash": 0,
            "checks": 0,
            "donation": actuals["donation"],
            "charge_house": 0,
            "offline_zon": 0,
            "vendor_coupons": 0,
            "paid_out": actuals["paid_out"],
            "paid_in": actuals["paid_in"],
        }
        closeout_payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": closeout_actuals,
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1000,
            "approve_final_pos": True,
        }
        folder = Path(__file__).parents[1] / "runtime" / f"activity_iif_{uuid4().hex}"
        folder.mkdir(parents=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = folder
        engine.LOG_DIR = folder
        engine.log.disabled = True
        try:
            try:
                iif_path = engine.generate_iif(
                    {},
                    {},
                    {},
                    date(2026, 8, 28),
                    bs_data={"donation": 250, "paid_out": 200},
                    paid_in_total=105,
                    closeout_payload=closeout_payload,
                    activity_payload=payload,
                )
            except TypeError as exc:
                self.fail(f"engine activity breakdown integration is missing: {exc}")
            text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            shutil.rmtree(folder)

        for expected in (
            "8506000 · Outreach - Donations\t\t100.00\tGiven to AME Zion Church for Food Pantry - Steoria\tMarketing",
            "1230000 · Miscellaneous Receivable\t\t-100.15\tPAID IN: 8/19's ESP Deposit - MR",
            "4444 · TBA Purchases\t\t-10.00\tPAID IN: Returned drawer cash",
            "1230000 · Miscellaneous Receivable\t\t184.35\tPAID OUT: 8/18's ESP Deposit - MR",
            "4444 · TBA Purchases\t\t5.25\tPAID OUT: Register correction",
            "8314000 · FE - Cash Over/Shorts\t\t50.00\tOver/Short per Closeout Sheet - Donation\tAdmin",
            "8314000 · FE - Cash Over/Shorts\t\t10.40\tOver/Short per Closeout Sheet - Paid Out\tAdmin",
            "8314000 · FE - Cash Over/Shorts\t\t5.15\tOver/Short per Closeout Sheet - Paid In\tAdmin",
        ):
            with self.subTest(expected=expected):
                self.assertIn(expected, text)
        self.assertNotIn("8506000 · Outreach - Donations\t\t200.00\t\t", text)
        self.assertNotIn("4444 · TBA Purchases\t\t-110.15\tPAID IN:", text)
        self.assertNotIn("4444 · TBA Purchases\t\t189.60\tPAID OUT:", text)

    def test_streamlit_engine_command_passes_activity_payload_file(self):
        source_path = Path(__file__).parents[1] / "streamlit_app.py"
        source_tree = ast.parse(source_path.read_text(encoding="utf-8"))
        command_node = next(
            node
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef) and node.name == "build_engine_command"
        )
        namespace = {"Path": Path, "date": date, "sys": sys}
        exec(
            compile(ast.Module(body=[command_node], type_ignores=[]), str(source_path), "exec"),
            namespace,
        )

        try:
            command = namespace["build_engine_command"](
                engine_path=Path("engine.py"),
                deposit_date=date(2026, 8, 28),
                membership_path=Path("members.json"),
                membership_mode="automatic",
                coupon_mode="quickbooks",
                coupon_closeout_total=None,
                coupon_ncg_total=None,
                coupon_mfg_total=None,
                activity_path=Path("activities.json"),
            )
        except TypeError as exc:
            self.fail(f"activity payload command integration is missing: {exc}")

        self.assertIn("--activity-breakdowns-file", command)
        self.assertEqual(
            command[command.index("--activity-breakdowns-file") + 1],
            "activities.json",
        )

    def test_engine_allows_mixed_activity_with_manual_closeout(self):
        from app import pos_to_quickbooks_v2 as engine

        payload = self.complete_payload()
        payload["paid_out"] = {"mode": "quickbooks", "rows": []}
        folder = Path(__file__).parents[1] / "runtime" / f"mixed_iif_{uuid4().hex}"
        folder.mkdir(parents=True)
        old_output_dir = engine.output_dir
        old_log_dir = engine.LOG_DIR
        old_log_disabled = engine.log.disabled
        engine.output_dir = folder
        engine.LOG_DIR = folder
        engine.log.disabled = True
        try:
            iif_path = engine.generate_iif(
                {},
                {},
                {},
                date(2026, 8, 28),
                bs_data={"donation": 250, "paid_out": 75},
                paid_in_total=105,
                activity_payload=payload,
                closeout_payload={"mode": "manual"},
            )
            text = iif_path.read_text(encoding="utf-8")
        finally:
            engine.output_dir = old_output_dir
            engine.LOG_DIR = old_log_dir
            engine.log.disabled = old_log_disabled
            shutil.rmtree(folder)

        self.assertIn("Given to AME Zion Church for Food Pantry - Steoria", text)
        self.assertIn("4444 · TBA Purchases\t\t75.00\tPAID OUT:", text)
        self.assertIn(
            "8314000 · FE - Cash Over/Shorts\t\t50.00\t"
            "Over/Short per Closeout Sheet - Donation\tAdmin",
            text,
        )
        self.assertIn(
            "8314000 · FE - Cash Over/Shorts\t\t5.15\t"
            "Over/Short per Closeout Sheet - Paid In\tAdmin",
            text,
        )

    def test_engine_rejects_mixed_activity_with_in_app_closeout(self):
        from app import pos_to_quickbooks_v2 as engine

        payload = self.complete_payload()
        payload["paid_out"] = {"mode": "quickbooks", "rows": []}
        closeout = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {
                "cash": 0,
                "checks": 0,
                "donation": 200,
                "charge_house": 0,
                "offline_zon": 0,
                "vendor_coupons": 0,
                "paid_out": 75,
                "paid_in": 110.15,
            },
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1,
            "approve_final_pos": True,
        }
        with self.assertRaisesRegex(ValueError, "every detected activity"):
            engine.generate_iif(
                {},
                {},
                {},
                date(2026, 8, 28),
                bs_data={"donation": 250, "paid_out": 75},
                paid_in_total=105,
                activity_payload=payload,
                closeout_payload=closeout,
            )

    def test_engine_requires_matching_locked_actuals(self):
        from app import pos_to_quickbooks_v2 as engine

        payload = self.complete_payload()

        mismatched_closeout = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {
                "cash": 0,
                "checks": 0,
                "donation": 199,
                "charge_house": 0,
                "offline_zon": 0,
                "vendor_coupons": 0,
                "paid_out": 189.60,
                "paid_in": 110.15,
            },
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1,
            "approve_final_pos": True,
        }
        with self.assertRaisesRegex(
            ValueError, "Donation Closeout actual must match its app breakdown"
        ):
            try:
                engine.generate_iif(
                    {},
                    {},
                    {},
                    date(2026, 8, 28),
                    activity_payload=payload,
                    closeout_payload=mismatched_closeout,
                )
            except TypeError as exc:
                self.fail(f"engine activity breakdown integration is missing: {exc}")


if __name__ == "__main__":
    unittest.main()
