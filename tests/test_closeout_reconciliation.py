from io import BytesIO
from pathlib import Path
from uuid import uuid4

import openpyxl
import unittest


STANDARD_ORDER = (
    "cash",
    "checks",
    "donation",
    "charge_house",
    "offline_zon",
    "vendor_coupons",
    "paid_in",
    "paid_out",
)


BS_VALUES = {
    901: 1250.00,
    902: 75.00,
    1122: 20.00,
    906: 45.00,
    934: -12.00,
    908: 188.25,
    1114: 47.06,
}


def workbook_bytes(
    *, paid_in_column=1, paid_in_amount=30.00, bs_values=None, preceding_amount_like_header=False
):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    balance_sheet = workbook.create_sheet("Daily BS")
    hash_sheet = workbook.create_sheet("Daily HASH")

    for code, amount in (bs_values or BS_VALUES).items():
        balance_sheet.append([code, None, None, None, amount])

    hash_header = ["Code", "Description", "Notes", "Type", "aMoUnT"]
    if preceding_amount_like_header:
        hash_header[1] = "Net Amount"
    hash_sheet.append(hash_header)
    hash_row = [None] * 5
    hash_row[paid_in_column - 1] = 34
    if preceding_amount_like_header:
        hash_row[1] = 999.99
    hash_row[4] = paid_in_amount
    hash_sheet.append(hash_row)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class CloseoutReconciliationTests(unittest.TestCase):
    def test_build_standard_reconciliation_returns_eight_rows_in_approved_order(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {
            "cash": 100.00,
            "checks": 50.00,
            "donation": 20.00,
            "charge_house": 10.00,
            "offline_zon": 12.00,
            "vendor_coupons": 181.50,
            "paid_out": 40.00,
            "paid_in": 30.00,
        }
        actuals = {
            "cash": 105.00,
            "checks": 45.00,
            "donation": 25.00,
            "charge_house": 8.00,
            "offline_zon": 0.00,
            "vendor_coupons": 188.25,
            "paid_out": 50.00,
            "paid_in": 35.00,
        }

        result = build_standard_reconciliation(baselines, actuals)

        self.assertEqual([row["key"] for row in result], list(STANDARD_ORDER))
        self.assertEqual(len(result), 8)
        self.assertTrue(
            all(
                set(row)
                == {
                    "key",
                    "label",
                    "baseline",
                    "actual",
                    "difference",
                    "detail_qb_effect",
                    "adjustment_account",
                    "adjustment_memo",
                    "adjustment_qb_effect",
                    "managed_externally",
                }
                for row in result
            )
        )

    def test_build_standard_reconciliation_calculates_fixture_effects_and_metadata(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {
            "cash": 100.00,
            "checks": 50.00,
            "donation": 20.00,
            "charge_house": 10.00,
            "offline_zon": 12.00,
            "vendor_coupons": 181.50,
            "paid_out": 40.00,
            "paid_in": 30.00,
        }
        actuals = {
            "cash": 105.00,
            "checks": 45.00,
            "donation": 25.00,
            "charge_house": 8.00,
            "offline_zon": 0.00,
            "vendor_coupons": 188.25,
            "paid_out": 50.00,
            "paid_in": 35.00,
        }

        rows = build_standard_reconciliation(baselines, actuals)
        by_key = {row["key"]: row for row in rows}

        self.assertEqual(by_key["cash"]["difference"], 5.00)
        self.assertIsNone(by_key["cash"]["detail_qb_effect"])
        self.assertEqual(by_key["cash"]["adjustment_qb_effect"], 5.00)
        self.assertEqual(
            by_key["checks"]["adjustment_memo"],
            "Over/Short per Closeout Sheet - Check",
        )
        self.assertEqual(by_key["donation"]["difference"], 5.00)
        self.assertEqual(by_key["donation"]["detail_qb_effect"], -25.00)
        self.assertEqual(by_key["donation"]["adjustment_qb_effect"], 5.00)
        self.assertEqual(by_key["paid_in"]["detail_qb_effect"], 35.00)
        self.assertEqual(by_key["paid_in"]["adjustment_qb_effect"], -5.00)
        self.assertTrue(by_key["vendor_coupons"]["managed_externally"])
        self.assertEqual(
            by_key["vendor_coupons"]["adjustment_memo"],
            "Over/Short per Closeout Sheet - Coupon",
        )

    def test_build_standard_reconciliation_rejects_negative_actual_for_every_category(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {key: 1.00 for key in STANDARD_ORDER}
        actuals = {key: 1.00 for key in STANDARD_ORDER}

        for key in STANDARD_ORDER:
            with self.subTest(key=key), self.assertRaisesRegex(
                ValueError, "zero or greater"
            ):
                invalid_actuals = dict(actuals)
                invalid_actuals[key] = -0.01
                build_standard_reconciliation(baselines, invalid_actuals)

    def test_build_standard_reconciliation_rejects_missing_baseline_or_actual_with_label(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {key: 1.00 for key in STANDARD_ORDER}
        actuals = {key: 1.00 for key in STANDARD_ORDER}
        labels = {
            "cash": "Cash",
            "checks": "Checks",
            "donation": "Donation",
            "charge_house": "Charge (House)",
            "offline_zon": "Offline Zon",
            "vendor_coupons": "Vendor Coupons",
            "paid_out": "Paid Out",
            "paid_in": "Paid In",
        }

        for key, label in labels.items():
            with self.subTest(mapping="baseline", key=key):
                missing_baseline = dict(baselines)
                missing_baseline.pop(key)
                with self.assertRaises(ValueError) as error:
                    build_standard_reconciliation(missing_baseline, actuals)
                self.assertIn(label, str(error.exception))
            with self.subTest(mapping="actual", key=key):
                missing_actual = dict(actuals)
                missing_actual.pop(key)
                with self.assertRaises(ValueError) as error:
                    build_standard_reconciliation(baselines, missing_actual)
                self.assertIn(label, str(error.exception))

    def test_build_standard_reconciliation_quantizes_decimal_sensitive_values(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {key: "0.00" for key in STANDARD_ORDER}
        actuals = {key: "0.00" for key in STANDARD_ORDER}
        baselines["donation"] = "10.009"
        actuals["donation"] = "10.004"

        donation = build_standard_reconciliation(baselines, actuals)[2]

        for field in (
            "baseline",
            "actual",
            "difference",
            "detail_qb_effect",
            "adjustment_qb_effect",
        ):
            self.assertIsInstance(donation[field], float)
        self.assertEqual(donation["baseline"], 10.01)
        self.assertEqual(donation["actual"], 10.00)
        self.assertEqual(donation["difference"], -0.01)
        self.assertEqual(donation["detail_qb_effect"], -10.00)
        self.assertEqual(donation["adjustment_qb_effect"], -0.01)

    def test_build_standard_reconciliation_rejects_negative_subcent_actuals_for_every_category(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {key: 1.00 for key in STANDARD_ORDER}
        actuals = {key: 1.00 for key in STANDARD_ORDER}

        for key in STANDARD_ORDER:
            with self.subTest(key=key), self.assertRaisesRegex(
                ValueError, "zero or greater"
            ):
                negative_subcent_actuals = dict(actuals)
                negative_subcent_actuals[key] = "-0.004"
                build_standard_reconciliation(baselines, negative_subcent_actuals)

    def test_build_standard_reconciliation_uses_canonical_order_for_reversed_mappings(self):
        from app.closeout_reconciliation import build_standard_reconciliation

        baselines = {key: float(index) for index, key in enumerate(reversed(STANDARD_ORDER), 1)}
        actuals = {key: float(index) for index, key in enumerate(reversed(STANDARD_ORDER), 1)}

        rows = build_standard_reconciliation(baselines, actuals)

        self.assertEqual([row["key"] for row in rows], list(STANDARD_ORDER))

    def test_standard_metadata_covers_exactly_canonical_closeout_keys(self):
        from app.closeout_reconciliation import STANDARD_METADATA, STANDARD_CLOSEOUT_ORDER

        self.assertEqual(set(STANDARD_METADATA), set(STANDARD_CLOSEOUT_ORDER))

    def test_read_closeout_baselines_returns_approved_key_order_and_values(self):
        from app.closeout_reconciliation import (
            STANDARD_CLOSEOUT_ORDER,
            read_closeout_baselines,
        )

        result = read_closeout_baselines(
            workbook_bytes(), "Daily BS", "Daily HASH"
        )

        self.assertEqual(STANDARD_CLOSEOUT_ORDER, STANDARD_ORDER)
        self.assertEqual(tuple(result), STANDARD_ORDER)
        self.assertEqual(result["cash"], 1250.00)
        self.assertEqual(result["checks"], 75.00)
        self.assertEqual(result["donation"], 20.00)
        self.assertEqual(result["charge_house"], 45.00)
        self.assertEqual(result["offline_zon"], 12.00)
        self.assertEqual(result["vendor_coupons"], 188.25)
        self.assertEqual(result["paid_out"], 47.06)
        self.assertEqual(result["paid_in"], 30.00)


    def test_default_closeout_actuals_preserve_baselines_except_counted_coupons_and_offline_zon(
        self,
    ):
        from app.closeout_reconciliation import default_closeout_actuals

        baselines = {
            "cash": 1250.00,
            "checks": 75.00,
            "donation": 20.00,
            "charge_house": 45.00,
            "offline_zon": 12.00,
            "vendor_coupons": 188.25,
            "paid_out": 47.06,
            "paid_in": 30.00,
        }

        result = default_closeout_actuals(baselines, 152.25)

        self.assertEqual(tuple(result), STANDARD_ORDER)
        self.assertEqual(result["cash"], 1250.00)
        self.assertEqual(result["offline_zon"], 0.00)
        self.assertEqual(result["vendor_coupons"], 152.25)
        self.assertEqual(baselines["offline_zon"], 12.00)
        self.assertEqual(baselines["vendor_coupons"], 188.25)


    def test_read_closeout_baselines_finds_paid_in_code_in_any_first_four_columns(self):
        from app.closeout_reconciliation import read_closeout_baselines

        for paid_in_column in range(1, 5):
            with self.subTest(paid_in_column=paid_in_column):
                result = read_closeout_baselines(
                    workbook_bytes(paid_in_column=paid_in_column),
                    "Daily BS",
                    "Daily HASH",
                )
                self.assertEqual(result["paid_in"], 30.00)


    def test_read_closeout_baselines_requires_exact_amount_header(self):
        from app.closeout_reconciliation import read_closeout_baselines

        result = read_closeout_baselines(
            workbook_bytes(preceding_amount_like_header=True),
            "Daily BS",
            "Daily HASH",
        )

        self.assertEqual(result["paid_in"], 30.00)


    def test_read_closeout_baselines_treats_blank_paid_in_amount_as_zero(self):
        """A blank HASH code 34 placeholder means there was no Paid In activity."""
        from app.closeout_reconciliation import read_closeout_baselines

        result = read_closeout_baselines(
            workbook_bytes(paid_in_amount=None),
            "Daily BS",
            "Daily HASH",
        )

        self.assertEqual(result["paid_in"], 0.00)


    def test_read_closeout_baselines_uses_hash_paid_in_fallback_amount(self):
        """Closeout must match the established HASH parser's shifted-column fallback."""
        from app.closeout_reconciliation import read_closeout_baselines

        workbook = openpyxl.Workbook()
        balance_sheet = workbook.active
        balance_sheet.title = "Daily BS"
        hash_sheet = workbook.create_sheet("Daily HASH")
        hash_sheet.append(
            ["Code", "Description", "Notes", "Qty", "Value", None, None, None, "Amount"]
        )
        hash_sheet.append([34, "Paid-Ins", None, 1, 184.35, None, None, None, None])
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        result = read_closeout_baselines(
            output.getvalue(),
            "Daily BS",
            "Daily HASH",
        )

        self.assertEqual(result["paid_in"], 184.35)


    def test_default_closeout_actuals_rebuilds_canonical_normalized_float_mapping(self):
        from app.closeout_reconciliation import default_closeout_actuals

        baselines = {
            "paid_in": "30.004",
            "paid_out": "47.064",
            "vendor_coupons": "188.254",
            "offline_zon": "12.004",
            "charge_house": "45.004",
            "donation": "20.004",
            "checks": "75.004",
            "cash": "1250.004",
            "extra": "999.999",
        }

        result = default_closeout_actuals(baselines, "-152.256")

        self.assertEqual(tuple(result), STANDARD_ORDER)
        self.assertTrue(all(isinstance(value, float) for value in result.values()))
        self.assertEqual(result["cash"], 1250.00)
        self.assertEqual(result["checks"], 75.00)
        self.assertEqual(result["donation"], 20.00)
        self.assertEqual(result["charge_house"], 45.00)
        self.assertEqual(result["offline_zon"], 0.00)
        self.assertEqual(result["vendor_coupons"], 152.26)
        self.assertEqual(result["paid_out"], 47.06)
        self.assertEqual(result["paid_in"], 30.00)


    def test_read_closeout_baselines_requires_exact_requested_sheet_names(self):
        from app.closeout_reconciliation import read_closeout_baselines

        cases = [
            ("Missing BS", "Daily HASH", "Balance Sheet.*Missing BS"),
            ("Daily BS", "Missing HASH", "HASH.*Missing HASH"),
        ]
        for bs_sheet_name, hash_sheet_name, message in cases:
            with self.subTest(bs_sheet_name=bs_sheet_name), self.assertRaisesRegex(
                ValueError, message
            ):
                read_closeout_baselines(
                    workbook_bytes(), bs_sheet_name, hash_sheet_name
                )


    def test_read_closeout_baselines_rejects_invalid_monetary_data(self):
        from app.closeout_reconciliation import read_closeout_baselines

        invalid_bs_values = dict(BS_VALUES)
        invalid_bs_values[901] = "not money"

        with self.assertRaisesRegex(ValueError, "valid.*amount|monetary"):
            read_closeout_baselines(
                workbook_bytes(bs_values=invalid_bs_values),
                "Daily BS",
                "Daily HASH",
            )

        with self.assertRaisesRegex(ValueError, "valid.*amount|monetary"):
            read_closeout_baselines(
                workbook_bytes(paid_in_amount="NaN"), "Daily BS", "Daily HASH"
            )


    def test_default_closeout_actuals_rejects_invalid_coupon_total(self):
        from app.closeout_reconciliation import default_closeout_actuals

        with self.assertRaisesRegex(ValueError, "valid.*amount|monetary"):
            default_closeout_actuals({"cash": 1.00}, "Infinity")


class CloseoutAdjustmentTests(unittest.TestCase):
    def closeout_payload(self, **overrides):
        payload = {
            "mode": "closeout",
            "reviewed": True,
            "actuals": {key: 10 for key in STANDARD_ORDER},
            "payroll": 0,
            "safe": {"type": "none", "amount": 0},
            "plants_purchase": 0,
            "custom_tba": [],
            "final_total": 1000,
            "approve_final_pos": False,
        }
        payload.update(overrides)
        return payload

    def test_build_misc_adjustments_uses_exact_mappings_and_iif_signs(self):
        from app.closeout_reconciliation import build_misc_adjustments

        rows = build_misc_adjustments(
            {
                "payroll": -4000,
                "safe": {"type": "overage", "amount": 25},
                "plants_purchase": 60,
                "custom_tba": [
                    {
                        "memo": "Unusual Closeout item",
                        "amount": 12,
                        "direction": "removes",
                    }
                ],
            }
        )

        self.assertEqual(
            [
                (row["account"], row["memo"], row["qb_effect"], row["iif_amount"])
                for row in rows
            ],
            [
                ("1140000 · Cash Drawers/Safe", "Payroll - Check Cashing", -4000.0, 4000.0),
                (
                    "8314000 · FE - Cash Over/Shorts",
                    "Safe Overage Cash added to deposit",
                    25.0,
                    -25.0,
                ),
                ("1130000 · Petty Cash", "Plants Dept - Market Purchases", -60.0, 60.0),
                ("4444 · TBA Purchases", "Unusual Closeout item", -12.0, 12.0),
            ],
        )

    def test_safe_shortage_and_custom_adds_post_negative_and_positive_effects(self):
        from app.closeout_reconciliation import build_misc_adjustments

        rows = build_misc_adjustments(
            {
                "safe": {"type": "shortage", "amount": 25},
                "custom_tba": [
                    {"memo": "Cash found", "amount": 12, "direction": "adds"}
                ],
            }
        )

        self.assertEqual([row["qb_effect"] for row in rows], [-25.0, 12.0])
        self.assertEqual([row["iif_amount"] for row in rows], [25.0, -12.0])

    def test_build_misc_adjustments_rejects_invalid_payroll_and_custom_rows(self):
        from app.closeout_reconciliation import build_misc_adjustments

        invalid_payloads = [
            {"payroll": 3999, "custom_tba": []},
            {
                "payroll": 0,
                "custom_tba": [{"memo": " ", "amount": 5, "direction": "adds"}],
            },
            {
                "payroll": 0,
                "custom_tba": [{"memo": "Item", "amount": 0, "direction": "adds"}],
            },
            {
                "payroll": 0,
                "custom_tba": [
                    {"memo": "Item", "amount": 5, "direction": "addition"}
                ],
            },
        ]

        for payload in invalid_payloads:
            with self.subTest(payload=payload), self.assertRaises(ValueError):
                build_misc_adjustments(payload)

    def test_normalize_closeout_payload_rejects_iif_delimiters_in_custom_tba_memos(self):
        from app.closeout_reconciliation import normalize_closeout_payload

        for delimiter, label in (("\t", "tab"), ("\r", "carriage return"), ("\n", "line feed")):
            payload = self.closeout_payload(
                custom_tba=[
                    {
                        "memo": f"Unsafe{delimiter}memo",
                        "amount": 5,
                        "direction": "adds",
                    }
                ]
            )
            with self.subTest(delimiter=label), self.assertRaisesRegex(
                ValueError, "Custom TBA memo cannot contain tabs or line breaks"
            ):
                normalize_closeout_payload(payload)

    def test_final_difference_requires_explicit_approval_with_exact_pos_line(self):
        from app.closeout_reconciliation import calculate_final_pos_adjustment

        unapproved = calculate_final_pos_adjustment(1000, 1025, approved=False)
        approved = calculate_final_pos_adjustment(1000, 1025, approved=True)

        self.assertEqual(unapproved["remaining"], 25.0)
        self.assertIsNone(unapproved["line"])
        self.assertTrue(unapproved["requires_approval"])
        self.assertEqual(
            approved["line"],
            {
                "kind": "final_pos",
                "account": "8314000 · FE - Cash Over/Shorts",
                "memo": "Over/Short per POS (to = POS total)",
                "qb_effect": 25.0,
                "iif_amount": -25.0,
            },
        )

    def test_final_balancing_omits_zero_and_rejects_nonpositive_total(self):
        from app.closeout_reconciliation import calculate_final_pos_adjustment

        zero = calculate_final_pos_adjustment("1000.004", "1000.004", approved=False)

        self.assertEqual(zero["provisional_total"], 1000.0)
        self.assertEqual(zero["final_total"], 1000.0)
        self.assertEqual(zero["remaining"], 0.0)
        self.assertIsNone(zero["line"])
        self.assertFalse(zero["requires_approval"])
        for final_total in (0, -0.01, "NaN"):
            with self.subTest(final_total=final_total), self.assertRaises(ValueError):
                calculate_final_pos_adjustment(1000, final_total, approved=True)

    def test_final_balancing_rejects_non_boolean_approval_values(self):
        from app.closeout_reconciliation import calculate_final_pos_adjustment

        for approved in ("false", 1, None):
            with self.subTest(approved=approved), self.assertRaisesRegex(
                ValueError, "approved.*boolean"
            ):
                calculate_final_pos_adjustment(1000, 1025, approved=approved)

    def test_normalize_manual_mode_returns_only_canonical_manual_payload(self):
        from app.closeout_reconciliation import normalize_closeout_payload

        self.assertEqual(
            normalize_closeout_payload({"mode": "manual", "unexpected": "ignored"}),
            {"mode": "manual"},
        )

    def test_closeout_payload_round_trips_to_distinct_json_files(self):
        from app.closeout_reconciliation import (
            load_closeout_payload_file,
            normalize_closeout_payload,
            write_closeout_payload_file,
        )

        payload = self.closeout_payload(
            payroll="-4000.004",
            safe={"type": "overage", "amount": "25.004"},
            plants_purchase="60.004",
            custom_tba=[
                {"memo": "  Unusual Closeout item  ", "amount": "12.004", "direction": "removes"}
            ],
        )
        folder = Path(__file__).parent
        first_path = write_closeout_payload_file(folder, payload)
        second_path = write_closeout_payload_file(folder, payload)
        try:
            self.assertNotEqual(first_path, second_path)
            self.assertEqual(first_path.parent, folder)
            self.assertEqual(
                load_closeout_payload_file(first_path), normalize_closeout_payload(payload)
            )
        finally:
            first_path.unlink(missing_ok=True)
            second_path.unlink(missing_ok=True)

    def test_normalize_closeout_payload_rejects_invalid_state_and_does_not_mutate_input(self):
        from app.closeout_reconciliation import normalize_closeout_payload

        payload = self.closeout_payload(
            actuals={key: "10.004" for key in reversed(STANDARD_ORDER)},
            safe={"type": "overage", "amount": "25.004"},
            custom_tba=[{"memo": "  Item  ", "amount": "1.004", "direction": "adds"}],
        )
        normalized = normalize_closeout_payload(payload)

        self.assertEqual(tuple(normalized["actuals"]), STANDARD_ORDER)
        self.assertEqual(normalized["actuals"]["cash"], 10.0)
        self.assertEqual(normalized["safe"], {"type": "overage", "amount": 25.0})
        self.assertEqual(normalized["custom_tba"][0]["memo"], "Item")
        self.assertEqual(payload["actuals"]["cash"], "10.004")
        self.assertEqual(payload["safe"]["amount"], "25.004")
        self.assertEqual(payload["custom_tba"][0]["memo"], "  Item  ")

        invalid_payloads = [
            self.closeout_payload(reviewed=False),
            self.closeout_payload(actuals={"cash": 10}),
            self.closeout_payload(approve_final_pos=1),
            "not a payload",
        ]
        for invalid_payload in invalid_payloads:
            with self.subTest(invalid_payload=invalid_payload), self.assertRaises(ValueError):
                normalize_closeout_payload(invalid_payload)

    def test_loader_rejects_malformed_json(self):
        from app.closeout_reconciliation import load_closeout_payload_file

        path = Path(__file__).parent / f"invalid_closeout_{uuid4()}.json"
        try:
            path.write_text("{not json", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "JSON"):
                load_closeout_payload_file(path)
        finally:
            path.unlink(missing_ok=True)
