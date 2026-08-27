from contextlib import contextmanager
from decimal import Decimal, InvalidOperation, ROUND_DOWN
import hashlib
import json
import os
from pathlib import Path
import time
from uuid import uuid4


MEMBER_SHARES_RECEIVABLE = "1260000 · Member Shares Receivable"
INTEREST_INCOME = "9104000 · Interest Income"
MEMBER_SHARES_EQUITY = "6100000 · Member Shares (Paid-In Equity)"

PLAN_TERMS = {
    "1 year": (Decimal("8.45"), Decimal("0.27")),
    "3 year": (Decimal("15.69"), Decimal("1.52")),
    "5 year": (Decimal("10.55"), Decimal("1.55")),
}

PLAN_DEPOSITS = {
    "1 year": Decimal("10.00"),
    "3 year": Decimal("15.00"),
    "5 year": Decimal("10.00"),
}

PLAN_MAX_PERIODS = {
    "1 year": 11,
    "3 year": 6,
    "5 year": 10,
}

PAYMENT_TYPES = {"Paid in full", "New plan", "Existing plan"}
HANDLING_MODES = {"automatic", "manual"}
HANDLING_CHOICES = {
    "Breakdown in app using the Ownership Payments sheet": "automatic",
    "Finish manually in QuickBooks": "manual",
}
PAYMENT_OPTIONS = {
    "Paid in full — $100": {"payment_type": "Paid in full", "plan": ""},
    "New plan — 1 year": {"payment_type": "New plan", "plan": "1 year"},
    "New plan — 3 year": {"payment_type": "New plan", "plan": "3 year"},
    "New plan — 5 year": {"payment_type": "New plan", "plan": "5 year"},
    "Existing plan — 1 year": {"payment_type": "Existing plan", "plan": "1 year"},
    "Existing plan — 3 year": {"payment_type": "Existing plan", "plan": "3 year"},
    "Existing plan — 5 year": {"payment_type": "Existing plan", "plan": "5 year"},
}


def payment_fields_from_option(option: str) -> dict:
    try:
        return dict(PAYMENT_OPTIONS[option])
    except KeyError:
        raise ValueError("Unknown membership payment option") from None


def membership_payment_from_entry(
    *,
    member_name: str,
    member_number_status: str | None,
    member_number: str,
    quickbooks_member_exists: bool | None = None,
    payment_option: str,
    amount: float,
    interest_periods: int | None = None,
) -> dict:
    if quickbooks_member_exists is None:
        raise ValueError("Select Yes or No for whether the member exists in QuickBooks")
    if member_number_status not in {"Yes", "No"}:
        raise ValueError("Select Yes or No for the member number question")
    payment = {
        "member_name": member_name if quickbooks_member_exists else "",
        "member_number": member_number if member_number_status == "Yes" else "",
        "member_number_pending": member_number_status == "No",
        "quickbooks_member_exists": quickbooks_member_exists,
        "payment_option": payment_option,
        "amount": amount,
        "interest_periods": interest_periods,
    }
    payment.update(payment_fields_from_option(payment_option))
    payment.pop("payment_option")
    return payment


def remove_membership_payment(payments: list[dict], position: int) -> list[dict]:
    if position < 0 or position >= len(payments):
        raise IndexError("Membership payment position is out of range")
    return [dict(payment) for index, payment in enumerate(payments) if index != position]


def membership_mode_from_choice(choice: str | None) -> str | None:
    if choice is None:
        return None
    try:
        return HANDLING_CHOICES[choice]
    except KeyError:
        raise ValueError("Unknown membership workflow choice") from None


def plan_reference_rows() -> list[dict]:
    rows = []
    for plan in ("1 year", "3 year", "5 year"):
        installment, interest = PLAN_TERMS[plan]
        deposit = PLAN_DEPOSITS[plan]
        payments = PLAN_MAX_PERIODS[plan]
        rows.append({
            "Plan": plan,
            "Deposit": float(deposit),
            "Total Paid": float((deposit + installment * payments).quantize(Decimal("0.01"))),
            "Payments": payments,
            "Installment": float(installment),
            "Principal": float((installment - interest).quantize(Decimal("0.01"))),
            "Interest": float(interest),
        })
    return rows


def prepare_membership_editor_rows(
    rows: list[dict], allow_interest_override: bool
) -> list[dict]:
    prepared_rows = [dict(row) for row in rows]
    for row in prepared_rows:
        row.pop("delete", None)
        payment_option = row.pop("payment_option", None)
        if isinstance(payment_option, str) and payment_option.strip():
            row.update(payment_fields_from_option(payment_option))
    if not allow_interest_override:
        for row in prepared_rows:
            row["interest_periods"] = None
    return prepared_rows


def normalize_membership_editor_rows(rows: list[dict]) -> tuple[list[dict], bool]:
    normalized_rows = [dict(row) for row in rows]
    refresh_required = False
    for row in normalized_rows:
        if row.get("payment_option") != "Paid in full — $100":
            continue
        try:
            current_amount = Decimal(str(row.get("amount")))
            is_hundred = current_amount.is_finite() and current_amount == Decimal("100.00")
        except (InvalidOperation, TypeError, ValueError):
            is_hundred = False
        if not is_hundred:
            row["amount"] = 100.00
            refresh_required = True
    return normalized_rows, refresh_required


def subscription_action_status(subscription_total: float) -> dict:
    total = abs(Decimal(str(subscription_total))).quantize(Decimal("0.01"))
    if total == Decimal("0.00"):
        return {
            "needs_action": False,
            "title": "No Subscription Revenue",
            "message": "No member-share action is needed for this deposit.",
        }
    return {
        "needs_action": True,
        "title": f"Subscription Revenue found: ${total:,.2f}",
        "message": (
            "Choose automatic splitting or finish manually in QuickBooks "
            "before building the deposit."
        ),
    }


def membership_editor_key(workbook_bytes: bytes, reset_counter: int) -> str:
    workbook_digest = hashlib.sha256(workbook_bytes).hexdigest()[:16]
    return f"membership_payments_{reset_counter}_{workbook_digest}"


def load_membership_payments_file(path: str | Path) -> list[dict]:
    payments = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(payments, list):
        raise ValueError("Membership payment file must contain a list of payments")
    return payments


def _remove_stale_membership_files(folder: Path, stale_seconds: int) -> None:
    cutoff = time.time() - stale_seconds
    for candidate in folder.glob("membership_payments_*.json"):
        try:
            if candidate.stat().st_mtime < cutoff:
                candidate.unlink()
        except (FileNotFoundError, OSError):
            pass


def write_membership_payments_file(
    folder: str | Path, payments: list[dict], stale_seconds: int = 3600
) -> Path:
    folder = Path(folder)
    folder.mkdir(parents=True, exist_ok=True)
    _remove_stale_membership_files(folder, stale_seconds)
    destination = folder / f"membership_payments_{uuid4().hex}.json"
    destination.write_text(
        json.dumps(payments, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    try:
        destination.chmod(0o600)
    except OSError:
        pass
    return destination


@contextmanager
def exclusive_run_lock(lock_path: str | Path, stale_seconds: int = 600):
    """Prevent shared date-based inputs and outputs from overlapping across processes."""
    _ = stale_seconds  # Kept for compatibility; OS locks are released automatically on crashes.
    lock_path = Path(lock_path)
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_file = lock_path.open("a+b")
    try:
        lock_file.seek(0, os.SEEK_END)
        if lock_file.tell() == 0:
            lock_file.write(b"\0")
            lock_file.flush()
        lock_file.seek(0)

        if os.name == "nt":
            import msvcrt

            msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl

            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except (BlockingIOError, OSError):
        lock_file.close()
        raise RuntimeError(
            "Another deposit is currently running. Wait for it to finish, then try again."
        ) from None

    try:
        yield
    finally:
        try:
            lock_file.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(lock_file.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)
        finally:
            lock_file.close()


def read_subscription_total(workbook_bytes: bytes, bs_sheet_name: str | None = None) -> float:
    from io import BytesIO

    import openpyxl

    workbook = openpyxl.load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        if bs_sheet_name:
            if bs_sheet_name not in workbook.sheetnames:
                raise ValueError(f"Balance Sheet tab '{bs_sheet_name}' was not found")
            sheet = workbook[bs_sheet_name]
        else:
            sheet_name = next(
                (name for name in workbook.sheetnames if name.upper().endswith(" BS")),
                None,
            )
            if sheet_name is None:
                raise ValueError("Balance Sheet tab was not found")
            sheet = workbook[sheet_name]

        for row in sheet.iter_rows(values_only=True):
            try:
                code = int(float(row[0]))
            except (TypeError, ValueError):
                continue
            if code == 3420:
                try:
                    amount = Decimal(str(row[4]))
                    if not amount.is_finite():
                        raise InvalidOperation
                    return float(abs(amount).quantize(Decimal("0.01")))
                except (IndexError, InvalidOperation, TypeError, ValueError):
                    raise ValueError(
                        "Subscription Revenue (BS code 3420) does not contain a valid amount"
                    ) from None
        return 0.0
    finally:
        workbook.close()


def _validate_payment(payment: dict) -> dict:
    quickbooks_member_exists = payment.get("quickbooks_member_exists", True) is not False
    if quickbooks_member_exists:
        raw_member_name = str(payment.get("member_name") or "")
        if any(delimiter in raw_member_name for delimiter in ("\t", "\r", "\n")):
            raise ValueError("Member name cannot contain tabs or line breaks")
        member_name = raw_member_name.strip()
        if not member_name:
            raise ValueError("Member name is required")
    else:
        member_name = ""

    payment_type = str(payment.get("payment_type") or "").strip()
    if payment_type not in PAYMENT_TYPES:
        raise ValueError("Payment type must be Paid in full, New plan, or Existing plan")

    raw_member_number = str(payment.get("member_number") or "")
    if any(delimiter in raw_member_number for delimiter in ("\t", "\r", "\n")):
        raise ValueError("Member number cannot contain tabs or line breaks")
    member_number = raw_member_number.strip()
    if member_number.startswith("#"):
        member_number = member_number[1:]
    pending_value = payment.get("member_number_pending", False)
    member_number_pending = (
        pending_value is True
        or str(pending_value).strip().lower() in {"true", "yes", "1"}
    )
    if payment_type == "Paid in full":
        member_number = ""
    elif member_number_pending:
        if member_number:
            raise ValueError(
                "Enter a member number or select Member # pending, not both"
            )
        member_number = "Pending"
    else:
        if not member_number:
            raise ValueError("Member number is required unless Member # pending is selected")
        if not member_number.isascii() or not member_number.isdigit():
            raise ValueError("Member number must contain digits only")

    try:
        amount = Decimal(str(payment.get("amount"))).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError("Amount must be a valid dollar amount") from None
    if not amount.is_finite() or amount <= 0:
        raise ValueError("Amount must be greater than zero")

    plan = str(payment.get("plan") or "").strip()
    if payment_type == "Paid in full":
        if amount != Decimal("100.00"):
            raise ValueError("Paid in full must be exactly $100.00")
        plan = ""
    elif plan not in PLAN_TERMS:
        raise ValueError("Plan must be 1 year, 3 year, or 5 year")

    if payment_type == "New plan" and amount < PLAN_DEPOSITS[plan]:
        raise ValueError(
            f"New {plan} plan payment must include the ${PLAN_DEPOSITS[plan]:.2f} deposit"
        )

    override = payment.get("interest_periods")
    if override is None:
        interest_periods = None
    else:
        try:
            override_value = Decimal(str(override))
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError("Interest periods must be a whole number") from None
        if not override_value.is_finite():
            interest_periods = None
        elif override_value != override_value.to_integral_value():
            raise ValueError("Interest periods must be a whole number")
        else:
            interest_periods = int(override_value)

    if interest_periods is not None and payment_type != "Paid in full":
        maximum = PLAN_MAX_PERIODS[plan]
        if not 0 <= interest_periods <= maximum:
            raise ValueError(f"Interest periods must be between 0 and {maximum}")

    return {
        "member_name": member_name,
        "member_number": member_number,
        "payment_type": payment_type,
        "plan": plan,
        "amount": amount,
        "interest_periods": interest_periods,
    }


def build_membership_lines(
    payments: list[dict],
    expected_subscription_total: float | None = None,
    handling_mode: str = "automatic",
) -> list[dict]:
    if handling_mode not in HANDLING_MODES:
        raise ValueError("Membership handling mode must be automatic or manual")

    if handling_mode == "manual":
        if expected_subscription_total is None:
            raise ValueError("Manual QuickBooks mode requires the Subscription Revenue total")
        manual_total = abs(Decimal(str(expected_subscription_total))).quantize(Decimal("0.01"))
        if manual_total == Decimal("0.00"):
            return []
        return [{
            "account": MEMBER_SHARES_EQUITY,
            "name": "",
            "memo": "Member Shares - Paid",
            "class_name": "",
            "amount": float(manual_total),
        }]

    validated_payments = [_validate_payment(payment) for payment in payments]
    if expected_subscription_total is not None:
        entered_total = sum(
            (payment["amount"] for payment in validated_payments), Decimal("0.00")
        ).quantize(Decimal("0.01"))
        expected_total = abs(Decimal(str(expected_subscription_total))).quantize(Decimal("0.01"))
        if entered_total != expected_total:
            if expected_total and not validated_payments:
                raise ValueError(
                    f"Subscription Revenue is ${expected_total:.2f}, but no membership "
                    "payments were supplied. Enter them in the app or run the script with "
                    "--membership-payments-file."
                )
            raise ValueError(
                f"Entered membership payments (${entered_total:.2f}) must equal "
                f"Subscription Revenue (${expected_total:.2f})"
            )

    lines = []
    for payment in validated_payments:
        amount = payment["amount"]
        if payment["payment_type"] == "Paid in full":
            lines.append(
                {
                    "account": MEMBER_SHARES_EQUITY,
                    "name": payment["member_name"],
                    "memo": "Member Shares - Paid",
                    "class_name": "",
                    "amount": float(amount.quantize(Decimal("0.01"))),
                }
            )
            continue
        installment, interest_per_period = PLAN_TERMS[payment["plan"]]
        installment_amount = amount
        if payment["payment_type"] == "New plan":
            installment_amount -= PLAN_DEPOSITS[payment["plan"]]
            lines.extend(
                [
                    {
                        "account": MEMBER_SHARES_EQUITY,
                        "name": payment["member_name"],
                        "memo": "Member Shares - Receivable",
                        "class_name": "",
                        "amount": 100.0,
                    },
                    {
                        "account": MEMBER_SHARES_RECEIVABLE,
                        "name": payment["member_name"],
                        "memo": "Member Shares - Receivable",
                        "class_name": "",
                        "amount": -100.0,
                    },
                ]
            )
        automatic_periods = min(
            int((installment_amount / installment).to_integral_value(rounding=ROUND_DOWN)),
            PLAN_MAX_PERIODS[payment["plan"]],
        )
        override = payment["interest_periods"]
        if override is None:
            periods = automatic_periods
        else:
            if override > automatic_periods:
                raise ValueError(
                    f"Interest periods cannot exceed the automatic count of {automatic_periods}"
                )
            periods = override
        interest = (interest_per_period * periods).quantize(Decimal("0.01"))
        principal = (amount - interest).quantize(Decimal("0.01"))
        member_number = payment["member_number"]
        memo = f"Share Installments - Paid #{member_number}"

        lines.append(
            {
                "account": MEMBER_SHARES_RECEIVABLE,
                "name": payment["member_name"],
                "memo": memo,
                "class_name": "",
                "amount": float(principal),
            }
        )
        if interest:
            lines.append(
                {
                    "account": INTEREST_INCOME,
                    "name": payment["member_name"],
                    "memo": memo,
                    "class_name": "Admin",
                    "amount": float(interest),
                }
            )
    return lines
