from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from io import BytesIO
import hashlib
import json
from pathlib import Path
from uuid import uuid4


STANDARD_CLOSEOUT_ORDER: tuple[str, ...] = (
    "cash",
    "checks",
    "donation",
    "charge_house",
    "offline_zon",
    "vendor_coupons",
    "paid_out",
    "paid_in",
)

_BS_CODE_TO_FIELD = {
    901: "cash",
    902: "checks",
    1122: "donation",
    906: "charge_house",
    934: "offline_zon",
    1334: "offline_zon",
    908: "vendor_coupons",
    1114: "paid_out",
}


STANDARD_METADATA = OrderedDict(
    [
        (
            "cash",
            {
                "label": "Cash",
                "memo": "Over/Short per Closeout Sheet - Cash",
                "detail_direction": None,
            },
        ),
        (
            "checks",
            {
                "label": "Checks",
                "memo": "Over/Short per Closeout Sheet - Check",
                "detail_direction": None,
            },
        ),
        (
            "donation",
            {
                "label": "Donation",
                "memo": "Over/Short per Closeout Sheet - Donation",
                "detail_direction": -1,
            },
        ),
        (
            "charge_house",
            {
                "label": "Charge (House)",
                "memo": "Over/Short per Closeout Sheet - Charge (House)",
                "detail_direction": -1,
            },
        ),
        (
            "offline_zon",
            {
                "label": "Offline Zon",
                "memo": "Over/Short per Closeout Sheet - Offline Zon",
                "detail_direction": -1,
            },
        ),
        (
            "vendor_coupons",
            {
                "label": "Vendor Coupons",
                "memo": "Over/Short per Closeout Sheet - Coupon",
                "detail_direction": -1,
                "managed_externally": True,
            },
        ),
        (
            "paid_out",
            {
                "label": "Paid Out",
                "memo": "Over/Short per Closeout Sheet - Paid Out",
                "detail_direction": -1,
            },
        ),
        (
            "paid_in",
            {
                "label": "Paid In",
                "memo": "Over/Short per Closeout Sheet - Paid In",
                "detail_direction": 1,
            },
        ),
    ]
)

_STANDARD_ADJUSTMENT_ACCOUNT = "8314000 · FE - Cash Over/Shorts"
_PAYROLL_VALUES = {Decimal("-4000.00"), Decimal("0.00"), Decimal("4000.00")}
_CENTS = Decimal("0.01")


def _money(value, label: str) -> Decimal:
    try:
        amount = Decimal(str(value)).quantize(_CENTS)
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{label} must be a valid monetary amount") from None
    if not amount.is_finite():
        raise ValueError(f"{label} must be a valid monetary amount")
    return amount


def _raw_decimal(value, label: str) -> Decimal:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{label} must be a valid monetary amount") from None
    if not amount.is_finite():
        raise ValueError(f"{label} must be a valid monetary amount")
    return amount


def _nonnegative_money(value, label: str) -> Decimal:
    if _raw_decimal(value, label) < 0:
        raise ValueError(f"{label} must be zero or greater")
    amount = _money(value, label)
    if amount < 0:
        raise ValueError(f"{label} must be zero or greater")
    return amount


def _adjustment(kind: str, account: str, memo: str, qb_effect) -> dict:
    effect = _money(qb_effect, memo)
    return {
        "kind": kind,
        "account": account,
        "memo": memo,
        "qb_effect": float(effect),
        "iif_amount": float(-effect),
    }


def build_standard_reconciliation(
    baselines: dict[str, float],
    actuals: dict[str, float],
) -> list[dict]:
    """Build the approved closeout reconciliation rows in canonical order."""
    normalized_baselines = {}
    normalized_actuals = {}
    for key in STANDARD_CLOSEOUT_ORDER:
        metadata = STANDARD_METADATA[key]
        label = metadata["label"]
        if key not in baselines:
            raise ValueError(f"Closeout {label} is missing")
        if key not in actuals:
            raise ValueError(f"Closeout {label} is missing")

        normalized_baselines[key] = abs(_money(baselines[key], f"{label} baseline"))
        normalized_actuals[key] = _nonnegative_money(
            actuals[key], f"{label} actual"
        )

    rows = []
    for key in STANDARD_CLOSEOUT_ORDER:
        metadata = STANDARD_METADATA[key]
        label = metadata["label"]
        memo = metadata["memo"]
        detail_direction = metadata.get("detail_direction")
        baseline = normalized_baselines[key]
        actual = normalized_actuals[key]
        difference = (actual - baseline).quantize(Decimal("0.01"))
        if detail_direction is None:
            detail_qb_effect = None
            adjustment_qb_effect = difference
        else:
            detail_qb_effect = (actual * detail_direction).quantize(Decimal("0.01"))
            adjustment_qb_effect = (
                (baseline * detail_direction) - (actual * detail_direction)
            ).quantize(Decimal("0.01"))

        rows.append(
            {
                "key": key,
                "label": label,
                "baseline": float(baseline),
                "actual": float(actual),
                "difference": float(difference),
                "detail_qb_effect": (
                    None if detail_qb_effect is None else float(detail_qb_effect)
                ),
                "adjustment_account": _STANDARD_ADJUSTMENT_ACCOUNT,
                "adjustment_memo": memo,
                "adjustment_qb_effect": float(adjustment_qb_effect),
                "managed_externally": bool(metadata.get("managed_externally")),
            }
        )
    return rows


def _code(value):
    try:
        return int(float(value))
    except (TypeError, ValueError, OverflowError):
        return None


def read_closeout_baselines(
    workbook_bytes: bytes,
    bs_sheet_name: str,
    hash_sheet_name: str,
) -> dict[str, float]:
    import openpyxl

    workbook = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=True,
        data_only=True,
    )
    try:
        if bs_sheet_name not in workbook.sheetnames:
            raise ValueError(f"Balance Sheet tab '{bs_sheet_name}' was not found")
        if hash_sheet_name not in workbook.sheetnames:
            raise ValueError(f"HASH tab '{hash_sheet_name}' was not found")

        baselines = {field: 0.0 for field in STANDARD_CLOSEOUT_ORDER}
        balance_sheet = workbook[bs_sheet_name]
        for row in balance_sheet.iter_rows(values_only=True):
            code = _code(row[0] if row else None)
            field = _BS_CODE_TO_FIELD.get(code)
            if field is None:
                continue
            try:
                amount = abs(_money(row[4], f"BS code {code} amount"))
            except IndexError:
                raise ValueError(
                    f"BS code {code} does not contain a valid monetary amount"
                ) from None
            baselines[field] = float(amount)

        hash_sheet = workbook[hash_sheet_name]
        amount_column = None
        for row in hash_sheet.iter_rows(
            min_row=1,
            max_row=min(hash_sheet.max_row, 20),
            values_only=True,
        ):
            for index, value in enumerate(row):
                label = str(value or "").strip().casefold()
                if label == "amount":
                    amount_column = index
                    break
            if amount_column is not None:
                break
        if amount_column is None:
            raise ValueError("HASH Amount header was not found in the first 20 rows")

        for row in hash_sheet.iter_rows(values_only=True):
            paid_in_row = False
            for value in row[:4]:
                if _code(value) == 34:
                    paid_in_row = True
                    break
            if not paid_in_row:
                continue
            try:
                raw_amount = row[amount_column]
            except IndexError:
                raise ValueError(
                    "HASH code 34 does not contain a valid monetary amount"
                ) from None
            if raw_amount is None or (
                isinstance(raw_amount, str) and not raw_amount.strip()
            ):
                continue
            amount = abs(_money(raw_amount, "HASH code 34 amount"))
            baselines["paid_in"] = float(amount)

        return baselines
    finally:
        workbook.close()


def default_closeout_actuals(
    baselines: dict[str, float],
    coupon_actual_total: float,
) -> dict[str, float]:
    counted_coupons = abs(_money(coupon_actual_total, "Vendor Coupons actual total"))
    actuals = {
        field: float(
            _money(baselines.get(field, 0.0), f"{field} baseline")
        )
        for field in STANDARD_CLOSEOUT_ORDER
    }
    actuals["offline_zon"] = 0.0
    actuals["vendor_coupons"] = float(counted_coupons)
    return actuals


def coupon_workflow_is_required(
    coupon_bs_total: float,
    closeout_choice: str | None,
) -> bool:
    """Show coupon counting when BS has coupons or Closeout is handled in-app."""
    baseline = abs(_money(coupon_bs_total, "BS Coupons Receivable total"))
    return bool(
        baseline
        or closeout_choice == "Breakdown in app using Closeout Sheet"
    )


def _normalize_misc_values(payload: dict) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("Closeout payload must be an object")

    payroll = _money(payload.get("payroll", 0), "Payroll")
    if payroll not in _PAYROLL_VALUES:
        raise ValueError("Payroll must be -4000.00, 0.00, or 4000.00")

    safe = payload.get("safe", {"type": "none", "amount": 0})
    if not isinstance(safe, dict):
        raise ValueError("Safe adjustment must be an object")
    safe_type = safe.get("type", "none")
    if safe_type not in {"none", "overage", "shortage"}:
        raise ValueError("Safe type must be none, overage, or shortage")
    safe_amount = _nonnegative_money(safe.get("amount", 0), "Safe amount")
    if safe_type == "none":
        if safe_amount != 0:
            raise ValueError("Safe amount must be zero when type is none")
    elif safe_amount <= 0:
        raise ValueError("Safe amount must be greater than zero")

    plants_purchase = _nonnegative_money(
        payload.get("plants_purchase", 0), "Plants purchase"
    )

    custom_tba = payload.get("custom_tba", [])
    if not isinstance(custom_tba, list):
        raise ValueError("Custom TBA must be a list")
    normalized_custom_tba = []
    for row in custom_tba:
        if not isinstance(row, dict):
            raise ValueError("Custom TBA row must be an object")
        memo = row.get("memo")
        if not isinstance(memo, str) or not memo.strip():
            raise ValueError("Custom TBA memo is required")
        if any(delimiter in memo for delimiter in ("\t", "\r", "\n")):
            raise ValueError("Custom TBA memo cannot contain tabs or line breaks")
        amount = _nonnegative_money(row.get("amount"), "Custom TBA amount")
        if amount <= 0:
            raise ValueError("Custom TBA amount must be greater than zero")
        direction = row.get("direction")
        if direction not in {"adds", "removes"}:
            raise ValueError("Custom TBA direction must be adds or removes")
        normalized_custom_tba.append(
            {"memo": memo.strip(), "amount": float(amount), "direction": direction}
        )

    return {
        "payroll": float(payroll),
        "safe": {"type": safe_type, "amount": float(safe_amount)},
        "plants_purchase": float(plants_purchase),
        "custom_tba": normalized_custom_tba,
    }


def build_misc_adjustments(payload: dict) -> list[dict]:
    """Build the approved miscellaneous adjustments in their input order."""
    normalized = _normalize_misc_values(payload)
    rows = []
    payroll = normalized["payroll"]
    if payroll:
        rows.append(
            _adjustment(
                "payroll",
                "1140000 · Cash Drawers/Safe",
                "Payroll - Check Cashing",
                payroll,
            )
        )

    safe = normalized["safe"]
    if safe["type"] == "overage":
        rows.append(
            _adjustment(
                "safe_overage",
                _STANDARD_ADJUSTMENT_ACCOUNT,
                "Safe Overage Cash added to deposit",
                safe["amount"],
            )
        )
    elif safe["type"] == "shortage":
        rows.append(
            _adjustment(
                "safe_shortage",
                _STANDARD_ADJUSTMENT_ACCOUNT,
                "Safe Shortage Cash Taken from Deposit",
                -safe["amount"],
            )
        )

    if normalized["plants_purchase"]:
        rows.append(
            _adjustment(
                "plants_purchase",
                "1130000 · Petty Cash",
                "Plants Dept - Market Purchases",
                -normalized["plants_purchase"],
            )
        )

    for custom in normalized["custom_tba"]:
        effect = custom["amount"] if custom["direction"] == "adds" else -custom["amount"]
        rows.append(
            _adjustment("custom_tba", "4444 · TBA Purchases", custom["memo"], effect)
        )
    return rows


def calculate_final_pos_adjustment(
    provisional_total: float,
    final_total: float,
    approved: bool,
) -> dict:
    provisional = _money(provisional_total, "Generated deposit")
    final = _money(final_total, "Final Closeout Sheet Deposit Total")
    if final <= 0:
        raise ValueError("Final Closeout Sheet Deposit Total must be greater than zero")
    if type(approved) is not bool:
        raise ValueError("approved must be a boolean")
    remaining = (final - provisional).quantize(_CENTS)
    line = None
    if remaining and approved:
        line = _adjustment(
            "final_pos",
            _STANDARD_ADJUSTMENT_ACCOUNT,
            "Over/Short per POS (to = POS total)",
            remaining,
        )
    return {
        "provisional_total": float(provisional),
        "final_total": float(final),
        "remaining": float(remaining),
        "line": line,
        "requires_approval": bool(remaining and not approved),
    }


def normalize_closeout_payload(payload: dict) -> dict:
    """Return a fresh, canonical Closeout payload suitable for persistence."""
    if not isinstance(payload, dict):
        raise ValueError("Closeout payload must be an object")
    mode = payload.get("mode")
    if mode == "manual":
        return {"mode": "manual"}
    if mode != "closeout":
        raise ValueError("Closeout mode must be manual or closeout")
    if payload.get("reviewed") is not True:
        raise ValueError("Paper Closeout Sheet review is required")

    actuals = payload.get("actuals")
    if not isinstance(actuals, dict):
        raise ValueError("Closeout actuals must be an object")
    normalized_actuals = {}
    for key in STANDARD_CLOSEOUT_ORDER:
        if key not in actuals:
            raise ValueError(f"Closeout {STANDARD_METADATA[key]['label']} is missing")
        normalized_actuals[key] = float(
            _nonnegative_money(actuals[key], f"{STANDARD_METADATA[key]['label']} actual")
        )

    normalized_misc = _normalize_misc_values(payload)
    final_total = _money(payload.get("final_total"), "Final Closeout Sheet Deposit Total")
    if final_total <= 0:
        raise ValueError("Final Closeout Sheet Deposit Total must be greater than zero")
    approve_final_pos = payload.get("approve_final_pos")
    if type(approve_final_pos) is not bool:
        raise ValueError("approve_final_pos must be a boolean")

    return {
        "mode": "closeout",
        "reviewed": True,
        "actuals": normalized_actuals,
        "payroll": normalized_misc["payroll"],
        "safe": normalized_misc["safe"],
        "plants_purchase": normalized_misc["plants_purchase"],
        "custom_tba": normalized_misc["custom_tba"],
        "final_total": float(final_total),
        "approve_final_pos": approve_final_pos,
    }


def build_closeout_form_payload(
    *,
    baselines: dict[str, float],
    actuals: dict[str, float],
    reviewed: bool,
    payroll: float,
    safe_type: str,
    safe_amount: float,
    plants_purchase: float,
    custom_tba: list[dict],
    final_total: float,
    approve_final_pos: bool,
) -> dict:
    """Build and validate the canonical payload produced by the employee form."""
    if reviewed is not True:
        raise ValueError(
            "Confirm that you reviewed the amounts against the paper Closeout Sheet"
        )
    reconciliation = build_standard_reconciliation(baselines, actuals)
    ordered_actuals = {
        row["key"]: row["actual"]
        for row in reconciliation
    }
    return normalize_closeout_payload(
        {
            "mode": "closeout",
            "reviewed": True,
            "actuals": ordered_actuals,
            "payroll": payroll,
            "safe": {"type": safe_type, "amount": safe_amount},
            "plants_purchase": plants_purchase,
            "custom_tba": custom_tba,
            "final_total": final_total,
            "approve_final_pos": approve_final_pos,
        }
    )


def closeout_input_fingerprint(
    payload: dict,
    review_context: dict | None = None,
) -> str:
    """Identify the reviewed financial inputs, excluding final POS approval."""
    normalized = normalize_closeout_payload(payload)
    financial_inputs = dict(normalized)
    financial_inputs.pop("approve_final_pos", None)
    fingerprint_inputs = {
        "closeout": financial_inputs,
        "review_context": review_context or {},
    }
    serialized = json.dumps(
        fingerprint_inputs,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()


def closeout_preview_is_fresh(
    payload: dict,
    saved_preview: dict | None,
    review_context: dict | None = None,
) -> bool:
    """Return whether a saved preview was produced from the current inputs."""
    if not isinstance(saved_preview, dict) or not isinstance(
        saved_preview.get("preview"), dict
    ):
        return False
    return saved_preview.get("input_fingerprint") == closeout_input_fingerprint(
        payload,
        review_context=review_context,
    )


def write_closeout_payload_file(folder: Path, payload: dict) -> Path:
    normalized_payload = normalize_closeout_payload(payload)
    target_folder = Path(folder)
    target_folder.mkdir(parents=True, exist_ok=True)
    path = target_folder / f"closeout_{uuid4()}.json"
    path.write_text(
        json.dumps(normalized_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    try:
        path.chmod(0o600)
    except OSError:
        pass
    return path


def load_closeout_payload_file(path: str | Path) -> dict:
    try:
        raw_payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        raise ValueError("Closeout payload file contains malformed JSON") from None
    except OSError:
        raise ValueError("Closeout payload file could not be read") from None
    try:
        return normalize_closeout_payload(raw_payload)
    except ValueError as error:
        raise ValueError(f"Closeout payload is invalid: {error}") from None
