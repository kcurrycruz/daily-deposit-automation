from decimal import Decimal, InvalidOperation


COUPON_CATEGORIES = ("NCG", "MFG", "VP", "MKTG", "SITKA")
NCG_DENOMINATIONS = (0.25, 0.50, 0.75, 1.00, 1.25, 1.50, 2.00, 3.00, 4.00, 5.00)


def read_coupon_receivable_total(
    workbook_bytes: bytes,
    bs_sheet_name: str | None = None,
) -> float:
    from io import BytesIO

    import openpyxl

    workbook = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=True,
        data_only=True,
    )
    try:
        if bs_sheet_name:
            if bs_sheet_name not in workbook.sheetnames:
                raise ValueError(f"Balance Sheet tab '{bs_sheet_name}' was not found")
            sheet = workbook[bs_sheet_name]
        else:
            sheet_name = next(
                (
                    name for name in workbook.sheetnames
                    if name.casefold().endswith(" bs")
                    and "xxxxxx" not in name.casefold()
                ),
                None,
            )
            if sheet_name is None:
                raise ValueError("Balance Sheet tab was not found")
            sheet = workbook[sheet_name]

        for row in sheet.iter_rows(values_only=True):
            try:
                code = int(float(row[0]))
            except (TypeError, ValueError):
                continue
            if code != 908:
                continue
            try:
                amount = Decimal(str(row[4]))
                if not amount.is_finite():
                    raise InvalidOperation
                return float(abs(amount).quantize(Decimal("0.01")))
            except (IndexError, InvalidOperation, TypeError, ValueError):
                raise ValueError(
                    "Coupons Receivable (BS code 908) does not contain a valid amount"
                ) from None
        return 0.0
    finally:
        workbook.close()


def _money(value, label: str) -> Decimal:
    try:
        amount = Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{label} must be a valid dollar amount") from None
    if not amount.is_finite() or amount < 0:
        raise ValueError(f"{label} must be zero or greater")
    return amount


def summarize_coupon_stacks(stacks: list[dict]) -> dict:
    category_totals = {
        category: Decimal("0.00") for category in COUPON_CATEGORIES
    }
    summarized_stacks = []

    for stack in stacks:
        category = str(stack.get("category", "")).strip().upper()
        if category not in category_totals:
            raise ValueError(f"Unknown coupon category: {category or '(blank)'}")

        subtotal = Decimal("0.00")
        normalized_amounts = []
        for value in stack.get("amounts", []):
            amount = _money(value, f"{category} coupon amount")
            if amount <= 0:
                raise ValueError(f"{category} coupon amounts must be greater than zero")
            subtotal += amount
            normalized_amounts.append(float(amount))
        subtotal = subtotal.quantize(Decimal("0.01"))
        category_totals[category] += subtotal

        raw_expected = stack.get("expected_total")
        expected = (
            None
            if raw_expected in (None, "")
            else _money(raw_expected, f"{category} written stack total")
        )
        difference = None if expected is None else (subtotal - expected).quantize(Decimal("0.01"))
        summarized_stacks.append({
            "id": stack.get("id"),
            "category": category,
            "label": str(stack.get("label", "")).strip(),
            "expected_total": None if expected is None else float(expected),
            "amounts": normalized_amounts,
            "subtotal": float(subtotal),
            "difference": None if difference is None else float(difference),
            "matches_expected": None if difference is None else difference == 0,
        })

    category_totals = {
        category: total.quantize(Decimal("0.01"))
        for category, total in category_totals.items()
    }
    ncg_total = category_totals["NCG"]
    mfg_total = sum(
        (category_totals[category] for category in ("MFG", "VP", "MKTG", "SITKA")),
        Decimal("0.00"),
    ).quantize(Decimal("0.01"))
    overall_total = (ncg_total + mfg_total).quantize(Decimal("0.01"))

    return {
        "category_totals": {
            category: float(total) for category, total in category_totals.items()
        },
        "ncg_total": float(ncg_total),
        "mfg_total": float(mfg_total),
        "overall_total": float(overall_total),
        "stacks": summarized_stacks,
    }


def add_coupon_amount(stacks: list[dict], stack_id: str, amount: float) -> list[dict]:
    normalized_amount = _money(amount, "Coupon amount")
    if normalized_amount <= 0:
        raise ValueError("Coupon amount must be greater than zero")

    updated = []
    found = False
    for stack in stacks:
        copied = dict(stack)
        copied["amounts"] = list(stack.get("amounts", []))
        if stack.get("id") == stack_id:
            copied["amounts"].append(float(normalized_amount))
            found = True
        updated.append(copied)
    if not found:
        raise ValueError("Coupon stack was not found")
    return updated


def remove_coupon_amount(
    stacks: list[dict],
    stack_id: str,
    position: int,
) -> list[dict]:
    updated = []
    found = False
    for stack in stacks:
        copied = dict(stack)
        amounts = list(stack.get("amounts", []))
        if stack.get("id") == stack_id:
            if position < 0 or position >= len(amounts):
                raise IndexError("Coupon position is out of range")
            amounts.pop(position)
            found = True
        copied["amounts"] = amounts
        updated.append(copied)
    if not found:
        raise ValueError("Coupon stack was not found")
    return updated


def reconcile_coupon_receivable(
    bs_total: float,
    *,
    mode: str,
    closeout_actual_total: float | None = None,
    ncg_total: float | None = None,
    mfg_total: float | None = None,
) -> dict:
    bs_amount = abs(_money(bs_total, "BS Coupons Receivable total"))
    if mode == "quickbooks":
        return {
            "bs_total": float(bs_amount),
            "closeout_actual_total": None,
            "ncg_total": float(bs_amount),
            "mfg_total": None,
            "difference": None,
        }
    if mode != "closeout":
        raise ValueError("Coupon handling mode must be quickbooks or closeout")

    closeout_amount = _money(
        closeout_actual_total,
        "Closeout Sheet Coupon Actual Total",
    )
    ncg_amount = _money(ncg_total, "NCG Coupons")
    mfg_amount = _money(mfg_total, "MFG Coupons")
    counted_total = (ncg_amount + mfg_amount).quantize(Decimal("0.01"))
    if counted_total != closeout_amount:
        raise ValueError(
            f"NCG Coupons (${ncg_amount:.2f}) + MFG Coupons (${mfg_amount:.2f}) "
            "must equal Closeout Sheet Coupon Actual Total "
            f"(${closeout_amount:.2f})"
        )

    difference = (closeout_amount - bs_amount).quantize(Decimal("0.01"))
    return {
        "bs_total": float(bs_amount),
        "closeout_actual_total": float(closeout_amount),
        "ncg_total": float(ncg_amount),
        "mfg_total": float(mfg_amount),
        "difference": float(difference),
    }
