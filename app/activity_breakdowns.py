from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import json
from pathlib import Path
from uuid import uuid4


ACTIVITY_ORDER = ("donation", "paid_out", "paid_in")
_MODES = {"app", "quickbooks"}
_CENTS = Decimal("0.01")


def activity_workflow_keys(source_totals: dict) -> tuple[str, ...]:
    if not isinstance(source_totals, dict):
        raise ValueError("Activity source totals must be an object")
    active = []
    for category in ACTIVITY_ORDER:
        try:
            amount = Decimal(str(source_totals.get(category, 0)))
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError(f"{category} source total must be a valid amount") from None
        if not amount.is_finite():
            raise ValueError(f"{category} source total must be a valid amount")
        if amount != 0:
            active.append(category)
    return tuple(active)


def activity_closeout_ready(payload: dict, source_totals: dict) -> bool:
    if not isinstance(payload, dict):
        return False
    for category in activity_workflow_keys(source_totals):
        section = payload.get(category)
        if not isinstance(section, dict) or section.get("mode") != "app":
            return False
    return True


def _source_code(value):
    try:
        return int(float(value))
    except (TypeError, ValueError, OverflowError):
        return None


def _source_number(value) -> float | None:
    try:
        amount = Decimal(str(value)).quantize(_CENTS)
    except (InvalidOperation, TypeError, ValueError):
        return None
    if not amount.is_finite():
        return None
    return float(amount)


def _source_amount(value) -> float | None:
    amount = _source_number(value)
    return None if amount is None else abs(amount)


def extract_hash_row_amount(
    row: tuple,
    amount_column: int,
    code: int,
) -> float | None:
    """Use the HASH Amount cell, then the established shifted-column fallback."""
    amount = _source_number(
        row[amount_column] if len(row) > amount_column else None
    )
    if amount is not None:
        return amount

    fallback_values = []
    for index, value in enumerate(row):
        if index < 3 or value is None:
            continue
        candidate = _source_number(value)
        if candidate is None or abs(candidate - code) < 0.000001:
            continue
        fallback_values.append(candidate)
    if not fallback_values:
        return None
    return fallback_values[1] if len(fallback_values) >= 2 else fallback_values[0]


def read_activity_source_totals(
    workbook_bytes: bytes,
    bs_sheet_name: str | None,
    hash_sheet_name: str | None,
) -> dict[str, float]:
    """Read only the three optional workflow totals without strict Closeout parsing.

    Missing sheets, unrelated malformed rows, and a HASH export without an exact
    Amount header leave the affected optional total at zero. The full Closeout
    workflow continues to use its stricter parser after an employee selects it.
    """
    import openpyxl

    totals = {category: 0.0 for category in ACTIVITY_ORDER}
    workbook = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=True,
        data_only=True,
    )
    try:
        if bs_sheet_name in workbook.sheetnames:
            balance_sheet = workbook[bs_sheet_name]
            fields = {1122: "donation", 1114: "paid_out"}
            for row in balance_sheet.iter_rows(values_only=True):
                code = _source_code(row[0] if row else None)
                category = fields.get(code)
                if category is None:
                    continue
                amount = _source_amount(row[4] if len(row) > 4 else None)
                if amount is not None:
                    totals[category] = amount

        if hash_sheet_name in workbook.sheetnames:
            hash_sheet = workbook[hash_sheet_name]
            amount_column = None
            for row in hash_sheet.iter_rows(
                min_row=1,
                max_row=min(hash_sheet.max_row, 20),
                values_only=True,
            ):
                for index, value in enumerate(row):
                    if str(value or "").strip().casefold() == "amount":
                        amount_column = index
                        break
                if amount_column is not None:
                    break
            if amount_column is not None:
                for row in hash_sheet.iter_rows(values_only=True):
                    if not any(_source_code(value) == 34 for value in row[:4]):
                        continue
                    amount = extract_hash_row_amount(row, amount_column, 34)
                    if amount is not None:
                        totals["paid_in"] = abs(amount)
        return totals
    finally:
        workbook.close()


def _money(value, label: str) -> Decimal:
    try:
        amount = Decimal(str(value)).quantize(_CENTS)
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{label} must be a valid monetary amount") from None
    if not amount.is_finite():
        raise ValueError(f"{label} must be a valid monetary amount")
    if amount <= 0:
        raise ValueError(f"{label} must be greater than zero")
    return amount


def _text(value, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{label} is required")
    cleaned = value.strip()
    if any(delimiter in cleaned for delimiter in ("\t", "\r", "\n")):
        raise ValueError(f"{label} cannot contain tabs or line breaks")
    return cleaned


def _date(value, label: str) -> str:
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, str) and value.strip():
        parsed = None
        for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                parsed = datetime.strptime(value.strip(), pattern).date()
                break
            except ValueError:
                continue
        if parsed is None:
            raise ValueError(f"{label} must be a valid date")
    else:
        raise ValueError(f"{label} is required")
    return parsed.isoformat()


def _normalize_donation_row(row: dict) -> dict:
    if not isinstance(row, dict):
        raise ValueError("Donation row must be an object")
    return {
        "given_to": _text(row.get("given_to"), "Given To"),
        "purpose": _text(row.get("purpose"), "For"),
        "manager": _text(row.get("manager"), "Manager Approval"),
        "amount": float(_money(row.get("amount"), "Donation amount")),
    }


def _normalize_paid_row(category: str, row: dict) -> dict:
    label = "Paid In" if category == "paid_in" else "Paid Out"
    if not isinstance(row, dict):
        raise ValueError(f"{label} row must be an object")
    row_type = row.get("type")
    if row_type not in {"esp", "other"}:
        raise ValueError(f"{label} type must be ESP Deposit or Other")
    normalized = {
        "type": row_type,
        "amount": float(_money(row.get("amount"), f"{label} amount")),
    }
    if row_type == "esp":
        normalized.update(
            {
                "original_date": _date(
                    row.get("original_date"), "Original ESP deposit date"
                ),
                "initials": _text(row.get("initials"), "Initials"),
            }
        )
    else:
        normalized["memo"] = _text(row.get("memo"), f"{label} memo")
    return normalized


def normalize_activity_section(category: str, section: dict) -> dict:
    if category not in ACTIVITY_ORDER:
        raise ValueError(f"Unknown activity breakdown: {category}")
    if not isinstance(section, dict):
        raise ValueError(f"{category} breakdown must be an object")
    mode = section.get("mode", "quickbooks")
    if mode not in _MODES:
        raise ValueError(f"{category} mode must be app or quickbooks")
    if mode == "quickbooks":
        return {"mode": mode, "rows": []}
    rows = section.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ValueError(f"{category} breakdown requires at least one row")
    row_builder = (
        _normalize_donation_row
        if category == "donation"
        else lambda row: _normalize_paid_row(category, row)
    )
    return {"mode": mode, "rows": [row_builder(row) for row in rows]}


def normalize_activity_payload(payload: dict | None) -> dict:
    if payload is None:
        payload = {}
    if not isinstance(payload, dict):
        raise ValueError("Activity breakdown payload must be an object")
    unknown = set(payload) - set(ACTIVITY_ORDER)
    if unknown:
        raise ValueError(f"Unknown activity breakdown: {sorted(unknown)[0]}")
    return {
        category: normalize_activity_section(
            category,
            payload.get(category, {"mode": "quickbooks", "rows": []}),
        )
        for category in ACTIVITY_ORDER
    }


def activity_actuals(payload: dict) -> dict[str, float | None]:
    normalized = normalize_activity_payload(payload)
    return {
        category: (
            round(sum(row["amount"] for row in section["rows"]), 2)
            if section["mode"] == "app"
            else None
        )
        for category, section in normalized.items()
    }


def _paid_memo(category: str, row: dict) -> str:
    prefix = "PAID IN" if category == "paid_in" else "PAID OUT"
    if row["type"] == "other":
        return f"{prefix}: {row['memo']}"
    original_date = date.fromisoformat(row["original_date"])
    short_date = f"{original_date.month}/{original_date.day}"
    return f"{prefix}: {short_date}'s ESP Deposit - {row['initials']}"


def build_activity_lines(payload: dict) -> dict[str, list[dict]]:
    normalized = normalize_activity_payload(payload)
    lines = {category: [] for category in ACTIVITY_ORDER}
    for row in normalized["donation"]["rows"]:
        lines["donation"].append(
            {
                "account": "8506000 · Outreach - Donations",
                "memo": (
                    f"Given to {row['given_to']} for {row['purpose']} - "
                    f"{row['manager']}"
                ),
                "class_name": "Marketing",
                "qb_effect": -row["amount"],
            }
        )
    for category in ("paid_out", "paid_in"):
        direction = 1 if category == "paid_in" else -1
        for row in normalized[category]["rows"]:
            lines[category].append(
                {
                    "account": (
                        "1230000 · Miscellaneous Receivable"
                        if row["type"] == "esp"
                        else "4444 · TBA Purchases"
                    ),
                    "memo": _paid_memo(category, row),
                    "class_name": "",
                    "qb_effect": round(direction * row["amount"], 2),
                }
            )
    return lines


def write_activity_payload_file(folder: Path, payload: dict) -> Path:
    normalized = normalize_activity_payload(payload)
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"activity_breakdowns_{uuid4().hex}.json"
    path.write_text(json.dumps(normalized, indent=2), encoding="utf-8")
    return path


def load_activity_payload_file(path: str | Path) -> dict:
    try:
        raw = json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Could not read activity breakdown file: {exc}") from None
    return normalize_activity_payload(raw)
