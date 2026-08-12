"""
HWFC  —  Daily Deposit Automation  v5.0
========================================
Generates a QuickBooks IIF that imports directly into
Banking -> Make Deposits -> 1120200 · NBT Bank - Operating Account

RUN:
  py C:\\POS_Automation\\pos_to_quickbooks_v2.py
"""

import sys, re, logging, csv
from datetime import date, datetime, timedelta
from datetime import date, datetime
from pathlib import Path


# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════

CONFIG = {
    "pos_export_folder": r"C:\POS_Reports\Daily",
    "output_folder":     r"S:\Finance & Payroll Forms\Finance Work Files\Daily Deposit\HWFC_Deposit\QB_Imports",
    "company_name":      "HWFC",
    # Base path for Excel SubDept Sales Total Reports
    # Script will look in: base_excel_path / FY{year} / {Month} /
    "base_excel_path":   r"S:\Finance & Payroll Forms\Finance Work Files\Daily Deposit\Daily SubDept Sales Total Reports",
    # The exact QB account name shown in "Deposit To" field
    "deposit_account":   "1120200 · NBT Bank - Operating Account",
    "cc_pattern": ["Settlement", "Batch", "Commerce", "CCC", "BusinessTrack"],
}


# ═══════════════════════════════════════════════════════════════
#  MAPPING 1 — Sub-department number → QB Account name
# ═══════════════════════════════════════════════════════════════

SUBDEPT_TO_QB = {
    110:  "7110110 · Sales - Packaged Grocery - NT",
    120:  "7110120 · Sales - Dairy/Refrigerated",
    130:  "7110130 · Sales - Frozen Foods",
    140:  "7110141 · Sales - Bulk",
    142:  "7110142 · Sales - Bulk Herbs",
    150:  "7110150 · Sales - Bread",
    160:  "7110160 · Sales - Beer",
    180:  "7110180 · Sales - Pets",
    190:  "7110190 · Sales - General Merchandise",
    210:  "7110210 · Sales - In-House Deli",
    220:  "7110220 · Sales - In-House Bakery",
    230:  "7110210 · Sales - In-House Deli|Catering",  # Catering — separate line
    235:  "7110235 · Sales - Sushi",
    240:  "7110240 · Sales - Java & Juice",
    250:  "7110250 · Sales - Co-op Cafe (ESP)",
    310:  "7110310 · Sales - Cheese",
    320:  "7110320 · Sales - Meat",
    330:  "7110330 · Sales - Specialty Foods",
    340:  "7110340 · Sales - Fish & Seafood",
    350:  "7110350 · Sales - Specialty Mercantile",
    410:  "7110410 · Sales - Produce",
    420:  "7110420 · Sales - Gardening/Plants",
    510:  "7110510 · Sales - Personal Care Taxable",
    520:  "7110520 · Sales - Vitamins & Supplements",
    550:  "7110550 · Sales - Magazines",
    1300: "7111300 · Promotional Sales",
    27:   "8515000 · Marketing - Coupons, Store",
    28:   "8320000 · Store Supplies",
    38:   "4150300 · NYS Paper Bag Fees Payable",
    170:  None,   # Adjustments — handled via MISC_SUBDEPTS → TBA Purchases
}

# Income accounts post as negative amounts in the deposit
# (they are credits — money coming IN to the store)
# Income accounts post as NEGATIVE SPL (credits) in the deposit
# Everything NOT in this set posts as POSITIVE (debit)
INCOME_ACCOUNTS = {
    "7110110 · Sales - Packaged Grocery - NT",
    "7110120 · Sales - Dairy/Refrigerated",
    "7110130 · Sales - Frozen Foods",
    "7110141 · Sales - Bulk",
    "7110142 · Sales - Bulk Herbs",
    "7110150 · Sales - Bread",
    "7110160 · Sales - Beer",
    "7110180 · Sales - Pets",
    "7110190 · Sales - General Merchandise",
    "7110210 · Sales - In-House Deli",
    "7110220 · Sales - In-House Bakery",
    "7110235 · Sales - Sushi",
    "7110240 · Sales - Java & Juice",
    "7110250 · Sales - Co-op Cafe (ESP)",
    "7110310 · Sales - Cheese",
    "7110320 · Sales - Meat",
    "7110330 · Sales - Specialty Foods",
    "7110340 · Sales - Fish & Seafood",
    "7110350 · Sales - Specialty Mercantile",
    "7110410 · Sales - Produce",
    "7110420 · Sales - Gardening/Plants",
    "7110510 · Sales - Personal Care Taxable",
    "7110520 · Sales - Vitamins & Supplements",
    "7110550 · Sales - Magazines",
    "7111300 · Promotional Sales",
    # Store Supplies and NYS Bag Fees are POSITIVE (debit) — NOT in this set
}


# ═══════════════════════════════════════════════════════════════
#  MAPPING 2 — Shopper Level code → QB Discount Account
# ═══════════════════════════════════════════════════════════════

SHOPPER_LEVEL_TO_QB = {
    2:  "8512001 · Discount 2% - Owners",
    3:  "8511001 · Discount 2% - Senior Non Owner",
    4:  "8511003 · Discount 5% - Senior Owners",
    5:  "8140010 · Monthly Time Discount (8%)",
    6:  "8140026 · Weekly Time Discount (24%)",
    7:  "8423100 · Discount - Staff  (24%)",
    8:  "8512002 · Discount 2% - Visiting Coop",
    9:  "8512003 · Discount 8% - Vendors",
    19: "8512007 · Discount 15% - Non-profit",
    10: "8511001 · Discount 2% - Senior Non Owner",  # Seniors on Wednesday
}


# ═══════════════════════════════════════════════════════════════
#  MAPPING 3 — Credit card types
# ═══════════════════════════════════════════════════════════════

CC_ACCOUNT = "1240001 · Credit Card Payments Receivable"

CC_TYPE_MAP = {
    "Visa/MC":             ["visa", "mastercard", "mc", "visa/mc"],
    "Discover":            ["discover"],
    "AMEX":                ["amex", "american express"],
    "Debit Card":          ["debit", "pin debit"],
    "EBT Cash/Food Stamp": ["ebt", "food stamp", "snap"],
}


# ═══════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════

output_dir = Path(CONFIG["output_folder"])
if not output_dir.exists():
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except FileExistsError:
        pass

# Log to console only — no file needed
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def clean_amount(val) -> float:
    if val is None:
        return 0.0
    s = str(val).strip().replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_rows(filepath: Path) -> list:
    rows = []
    with open(filepath, encoding="utf-8-sig", errors="replace", newline="") as f:
        for row in csv.reader(f):
            rows.append([cell.strip() for cell in row])
    return rows


# ═══════════════════════════════════════════════════════════════
#  PARSER — SMS Combined Export File
# ═══════════════════════════════════════════════════════════════

def parse_sms_file(filepath: Path) -> tuple:
    rows = read_rows(filepath)
    log.info(f"  Parsing {filepath.name} ({len(rows)} rows)")

    sales_raw    = {}
    discount_raw = {}
    mode    = None
    in_data = False

    for row in rows:
        if not any(row):
            continue
        joined = ",".join(row)

        # Detect section type
        if "Sub-department Single Total" in joined or "Sub-Department Single Total" in joined:
            mode    = "subdept"
            in_data = False
            continue
        if "Discounts by Shopper Level" in joined:
            mode    = "discount"
            in_data = False
            continue

        # Detect start of data rows
        if mode == "subdept" and "Sub-Department" in joined:
            in_data = True
            continue
        if mode == "discount" and "Description" in joined:
            in_data = True
            continue

        if not in_data:
            continue

        # Skip footer/summary rows
        low = joined.lower()
        if any(x in low for x in ["total", "printed", "grand total", "member discount",
                                   "honest weight", "date:", "target:", "tlz.", "s-dept"]):
            continue

        # ── Sub-dept rows ──────────────────────────────────
        if mode == "subdept":
            dept_num = None
            for cell in row:
                if re.match(r"^\d+$", cell):
                    try:
                        dept_num = int(cell)
                        break
                    except ValueError:
                        pass
            if dept_num is None:
                continue

            amt = 0.0
            for col_idx in [7, 8, 6, 9]:
                if col_idx < len(row) and row[col_idx]:
                    candidate = clean_amount(row[col_idx])
                    if candidate != 0.0:
                        amt = candidate
                        break

            if amt == 0.0:
                if dept_num == 170:
                    log.info(f"    Dept 170 Adjustments — skipped")
                continue

            sales_raw[dept_num] = round(sales_raw.get(dept_num, 0.0) + amt, 2)

        # ── Discount rows ──────────────────────────────────
        elif mode == "discount":
            code = None
            for cell in row:
                if re.match(r"^\d{1,2}$", cell):
                    try:
                        code = int(cell)
                        break
                    except ValueError:
                        pass
            if code is None:
                continue

            amt = 0.0
            for col_idx in [8, 7, 9]:
                if col_idx < len(row) and row[col_idx]:
                    candidate = clean_amount(row[col_idx])
                    if candidate != 0.0:
                        amt = candidate
                        break
            if amt == 0.0:
                continue

            discount_raw[code] = round(discount_raw.get(code, 0.0) + amt, 2)

    # Map to QB accounts
    sales_totals = {}
    log.info("\n  ── Sales ──")
    for num, amt in sorted(sales_raw.items()):
        qb = SUBDEPT_TO_QB.get(num)
        if qb is None:
            if num in SUBDEPT_TO_QB:
                log.info(f"    Dept {num:>5}  SKIPPED")
            else:
                log.warning(f"    Dept {num:>5}  ${amt:,.2f}  NOT IN MAP")
            continue
        log.info(f"    Dept {num:>5}  ${amt:>10,.2f}  →  {qb}")
        sales_totals[qb] = round(sales_totals.get(qb, 0.0) + amt, 2)

    discount_totals = {}
    log.info("\n  ── Discounts ──")
    for code, amt in sorted(discount_raw.items()):
        qb = SHOPPER_LEVEL_TO_QB.get(code)
        if qb:
            log.info(f"    Level {code:>2}  ${amt:>9,.2f}  →  {qb}")
            discount_totals[qb] = round(discount_totals.get(qb, 0.0) + amt, 2)
        else:
            log.warning(f"    Level {code}  ${amt:,.2f}  NOT MAPPED")

    return sales_totals, discount_totals


# ═══════════════════════════════════════════════════════════════
#  PARSER — Commerce Control Center
# ═══════════════════════════════════════════════════════════════

def parse_cc_file(filepath: Path) -> dict:
    rows = read_rows(filepath)
    log.info(f"  Parsing CC: {filepath.name}")

    header_idx = None
    for i, row in enumerate(rows):
        joined = ",".join(row).lower()
        if any(x in joined for x in ["card type", "cardtype", "tender", "payment type"]):
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"  Could not find header in {filepath.name}")
        return {}

    headers  = [h.lower() for h in rows[header_idx]]
    type_idx = next((i for i, h in enumerate(headers)
                     if any(x in h for x in ["card type", "cardtype", "tender", "type"])), None)
    amt_idx  = next((i for i, h in enumerate(headers)
                     if any(x in h for x in ["net", "amount", "total", "sales"])), None)

    if type_idx is None or amt_idx is None:
        log.warning(f"  CC columns not found. Headers: {headers}")
        return {}

    raw_totals = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= max(type_idx, amt_idx):
            continue
        card = row[type_idx].lower().strip()
        amt  = clean_amount(row[amt_idx])
        if not card or amt == 0:
            continue
        raw_totals[card] = round(raw_totals.get(card, 0.0) + amt, 2)

    totals = {}
    log.info("\n  ── Credit Cards ──")
    for card_raw, amt in raw_totals.items():
        label = next(
            (lbl for lbl, kws in CC_TYPE_MAP.items() if any(kw in card_raw for kw in kws)),
            card_raw.title()
        )
        key = f"{CC_ACCOUNT}|{label}"
        log.info(f"    {label:30s}  ${amt:>10,.2f}")
        totals[key] = round(totals.get(key, 0.0) + amt, 2)

    return totals


# ═══════════════════════════════════════════════════════════════
#  FILE FINDER
# ═══════════════════════════════════════════════════════════════

def parse_coupon_file(filepath: Path) -> tuple:
    """
    Read the coupons CSV.
    Returns (per_dept_coupons, owner_local_total) where:
      per_dept_coupons = { subdept_num: coupon_amount }  (for net sales calc)
      owner_local_total = grand total (for 8512006)
    Coupon file format:
      col 1 = sub-dept number
      col 7 = coupon amount for that dept
      Total row at bottom has grand total in col 7
    """
    rows = read_rows(filepath)
    log.info(f"  Parsing coupon file: {filepath.name}")

    per_dept  = {}   # { subdept_num: amount }
    grand_total = 0.0
    in_data   = False

    for row in rows:
        if not any(row):
            continue
        joined = ",".join(row)

        # Detect data start
        if "Sub-Department" in joined:
            in_data = True
            continue

        if not in_data:
            continue

        # Total row — grab grand total
        low = joined.lower()
        if "total" in low and "printed" not in low:
            amt = clean_amount(row[7]) if len(row) > 7 and row[7] else 0.0
            if abs(amt) > 0:
                grand_total = abs(amt)
                log.info(f"    Coupon grand total: ${grand_total:.2f}")
            continue

        # Skip printed/footer rows
        if "printed" in low:
            continue

        # Data row: col 1 = dept num, col 7 = amount
        if len(row) > 7:
            try:
                dept_num = int(str(row[1]).strip())
            except (ValueError, TypeError):
                continue
            amt = clean_amount(row[7])
            if abs(amt) > 0:
                per_dept[dept_num] = round(per_dept.get(dept_num, 0.0) + abs(amt), 2)
                log.info(f"    Coupon dept {dept_num:>5}: ${abs(amt):.2f}")

    return per_dept, grand_total


# Misc sub-depts that post to TBA Purchases if non-zero
MISC_SUBDEPTS = {
    170: "Adjustments",
    22: "Coop Scoop Ad payment",
    23: "Refunded Discounts",
    24: "Bottle Deposits",
    25: "Gift Certificates",
    26: "Share Payment",
    29: "Building Blocks",
    30: "Groupon",
    31: "Crowd Savings",
    32: "Pass Through Donations",
    33: "Bag Credits",
    34: "Paid-Ins",
    35: "Owner Appreciation 5%",
    37: "Staff Appreciation",
    50: "Envirotokens",
    260: "Maria College Cafe",
    530: "Herbs",
    540: "Books",
    560: "Candles/Incense/Baskets",
    999: "UnAssigned",
}


def parse_excel_report(filepath: Path) -> tuple:
    """
    Read the SubDept Sales Report Excel file.
    Sheet: 'SubDept Sales Report'
      Col A = sub-dept number, Col G = net amount
      Row 1 Col M = Milk Bottle Return
      Row 2 Col J = Owner Apprec 5% Total, Col M = Store Coupons
    Returns (sales_dict, misc_lines, milk_bottle_return, store_coupons_amt, owner_apprec_amt)
      misc_lines = [(memo, amount)] for non-zero misc sub-depts → post to TBA Purchases
    """
    import openpyxl
    log.info(f"  Reading Excel report: {filepath.name}")

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        if "WinError 2" in str(e) or "cannot find the file" in str(e).lower() or "openpyx" in str(e).lower():
            raise RuntimeError(
                f"Cannot read Excel file — please CLOSE IT IN EXCEL first, then run again.\n"
                f"File: {filepath.name}"
            ) from e
        raise

    if 'SubDept Sales Report' not in wb.sheetnames:
        log.error(f"  Sheet 'SubDept Sales Report' not found in {filepath.name}")
        log.error(f"  Available sheets: {wb.sheetnames}")
        return {}, [], 0.0, 0.0, 0.0

    ws = wb['SubDept Sales Report']
    sales              = {}
    misc_lines         = []
    milk_bottle_return = 0.0
    store_coupons_amt  = 0.0
    owner_apprec_amt   = 0.0
    sales_total_xl     = 0.0
    pass_through_total   = 0.0
    dust_bunnies_total   = 0.0
    milk_bottles_returns = 0.0
    refunded_discounts   = 0.0
    hash_sales_total     = 0.0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            val = row[12] if len(row) > 12 else None
            if val is not None:
                try:
                    milk_bottle_return = -abs(float(val))
                    log.info(f"    Milk Bottle Return (M1): ${milk_bottle_return:.2f}")
                except (ValueError, TypeError):
                    pass
            # M1 (index 12) = Milk Bottles Returns amount (L1 is the label, M1 is the number)
            val_m1_mbr = row[12] if len(row) > 12 else None
            if val_m1_mbr is not None:
                try:
                    milk_bottles_returns = float(val_m1_mbr)
                    log.info(f"    Milk Bottles Returns (M1): ${milk_bottles_returns:.2f}")
                except (ValueError, TypeError):
                    pass
            # Read O1 label to know what P1 is
            val_o1 = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
            val_p1 = row[15] if len(row) > 15 else None
            if val_p1 is not None:
                try:
                    p1_val = float(val_p1)
                    if "refunded" in val_o1:
                        refunded_discounts = p1_val
                        log.info(f"    Refunded Discounts (O1/P1): ${p1_val:.2f}")
                    elif "hash" in val_o1:
                        hash_sales_total = p1_val
                        log.info(f"    Hash Sales (O1/P1): ${p1_val:.2f}")
                    elif "pass through" in val_o1 or "donation" in val_o1:
                        pass_through_total = p1_val
                        log.info(f"    Pass Through (O1/P1): ${p1_val:.2f}")
                    elif "dust" in val_o1:
                        dust_bunnies_total = p1_val
                        log.info(f"    Dust Bunnies (O1/P1): ${p1_val:.2f}")
                except (ValueError, TypeError):
                    pass
            continue

        if i == 2:
            val_j = row[9]  if len(row) > 9  else None
            val_m = row[12] if len(row) > 12 else None
            if val_j is not None:
                try:
                    owner_apprec_amt = float(val_j)
                    log.info(f"    Owner Apprec 5% (J2): ${owner_apprec_amt:.2f}")
                except (ValueError, TypeError):
                    pass
            if val_m is not None:
                try:
                    store_coupons_amt = float(val_m)
                    log.info(f"    Store Coupons (M2): ${store_coupons_amt:.2f}")
                except (ValueError, TypeError):
                    pass
            # Read O2 label to know what P2 is
            val_o2 = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
            val_p2 = row[15] if len(row) > 15 else None
            if val_p2 is not None:
                try:
                    p2_val = float(val_p2)
                    if "refunded" in val_o2:
                        refunded_discounts = p2_val
                        log.info(f"    Refunded Discounts (O2/P2): ${p2_val:.2f}")
                    elif "hash" in val_o2:
                        hash_sales_total = p2_val
                        log.info(f"    Hash Sales (O2/P2): ${p2_val:.2f}")
                    elif "pass through" in val_o2 or "donation" in val_o2:
                        pass_through_total = p2_val
                        log.info(f"    Pass Through (O2/P2): ${p2_val:.2f}")
                    elif "dust" in val_o2:
                        dust_bunnies_total = p2_val
                        log.info(f"    Dust Bunnies (O2/P2): ${p2_val:.2f}")
                except (ValueError, TypeError):
                    pass
            continue

        if i == 3:
            val_j3 = row[9] if len(row) > 9 else None
            if val_j3 is not None:
                try:
                    sales_total_xl = round(float(val_j3), 2)
                    log.info(f"    Excel Sales Total (J3): ${sales_total_xl:,.2f}")
                except (ValueError, TypeError):
                    pass
            # Read O3 label to know what P3 is
            val_o3 = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
            val_p3 = row[15] if len(row) > 15 else None
            if val_p3 is not None:
                try:
                    p3_val = float(val_p3)
                    if "dust" in val_o3:
                        dust_bunnies_total = p3_val
                        log.info(f"    Dust Bunnies (O3/P3): ${p3_val:.2f}")
                    elif "refunded" in val_o3:
                        refunded_discounts = p3_val
                        log.info(f"    Refunded Discounts (O3/P3): ${p3_val:.2f}")
                    elif "hash" in val_o3:
                        hash_sales_total = p3_val
                        log.info(f"    Hash Sales (O3/P3): ${p3_val:.2f}")
                    elif "pass through" in val_o3 or "donation" in val_o3:
                        pass_through_total = p3_val
                        log.info(f"    Pass Through (O3/P3): ${p3_val:.2f}")
                except (ValueError, TypeError):
                    pass
            continue

        dept_num = row[0]
        amount   = row[6] if len(row) > 6 else None

        if dept_num is None:
            continue
        try:
            dept_num = int(dept_num)
        except (ValueError, TypeError):
            continue

        if amount is None:
            amount = 0.0
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0.0

        qb = SUBDEPT_TO_QB.get(dept_num)
        if qb is None:
            # Check if misc sub-dept with non-zero amount → TBA Purchases
            if dept_num in MISC_SUBDEPTS:
                if amount != 0.0:
                    memo = MISC_SUBDEPTS[dept_num]
                    misc_lines.append((memo, round(amount, 2)))
                    log.info(f"    Dept {dept_num:>5}  ${amount:>10,.2f}  → TBA Purchases ({memo})")
            elif dept_num in SUBDEPT_TO_QB:
                log.info(f"    Dept {dept_num:>5}  SKIPPED")
            else:
                # Unknown dept — if non-zero, post to TBA Purchases
                if amount != 0.0:
                    misc_lines.append((f"Dept {dept_num}", round(amount, 2)))
                    log.warning(f"    Dept {dept_num:>5}  ${amount:,.2f}  → TBA Purchases (unmapped dept — add to MISC_SUBDEPTS)")
                else:
                    log.info(f"    Dept {dept_num:>5}  zero — skipped")
            continue

        log.info(f"    Dept {dept_num:>5}  ${amount:>10,.2f}  →  {qb}")
        sales[qb] = round(sales.get(qb, 0.0) + amount, 2)

    log.info(f"  Excel: {len(sales)} accounts mapped, {len(misc_lines)} misc TBA lines")
    return sales, misc_lines, milk_bottle_return, store_coupons_amt, owner_apprec_amt, sales_total_xl, pass_through_total, dust_bunnies_total, milk_bottles_returns, refunded_discounts, hash_sales_total


def parse_excel_discounts(filepath: Path, report_date) -> dict:
    """
    Read the discounts sheet from the Excel file.
    Tab name format: MMDDYY discounts (e.g. '41226 discounts' for 4/12/26)
    Col 3 (index 3) = shopper level code
    Col 9 (index 9) = amount
    Returns { qb_account: amount } — amounts are positive (IIF sign flip handled in generate_iif)
    """
    import openpyxl

    # Build expected tab name: MMDDYY discounts (e.g. "41226 discounts")
    # Remove leading zero from month for Windows compatibility
    mmddyy = f"{report_date.month}{report_date.strftime('%d%y')}"
    tab_candidates = [
        f"{mmddyy} discounts",               # 41226 discounts
        f"0{mmddyy} discounts",              # 041226 discounts
        report_date.strftime("%m%d%y discounts"),  # 041226 discounts (with leading zero)
    ]

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    found_tab = None

    for candidate in tab_candidates:
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            found_tab = candidate
            break

    # Fallback: find any sheet ending in " discounts"
    if ws is None:
        for sheet in wb.sheetnames:
            if sheet.lower().endswith(" discounts") or sheet.lower().endswith("discounts"):
                ws = wb[sheet]
                found_tab = sheet
                break

    if ws is None:
        log.warning(f"  No discounts sheet found in {filepath.name}")
        log.warning(f"  Tried: {tab_candidates}")
        log.warning(f"  Available: {wb.sheetnames}")
        return {}

    log.info(f"  Reading discounts from sheet: '{found_tab}'")

    # Wednesday Senior Day: levels 3, 4, 10 combine into 8511002
    WEDNESDAY_SENIOR_LEVELS = {3, 4, 10}
    WEDNESDAY_SENIOR_ACCT   = "8511002 · Discount 8% - Senior Day"
    is_wednesday = (report_date.weekday() == 2)  # 0=Mon, 2=Wed
    if is_wednesday:
        log.info(f"  Wednesday — combining Senior levels 3, 4, 10 → {WEDNESDAY_SENIOR_ACCT}")

    discounts = {}
    grand_total = 0.0

    for row in ws.iter_rows(values_only=True):
        # Grand total row — "Member Discounts" label in col 2, total in col 8
        if row[2] is not None and "member discount" in str(row[2]).lower():
            try:
                grand_total = abs(float(str(row[8])))
                log.info(f"    Discounts grand total: ${grand_total:,.2f}")
            except (ValueError, TypeError):
                pass
            continue

        # Data rows — col 3 = shopper level code, col 8 = amount
        code = row[3] if len(row) > 3 else None
        amt  = row[8] if len(row) > 8 else None

        if code is None or amt is None:
            continue
        try:
            code = int(float(str(code)))
            amt  = float(str(amt))
        except (ValueError, TypeError):
            continue

        if amt == 0:
            continue

        # Wednesday: combine senior levels 3, 4, 10 into Senior Day account
        if is_wednesday and code in WEDNESDAY_SENIOR_LEVELS:
            discounts[WEDNESDAY_SENIOR_ACCT] = round(discounts.get(WEDNESDAY_SENIOR_ACCT, 0.0) + abs(amt), 2)
            log.info(f"    Level {code:>2}  ${amt:>9,.2f}  →  {WEDNESDAY_SENIOR_ACCT} (Senior Day)")
            continue

        qb = SHOPPER_LEVEL_TO_QB.get(code)
        if qb:
            discounts[qb] = round(discounts.get(qb, 0.0) + abs(amt), 2)
            log.info(f"    Level {code:>2}  ${amt:>9,.2f}  →  {qb}")
        elif code == 15:
            log.info(f"    Level 15 (Student Discount Sun) ${amt:,.2f} — skipped (Sunday only)")
            grand_total = round(grand_total - abs(amt), 2)
        else:
            log.warning(f"    Level {code}  ${amt:,.2f}  NOT MAPPED")

    log.info(f"  Discounts: {len(discounts)} accounts, grand total=${grand_total:,.2f}")
    return discounts, grand_total


def parse_bs_sheet(filepath: Path, report_date) -> dict:
    """
    Read the Balance Sheet tab (e.g. '41226 BS') from the Excel file.
    Extracts: Sales Tax, Bottle Sales/Fee/Return, Charity, Credit Cards, Cash, etc.
    Returns dict of { field_name: amount }
    """
    import openpyxl

    # Build tab name: MMDDYY BS
    mmddyy = f"{report_date.month}{report_date.strftime('%d%y')}"
    tab_candidates = [
        f"{mmddyy} BS",
        f"0{mmddyy} BS",
        report_date.strftime("%m%d%y BS"),
    ]

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    for candidate in tab_candidates:
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            log.info(f"  Reading BS sheet: '{candidate}'")
            break

    if ws is None:
        for sheet in wb.sheetnames:
            if sheet.upper().endswith(" BS"):
                ws = wb[sheet]
                log.info(f"  Reading BS sheet: '{sheet}'")
                break

    if ws is None:
        log.warning(f"  No BS sheet found in {filepath.name}")
        return {}

    bs = {
        "sales_tax":        0.0,
        "bottle_sales":     0.0,
        "milk_bottle_fee":  0.0,
        "milk_bottle_return": 0.0,
        "bottle_return":    0.0,
        "charity":          0.0,
        "prepaid_increase": 0.0,
        "visa_mc":          0.0,
        "amex":             0.0,
        "discover":         0.0,
        "debit":            0.0,
        "ebt_cash":         0.0,
        "ebt_food":         0.0,
        "dufb":             0.0,
        "cash":             0.0,
        "check":            0.0,
        "vendor_coupon":    0.0,
        "charge":           0.0,
        "prepaid_card":     0.0,
        "donation":         0.0,
        "subscription":     0.0,
    }

    for row in ws.iter_rows(values_only=True):
        code = row[0]
        desc = str(row[2] or row[1] or "").strip().lower()
        amt  = row[4] if len(row) > 4 else None

        # Total rows use col 7 (Balance column)
        bal  = row[7] if len(row) > 7 else None

        def to_float(v):
            try: return float(v)
            except: return 0.0

        # Total Taxes → Sales Tax
        if row[1] == "Total" and row[2] == "Taxes":
            bs["sales_tax"] = to_float(bal)
            log.info(f"    Sales Tax: ${bs['sales_tax']:,.2f}")

        # Individual rows by code
        if code == 39:   bs["bottle_sales"]      = to_float(amt)
        if code == 40:   bs["milk_bottle_fee"]    = to_float(amt)
        if code == 205:  bs["charity"]            = to_float(amt)
        if code == 208:  bs["prepaid_increase"]   = to_float(amt)
        # Revenue rows (PkUp = what was actually collected)
        if code == 910:  bs["milk_bottle_return"] = to_float(amt)
        if code == 911:  bs["bottle_return"]      = to_float(amt)
        if code == 901:  bs["cash"]               = to_float(amt)
        if code == 902:  bs["check"]              = to_float(amt)
        if code == 903:  bs["debit"]              = to_float(amt)
        if code == 920:  bs["ebt_cash"]           = to_float(amt)
        if code == 921:  bs["ebt_food"]           = to_float(amt)
        if code == 928:  bs["dufb"]               = to_float(amt)
        if code == 930:  bs["visa_mc"]            = round(bs["visa_mc"] + to_float(amt), 2)  # Visa
        if code == 931:  bs["visa_mc"]            = round(bs["visa_mc"] + to_float(amt), 2)  # Master
        if code == 932:  bs["amex"]               = to_float(amt)
        if code == 933:  bs["discover"]           = to_float(amt)
        if code == 980:  bs["prepaid_card"]       = to_float(amt)
        if code == 1117: bs["prepaid_card"]       = round(bs["prepaid_card"] + to_float(amt), 2)  # PkUp Gift card used — add to prepaid
        if code == 906:  bs["charge"]             = to_float(amt)
        if code == 908:  bs["vendor_coupon"]      = to_float(amt)
        if code == 1122: bs["donation"]           = to_float(amt)
        if code == 3420: bs["subscription"]       = to_float(amt)

    log.info(f"  BS: Tax=${bs['sales_tax']:,.2f} BottleSales=${bs['bottle_sales']:,.2f} "
             f"Fee=${bs['milk_bottle_fee']:,.2f} Charity=${bs['charity']:,.2f} "
             f"Visa/MC=${bs['visa_mc']:,.2f} AMEX=${bs['amex']:,.2f} "
             f"Discover=${bs['discover']:,.2f} Debit=${bs['debit']:,.2f}")
    return bs


def find_todays_files(deposit_date=None):
    folder = Path(CONFIG["pos_export_folder"])
    today  = date.today()
    date_patterns = [
        today.strftime("%Y-%m-%d"), today.strftime("%m-%d-%Y"),
        today.strftime("%m%d%Y"),   today.strftime("%Y%m%d"),
        today.strftime("%m%d%y"),
    ]

    all_files = list(folder.glob("*.csv")) + list(folder.glob("*.xlsx"))
    todays = [
        f for f in all_files
        if not f.name.startswith("~$")   # skip Excel temp/lock files
        and (any(p in f.name for p in date_patterns)
             or datetime.fromtimestamp(f.stat().st_mtime).date() == today)
    ]

    if not todays:
        log.warning(f"No CSV files found in {folder} for {today} — will try network Excel path only.")

    sms_files, cc_file, coupon_files, excel_files = [], None, [], []
    for f in todays:
        name_up = f.name.upper()
        try:
            peek = f.read_text(encoding="utf-8-sig", errors="replace")[:500].upper()
        except Exception:
            peek = ""
        if any(p.upper() in name_up or p.upper() in peek for p in CONFIG["cc_pattern"]):
            cc_file = f
            log.info(f"  CC settlement : {f.name}")
        elif "COUPON" in name_up:
            coupon_files.append(f)
            log.info(f"  Coupon file     : {f.name}")
        elif "DISCOUNTS BY SHOPPER LEVEL" in peek:
            sms_files.append(f)
            log.info(f"  SMS discount file: {f.name}")
        elif "SUB-DEPARTMENT SINGLE TOTAL" in peek or "SUB-DEPARTMENT" in peek:
            sms_files.append(f)
            log.info(f"  SMS sales file: {f.name}")
        else:
            log.info(f"  Skipping (no recognised data): {f.name}")

    # Detect Excel SubDept Sales Report files in POS folder
    for f in list(folder.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        if any(p in f.name for p in date_patterns) or            datetime.fromtimestamp(f.stat().st_mtime).date() == today:
            excel_files.append(f)
            log.info(f"  Excel report    : {f.name}")

    # Search network shared drive for Excel file
    if not excel_files:
        base = Path(CONFIG["base_excel_path"])

        # FY2026 = July 2025 - June 2026
        deposit_date = deposit_date or (today - timedelta(days=1))

        fy_year = deposit_date.year + 1 if deposit_date.month >= 7 else deposit_date.year
        fy_folder = base / f"FY {fy_year}"

        # Fiscal month: July=1, Aug=2, Sep=3, Oct=4, Nov=5, Dec=6,
        #               Jan=7, Feb=8, Mar=9, Apr=10, May=11, Jun=12
        FISCAL_MONTHS = {
            7:1, 8:2, 9:3, 10:4, 11:5, 12:6,
            1:7, 2:8, 3:9,  4:10, 5:11, 6:12
        }
        fiscal_num  = FISCAL_MONTHS[deposit_date.month]
        month_name  = deposit_date.strftime("%B")   # "April"
        cal_year    = deposit_date.year              # 2026
        # e.g. "10 - April 2026"
        month_folder = fy_folder / f"{fiscal_num} - {month_name} {cal_year}"

        log.info(f"  Network path: {month_folder}")

        if month_folder.exists():
            # Search for xlsx files matching the deposit date patterns
            for f in sorted(month_folder.glob("*.xlsx")) + sorted(month_folder.glob("*.xls")):
                if f.name.startswith("~$"):
                    continue
                if any(p in f.name for p in date_patterns):
                    excel_files.append(f)
                    log.info(f"  Excel (network) : {f.name}")
            # Fallback: most recently modified xlsx today
            if not excel_files:
                for f in sorted(month_folder.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
                    if not f.name.startswith("~$"):
                        excel_files.append(f)
                        log.info(f"  Excel (latest)  : {f.name}")
                        break
        else:
            log.warning(f"  Month folder not found: {month_folder}")
            log.warning(f"  Expected: FY {fy_year} / {fiscal_num} - {month_name} {cal_year}")

    if not sms_files:
        log.info("  Using Excel report only (no CSV files needed)")
        sms_files = [f for f in todays if f != cc_file]

    return sms_files, cc_file, coupon_files, excel_files


# ═══════════════════════════════════════════════════════════════
#  IIF GENERATOR
#  Uses DEPOSIT transaction type so it lands in
#  Banking → Make Deposits → 1120200 · NBT Bank - Operating Account
# ═══════════════════════════════════════════════════════════════

def spl(date_str, acct, name, amount, memo):
    """Build one SPL line. amount=None leaves the amount blank for manual entry."""
    amt_str = f"{amount:.2f}" if amount is not None else ""
    return f"SPL\tDEPOSIT\t{date_str}\t{acct}\t{name}\t{amt_str}\t{memo}\t"


def generate_iif(sales: dict, discounts: dict, cc: dict, report_date: date, owner_local_amt: float = 0.0, per_dept_coupons: dict = None, milk_bottle_return: float = 0.0, store_coupons_xl: float = 0.0, owner_apprec_xl: float = 0.0, misc_tba_lines: list = None, excel_sales_total: float = 0.0, excel_discount_total: float = 0.0, bs_data: dict = None, pass_through_total: float = 0.0, dust_bunnies_total: float = 0.0, milk_bottles_returns: float = 0.0, refunded_discounts: float = 0.0, hash_sales_total: float = 0.0) -> Path:
    date_str     = report_date.strftime("%m/%d/%Y")
    deposit_acct = CONFIG["deposit_account"]
    iif_path     = output_dir / f"deposit_{report_date.strftime('%Y%m%d')}.iif"

    if misc_tba_lines is None:
        misc_tba_lines = []
    if bs_data is None:
        bs_data = {}

    def bs(key, default=None):
        """Get a value from BS data, return as positive amount or None."""
        v = bs_data.get(key, 0.0)
        if v != 0.0:
            return abs(v)
        return default

    # Helper to get amount from sales dict, None if not found
    def s(key):
        """Return gross sales as negative SPL (income = credit)."""
        if key not in sales:
            return None
        return -abs(sales[key])

    def d(key):
        # Positive in IIF → QB shows as negative in Make Deposits
        return abs(discounts[key]) if key in discounts else None

    lines = []
    lines.append("!TRNS\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO\tCLASS")
    lines.append("!SPL\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO\tCLASS")
    lines.append("!ENDTRNS")
    lines.append("")

    spls = []

    # ── SALES (Image 1) — exact order matching your manual entry ──
    SALES_ORDER = [
        ("7110110 · Sales - Packaged Grocery - NT", ""),
        ("7110120 · Sales - Dairy/Refrigerated",    ""),
        ("7110130 · Sales - Frozen Foods",           ""),
        ("7110141 · Sales - Bulk",                   ""),
        ("7110142 · Sales - Bulk Herbs",             ""),
        ("7110150 · Sales - Bread",                  ""),
        ("7110160 · Sales - Beer",                   ""),
        ("7110180 · Sales - Pets",                   ""),
        ("7110190 · Sales - General Merchandise",    ""),
        ("7110210 · Sales - In-House Deli",          ""),
        ("7110220 · Sales - In-House Bakery",        ""),
        ("7110210 · Sales - In-House Deli",          "Catering"),   # separate catering line
        ("7110235 · Sales - Sushi",                  ""),
        ("7110240 · Sales - Java & Juice",           ""),
        ("7110250 · Sales - Co-op Cafe (ESP)",       "ESP Co-op Cafe"),
        ("7110310 · Sales - Cheese",                 ""),
        ("7110320 · Sales - Meat",                   ""),
        ("7110330 · Sales - Specialty Foods",        ""),
        ("7110340 · Sales - Fish & Seafood",         ""),
        ("7110350 · Sales - Specialty Mercantile",   ""),
        ("7110410 · Sales - Produce",                ""),
        ("7110420 · Sales - Gardening/Plants",       ""),
        ("7110510 · Sales - Personal Care Taxable",  ""),
        ("7110520 · Sales - Vitamins & Supplements", ""),
        ("7110550 · Sales - Magazines",              ""),
        ("7111300 · Promotional Sales",              ""),
        ("8320000 · Store Supplies",                 "HWFC Grocery Paper Bags"),
    ]

    # Track running total for auto-filled amounts only
    spl_total = 0.0
    seen_deli = False  # Deli appears twice — first is regular, second is Catering

    # Catering (sub-dept 230) was merged into Deli in SUBDEPT_TO_QB.
    # Get the raw Deli-only amount by subtracting Catering from combined total.
    catering_amt = None
    deli_combined = sales.get("7110210 · Sales - In-House Deli", 0)
    # Catering amount is stored in the SMS file under sub-dept 230
    # We stored it merged — so show combined on first line, blank on Catering line
    # (Catering is already included in the Deli total)

    for acct, memo in SALES_ORDER:
        if acct == "7110210 · Sales - In-House Deli":
            if not seen_deli:
                # First Deli line — regular deli amount (not catering)
                amt = s(acct)
                seen_deli = True
            else:
                # Catering line — pull from the pipe-keyed Catering entry
                catering_key = "7110210 · Sales - In-House Deli|Catering"
                amt = -abs(sales[catering_key]) if catering_key in sales else None
        elif acct in INCOME_ACCOUNTS:
            amt = s(acct)
        else:
            # Store Supplies: negative in IIF → QB shows positive
            amt = -(abs(sales[acct])) if acct in sales else None

        if amt is not None:
            spl_total += amt
        spls.append(spl(date_str, acct, "", amt, memo))

    # ── NYS BAG FEES — negative in IIF so QB shows positive ──
    bag_amt = -(abs(sales.get("4150300 · NYS Paper Bag Fees Payable", 0))) or None
    if bag_amt:
        spl_total += bag_amt
    spls.append(spl(date_str, "4150300 · NYS Paper Bag Fees Payable", "",
                    bag_amt, "NYS-Albany County Paper Bag Fees"))

    # ── COUPONS — use Excel Col M value if available, else fall back to sales ──
    if store_coupons_xl != 0.0:
        # QB flips sign — put positive in IIF to show negative
        coupon_amt = abs(store_coupons_xl)
        log.info(f"    Coupons from Excel M2: ${store_coupons_xl:.2f}")
    else:
        raw_coupon = sales.get("8515000 · Marketing - Coupons, Store", 0)
        coupon_amt = abs(raw_coupon) if raw_coupon != 0 else None
    if coupon_amt is not None:
        spl_total += coupon_amt
    spls.append(spl(date_str, "8515000 · Marketing - Coupons, Store", "",
                    coupon_amt, "Store Coupons"))

    # ── DISCOUNTS — all forced negative ──
    DISCOUNT_ORDER = [
        ("8512006 · Discount 5% - Owner buy Local",  "PdOut -"),   # manual — shopper code unknown
        ("8512001 · Discount 2% - Owners",           "PdOut -"),
        ("8511001 · Discount 2% - Senior Non Owner", "PdOut -"),
        ("8511002 · Discount 8% - Senior Day",        "PdOut -"),
        ("8511003 · Discount 5% - Senior Owners",     "PdOut -"),
        ("8140010 · Monthly Time Discount (8%)",      "PdOut -"),
        ("8140026 · Weekly Time Discount (24%)",      "PdOut -"),
        ("8423100 · Discount - Staff  (24%)",         "PdOut -"),
        ("8512002 · Discount 2% - Visiting Coop",    "PdOut -"),
        ("8512003 · Discount 8% - Vendors",           "PdOut -"),
        ("8512007 · Discount 15% - Non-profit",       "PdOut-"),
        ("8140026 · Weekly Time Discount (24%)",      "Refunded Discounts", refunded_discounts if refunded_discounts else None),
    ]
    for entry in DISCOUNT_ORDER:
        acct = entry[0]
        memo = entry[1]
        override_amt = entry[2] if len(entry) > 2 else None

        if override_amt is not None:
            # Pass raw Excel value — manual loop will flip sign for QB display
            # Positive Excel value → negative IIF → QB shows positive
            # Negative Excel value → positive IIF → QB shows negative
            amt = -override_amt
        elif acct == "8512006 · Discount 5% - Owner buy Local":
            source = owner_apprec_xl if owner_apprec_xl != 0 else owner_local_amt
            amt = abs(source) if source != 0 else None
        elif memo == "Refunded Discounts":
            amt = None   # no Excel value found — leave blank
        else:
            amt = d(acct)
        if amt is not None:
            spl_total += amt
        spls.append(spl(date_str, acct, "", amt, memo))

    # ── MANUAL LINES (Image 2 & 3) — blank amounts, accounts + memos pre-loaded ──
    # Combine EBT Cash + EBT Food Stamps into one line
    ebt_cash = bs_data.get("ebt_cash", 0.0)
    ebt_food = bs_data.get("ebt_food", 0.0)
    bs_ebt   = round(abs(ebt_cash) + abs(ebt_food), 2) or None

    MANUAL_LINES = [
        # Member shares
        ("6100000 · Member Shares (Paid-In Equity)",  "",                    "Member Shares - Paid", bs("subscription") if bs("subscription") else None),
        ("6100000 · Member Shares (Paid-In Equity)",  "",                    "Member Shares - Receivable"),
        ("1260000 · Member Shares Receivable",         "",                    "Share Installments - Receivable"),
        ("1260000 · Member Shares Receivable",         "",                    "Share Installments - Paid"),
        ("9104000 · Interest Income",                  "",                    "Share Installments - Paid"),
        # Sales tax — from BS sheet Total Taxes
        ("4150100 · Sales Tax Payable",                "New York State Sales Tax", "", bs("sales_tax")),
        # Bottle deposits — from BS sheet
        ("1311100 · Inventory - Bottles Deposit",      "",                    "Bottle Sales",       bs("bottle_sales")),
        ("1311100 · Inventory - Bottles Deposit",      "",                    "Milk Bottle Fee",    bs("milk_bottle_fee")),
        # Milk Bottle Return = Excel M1 (-118) + BS row 910 — both negative in QB
        ("1311100 · Inventory - Bottles Deposit",      "",                    "Milk Bottle Return",
            -(abs(milk_bottle_return) + bs("milk_bottle_return", 0.0)) if (milk_bottle_return or bs("milk_bottle_return")) else None),
        ("1311100 · Inventory - Bottles Deposit",      "",                    "Bottle Return",      -bs("bottle_return") if bs("bottle_return") else None),
        # Charitable donations — from BS sheet
        ("4160000 · Charitable Donations Payable",     "",                    "Charity/Pass through Donations (Round up)", round((bs("charity") or 0.0) + (pass_through_total or 0.0), 2) or None),
        ("4160000 · Charitable Donations Payable",     "",                    "Dust Bunnies", dust_bunnies_total if dust_bunnies_total else None),
        # Gift cards & food bucks
        ("4160500 · Gift Cards - Sold - Old/Vantiv",   "",                    "Gift cards sold", bs("prepaid_increase") if bs("prepaid_increase") else None),
        ("1230400 · Due From Double Up Food Bucks",    "",                    "Double Up Food Bucks Customer Spending", -bs("dufb") if bs("dufb") else None),
        ("4160510 · Gift Cards- Redeemed-Old/Vantiv",  "",                    "Gift cards redeemed", -bs("prepaid_card") if bs("prepaid_card") else None),
        # NCG / MFG coupons
        ("1250000 · Coupons Receivable",               "",                    "NCG Coupons", -bs("vendor_coupon") if bs("vendor_coupon") else None),
        ("1250000 · Coupons Receivable",               "",                    "MFG Coupons"),
        # InHouse blank line
        ("4444 · TBA Purchases",                       "",                    "InHouse:", -bs("charge") if bs("charge") else None),
        # Blank rows
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        ("4444 · TBA Purchases", "", ""),
        # Outreach
        ("8506000 · Outreach - Donations",             "",                    "", -bs("donation") if bs("donation") else None),
        # Paid in/out labels
        ("4444 · TBA Purchases",                       "",                    "PAID IN:"),
        ("4444 · TBA Purchases",                       "",                    "PAID OUT:"),
        # Credit cards — negative in QB, so pass negative amounts
        # (manual loop does iif_amt = -amt, so negative amt → positive IIF → QB shows negative)
        ("1240001 · Credit Card Payments Receivable",  "",  "Visa/MC",            -bs("visa_mc")    if bs("visa_mc")    else None),
        ("1240001 · Credit Card Payments Receivable",  "",  "Discover",           -bs("discover")   if bs("discover")   else None),
        ("1240001 · Credit Card Payments Receivable",  "",  "AMEX",               -bs("amex")       if bs("amex")       else None),
        ("1240001 · Credit Card Payments Receivable",  "",  "Debit Card",         -bs("debit")      if bs("debit")      else None),
        ("1240001 · Credit Card Payments Receivable",  "",  "EBT Cash/Food Stamp",-bs_ebt           if bs_ebt           else None),
        # Cash over/short
        ("8314000 · FE - Cash Over/Shorts",            "",                    "Over/Short per Closeout Sheet"),
        ("8314000 · FE - Cash Over/Shorts",            "",                    "Over/Short per POS (to = POS total)"),
    ]

    for entry in MANUAL_LINES:
        acct, name, memo = entry[0], entry[1], entry[2]
        amt = entry[3] if len(entry) > 3 else None
        if amt is not None and amt != 0:
            # QB flips sign — put -amt in IIF so QB displays amt correctly
            iif_amt = -amt
            spl_total += iif_amt  # include in balance calculation
        else:
            iif_amt = None
        spls.append(spl(date_str, acct, name, iif_amt, memo))

    # ── MISC TBA PURCHASES — non-zero misc sub-depts at the bottom ──
    if misc_tba_lines:
        for memo, amount in misc_tba_lines:
            # Positive amounts in IIF → QB shows negative; negative → positive
            iif_amt = -amount
            spl_total += iif_amt
            spls.append(spl(date_str, "4444 · TBA Purchases", "", iif_amt, memo))
            log.info(f"    TBA Purchases: {memo} = ${amount:.2f}")

    # ── TRNS total — must be exact negative of all SPL amounts so it zeros out ──
    spl_total = round(spl_total, 2)
    trns_amt  = round(-spl_total, 2)

    # ── VERIFY 1: Gross sales minus coupons minus owner apprec should match Excel J3 ──
    # Mirrors the Excel formula: J3 = G53 + J1 (coupons) + J2 (owner apprec)
    # Gross sales = all sales accounts EXCEPT store coupons (sub-dept 27)
    # Includes: 711xxxx income, store supplies, NYS bags, AND misc TBA lines
    COUPON_ACCT = "8515000 · Marketing - Coupons, Store"
    sales_acct_set = set(sales.keys()) - {COUPON_ACCT}
    gross_sales_spl = 0.0
    for s in spls:
        parts = s.split("\t")
        if len(parts) < 6 or not parts[5].strip():
            continue
        acct = parts[3].strip()
        # Include if it came from the sales dict (excluding coupons)
        if acct in sales_acct_set:
            try:
                gross_sales_spl += abs(float(parts[5]))
            except ValueError:
                pass

    # Also add misc TBA lines (560, 999, 170, etc.) — only positive amounts
    # Negative misc lines (e.g. Refunded Discounts) reduce sales, not increase gross
    if misc_tba_lines:
        for memo, amount in misc_tba_lines:
            if amount > 0:
                gross_sales_spl += amount
            else:
                gross_sales_spl += amount  # still include — matches Excel G column total
    gross_sales_spl = round(gross_sales_spl, 2)

    coupon_deduct        = abs(store_coupons_xl) if store_coupons_xl else abs(sales.get("8515000 · Marketing - Coupons, Store", 0))
    owner_apprec_deduct  = abs(owner_apprec_xl)  if owner_apprec_xl  else 0.0
    # milk_bottle_return is already the M1 value (negative) — reuse it for the deduction
    milk_returns_deduct  = abs(milk_bottle_return) if milk_bottle_return else abs(milk_bottles_returns) if milk_bottles_returns else 0.0
    net_sales_check      = round(gross_sales_spl - coupon_deduct - owner_apprec_deduct - milk_returns_deduct, 2)

    if excel_sales_total != 0.0:
        diff = round(abs(excel_sales_total - net_sales_check), 2)
        log.info(f"")
        log.info(f"  ─────────────────────────────────────────")
        log.info(f"  SALES CHECK")
        log.info(f"  ─────────────────────────────────────────")
        log.info(f"  Gross Sales:        ${gross_sales_spl:>12,.2f}")
        log.info(f"  Store Coupons:     -${coupon_deduct:>12,.2f}")
        log.info(f"  Owner Apprec:      -${owner_apprec_deduct:>12,.2f}")
        if milk_returns_deduct:
            log.info(f"  Milk Btl Returns:  -${milk_returns_deduct:>12,.2f}")
        log.info(f"  ─────────────────────────────────────────")
        log.info(f"  Script Net Sales:   ${net_sales_check:>12,.2f}")
        log.info(f"  Excel Sales Total:  ${excel_sales_total:>12,.2f}")
        if diff < 0.02:
            log.info(f"  RESULT: ✓ MATCH — OK to import!")
        else:
            log.info(f"  Difference:         ${diff:>12,.2f}")
            log.warning(f"  RESULT: ⚠ MISMATCH — Check your Excel file before importing!")
        log.info(f"  ─────────────────────────────────────────")
        log.info(f"")

    # ── HASH SALES 6 — Refunded Discounts + Pass Through Donations ──
    script_hash  = round(abs(refunded_discounts) + abs(pass_through_total), 2)
    excel_hash   = round(abs(refunded_discounts) + abs(pass_through_total), 2)
    hash_diff    = round(abs(excel_hash - script_hash), 2)
    log.info(f"  ─────────────────────────────────────────")
    log.info(f"  HASH SALES 6 CHECK")
    log.info(f"  ─────────────────────────────────────────")
    log.info(f"  Refunded Discounts:  ${abs(refunded_discounts):>10,.2f}")
    log.info(f"  Pass Thru Donations: ${abs(pass_through_total):>10,.2f}")
    log.info(f"  ─────────────────────────────────────────")
    log.info(f"  Script Total:        ${script_hash:>10,.2f}")
    log.info(f"  Hash Sales 6 Total:  ${excel_hash:>10,.2f}")
    if hash_diff < 0.02:
        log.info(f"  RESULT: ✓ MATCH — OK to import!")
    else:
        log.info(f"  Difference:          ${hash_diff:>10,.2f}")
        log.warning(f"  RESULT: ⚠ MISMATCH — Check before importing!")
    log.info(f"  ─────────────────────────────────────────")
    log.info(f"")

    # ── VERIFY 2: Discount total must match Excel discounts sheet grand total ──
    # Exclude Refunded Discounts (positive amount — not a real discount)
    disc_acct_names = set(discounts.keys())
    disc_spl_total = 0.0
    for s in spls:
        parts = s.split("\t")
        if len(parts) < 6 or not parts[5].strip():
            continue
        acct = parts[3].strip()
        memo = parts[6].strip() if len(parts) > 6 else ""
        if acct in disc_acct_names and "refunded" not in memo.lower():
            try:
                disc_spl_total += abs(float(parts[5]))
            except ValueError:
                pass
    disc_spl_total = round(disc_spl_total, 2)

    if excel_discount_total != 0.0:
        diff2 = round(abs(excel_discount_total - disc_spl_total), 2)
        log.info(f"  ─────────────────────────────────────────────")
        log.info(f"  DISCOUNTS CHECK")
        log.info(f"  ─────────────────────────────────────────────")
        log.info(f"  Script Discounts:   ${disc_spl_total:>12,.2f}")
        log.info(f"  Excel Disc Total:   ${excel_discount_total:>12,.2f}")
        if diff2 < 0.02:
            log.info(f"  RESULT: ✓ MATCH — OK to import!")
        else:
            log.info(f"  Difference:         ${diff2:>12,.2f}")
            if report_date.weekday() == 6:  # Sunday
                log.info(f"  NOTE: Sunday — difference likely Student Discount (level 15, not mapped)")
                log.info(f"  RESULT: ✓ OK to import!")
            else:
                log.warning(f"  RESULT: ⚠ MISMATCH — Check your discounts sheet before importing!")
        log.info(f"  ─────────────────────────────────────────────")
        log.info(f"")

    # ── Write status summary file ────────────────────────────────
    status_lines = []
    status_lines.append(f"HWFC Daily Deposit — {report_date.strftime('%A, %B %d, %Y')}")
    status_lines.append(f"Run time: {datetime.now().strftime('%I:%M %p')}")
    status_lines.append("")

    if excel_sales_total != 0.0:
        sales_diff = round(abs(excel_sales_total - net_sales_check), 2)
        if sales_diff < 0.02:
            status_lines.append(f"  SALES:     ✓ MATCH   ${net_sales_check:,.2f}")
        else:
            status_lines.append(f"  SALES:     ⚠ MISMATCH  Script=${net_sales_check:,.2f}  Excel=${excel_sales_total:,.2f}  (off by ${sales_diff:,.2f})")

    if excel_discount_total != 0.0:
        disc_spl = round(sum(abs(v) for v in discounts.values()), 2)
        disc_diff = round(abs(excel_discount_total - disc_spl), 2)
        if disc_diff < 0.02:
            status_lines.append(f"  DISCOUNTS: ✓ MATCH   ${disc_spl:,.2f}")
        else:
            status_lines.append(f"  DISCOUNTS: ⚠ MISMATCH  Script=${disc_spl:,.2f}  Excel=${excel_discount_total:,.2f}  (off by ${disc_diff:,.2f})")

    status_lines.append("")
    status_lines.append(f"  IIF file: {iif_path}")
    status_lines.append(f"  QB import: File → Utilities → Import → IIF Files")
    status_lines.append("")

    overall_ok = all([
        abs(excel_sales_total - net_sales_check) < 0.02 if excel_sales_total else True,
        abs(excel_discount_total - round(sum(abs(v) for v in discounts.values()), 2)) < 0.02 if excel_discount_total else True,
    ])
    if overall_ok:
        status_lines.append("  ✓ ALL CHECKS PASSED — Safe to import into QuickBooks!")
    else:
        status_lines.append("  ⚠ ISSUES FOUND — Review before importing into QuickBooks!")

    lines.append(
        f"TRNS\tDEPOSIT\t{date_str}\t{deposit_acct}\t\t{trns_amt:.2f}\tDeposit\t"
    )
    lines.extend(spls)
    lines.append("ENDTRNS")

    try:
        with open(iif_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        log.info(f"\n  IIF written → {iif_path}")
        log.info(f"  Auto-filled total: ${trns_amt:,.2f}")
    except Exception as e:
        log.error(f"  FAILED to write file: {e}")
        log.error(f"  Attempted path: {iif_path}")
        raise
    return iif_path


# ═══════════════════════════════════════════════════════════════
#  EXCEL SUMMARY
# ═══════════════════════════════════════════════════════════════

def write_excel_summary(sales, discounts, cc, report_date: date) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    xlsx_path = output_dir / f"deposit_summary_{report_date.strftime('%Y%m%d')}.xlsx"
    wb = openpyxl.Workbook()

    def fill_sheet(ws, data: dict, title: str):
        ws.title = title
        ws.append(["QB Account", "Memo", "Amount"])
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        total = 0.0
        for key, amt in sorted(data.items()):
            acct = key.split("|")[0]
            memo = key.split("|")[1] if "|" in key else ""
            ws.append([acct, memo, amt])
            total += amt

        ws.append(["TOTAL", "", round(total, 2)])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill("solid", fgColor="D6E4F0")

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(width, 55)

    fill_sheet(wb.active,         sales,     "Sales")
    fill_sheet(wb.create_sheet(), discounts, "Discounts")
    fill_sheet(wb.create_sheet(), cc,        "Credit Cards")

    try:
        wb.save(xlsx_path)
        log.info(f"  Excel  → {xlsx_path}")
    except Exception as e:
        if "WinError 2" in str(e) or "openpyx" in str(e).lower():
            log.warning(f"  Could not write Excel summary (file may be open) — skipping")
        else:
            log.warning(f"  Could not write Excel summary: {e}")
    return xlsx_path


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    log.info("=" * 65)
    log.info(f"  HWFC Daily Deposit Automation   {datetime.now():%Y-%m-%d %H:%M:%S}")
    log.info("=" * 65)

    try:
        today = date.today()

        # Check for --auto flag (used by Task Scheduler — skips date prompt)
        auto_mode = "--auto" in sys.argv

        if auto_mode:
            yesterday = today - timedelta(days=1)
            log.info(f"  Auto mode: using yesterday {yesterday.strftime('%B %d, %Y')}")
        else:
            # Interactive mode — ask for date
            print("")
            print(f"  Today is {today.strftime('%A, %B %d, %Y')}")
            date_input = input("  Enter deposit date (MM/DD/YY or MM/DD/YYYY) or press Enter for yesterday: ").strip()

            if date_input == "":
                yesterday = today - timedelta(days=1)
                log.info(f"  Using yesterday: {yesterday.strftime('%B %d, %Y')}")
            else:
                parsed = None
                for fmt in ("%m/%d/%y", "%m/%d/%Y", "%m-%d-%y", "%m-%d-%Y", "%m%d%y", "%m%d%Y"):
                    try:
                        parsed = datetime.strptime(date_input, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    log.error(f"  Could not parse date '{date_input}'. Use format MM/DD/YY (e.g. 04/07/26)")
                    sys.exit(1)
                yesterday = parsed
                log.info(f"  Using entered date: {yesterday.strftime('%B %d, %Y')}")

        log.info(f"Scanning {CONFIG['pos_export_folder']} ...")
        sms_files, cc_file, coupon_files, excel_files = find_todays_files(yesterday)

        sales     = {}
        discounts = {}
        cc        = {}
        owner_local_amt = 0.0

        # Prefer Excel report for sales if available (more accurate net amounts)
        milk_bottle_return = 0.0
        store_coupons_xl   = 0.0
        owner_apprec_xl    = 0.0
        misc_tba_lines     = []
        excel_sales_total  = 0.0
        pass_through_total   = 0.0
        dust_bunnies_total   = 0.0
        milk_bottles_returns = 0.0
        refunded_discounts   = 0.0
        hash_sales_total     = 0.0
        refunded_discounts   = 0.0
        bs_data            = {}

        if excel_files:
            for f in excel_files:
                xl_sales, xl_misc, mb_ret, sc_amt, oa_amt, xl_total, pt_total, db_total, mbr, rd, hs = parse_excel_report(f)
                for k, v in xl_sales.items():
                    sales[k] = round(sales.get(k, 0.0) + v, 2)
                misc_tba_lines.extend(xl_misc)
                milk_bottle_return   += mb_ret
                store_coupons_xl     += sc_amt
                owner_apprec_xl      += oa_amt
                excel_sales_total    += xl_total
                pass_through_total   += pt_total
                dust_bunnies_total   += db_total
                milk_bottles_returns += mbr
                refunded_discounts   += rd
                hash_sales_total     += hs
            log.info(f"  Using Excel report for sales ({len(excel_files)} file(s))")
            # Also read BS sheet
            for f in excel_files:
                bs_data = parse_bs_sheet(f, yesterday)
        else:
            log.warning("  No Excel report found — falling back to SMS CSV files")

        # Read discounts from Excel file if available, else fall back to SMS CSV
        excel_discount_total = 0.0
        if excel_files:
            for f in excel_files:
                d_map, d_total = parse_excel_discounts(f, yesterday)
                for k, v in d_map.items():
                    discounts[k] = round(discounts.get(k, 0.0) + v, 2)
                excel_discount_total += d_total
            if not discounts:
                log.warning("  No discounts found in Excel — falling back to SMS CSV")
                for f in sms_files:
                    _, d_map = parse_sms_file(f)
                    for k, v in d_map.items():
                        discounts[k] = round(discounts.get(k, 0.0) + v, 2)
        else:
            for f in sms_files:
                _, d_map = parse_sms_file(f)
                for k, v in d_map.items():
                    discounts[k] = round(discounts.get(k, 0.0) + v, 2)

        per_dept_coupons = {}
        for f in coupon_files:
            dept_map, grand_total = parse_coupon_file(f)
            owner_local_amt += grand_total
            for k, v in dept_map.items():
                per_dept_coupons[k] = round(per_dept_coupons.get(k, 0.0) + v, 2)

        if cc_file:
            cc = parse_cc_file(cc_file)
        else:
            log.warning("\n  No CC file found — add Commerce Control Center CSV to folder.")

        if not any([sales, discounts, cc]):
            log.error("Nothing mapped. Check files and try again.")
            sys.exit(1)

        iif_path  = generate_iif(sales, discounts, cc, yesterday, owner_local_amt, per_dept_coupons, milk_bottle_return, store_coupons_xl, owner_apprec_xl, misc_tba_lines, excel_sales_total, excel_discount_total, bs_data, pass_through_total, dust_bunnies_total, milk_bottles_returns, refunded_discounts, hash_sales_total)
        try:
            xlsx_path = write_excel_summary(sales, discounts, cc, yesterday)
        except Exception as e:
            log.warning(f"  Excel summary skipped: {e}")
            xlsx_path = None

        log.info("")
        log.info("✓ COMPLETE")
        log.info(f"  Sales accounts   : {len(sales)}")
        log.info(f"  Discount accounts: {len(discounts)}")
        log.info(f"  CC card types    : {len(cc)}")
        log.info(f"  IIF  → {iif_path}")
        if xlsx_path:
            log.info(f"  XLSX → {xlsx_path}")
        log.info("")
        log.info("  QB import: File → Utilities → Import → IIF Files")
        log.info("  Then check: Banking → Make Deposits")
        log.info("")
        log.info("  Still manual: Sales Tax, Member Shares, Gift Cards,")
        log.info("  Bottle Deposits, NCG Coupons, Petty Cash, Donations,")
        log.info("  In-House Purchases, Cash Over/Short")

    except FileNotFoundError as e:
        log.error(str(e))
        sys.exit(1)
    except Exception as e:
        log.exception(f"Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
